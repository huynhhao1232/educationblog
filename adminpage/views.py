import logging
import os
import re
from pathlib import Path
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from homepage.models import *
import json
from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from django.conf import settings
from django.utils import timezone
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from decimal import Decimal, InvalidOperation
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from copy import copy as shallow_copy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import zipfile
import calendar
from collections import defaultdict
from django.db import models as django_models
from adminpage.models import *
from adminpage.services import get_daily_periods_for_schedule, get_hoso_extra_for_schedule
from django.conf import settings
from django.contrib import messages
from django.db import transaction
import unicodedata
import random
# Create your views here.

def natural_sort_key(text):
    """Hàm helper để sắp xếp tự nhiên (natural sort) cho tên lớp"""
    if not text:
        return (0, '')
    # Tách chuỗi thành các phần số và chữ
    def convert(text_part):
        return int(text_part) if text_part.isdigit() else text_part.lower()

    return [convert(c) for c in re.split(r'(\d+)', str(text))]

def _extract_grade_from_class_name(class_name):
    """Lấy 2 ký tự đầu của tên lớp làm khối (vd: '12A1' -> 12)."""
    if not class_name:
        return None
    prefix = str(class_name)[:2]
    digits = ''.join(ch for ch in prefix if ch.isdigit())
    try:
        return int(digits) if digits else None
    except ValueError:
        return None


def _vn_normalize_no_diacritics(text):
    """Chuẩn hoá chuỗi: bỏ dấu tiếng Việt, lower-case, bỏ khoảng trắng dư."""
    if not text:
        return ""
    text = str(text).strip()
    # 'Đ/đ' là ký tự riêng trong Unicode (không phải D + dấu),
    # nên cần quy về D/d trước khi strip combining marks để sắp xếp đúng theo tiếng Việt.
    text = text.replace("Đ", "D").replace("đ", "d")
    nfkd = unicodedata.normalize("NFKD", text)
    without_diacritics = "".join(c for c in nfkd if not unicodedata.combining(c))
    return without_diacritics.lower().strip()


# Thứ tự chữ cái tiếng Việt (có phân biệt ă/â, ô/ơ, ê, ư, đ).
_VI_ALPHA_ORDER = [
    "a", "ă", "â",
    "b",
    "c",
    "d", "đ",
    "e", "ê",
    "g",
    "h",
    "i",
    "k",
    "l",
    "m",
    "n",
    "o", "ô", "ơ",
    "p",
    "q",
    "r",
    "s",
    "t",
    "u", "ư",
    "v",
    "x",
    "y",
]
_VI_ALPHA_RANK = {ch: i for i, ch in enumerate(_VI_ALPHA_ORDER, start=1)}

# Dấu thanh (chỉ dùng khi cùng chữ cái): ngang < sắc < huyền < hỏi < ngã < nặng
_VI_TONE_RANK = {
    "": 0,
    "\u0301": 1,  # acute  sắc
    "\u0300": 2,  # grave  huyền
    "\u0309": 3,  # hook above hỏi
    "\u0303": 4,  # tilde ngã
    "\u0323": 5,  # dot below nặng
}

# Dấu chữ (để phân biệt a/ă/â, o/ô/ơ, e/ê, u/ư)
_VI_LETTER_MARKS = {
    # a
    ("a", "\u0306"): "ă",  # breve
    ("a", "\u0302"): "â",  # circumflex
    # e
    ("e", "\u0302"): "ê",
    # o
    ("o", "\u0302"): "ô",
    ("o", "\u031b"): "ơ",  # horn
    # u
    ("u", "\u031b"): "ư",  # horn
}


def _vn_collation_key(text: str):
    """
    Key so sánh theo bảng chữ cái tiếng Việt:
    - Phân biệt a/ă/â, d/đ, e/ê, o/ô/ơ, u/ư
    - Nếu cùng chữ cái thì xét dấu thanh (ngang < sắc < huyền < hỏi < ngã < nặng)

    Trả về tuple[int,...] để dùng trong sorted(key=...).
    """
    if not text:
        return ()
    out: list[int] = []
    for raw_ch in str(text).strip().lower():
        if raw_ch == "đ":
            out.append(_VI_ALPHA_RANK["đ"] * 10)
            continue

        nfd = unicodedata.normalize("NFD", raw_ch)
        base = nfd[0]
        marks = [c for c in nfd[1:] if unicodedata.combining(c)]

        # tone: lấy 1 dấu thanh nếu có
        tone = ""
        for m in marks:
            if m in _VI_TONE_RANK:
                tone = m
                break

        # letter variant: lấy dấu chữ (breve/circumflex/horn) nếu có
        letter = base
        for m in marks:
            mapped = _VI_LETTER_MARKS.get((base, m))
            if mapped:
                letter = mapped
                break

        r = _VI_ALPHA_RANK.get(letter)
        if r is not None:
            out.append(r * 10 + _VI_TONE_RANK.get(tone, 0))
        else:
            # ký tự không thuộc alphabet VI: để sau, nhưng vẫn deterministic
            out.append(10000 + ord(raw_ch))
    return tuple(out)


def _exam_number_sort_key(exam_number: str | None, fallback: str = ""):
    """Sort key cho SBD: ưu tiên số (int), fallback sang chuỗi."""
    val = (exam_number or "").strip()
    try:
        return (0, int(val))
    except Exception:
        return (1, val or fallback)


def _exam_student_room_order_tuple(stu):
    """Thứ tự trong phòng/ca: thường trước (SBD → tên), hòa nhập cuối."""
    return (
        1 if getattr(stu, "is_integration", False) else 0,
        _exam_number_sort_key(getattr(stu, "exam_number", None), getattr(stu, "student_code", "")),
        _name_sort_key_vi(getattr(stu, "full_name", "") or ""),
    )


def _exam_room_seat_order_key(a):
    """Bọc StudentExamAssignment → cùng quy tắc `_exam_student_room_order_tuple`."""
    return _exam_student_room_order_tuple(a.student)


def _exam_assignment_export_list_order_key(a):
    """Thứ tự dòng trên danh sách xuất Excel: SBD tăng dần (số), rồi tên."""
    stu = a.student
    return (
        _exam_number_sort_key(getattr(stu, 'exam_number', None), getattr(stu, 'student_code', '') or ''),
        _name_sort_key_vi(getattr(stu, 'full_name', '') or ''),
        getattr(stu, 'pk', 0),
    )


def _name_sort_key_vi(full_name: str):
    """
    Sắp xếp theo yêu cầu:
    - Tên (cuối) trước
    - Nếu trùng tên thì tới tên đệm
    - Nếu trùng cả tên đệm thì tới họ
    (đều bỏ dấu, không phân biệt hoa thường)
    """
    if not full_name:
        return ("", "", "")
    parts = str(full_name).strip().split()
    if not parts:
        return ("", "", "")
    first_name = parts[-1]
    last_name = parts[0]
    middle_name = " ".join(parts[1:-1]) if len(parts) > 2 else (" ".join(parts[1:-1]) if len(parts) > 1 else "")
    return (
        _vn_collation_key(first_name),
        _vn_collation_key(middle_name),
        _vn_collation_key(last_name),
    )


def _assign_exam_numbers_for_all_students(
    start_serial: int = 1,
    campus_start_serials: dict[int, int] | None = None,
):
    """
    Đánh số báo danh cho ExamRoomStudent (import danh sách HV / nút đánh lại SBD):
    - Mã SBD: 6 số (1 chữ số cơ sở + 5 chữ số STT).
    - Có thể truyền ``start_serial`` để chọn số thứ tự bắt đầu mặc định trong mỗi cơ sở.
    - Có thể truyền ``campus_start_serials`` để override số bắt đầu theo từng cơ sở.
    - Thứ tự trong từng cơ sở: **khối** (tăng dần: 10 → 11 → …) → trong mỗi khối **buổi**
      (ký tự thứ 3 mã HV: 0 = sáng trước, 1 = tối sau) → tổ hợp (ký tự 4) → họ tên.
    - **Không** phân biệt thường / hòa nhập. Sau khi **xếp phòng** (import / lấy HV / gán dashboard),
      nếu cơ sở có HV hòa nhập thì đánh lại SBD theo ca → tổ hợp môn → phòng
      (xem `_sync_exam_sbd_campus_after_room_assignment`).
    - **Xoá hoặc thêm** HV trong chi tiết phòng thi gọi sync với ``force=True`` — luôn đánh lại SBD cả cơ sở
      theo cùng quy tắc, kể cả khi không có HV hòa nhập.
    """
    from homepage.models import Campus, ExamRoomStudent

    try:
        start_serial = int(start_serial)
    except (TypeError, ValueError):
        start_serial = 1
    if start_serial < 1:
        start_serial = 1
    campus_start_serials = campus_start_serials or {}

    # Dùng chung map thứ hạng cơ sở toàn module để nhất quán với dashboard.
    # Lưu ý: chỉ cập nhật trường exam_number, không xoá/sửa dữ liệu học viên hiện có.
    campus_sbd_digit_map = dict(_EXAM_SORT_CAMPUS_RANK)
    code_upper_to_digit = {k.upper(): v for k, v in campus_sbd_digit_map.items()}

    # Thứ tự cơ sở theo số đầu SBD tăng dần (1..8), cùng prefix thì theo mã cơ sở.
    # Ví dụ DS và ĐS đều map "7": xử lý ổn định để không bỏ sót cơ sở.
    campus_order = [
        code
        for code, _digit in sorted(
            campus_sbd_digit_map.items(),
            key=lambda item: (int(item[1]), item[0]),
        )
    ]
    campuses = list(Campus.objects.all().order_by("code"))
    campuses_by_code_upper = {(c.code or "").upper(): c for c in campuses}
    target_campus_ids = {
        c.id
        for code_key in campus_order
        for c in [campuses_by_code_upper.get(code_key.upper())]
        if c is not None
    }

    # Hai pha để tránh va chạm UNIQUE giữa các cơ sở:
    # pha 1: clear toàn bộ exam_number ở các cơ sở cần đánh lại;
    # pha 2: gán lại theo thứ tự map.
    with transaction.atomic():
        if target_campus_ids:
            ExamRoomStudent.objects.filter(campus_id__in=target_campus_ids).update(exam_number=None)

        # Chỉ xử lý cơ sở có trong map, theo thứ tự campus_order
        for code_key in campus_order:
            prefix_digit = campus_sbd_digit_map.get(code_key) or code_upper_to_digit.get(code_key.upper())
            if not prefix_digit:
                continue
            campus = campuses_by_code_upper.get(code_key.upper())
            if not campus:
                continue
            students_campus = list(ExamRoomStudent.objects.filter(campus=campus))
            if not students_campus:
                continue
            campus_start = campus_start_serials.get(campus.id, start_serial)
            if campus_start < 1:
                campus_start = 1

            def sort_key(stu):
                code = stu.student_code or ""
                shift_digit = code[2] if len(code) >= 3 else "0"
                group_digit = code[3] if len(code) >= 4 else "0"
                try:
                    shift_val = int(shift_digit)
                except ValueError:
                    shift_val = 9
                grade = _extract_grade_from_class_name(stu.class_name) or 0
                try:
                    group_val = int(group_digit)
                except ValueError:
                    group_val = 9
                name_key = _name_sort_key_vi(stu.full_name or "")
                # Khối trước (10 hết rồi mới 11), trong khối: sáng trước tối sau → tổ hợp → tên.
                return (grade, shift_val, group_val, name_key)

            students_sorted = sorted(students_campus, key=sort_key)
            for idx, stu in enumerate(students_sorted, start=campus_start):
                stu.exam_number = f"{prefix_digit}{idx:05d}"
            ExamRoomStudent.objects.bulk_update(students_sorted, ["exam_number"])


_EXAM_SORT_CAMPUS_RANK = {
    "AT": "1",
    "BS": "2",
    "CS": "3",
    "VH": "4",
    "KT": "5",
    "CN": "6",
    "ĐS": "7",
    "DS": "7",  # phòng khi dữ liệu dùng DS không dấu
    "HT": "8",
}

def _exam_hv_shift_digit_from_code(student_code: str) -> str:
    """Ký tự thứ 3 trong mã HV: 0 = buổi sáng, 1 = buổi tối (quy ước import)."""
    if not student_code or len(student_code) < 3:
        return ''
    return student_code[2]


def _exam_room_import_cell_is_integration(val) -> bool:
    """Ô cột Hoà nhập khi import danh sách xếp phòng: có chữ HN (không phân biệt hoa thường) → hoà nhập."""
    if val is None:
        return False
    s = str(val).strip().upper()
    if not s:
        return False
    return 'HN' in ''.join(s.split())


_EXAM_SORT_SHIFT_ORDER = ('sang', 'chieu', 'toi')


def _get_room_grid_for_shift(room, shift: str | None) -> tuple[int, int]:
    """Lấy kích thước phòng theo ca; fallback về kích thước mặc định của phòng."""
    cfg = None
    if room and shift:
        cfg = ExamRoomShiftConfig.objects.filter(exam_room=room, shift=shift).first()
    if cfg:
        rows = int(getattr(cfg, 'row_count', 0) or 0)
        cols = int(getattr(cfg, 'col_count', 0) or 0)
    else:
        rows = int(getattr(room, 'row_count', 0) or 0)
        cols = int(getattr(room, 'col_count', 0) or 0)
    return rows, cols


def _get_room_capacity_for_shift(room, shift: str | None) -> int:
    rows, cols = _get_room_grid_for_shift(room, shift)
    return max(rows * cols, 0)


def _exam_sbd_prefix_digit_for_campus(campus) -> str | None:
    if not campus or not (campus.code or "").strip():
        return None
    campus_sbd_digit_map = dict(_EXAM_SORT_CAMPUS_RANK)
    code_upper_to_digit = {k.upper(): v for k, v in campus_sbd_digit_map.items()}
    c = (campus.code or "").upper()
    return campus_sbd_digit_map.get(campus.code) or code_upper_to_digit.get(c)


def _exam_campus_current_start_serial(campus, prefix: str) -> int:
    """
    Lấy mốc serial hiện có nhỏ nhất của cơ sở theo prefix SBD (vd 100321 -> 321).
    Dùng để giữ nguyên "số bắt đầu" mà người dùng đã đánh trước đó khi sync sau xếp phòng.
    """
    nums = (
        ExamRoomStudent.objects.filter(campus=campus, exam_number__isnull=False)
        .exclude(exam_number="")
        .values_list("exam_number", flat=True)
    )
    serials: list[int] = []
    for n in nums:
        s = (str(n or "")).strip()
        if len(s) != 6 or not s.isdigit():
            continue
        if not s.startswith(prefix):
            continue
        serials.append(int(s[1:]))
    if not serials:
        return 1
    return max(min(serials), 1)


def _exam_reorder_seats_in_room_by_sbd(room, shift: str) -> None:
    """Gán lại seat_number 1..n: thường trước (theo SBD), hòa nhập cuối phòng."""
    with transaction.atomic():
        autos = list(
            StudentExamAssignment.objects.select_for_update()
            .select_related("student")
            .filter(exam_room=room, shift=shift)
        )
        autos.sort(key=_exam_room_seat_order_key)
        seat_no = 1
        for a in autos:
            a.seat_number = seat_no
            seat_no += 1
        if autos:
            StudentExamAssignment.objects.bulk_update(autos, ["seat_number"])


def _assign_exam_numbers_for_campus_sync_room_display(campus, *, force: bool = False) -> None:
    """
    Đánh lại SBD toàn cơ sở theo trình tự:
    **khối lớp** → **ca thi** → **tổ hợp môn** (mã SubjectGroup) → **phòng thi**.
    Trong cùng ca và tổ hợp: các phòng xếp theo học viên *đại diện* trong phòng (nhỏ nhất theo
    **tên → đệm → họ**, cùng quy tắc `_name_sort_key_vi`); nếu trùng thì theo tên phòng (natural sort).
    Trong từng phòng: số ghế → thường trước, hòa nhập cuối (SBD cũ → tên).
    HV chưa có phòng: cuối cùng, theo khối/buổi/tổ hợp/tên (như import DS).
    Không đổi phòng/ghế/layout — chỉ cập nhật exam_number.

    Mặc định chỉ chạy khi cơ sở có ít nhất một HV hòa nhập; ``force=True`` bỏ qua điều kiện này
    (dùng sau xoá/thêm HV khỏi phòng thi).
    """
    from homepage.models import ExamRoomStudent, StudentExamAssignment

    if not force and not ExamRoomStudent.objects.filter(campus=campus, is_integration=True).exists():
        return

    prefix = _exam_sbd_prefix_digit_for_campus(campus)
    if not prefix:
        return
    start_serial = _exam_campus_current_start_serial(campus, prefix)

    seen: set[str] = set()
    ordered_students: list = []

    autos_all = list(
        StudentExamAssignment.objects.filter(
            student__campus_id=campus.id,
            exam_room_id__isnull=False,
        ).select_related("student", "student__subject_group", "exam_room")
    )
    by_grade: dict[int, list] = defaultdict(list)
    for a in autos_all:
        grade = _extract_grade_from_class_name(a.student.class_name) or 99
        by_grade[grade].append(a)

    for grade in sorted(by_grade.keys()):
        grade_autos = by_grade[grade]
        for shift in _EXAM_SORT_SHIFT_ORDER:
            autos = [a for a in grade_autos if a.shift == shift]
            if not autos:
                continue
            by_sg: dict[str, list] = defaultdict(list)
            for a in autos:
                sg_code = (a.student.subject_group.code if a.student.subject_group else "") or ""
                by_sg[sg_code].append(a)
            for sg_code in sorted(by_sg.keys(), key=lambda c: (not bool(c), c)):
                bucket = by_sg[sg_code]
                by_room: dict[int, list] = defaultdict(list)
                for a in bucket:
                    by_room[a.exam_room_id].append(a)
                room_rank = []
                for rid, room_autos in by_room.items():
                    room_autos_for_compare = sorted(
                        room_autos,
                        key=_exam_room_seat_order_key,
                    )
                    first_auto = room_autos_for_compare[0]
                    first_name_key = _name_sort_key_vi(first_auto.student.full_name or "")
                    room = room_autos[0].exam_room
                    rname = natural_sort_key((room.name or "") if room else "")
                    room_rank.append((first_name_key, rname, rid))
                room_rank.sort(key=lambda t: (t[0], t[1]))
                for _name_key, _rname, rid in room_rank:
                    room_autos = sorted(
                        by_room[rid],
                        key=_exam_room_seat_order_key,
                    )
                    for a in room_autos:
                        pk = a.student.student_code
                        if pk not in seen:
                            seen.add(pk)
                            ordered_students.append(a.student)

    all_stu = list(ExamRoomStudent.objects.filter(campus=campus))
    unassigned = [s for s in all_stu if s.student_code not in seen]

    def _pool_key(stu):
        code = stu.student_code or ""
        shift_digit = code[2] if len(code) >= 3 else "0"
        group_digit = code[3] if len(code) >= 4 else "0"
        try:
            shift_val = int(shift_digit)
        except ValueError:
            shift_val = 9
        grade = _extract_grade_from_class_name(stu.class_name) or 0
        try:
            group_val = int(group_digit)
        except ValueError:
            group_val = 9
        return (grade, shift_val, group_val, _name_sort_key_vi(stu.full_name or ""))

    unassigned.sort(key=_pool_key)
    final_order = ordered_students + unassigned
    if not final_order:
        return

    with transaction.atomic():
        ExamRoomStudent.objects.filter(campus=campus).update(exam_number=None)
        for idx, stu in enumerate(final_order, start=start_serial):
            stu.exam_number = f"{prefix}{idx:05d}"
        ExamRoomStudent.objects.bulk_update(final_order, ["exam_number"])


def _exam_sort_filtered_student_rows(request_get):
    """
    Danh sách học viên sau khi áp dụng cùng bộ lọc và thứ tự sắp xếp như trang sắp xếp phòng thi.
    Dùng cho dashboard (phân trang) và export Excel theo filter.

    Thứ tự: cơ sở → SBD → ca/nhóm mã HV → khối → tên.

    Gán phòng hiển thị theo ca:
    - Một học viên có phòng ở bất kỳ ca nào được coi là đã có phòng; cột Phòng ưu tiên
      ca Sáng → Chiều → Tối (một dòng một phòng đại diện).
    """
    students_qs = ExamRoomStudent.objects.select_related('campus', 'subject_group').all()
    filter_campus = request_get.get('campus_id', '')
    filter_grade = request_get.get('grade', '')
    filter_subject = request_get.get('subject_group', '')
    filter_status = request_get.get('status', '')
    filter_class = (request_get.get('class_name') or '').strip()
    filter_class_lower = filter_class.lower()
    filter_name = (request_get.get('full_name') or '').strip()
    filter_student_status = (request_get.get('student_status') or '').strip()

    if filter_campus:
        students_qs = students_qs.filter(campus_id=filter_campus)

    students = list(students_qs)

    all_assignments = list(
        StudentExamAssignment.objects.select_related('exam_room').all()
    )
    by_student = {}
    for a in all_assignments:
        by_student.setdefault(a.student_id, []).append(a)

    def _pick_display_assignment(sid):
        lst = by_student.get(sid, [])
        for sh in _EXAM_SORT_SHIFT_ORDER:
            for x in lst:
                if x.shift == sh:
                    return x
        return lst[0] if lst else None

    assignment_map = {sid: _pick_display_assignment(sid) for sid in by_student}

    rows = []
    for s in students:
        grade = _extract_grade_from_class_name(s.class_name)
        subject_code = s.subject_group.code if s.subject_group else ''
        if filter_grade and str(grade or '') != filter_grade:
            continue
        if filter_subject and subject_code != filter_subject:
            continue
        if filter_class:
            if (s.class_name or '').strip().lower() != filter_class_lower:
                continue
        if filter_name:
            if _vn_normalize_no_diacritics(filter_name) not in _vn_normalize_no_diacritics(s.full_name):
                continue
        if filter_student_status == 'integration' and not getattr(s, 'is_integration', False):
            continue
        if filter_student_status == 'normal' and getattr(s, 'is_integration', False):
            continue
        a = assignment_map.get(s.student_code)
        if filter_status == 'assigned' and not a:
            continue
        if filter_status == 'unassigned' and a:
            continue
        rows.append({
            'student': s,
            'grade': grade,
            'subject_group_code': subject_code,
            'assignment': a,
        })

    def _dashboard_sort_key(row):
        stu = row['student']
        campus_code = getattr(stu.campus, 'code', '') or ''
        campus_rank = _EXAM_SORT_CAMPUS_RANK.get(campus_code, 9)
        code = stu.student_code or ''
        shift_digit = code[2] if len(code) >= 3 else '0'
        group_digit = code[3] if len(code) >= 4 else '0'
        try:
            shift_val = int(shift_digit)
        except ValueError:
            shift_val = 9
        grade_val = row['grade'] or 0
        try:
            group_val = int(group_digit)
        except ValueError:
            group_val = 9
        name_key = _name_sort_key_vi(stu.full_name)
        return (
            campus_rank,
            _exam_number_sort_key(getattr(stu, 'exam_number', None), code),
            shift_val,
            grade_val,
            group_val,
            name_key,
        )

    rows.sort(key=_dashboard_sort_key)
    return rows



def exam_room_sort_dashboard(request):
    """Trang sắp xếp phòng thi (tối giản): danh sách học viên + gán vào phòng theo ca."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    filter_campus = request.GET.get('campus_id', '')
    filter_grade = request.GET.get('grade', '')
    filter_subject = request.GET.get('subject_group', '')
    filter_status = request.GET.get('status', '')
    filter_class = (request.GET.get('class_name') or '').strip()
    filter_name = (request.GET.get('full_name') or '').strip()
    filter_student_status = (request.GET.get('student_status') or '').strip()
    if filter_student_status not in ('', 'integration', 'normal'):
        filter_student_status = ''

    rows = _exam_sort_filtered_student_rows(request.GET)

    # Phân trang: mỗi trang 50 học viên
    try:
        page = int(request.GET.get('page', '1'))
    except ValueError:
        page = 1
    if page < 1:
        page = 1
    page_size = 50
    total = len(rows)
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    paged_rows = rows[start:end]

    # Tạo range đơn giản cho phân trang trong template
    range_1_to_total_pages = list(range(1, total_pages + 1)) if total_pages > 1 else [1]

    context = {
        'sorted_students': paged_rows,
        'campuses': Campus.objects.all().order_by('code'),
        'subject_group_choices': SubjectGroup.objects.all().order_by('code'),
        'filter_campus': filter_campus,
        'filter_grade': filter_grade,
        'filter_subject': filter_subject,
        'filter_status': filter_status,
        'filter_student_status': filter_student_status,
        'filter_class': filter_class,
        'filter_name': filter_name,
        'page': page,
        'total_pages': total_pages,
        'total_students': total,
        'range_1_to_total_pages': range_1_to_total_pages,
        'rooms': ExamRoom.objects.select_related('campus').order_by('campus__code', 'name'),
    }
    return render(request, 'adminpageSIMCODE/exam_room_sort_dashboard.html', context)


@transaction.atomic
def exam_room_move_students(request):
    """POST: gán học viên (chưa có phòng trong ca đó) vào phòng thi theo ca."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:exam_room_sort_dashboard')

    shift = request.POST.get('shift', 'sang')
    room_id = request.POST.get('room_id')
    student_codes = request.POST.getlist('student_codes')
    if not room_id or not student_codes:
        messages.error(request, 'Thiếu phòng thi hoặc danh sách học viên.')
        return redirect('adminpage:exam_room_sort_dashboard')

    try:
        room = ExamRoom.objects.get(id=int(room_id))
    except (ValueError, ExamRoom.DoesNotExist):
        messages.error(request, 'Phòng thi không hợp lệ.')
        return redirect('adminpage:exam_room_sort_dashboard')

    already = set(
        StudentExamAssignment.objects.filter(shift=shift).values_list('student_id', flat=True)
    )
    current_count = StudentExamAssignment.objects.filter(exam_room=room, shift=shift).count()
    shift_capacity = _get_room_capacity_for_shift(room, shift)
    capacity_left = shift_capacity - current_count
    if capacity_left <= 0:
        messages.warning(request, f'Phòng {room.name} đã đủ {shift_capacity} chỗ ở ca {shift}.')
        return redirect('adminpage:exam_room_sort_dashboard')

    to_assign = [c for c in student_codes if c and c not in already][:capacity_left]
    if not to_assign:
        messages.warning(request, 'Không có học viên nào chưa có phòng trong ca này.')
        return redirect('adminpage:exam_room_sort_dashboard')

    used_seats = set(
        StudentExamAssignment.objects.filter(exam_room=room, shift=shift)
        .values_list('seat_number', flat=True)
    )
    next_seat = 1
    assigned = 0
    assigned_students: list[ExamRoomStudent] = []
    for code in to_assign:
        while next_seat in used_seats:
            next_seat += 1
        try:
            s = ExamRoomStudent.objects.get(student_code=code)
        except ExamRoomStudent.DoesNotExist:
            continue
        StudentExamAssignment.objects.update_or_create(
            student=s,
            shift=shift,
            defaults={'exam_room': room, 'seat_number': next_seat},
        )
        used_seats.add(next_seat)
        assigned_students.append(s)
        assigned += 1
        next_seat += 1

    msg = (
        f'Đã gán {assigned} học viên vào phòng {room.name} '
        f'(ca {dict(StudentExamAssignment.SHIFT_CHOICES).get(shift, shift)}).'
    )
    if assigned > 0:
        _sync_exam_sbd_campus_after_room_assignment(room.campus)
        msg += (
            ' Nếu cơ sở có học viên hòa nhập, SBD đã được đánh lại theo ca → tổ hợp môn → phòng (thứ tự phòng: ưu tiên tên trước, HV đại diện mỗi phòng).'
        )
    messages.success(request, msg)
    return redirect('adminpage:exam_room_sort_dashboard')


def exam_room_export_excel(request):
    """
    Export danh sách học viên ra Excel đúng bộ lọc + thứ tự như trang sắp xếp phòng thi
    (cơ sở, khối, tổ hợp, lớp, tên, tình trạng). Không đụng đăng ký thi tốt nghiệp.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    rows = _exam_sort_filtered_student_rows(request.GET)

    wb = Workbook()
    ws = wb.active
    ws.title = 'DS Xếp phòng thi'
    header = [
        'STT',
        'Số báo danh',
        'Mã HV',
        'Họ và tên',
        'Lớp',
        'Khối',
        'Cơ sở',
        'Tổ hợp',
        'Hoà nhập',
        'Tình trạng phòng',
        'Phòng (theo ca hiển thị)',
    ]
    ws.append(header)
    for idx, row in enumerate(rows, start=1):
        s = row['student']
        grade = row['grade']
        subject_code = row['subject_group_code']
        a = row['assignment']
        sbd = s.exam_number or ''
        if a and a.exam_room:
            room_str = f'{a.exam_room.name} ({a.get_shift_display()})'
            status_txt = 'Đã có phòng'
        else:
            room_str = ''
            status_txt = 'Chưa có phòng'
        hoan = 'Có' if getattr(s, 'is_integration', False) else 'Không'
        ws.append(
            [
                idx,
                sbd,
                s.student_code,
                s.full_name,
                s.class_name or '',
                grade or '',
                s.campus.name if s.campus else '',
                subject_code,
                hoan,
                status_txt,
                room_str,
            ]
        )
    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="ds_xep_phong_thi.xlsx"'
    wb.save(resp)
    return resp

@transaction.atomic
def exam_room_delete_all_students(request):
    """
    Xoá danh sách học viên dùng cho xếp phòng thi:
    - Xoá StudentExamAssignment, ExamSubjectSeat, ExamRoomSubject (để chi tiết phòng thi reset môn thi, sơ đồ ghế).
    - Xoá ExamRoomStudent (tách riêng, không đụng tới đăng ký thi tốt nghiệp).
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:exam_room_sort_dashboard')

    ExamSubjectSeat.objects.all().delete()
    ExamRoomSubject.objects.all().delete()
    StudentExamAssignment.objects.all().delete()
    deleted_count = ExamRoomStudent.objects.count()
    ExamRoomStudent.objects.all().delete()
    messages.success(request, f'Đã xoá {deleted_count} học viên (xếp phòng thi) và toàn bộ dữ liệu xếp phòng thi.')
    return redirect('adminpage:exam_room_sort_dashboard')


@transaction.atomic
def exam_room_reassign_sbd(request):
    """POST: Đánh lại số báo danh cho toàn bộ ExamRoomStudent (theo quy tắc hiện tại)."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:exam_room_sort_dashboard')

    raw_start = (request.POST.get('sbd_start') or '1').strip()
    try:
        start_serial = int(raw_start)
    except ValueError:
        messages.error(request, 'Số bắt đầu SBD không hợp lệ. Vui lòng nhập số nguyên >= 1.')
        return redirect('adminpage:exam_room_sort_dashboard')
    if start_serial < 1:
        messages.error(request, 'Số bắt đầu SBD phải lớn hơn hoặc bằng 1.')
        return redirect('adminpage:exam_room_sort_dashboard')
    campus_start_serials: dict[int, int] = {}
    for campus in Campus.objects.all().only('id'):
        key = f'sbd_start_campus_{campus.id}'
        raw_val = (request.POST.get(key) or '').strip()
        if not raw_val:
            continue
        try:
            val = int(raw_val)
        except ValueError:
            messages.error(request, f'Số bắt đầu theo cơ sở không hợp lệ (campus_id={campus.id}).')
            return redirect('adminpage:exam_room_sort_dashboard')
        if val < 1:
            messages.error(request, f'Số bắt đầu theo cơ sở phải >= 1 (campus_id={campus.id}).')
            return redirect('adminpage:exam_room_sort_dashboard')
        campus_start_serials[campus.id] = val

    from urllib.parse import urlencode
    params = {
        'page': request.POST.get('return_page') or '1',
        'campus_id': request.POST.get('return_campus_id', ''),
        'grade': request.POST.get('return_grade', ''),
        'subject_group': request.POST.get('return_subject_group', ''),
        'status': request.POST.get('return_status', ''),
        'student_status': request.POST.get('return_student_status', ''),
        'class_name': request.POST.get('return_class_name', ''),
        'full_name': request.POST.get('return_full_name', ''),
        '_': str(int(timezone.now().timestamp())),
    }

    try:
        _assign_exam_numbers_for_all_students(
            start_serial=start_serial,
            campus_start_serials=campus_start_serials,
        )
        messages.success(
            request,
            (
                f'Đã đánh lại số báo danh cho toàn bộ học viên '
                f'(mặc định bắt đầu từ {start_serial}; có {len(campus_start_serials)} cơ sở nhập riêng).'
            ),
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception('Đánh lại SBD: %s', e)
        messages.error(request, 'Đánh lại số báo danh gặp lỗi. Vui lòng xem log server.')

    url = reverse('adminpage:exam_room_sort_dashboard') + '?' + urlencode(params)
    return redirect(url)


@transaction.atomic
def exam_room_delete_all_rooms(request):
    """
    Xoá toàn bộ phòng thi (hoặc chỉ phòng tại một cơ sở) cùng dữ liệu xếp phòng:
    StudentExamAssignment, ExamSubjectSeat, ExamRoomSubject.
    Không xoá ExamRoomStudent (danh sách học viên trên trang Sắp xếp phòng thi).
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        messages.warning(request, 'Chỉ chấp nhận POST để xoá phòng thi.')
        return _redirect_exam_room_manage_preserve_campus(request)

    scope = (request.POST.get('scope') or 'all').strip()
    cid_raw = (request.POST.get('campus_id') or '').strip()

    if scope == 'campus':
        if not cid_raw.isdigit() or not Campus.objects.filter(id=int(cid_raw)).exists():
            messages.error(request, 'Thiếu hoặc sai cơ sở để xoá phòng theo cơ sở.')
            return _redirect_exam_room_manage_preserve_campus(request)
        campus_pk = int(cid_raw)
        n_rooms = ExamRoom.objects.filter(campus_id=campus_pk).count()
        StudentExamAssignment.objects.filter(exam_room__campus_id=campus_pk).delete()
        ExamSubjectSeat.objects.filter(exam_room__campus_id=campus_pk).delete()
        ExamRoomSubject.objects.filter(exam_room__campus_id=campus_pk).delete()
        ExamRoom.objects.filter(campus_id=campus_pk).delete()
        campus = Campus.objects.get(id=campus_pk)
        messages.success(
            request,
            f'Đã xoá {n_rooms} phòng thi tại {campus.name} và toàn bộ dữ liệu xếp phòng liên quan.',
        )
    else:
        n_rooms = ExamRoom.objects.count()
        StudentExamAssignment.objects.all().delete()
        ExamSubjectSeat.objects.all().delete()
        ExamRoomSubject.objects.all().delete()
        ExamRoom.objects.all().delete()
        messages.success(
            request,
            f'Đã xoá toàn bộ {n_rooms} phòng thi và dữ liệu xếp phòng liên quan (mọi cơ sở).',
        )

    return _redirect_exam_room_manage_preserve_campus(request)


def exam_room_delete_student(request, student_code):
    """Xoá 1 học viên khỏi danh sách xếp phòng thi (chỉ tác động dữ liệu xếp phòng)."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        messages.error(request, 'Tài khoản không hợp lệ.')
        return redirect('adminpage:exam_room_sort_dashboard')

    if request.method != 'POST':
        messages.warning(request, 'Thao tác xoá cần gửi POST. Vui lòng dùng nút Xoá trên bảng.')
        return redirect('adminpage:exam_room_sort_dashboard')

    try:
        student = ExamRoomStudent.objects.get(student_code=student_code)
    except ExamRoomStudent.DoesNotExist:
        messages.error(request, 'Không tìm thấy học viên.')
        return redirect('adminpage:exam_room_sort_dashboard')

    if StudentExamAssignment.objects.filter(student=student).exists():
        messages.error(
            request,
            f'Không thể xoá học viên {student.full_name}: đã được xếp vào phòng thi. '
            'Vui lòng xoá khỏi phòng thi trước (vào chi tiết phòng thi để thao tác).'
        )
        from urllib.parse import urlencode
        params = {
            'page': request.POST.get('return_page') or '1',
            'campus_id': request.POST.get('return_campus_id', ''),
            'grade': request.POST.get('return_grade', ''),
            'subject_group': request.POST.get('return_subject_group', ''),
            'status': request.POST.get('return_status', ''),
            'student_status': request.POST.get('return_student_status', ''),
            'class_name': request.POST.get('return_class_name', ''),
            'full_name': request.POST.get('return_full_name', ''),
        }
        url = reverse('adminpage:exam_room_sort_dashboard') + '?' + urlencode(params)
        return redirect(url)

    full_name = student.full_name
    # Xoá trong block atomic riêng để commit ngay; đánh lại SBD chạy sau, lỗi không rollback phần xoá
    with transaction.atomic():
        ExamSubjectSeat.objects.filter(student=student).delete()
        StudentExamAssignment.objects.filter(student=student).delete()
        student.delete()
    try:
        with transaction.atomic():
            _assign_exam_numbers_for_all_students()
        messages.success(request, f'Đã xoá học viên {full_name}. Số báo danh đã được đánh lại.')
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception('Đánh lại SBD sau khi xoá học viên: %s', e)
        messages.success(request, f'Đã xoá học viên {full_name}. Lưu ý: Đánh lại số báo danh gặp lỗi.')
    # Redirect về đúng trang và bộ lọc trước khi xoá
    from urllib.parse import urlencode
    params = {
        'page': request.POST.get('return_page') or '1',
        'campus_id': request.POST.get('return_campus_id', ''),
        'grade': request.POST.get('return_grade', ''),
        'subject_group': request.POST.get('return_subject_group', ''),
        'status': request.POST.get('return_status', ''),
        'student_status': request.POST.get('return_student_status', ''),
        'class_name': request.POST.get('return_class_name', ''),
        'full_name': request.POST.get('return_full_name', ''),
        '_': str(int(timezone.now().timestamp())),
    }
    url = reverse('adminpage:exam_room_sort_dashboard') + '?' + urlencode(params)
    return redirect(url)



def adminpage(request):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            users = User.objects.all()
            context = {'users': users}
            return render(request, 'adminpageSIMCODE/user.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:Login')

def getCategory(request):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST':
                action = int(request.POST.get('action'))
                category_name = request.POST.get('category_name')
                category_enable = int(request.POST.get('enableHidden'))
                category_id = request.POST.get('category_id', None)
                if category_enable == 1:
                    category_enable = True
                else:
                    category_enable = False
                if action == 0:
                    category = Category.objects.create(name = category_name, enable = category_enable)
                category.save()
                return redirect('adminpage:category')
            categories = Category.objects.all()
            context = {'categories': categories}
            return render(request, 'adminpageSIMCODE/category.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:Login')

def get_post_data(request, post_id):
    post = get_object_or_404(Post, id = post_id)
    uploadedfiles = UploadedFile.objects.filter(post = post)
    data = {'files': [{'name': f.pdf_file.name, 'url': f.pdf_file.url} for f in uploadedfiles],'id': post.id,'name': post.title,'image': post.image_file.url, 'enable': post.enable, 'content': post.content}
    return JsonResponse(data)

def getCategoryDetail(request, category_id):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST':
                action = int(request.POST.get('action'))
                if action == 1:
                    category_name = request.POST.get('category_name')
                    category_enable = int(request.POST.get('enableHidden'))
                    category_id = request.POST.get('category_id', None)
                    if category_enable == 1:
                        category_enable = True
                    else:
                        category_enable = False
                    category = Category.objects.get(id = category_id)
                    category.name = category_name
                    category.enable = category_enable
                    category.save()
                elif action == 0:
                    chapter_name = request.POST.get('chapter_name')
                    content = request.POST.get('ckeditor1', None)
                    image = request.FILES.get('topic-image', None)
                    files = request.FILES.getlist('topic-files', None)
                    category = Category.objects.get(id = category_id)
                    post = Post.objects.create(title = chapter_name, content = content, image_file = image, category = category)
                    post.save()
                    for file in files:
                        f = UploadedFile.objects.create(post=post, pdf_file=file)
                        f.save()
                elif action == 2:
                    post_id = request.POST.get('post-id')
                    post = get_object_or_404(Post, id=post_id)
                    category_enable = int(request.POST.get('enableHidden'))
                    if category_enable == 1:
                        category_enable = True
                    else:
                        category_enable = False
                    post.title = request.POST.get('chapter_name', post.title)
                    post.content = request.POST.get('ckeditor1', post.content)
                    post.enable = category_enable

        # Upload ảnh mới
                    if 'image_file' in request.FILES:
                        post.image_file = request.FILES['topic-image']

        # Xóa các tệp PDF cũ và thêm tệp PDF mới
                    existing_files = request.POST.get('existing-files', '[]')
     # Chuyển từ JSON string sang list
                    try:
                        existing_files = json.loads(existing_files)
                    except json.JSONDecodeError:
                        existing_files = []

# Lấy các file mới từ input file
                    new_files = request.FILES.getlist('topic-files')

# Nếu có file mới, xóa các file cũ không còn được chọn
                    if new_files:
                        for uploaded_file in post.uploadedfile_set.all():
        # Nếu file cũ không có trong danh sách file hiện tại, xóa nó
                            if uploaded_file.pdf_file.name not in existing_files:
                                uploaded_file.delete()

    # Thêm file mới vào cơ sở dữ liệu
                        for new_file in new_files:
                            UploadedFile.objects.create(post=post, pdf_file=new_file)
                    else:
    # Nếu không có file mới, không làm gì với các file cũ
                        pass
                    post.save()
                elif action == 3:
                    post_id = request.POST.get('post-id')
                    post = get_object_or_404(Post, id=post_id)
                    # Xóa các file đính kèm trước để tránh lỗi ràng buộc khóa ngoại
                    UploadedFile.objects.filter(post=post).delete()
                    post.delete()



                return redirect('adminpage:categorydetail', category_id = category_id )
            category = Category.objects.get(id = category_id)
            posts = Post.objects.filter(category__id = category_id)
            context = {'category': category, 'posts':posts}
            return render(request, 'adminpageSIMCODE/categorydetail.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:Login')

def get_PB(request):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST':
                action = int(request.POST.get('action'))
                category_name = request.POST.get('category_name')
                category_enable = int(request.POST.get('enableHidden'))
                category_id = request.POST.get('category_id', None)
                if category_enable == 1:
                    category_enable = True
                else:
                    category_enable = False
                if action == 0:
                    pb = PhongBan.objects.create(name = category_name, enable = category_enable)
                else:
                    pb = PhongBan.objects.get(id = category_id)
                    pb.name = category_name
                    pb.enable = category_enable
                pb.save()
                return redirect('adminpage:phongban')
            pbs = PhongBan.objects.all()
            context = {'pbs': pbs}
            return render(request, 'adminpageSIMCODE/phongban.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:Login')
def get_CCTC(request, pb_id):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST':
                action = int(request.POST.get('action'))
                chapter_name = request.POST.get('chapter_name')
                role1 = request.POST.get('role1', None)
                role2 = request.POST.get('role2')
                chuyenmon = request.POST.get('chuyenmon', None)
                namsinh = request.POST.get('namsinh')
                sex = int(request.POST.get('sex'))
                bac = request.POST.get('bac')
                image_file = request.FILES.get('topic-image', None)
                enable = int(request.POST.get('enableHidden'))
                if enable == 1:
                    enable = True
                else:
                    enable = False
                if sex == 1:
                    sex = True
                else:
                    sex = False
                if action == 0:
                    gv = GV.objects.create(name = chapter_name, role1 = role1, role2 = role2, sex = sex, chuyenmon = chuyenmon, namsinh = namsinh, bac = bac, image_file = image_file, enable = enable)
                    gv.save()
                    phongban = PhongBan.objects.get(id = pb_id)
                    pb_gv = PB_GV.objects.create(phongban = phongban, gv = gv)
                    pb_gv.save()
                else:
                    gv_id = int(request.POST.get('post-id'))
                    gv = get_object_or_404(GV, id = gv_id)
                    gv.name = chapter_name
                    gv.role1 = role1
                    gv.role2 = role2
                    gv.chuyenmon = chuyenmon
                    gv.namsinh = namsinh
                    gv.sex = sex
                    gv.enable = enable
                    gv.bac = bac
                    # request.FILES keys are form field names (name="topic-image"), not model field names
                    if image_file:
                        gv.image_file = image_file
                    gv.save()
                return redirect('adminpage:CCTC', pb_id = pb_id)
            phongban = get_object_or_404(PhongBan, id=pb_id)
            gvs = PB_GV.objects.filter(phongban__id = pb_id)
            context = {'gvs': gvs, 'category': phongban}
            return render(request, 'adminpageSIMCODE/cctc.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:Login')
def get_Admission(request):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            # Get all admission records
            admission_list = AdmissionForm.objects.all()

            # Apply filters
            name_filter = request.GET.get('name', '')
            gender_filter = request.GET.get('gender', '')
            status_filter = request.GET.get('status', '')
            campus_filter = request.GET.get('campus', '')
            shift_filter = request.GET.get('shift', '')
            subject_group_filter = request.GET.get('subject_group', '')
            exam_score_order = request.GET.get('exam_score', '')
            avg_score_order = request.GET.get('avg_score', '')
            graduation_year_filter = request.GET.get('graduation_year', '')
            conduct_filter = request.GET.get('conduct', '')
            from_date_filter = request.GET.get('from_date', '')
            to_date_filter = request.GET.get('to_date', '')

            if name_filter:
                admission_list = admission_list.filter(full_name__icontains=name_filter)

            if gender_filter:
                admission_list = admission_list.filter(gender=gender_filter)

            if status_filter:
                status_bool = status_filter == '1'
                admission_list = admission_list.filter(enable=status_bool)

            if campus_filter:
                admission_list = admission_list.filter(campus_id=campus_filter)

            if shift_filter:
                admission_list = admission_list.filter(shift_id=shift_filter)

            if subject_group_filter:
                admission_list = admission_list.filter(subject_group_id=subject_group_filter)

            if exam_score_order:
                order_by = '-exam_score' if exam_score_order == 'desc' else 'exam_score'
                admission_list = admission_list.order_by(order_by)

            if avg_score_order:
                order_by = '-avg_score' if avg_score_order == 'desc' else 'avg_score'
                admission_list = admission_list.order_by(order_by)

            if graduation_year_filter:
                admission_list = admission_list.filter(graduation_year=graduation_year_filter)

            if conduct_filter:
                admission_list = admission_list.filter(conduct=conduct_filter)
            if from_date_filter:
                admission_list = admission_list.filter(created_at__date__gte=from_date_filter)
            if to_date_filter:
                admission_list = admission_list.filter(created_at__date__lte=to_date_filter)

            # Set up pagination
            paginator = Paginator(admission_list, 10)  # Show 10 records per page
            page = request.GET.get('page', 1)

            try:
                ad = paginator.page(page)
            except PageNotAnInteger:
                ad = paginator.page(1)
            except EmptyPage:
                ad = paginator.page(paginator.num_pages)

            # Get data for filter dropdowns
            from homepage.models import Campus, Shift, SubjectGroup
            campuses = Campus.objects.all()
            shifts = Shift.objects.all()
            subject_groups = SubjectGroup.objects.all()
            # Graduation years for dropdown
            graduation_years = AdmissionForm.objects.order_by('graduation_year').values_list('graduation_year', flat=True).distinct()

            # Lấy thông báo từ session nếu có
            conduct_update_result = request.session.pop('conduct_update_result', None)

            context = {
                'ad': ad,
                'new_users': admission_list.count(),
                # Pass filter values back to template
                'name_filter': name_filter,
                'graduation_year_filter': graduation_year_filter,
                'status_filter': status_filter,
                'campus_filter': campus_filter,
                'shift_filter': shift_filter,
                'subject_group_filter': subject_group_filter,
                'exam_score_order': exam_score_order,
                'avg_score_order': avg_score_order,
                # Pass data for filter dropdowns
                'campuses': campuses,
                'shifts': shifts,
                'subject_groups': subject_groups,
                'graduation_years': graduation_years,
                'conduct_filter': conduct_filter,
                'from_date_filter': from_date_filter,
                'to_date_filter': to_date_filter,
                'conduct_update_result': conduct_update_result,
            }
            return render(request, 'adminpageSIMCODE/admission.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:Login')
def get_Letter(request, admission_id):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            # Get admission record
            admission = get_object_or_404(AdmissionForm, id=admission_id)

            if request.method == 'POST':
                action = request.POST.get('action', 'update')
                if action == 'delete':
                    admission.delete()
                    return redirect('adminpage:admission')
                try:
                    # Kiểm tra trạng thái trước khi cập nhật
                    old_status = admission.enable
                    new_status = request.POST.get('enable') == '1'

                    # Nếu đã duyệt rồi thì không cho phép thay đổi trạng thái
                    if old_status:
                        new_status = True

                    # Update admission record
                    admission.full_name = request.POST.get('full_name')
                    admission.gender = request.POST.get('gender')
                    admission.birthday = request.POST.get('birthday')
                    admission.ethnicity = request.POST.get('ethnicity')
                    admission.religion = request.POST.get('religion')
                    admission.email = request.POST.get('email')
                    admission.id_number = request.POST.get('id_number')
                    admission.id_issued_date = request.POST.get('id_issued_date')
                    admission.id_issued_place = request.POST.get('id_issued_place')

                    # Handle phone number
                    admission.phone = request.POST.get('phone')

                    # Handle CCCD image
                    cccd_image = request.FILES.get('cccd_image')
                    if cccd_image:
                        # Delete old image if exists
                        if admission.cccd_image:
                            try:
                                os.remove(os.path.join(settings.MEDIA_ROOT, str(admission.cccd_image)))
                            except:
                                pass
                        admission.cccd_image = cccd_image

                    # Handle School Record image
                    school_record_image = request.FILES.get('school_record_image')
                    if school_record_image:
                        # Delete old image if exists
                        if admission.school_record_image:
                            try:
                                os.remove(os.path.join(settings.MEDIA_ROOT, str(admission.school_record_image)))
                            except:
                                pass
                        admission.school_record_image = school_record_image

                    # Các trường địa chỉ
                    admission.cccd_province = request.POST.get('cccd_province')
                    admission.cccd_district = request.POST.get('cccd_district')
                    admission.cccd_ward = request.POST.get('cccd_ward')
                    admission.cccd_town = request.POST.get('cccd_town')

                    admission.hometown_province = request.POST.get('hometown_province')

                    admission.birth_reg_province = request.POST.get('birth_reg_province')
                    admission.birth_reg_district = request.POST.get('birth_reg_district')
                    admission.birth_reg_ward = request.POST.get('birth_reg_ward')
                    admission.birth_reg_town = request.POST.get('birth_reg_town')

                    admission.birth_place_province = request.POST.get('birth_place_province')
                    admission.birth_place_district = request.POST.get('birth_place_district')
                    admission.birth_place_ward = request.POST.get('birth_place_ward')
                    admission.birth_place_facility = request.POST.get('birth_place_facility')

                    admission.current_province = request.POST.get('current_province')
                    admission.current_district = request.POST.get('current_district')
                    admission.current_ward = request.POST.get('current_ward')

                    # Thông tin học vấn
                    admission.graduation_school = request.POST.get('graduation_school')
                    admission.graduation_rank = request.POST.get('graduation_rank')
                    admission.conduct = request.POST.get('conduct')
                    admission.current_job = request.POST.get('current_job')

                    # Thông tin phụ huynh
                    admission.father_name = request.POST.get('father_name')
                    admission.father_job = request.POST.get('father_job')
                    admission.father_birth = request.POST.get('father_birth')
                    admission.father_phone = request.POST.get('father_phone')

                    admission.mother_name = request.POST.get('mother_name')
                    admission.mother_job = request.POST.get('mother_job')
                    admission.mother_birth = request.POST.get('mother_birth')
                    admission.mother_phone = request.POST.get('mother_phone')

                    # Xử lý điểm dựa vào năm tốt nghiệp
                    graduation_year = request.POST.get('graduation_year')

                    # Process exam_score and avg_score - only update if field is present in form
                    if 'exam_score' in request.POST:  # Only update if field is present in form
                        exam_score_str = request.POST.get('exam_score')
                        try:
                            admission.exam_score = Decimal(exam_score_str.strip()) if exam_score_str and exam_score_str.strip() else None
                        except InvalidOperation:
                            admission.exam_score = None

                    if 'avg_score' in request.POST:  # Only update if field is present in form
                        avg_score_str = request.POST.get('avg_score')
                        try:
                            admission.avg_score = Decimal(avg_score_str.strip()) if avg_score_str and avg_score_str.strip() else None
                        except InvalidOperation:
                            admission.avg_score = None

                    # Handle math_score and literature_score (always applicable)
                    math_score_str = request.POST.get('math_score')
                    try:
                        admission.math_score = Decimal(math_score_str.strip()) if math_score_str and math_score_str.strip() else None
                    except InvalidOperation:
                        admission.math_score = None

                    literature_score_str = request.POST.get('literature_score')
                    try:
                        admission.literature_score = Decimal(literature_score_str.strip()) if literature_score_str and literature_score_str.strip() else None
                    except InvalidOperation:
                        admission.literature_score = None

                    # Handle THCS scores (lớp 6, 7, 8, 9)
                    # Lớp 6
                    math_score_6_str = request.POST.get('math_score_6')
                    try:
                        admission.math_score_6 = Decimal(math_score_6_str.strip()) if math_score_6_str and math_score_6_str.strip() else None
                    except InvalidOperation:
                        admission.math_score_6 = None

                    literature_score_6_str = request.POST.get('literature_score_6')
                    try:
                        admission.literature_score_6 = Decimal(literature_score_6_str.strip()) if literature_score_6_str and literature_score_6_str.strip() else None
                    except InvalidOperation:
                        admission.literature_score_6 = None

                    # Lớp 7
                    math_score_7_str = request.POST.get('math_score_7')
                    try:
                        admission.math_score_7 = Decimal(math_score_7_str.strip()) if math_score_7_str and math_score_7_str.strip() else None
                    except InvalidOperation:
                        admission.math_score_7 = None

                    literature_score_7_str = request.POST.get('literature_score_7')
                    try:
                        admission.literature_score_7 = Decimal(literature_score_7_str.strip()) if literature_score_7_str and literature_score_7_str.strip() else None
                    except InvalidOperation:
                        admission.literature_score_7 = None

                    # Lớp 8
                    math_score_8_str = request.POST.get('math_score_8')
                    try:
                        admission.math_score_8 = Decimal(math_score_8_str.strip()) if math_score_8_str and math_score_8_str.strip() else None
                    except InvalidOperation:
                        admission.math_score_8 = None

                    literature_score_8_str = request.POST.get('literature_score_8')
                    try:
                        admission.literature_score_8 = Decimal(literature_score_8_str.strip()) if literature_score_8_str and literature_score_8_str.strip() else None
                    except InvalidOperation:
                        admission.literature_score_8 = None

                    # Lớp 9
                    math_score_9_str = request.POST.get('math_score_9')
                    try:
                        admission.math_score_9 = Decimal(math_score_9_str.strip()) if math_score_9_str and math_score_9_str.strip() else None
                    except InvalidOperation:
                        admission.math_score_9 = None

                    literature_score_9_str = request.POST.get('literature_score_9')
                    try:
                        admission.literature_score_9 = Decimal(literature_score_9_str.strip()) if literature_score_9_str and literature_score_9_str.strip() else None
                    except InvalidOperation:
                        admission.literature_score_9 = None

                    # Cập nhật trạng thái
                    admission.enable = new_status

                    # Cập nhật ca học (shift)
                    shift_id = request.POST.get('shift')
                    if shift_id:
                        admission.shift_id = shift_id

                    # Lưu thay đổi
                    admission.save()

                    # Nếu trạng thái thay đổi từ chưa duyệt sang đã duyệt và có email
                    if not old_status and new_status and admission.email:
                        try:
                            # Chuẩn bị nội dung email
                            html_message = render_to_string('adminpageSIMCODE/email_template.html', {
                                'full_name': admission.full_name,
                                'graduation_year': admission.graduation_year,
                                'campus': admission.campus.name,
                                'shift': admission.shift.name,
                                'subject_group': admission.subject_group.code
                            })
                            plain_message = strip_tags(html_message)

                            # Gửi email
                            send_mail(
                                subject='Thông báo xét duyệt hồ sơ nhập học',
                                message=plain_message,
                                from_email=settings.EMAIL_HOST_USER,
                                recipient_list=[admission.email],
                                html_message=html_message,
                                fail_silently=False,
                            )
                        except Exception as e:
                            print(f"Error sending email: {e}")
                            # Continue with success response even if email fails

                    return JsonResponse({'status': 'success'})
                except Exception as e:
                    print(f"Error updating admission: {e}")
                    return JsonResponse({'status': 'error', 'message': str(e)})

            context = {
                'admission': admission,
                'subject_groups': SubjectGroup.objects.all(),
                'shifts': Shift.objects.all()
            }
            return render(request, 'adminpageSIMCODE/letter.html', context)
    return redirect('adminpage:login')
# def Phaser(request):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             if request.method == "POST":
#                 phaser_name = request.POST.get('phaser_name')
#                 phaser_type = request.POST.get('phaser_type')
#                 phaser_path = request.POST.get('phaser_path')
#                 phaser_content = request.POST.get('ckeditor1')
#                 if phaser_type == "1":
#                     imgRoot = os.path.join(settings.MEDIA_ROOT, 'game_pic/')
#                     imgUrl = os.path.join(settings.MEDIA_URL, 'game_pic/')
#                     gameUrl = os.path.join(settings.MEDIA_URL,'game/')
#                     imgRoot=imgRoot.replace("\\","/")
#                     fileimage = request.FILES['file-image']
#                     fss = FileSystemStorage(location=imgRoot)
#                     saveimg = fss.save(fileimage.name,fileimage)
#                     game= Game.objects.create(game_name = phaser_name, game_picture = imgUrl+saveimg,game_hyperlink=gameUrl+phaser_path,game_content = phaser_content)
#                     game.save()
#                     return redirect('Phaser')
#                 else:
#                     imgRoot = os.path.join(settings.MEDIA_ROOT, 'simulation_pic/')
#                     imgUrl = os.path.join(settings.MEDIA_URL, 'simulation_pic/')
#                     simUrl = os.path.join(settings.MEDIA_URL,'simulation/')
#                     imgRoot=imgRoot.replace("\\","/")
#                     fileimage = request.FILES['file-image']
#                     fss = FileSystemStorage(location=imgRoot)
#                     saveimg = fss.save(fileimage.name,fileimage)
#                     simulation= Simulation.objects.create(simulation_name = phaser_name, simulation_picture = imgUrl+saveimg,simulation_hyperlink=simUrl+phaser_path,simulation_content = phaser_content)
#                     simulation.save()
#                     return redirect('Phaser')

#             else:
#                 # courses = Course.objects.all()
#                 # grades = Grade.objects.all()
#                 # subjects = Subject.objects.all()
#                 # context = {'courses': courses, 'grades': grades, 'subjects': subjects}
#                 return render(request, 'adminpageSIMCODE/Phaser.html')
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')
# def course(request):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             if request.method == "POST":
#                 action = int(request.POST.get('action'))
#                 course_name = request.POST.get('course_name')
#                 grade_id = request.POST.get('grade')
#                 subject_id = request.POST.get('subject')
#                 course_desc = request.POST.get('ckeditor1')
#                 course_enable = request.POST.get('enableHidden')
#                 file = request.FILES.get('topic-file', None)
#                 grade = Grade.objects.get(grade_id = grade_id)
#                 subject = Subject.objects.get(subject_id = subject_id)
#                 print(action)
#                 if file is not None:
#                     folder_name = 'course_pic'
#                     fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, folder_name))
#                     filename = fs.save(file.name, file)
#                     saved_file_path = os.path.join(settings.MEDIA_URL, folder_name, filename)
#                 else:
#                     saved_file_path = None
#                 if action == 0:
#                     if saved_file_path is not None:
#                         course = Course.objects.create(course_name = course_name, course_desc = course_desc, course_enable = course_enable, grade = grade, subject = subject, course_picture = saved_file_path)
#                         course.save()
#                     else:
#                         course = Course.objects.create(course_name = course_name, course_desc = course_desc, course_enable = course_enable, grade = grade, subject = subject)
#                         course.save()

#                 else:
#                     course_id = request.POST.get('courseHidden')
#                     course = Course.objects.get(course_id = course_id)
#                     if saved_file_path is not None:
#                         print('a')

#                         course.course_name = course_name
#                         course.course_desc = course_desc
#                         course.course_picture = saved_file_path
#                         course.course_enable = course_enable
#                         course.grade = grade
#                         course.subject = subject
#                     else:
#                         course.course_name = course_name
#                         course.course_desc = course_desc
#                         course.course_picture = saved_file_path
#                         course.course_enable = course_enable
#                         course.grade = grade
#                         course.subject = subject
#                     course.save()

#                 return redirect('course')
#             else:
#                 courses = Course.objects.all()
#                 grades = Grade.objects.all()
#                 subjects = Subject.objects.all()
#                 context = {'courses': courses, 'grades': grades, 'subjects': subjects}
#                 return render(request, 'adminpageSIMCODE/course.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')

# def activity(request, course_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             course = Course.objects.get(course_id = course_id)
#             if request.method == 'POST':
#                 course_name = request.POST.get('course_name', None)
#                 if course_name is not None:
#                     gradeUpdate = request.POST.get('gradeUpdate')
#                     subjectUpdate = request.POST.get('subjectUpdate')
#                     enableHidden = int(request.POST.get('enableHidden'))

#                     grade = Grade.objects.get(grade_id = gradeUpdate)
#                     subject = Subject.objects.get(subject_id = subjectUpdate)
#                     course.course_name = course_name
#                     course.grade = grade
#                     course.subject = subject
#                     course.course_enable = (True if enableHidden == 1 else False)
#                     course.save()
#                 else:
#                     activity_name = request.POST.get('activityName')
#                     activity_order = request.POST.get('activityOrder')
#                     activity_type = request.POST.get('activityType')
#                     activityType = ActivityType.objects.get(activitytype_id = activity_type)
#                     activity = Activity.objects.create(activity_name = activity_name, activity_order = activity_order,  activity_enable = False, activitytype = activityType, course = course)
#                     activity.save()

#                 return redirect('activity', course_id=course_id)

#             else:
#                 grades = Grade.objects.all()
#                 subjects = Subject.objects.all()
#                 types = ActivityType.objects.all()
#                 activities = Activity.objects.filter(course = course)
#                 context = {'course': course, 'activities': activities, 'grades': grades, 'subjects': subjects, 'types': types}
#                 return render(request, 'adminpageSIMCODE/activity.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     else:
#         return redirect('homepage:Register')

# def activitydetail(request, activity_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':

#             activity = Activity.objects.get(activity_id = activity_id)
#             if request.method == "POST":
#                 activity_name = request.POST.get('activityName', None)

#                 if activity_name is not None:
#                     activity_order = request.POST.get('activityOrder')
#                     activity_type = request.POST.get('activityType')
#                     enableHidden = int(request.POST.get('enableHidden'))
#                     activityType = ActivityType.objects.get(activitytype_id = activity_type)
#                     activity.activity_name = activity_name
#                     activity.activity_order = activity_order
#                     activity.activitytype = activityType
#                     activity.activity_enable = (True if enableHidden == 1 else False)
#                     activity.save()
#                 else:
#                     name = request.POST.get('Name')
#                     text = request.POST.get('ckeditor1')
#                     id = request.POST.get('dataHidden', None)
#                     link = request.POST.get('link')
#                     if id is '':

#                         if activity.activitytype.activitytype_id == 1:
#                             order = request.POST.get('order')
#                             theoryUrl = os.path.join(settings.MEDIA_URL,'theory/')
#                             theory = Theory.objects.create(theory_name = name, theory_hyperlink = theoryUrl + link,  theory_order = order, activity = activity)
#                             theory.save()
#                         else:

#                             file = request.FILES.get('topic-file')
#                             folder_name = 'activity_pic'
#                             fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, folder_name))
#                             filename = fs.save(file.name, file)
#                             saved_file_path = os.path.join(settings.MEDIA_URL, folder_name, filename)

#                             if activity.activitytype.activitytype_id == 2:
#                                 gameUrl = os.path.join(settings.MEDIA_URL,'game/')
#                                 game = Game.objects.create(game_name = name, game_content = text, game_picture = saved_file_path, game_hyperlink = gameUrl + link, activity = activity)
#                                 game.save()
#                             elif activity.activitytype.activitytype_id == 3:
#                                 simUrl = os.path.join(settings.MEDIA_URL,'simulation/')
#                                 simulation = Simulation.objects.create(simulation_name = name, simulation_content = text, simulation_picture = saved_file_path, simulation_hyperlink = simUrl + link, activity = activity)
#                                 simulation.save()

#                     else:
#                         enableHidden = int(request.POST.get('enableHidden'))
#                         if activity.activitytype.activitytype_id == 1:
#                             order = request.POST.get('order')
#                             theoryUrl = os.path.join(settings.MEDIA_URL,'theory/')
#                             theory = Theory.objects.get(theory_id = id)
#                             theory.theory_name = name
#                             theory.theory_hyperlink = theoryUrl + link
#                             theory.theory_order = order
#                             theory.theory_enable = enableHidden
#                             theory.save()
#                         else:
#                             # link = request.POST.get('link')
#                             file = request.FILES.get('topic-file', None)
#                             if file is not None:
#                                 folder_name = 'activity_pic'
#                                 fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, folder_name))
#                                 filename = fs.save(file.name, file)
#                                 saved_file_path = os.path.join(settings.MEDIA_URL, folder_name, filename)
#                             else:
#                                 saved_file_path = None
#                             if activity.activitytype.activitytype_id == 2:
#                                 gameUrl = os.path.join(settings.MEDIA_URL,'game/')
#                                 game = Game.objects.get(game_id = id)
#                                 game.game_name = name
#                                 game.game_content = text
#                                 game.game_hyperlink = gameUrl + link
#                                 game.game_enable = enableHidden
#                                 if saved_file_path != game.game_picture and saved_file_path is not None:
#                                     game.game_picture = saved_file_path
#                                 game.save()
#                             elif activity.activitytype.activitytype_id == 3:
#                                 simUrl = os.path.join(settings.MEDIA_URL,'simulation/')
#                                 simulation = Simulation.objects.get(simulation_id = id)
#                                 simulation.simulation_name = name
#                                 simulation.simulation_content = text
#                                 simulation.simulation_hyperlink = simUrl + link
#                                 simulation.simulation_enable = enableHidden
#                                 if saved_file_path != simulation.simulation_picture and saved_file_path is not None:
#                                     simulation.simulation_picture = saved_file_path
#                                 simulation.save()


#                 return redirect('activitydetail', activity_id=activity_id)

#             else:
#                 types = ActivityType.objects.all()
#                 if activity.activitytype.activitytype_id == 1:
#                     ts = Theory.objects.filter(activity = activity).order_by('theory_order')
#                 elif activity.activitytype.activitytype_id == 2:
#                     ts = Game.objects.filter(activity = activity)
#                 elif activity.activitytype.activitytype_id == 3:
#                     ts = Simulation.objects.filter(activity = activity)
#                 context = {'activity':activity,  'types': types, 'ts': ts}
#                 return render(request, 'adminpageSIMCODE/activitydetail.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     else:
#         return redirect('homepage:Register')



# # def coursedetail(request, course_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             course = Course.objects.get(course_id = course_id)
#             if request.method == 'POST':
#                 chapter_name = request.POST.get('chapter_name', None)
#                 if chapter_name is not None:
#                     chapter_order = request.POST.get('chapterOrder')
#                     chapter = Chapter.objects.create(chapter_name = chapter_name, chapter_order = chapter_order, chapter_enable=False, course = course)
#                     chapter.save()
#                     return redirect('coursedetail', course_id=course.course_id)
#                 else:
#                     courseNameUpdate = request.POST.get('course_name')
#                     gradeUpdate = request.POST.get('gradeUpdate')
#                     subjectUpdate = request.POST.get('subjectUpdate')
#                     enableHidden = int(request.POST.get('enableHidden'))
#                     if(enableHidden == 0):
#                         course.course_name = courseNameUpdate
#                         grade = Grade.objects.get(grade_id = gradeUpdate)
#                         subject = Subject.objects.get(subject_id = subjectUpdate)
#                         course.grade = grade
#                         course.subject = subject
#                         course.course_enable = False

#                     else:
#                         course.course_name = courseNameUpdate
#                         grade = Grade.objects.get(grade_id = gradeUpdate)
#                         subject = Subject.objects.get(subject_id = subjectUpdate)
#                         course.grade = grade
#                         course.subject = subject
#                         course.course_enable = True
#                     course.save()
#                     return redirect('coursedetail', course_id=course.course_id)
#             else:
#                 course = Course.objects.get(course_id = course_id)
#                 chapters = course.chapter_set.all().order_by('chapter_order')
#                 grades = Grade.objects.all()
#                 subjects = Subject.objects.all()
#                 context = {'course': course, 'chapters':chapters, 'grades': grades, 'subjects':subjects}
#                 return render(request, 'adminpageSIMCODE/coursedetail.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')

# def chapterdetail(request, chapter_id, course_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             chapter = Chapter.objects.get(chapter_id = chapter_id)
#             course = Course.objects.get(course_id = course_id)
#             if request.method == "POST":
#                 chapter_name = request.POST.get('chapterName', None)
#                 if chapter_name is not None:
#                     chapter_order = request.POST.get('chapterOrder')
#                     chapter_enable = request.POST.get('enableHidden')
#                     chapter.chapter_name = chapter_name
#                     chapter.chapter_order = chapter_order
#                     chapter.chapter_enable = chapter_enable
#                     chapter.save()
#                     return redirect('chapterdetail', chapter_id=chapter_id, course_id=course_id)
#                 else:
#                     lesson_name = request.POST.get('lessonName')
#                     lesson_order = request.POST.get('lessonOrder')
#                     lesson = Lesson.objects.create(lesson_name = lesson_name, lesson_order = lesson_order, lesson_enable = False, chapter = chapter)
#                     lesson.save()
#                     return redirect('chapterdetail', chapter_id=chapter_id, course_id=course_id)
#             else:
#                 lessons = chapter.lesson_set.all().order_by('lesson_order')
#                 context = {'chapter' : chapter, 'course':course, 'lessons': lessons}
#                 return render(request, 'adminpageSIMCODE/chapterdetail.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')

# # def lessondetail(request, chapter_id, lesson_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             chapter = Chapter.objects.get(chapter_id = chapter_id)
#             lesson = Lesson.objects.get(lesson_id = lesson_id)
#             types = ActivityType.objects.all()
#             if request.method == "POST":
#                 lesson_name = request.POST.get('lessonName', None)
#                 if lesson_name is not None:
#                     lesson_order = request.POST.get('lessonOrder')
#                     lesson_enable = request.POST.get('enableHidden')
#                     lesson.lesson_name = lesson_name
#                     lesson.lesson_order = lesson_order
#                     lesson.lesson_enable = lesson_enable
#                     lesson.save()
#                     return redirect('lessondetail', chapter_id=chapter_id, lesson_id=lesson_id)
#                 else:
#                     activity_name = request.POST.get('activityName')
#                     activity_order = request.POST.get('activityOrder')
#                     activity_type = request.POST.get('activityType')
#                     activity_desc = request.POST.get('ckeditor1')
#                     activityType = ActivityType.objects.get(activitytype_id = activity_type)
#                     activity = Activity.objects.create(activity_name = activity_name, activity_order = activity_order, activity_content = activity_desc, activity_enable = False, activitytype = activityType, lesson = lesson)
#                     activity.save()
#                     return redirect('lessondetail', chapter_id=chapter_id, lesson_id=lesson_id)
#             else:
#                 activities = lesson.activity_set.all().order_by('activity_order')
#                 context = {'chapter': chapter, 'lesson': lesson, 'activities': activities, 'types': types}
#                 return render(request, 'adminpageSIMCODE/lessondetail.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')

# def posttype(request):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             posttypes = PostType.objects.all()
#             if request.method == 'POST':
#                 posttype_name = request.POST.get('posttype_name')
#                 posttype = PostType.objects.create(posttype_name = posttype_name)
#                 posttype.save()
#                 return redirect('posttype')
#             else:
#                 context = {'posttypes': posttypes}
#                 return render(request, 'adminpageSIMCODE/posttype.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')


# def forumpost(request, posttype_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             posttype = PostType.objects.get(posttype_id = posttype_id)
#             posts = Post.objects.filter(posttype = posttype)
#             if request.method == 'POST':
#                 pass
#             else:
#                 context = {'posttype': posttype, 'posts': posts}
#                 return render(request, 'adminpageSIMCODE/forumpost.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')

# def forumanswer(request, post_id):
#     if request.user.is_authenticated:
#         account = Account.objects.get(user = request.user)
#         accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
#         if accounttype.accounttype_role == 'admin':
#             post = Post.objects.get(post_id = post_id)
#             comments = Comment.objects.filter(post = post, comment_parent = None)
#             if request.method == 'POST':
#                 pass
#             else:
#                 context = {'post': post, 'comments': comments}
#                 return render(request, 'adminpageSIMCODE/forumanswer.html', context)
#         else:
#             return redirect('homepage:Homepage')
#     return redirect('homepage:Register')

def forum_manage(request):
    """Quản lý bài đăng diễn đàn (model PostForum)."""
    if request.user.is_authenticated:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST':
                action = request.POST.get('action')
                raw_id = request.POST.get('post_id')
                if raw_id and action in ('toggle', 'delete'):
                    post_id = int(raw_id)
                    if action == 'toggle':
                        pf = get_object_or_404(PostForum, id=post_id)
                        pf.enable = not pf.enable
                        pf.save()
                    elif action == 'delete':
                        pf = get_object_or_404(PostForum, id=post_id)
                        pf.delete()
                return redirect('adminpage:forum_manage')
            forum_posts = PostForum.objects.all().order_by('-createdate')
            return render(request, 'adminpageSIMCODE/forum_manage.html', {'forum_posts': forum_posts})
        return redirect('homepage:Homepage')
    return redirect('homepage:Login')

def upload_image(request):
    if request.method == 'POST':

        file = request.FILES.get('upload')
        # Lưu tệp vào thư mục media
        folder_name = 'theory_pic'
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, folder_name))
        filename = fs.save(file.name, file)
        saved_file_path = os.path.join(settings.MEDIA_URL, folder_name, filename)
        print(saved_file_path)
        # Trả về đường dẫn tới tệp đã tải lên
        return JsonResponse({
  "uploaded": 1,
  "url":saved_file_path
})
    return JsonResponse({'error': 'Invalid request'})

def load_vn_address_data():
    try:
        data_path = os.path.join(settings.BASE_DIR, 'adminpage', 'static', 'data.json')
        with open(data_path, encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading VN address data: {e}")
        return []

VN_ADDRESS_DATA = load_vn_address_data()

def get_province_name(code):
    try:
        if not code or code == 'None' or code == '':
            return ''
        for province in VN_ADDRESS_DATA:
            if province.get('Id') == code:
                return province.get('Name', code)
        return code
    except Exception as e:
        print(f"Error in get_province_name: {e}")
        return code if code else ''

def get_district_name(province_code, district_code):
    try:
        if not province_code or province_code == 'None' or province_code == '' or not district_code or district_code == 'None' or district_code == '':
            return ''
        for province in VN_ADDRESS_DATA:
            if province.get('Id') == province_code:
                for district in province.get('Districts', []):
                    if district.get('Id') == district_code:
                        return district.get('Name', district_code)
        return district_code
    except Exception as e:
        print(f"Error in get_district_name: {e}")
        return district_code if district_code else ''

def get_ward_name(province_code, district_code, ward_code):
    try:
        if not province_code or province_code == 'None' or province_code == '' or not district_code or district_code == 'None' or district_code == '' or not ward_code or ward_code == 'None' or ward_code == '':
            return ''
        for province in VN_ADDRESS_DATA:
            if province.get('Id') == province_code:
                for district in province.get('Districts', []):
                    if district.get('Id') == district_code:
                        for ward in district.get('Wards', []):
                            if ward.get('Id') == ward_code:
                                return ward.get('Name', ward_code)
        return ward_code
    except Exception as e:
        print(f"Error in get_ward_name: {e}")
        return ward_code if ward_code else ''

def export_approved_admissions(request):
    """Export approved admissions to Excel file."""
    if not request.user.is_authenticated:
        return redirect('login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    # Get all admissions (not just approved ones, apply filters)
    admissions = AdmissionForm.objects.all()

    # Apply filters from GET parameters
    name_filter = request.GET.get('name', '')
    gender_filter = request.GET.get('gender', '')
    status_filter = request.GET.get('status', '')
    campus_filter = request.GET.get('campus', '')
    shift_filter = request.GET.get('shift', '')
    subject_group_filter = request.GET.get('subject_group', '')
    exam_score_order = request.GET.get('exam_score', '')
    avg_score_order = request.GET.get('avg_score', '')
    graduation_year_filter = request.GET.get('graduation_year', '')
    conduct_filter = request.GET.get('conduct', '')
    from_date_filter = request.GET.get('from_date', '')
    to_date_filter = request.GET.get('to_date', '')

    if name_filter:
        admissions = admissions.filter(full_name__icontains=name_filter)

    if gender_filter:
        admissions = admissions.filter(gender=gender_filter)

    if status_filter:
        status_bool = status_filter == '1'
        admissions = admissions.filter(enable=status_bool)

    if campus_filter:
        admissions = admissions.filter(campus_id=campus_filter)

    if shift_filter:
        admissions = admissions.filter(shift_id=shift_filter)

    if subject_group_filter:
        admissions = admissions.filter(subject_group_id=subject_group_filter)

    if exam_score_order:
        order_by = '-exam_score' if exam_score_order == 'desc' else 'exam_score'
        admissions = admissions.order_by(order_by)

    if avg_score_order:
        order_by = '-avg_score' if avg_score_order == 'desc' else 'avg_score'
        admissions = admissions.order_by(order_by)

    if graduation_year_filter:
        admissions = admissions.filter(graduation_year=graduation_year_filter)

    if conduct_filter:
        admissions = admissions.filter(conduct=conduct_filter)
    if from_date_filter:
        admissions = admissions.filter(created_at__date__gte=from_date_filter)
    if to_date_filter:
        admissions = admissions.filter(created_at__date__lte=to_date_filter)

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Set title based on filters
    title = "Danh sách học sinh"
    if status_filter == '1':
        title += " đã duyệt"
    elif status_filter == '0':
        title += " chưa duyệt"
    else:
        title += " (tất cả)"

    ws.title = title

    # Define headers
    headers = [
        'Họ và tên', 'Giới tính', 'Ngày sinh', 'Dân tộc', 'Tôn giáo',
        'Email', 'CCCD/CMND', 'Ngày cấp', 'Nơi cấp',
        'Địa chỉ thường trú', 'Quê quán',
        'Nơi đăng ký khai sinh', 'Nơi sinh', 'Nơi ở hiện tại',
        'Họ tên cha', 'Nghề nghiệp cha', 'Năm sinh cha', 'SĐT cha',
        'Họ tên mẹ', 'Nghề nghiệp mẹ', 'Năm sinh mẹ', 'SĐT mẹ',
        'Trường đang học', 'Công việc hiện tại',
        'Điểm thi chuyển cấp', 'Điểm TB Toán, Văn',
        'Điểm Toán lớp 6', 'Điểm Văn lớp 6',
        'Điểm Toán lớp 7', 'Điểm Văn lớp 7',
        'Điểm Toán lớp 8', 'Điểm Văn lớp 8',
        'Điểm Toán lớp 9', 'Điểm Văn lớp 9',
        'Hạnh kiểm', 'Học lực', 'Năm tốt nghiệp',
        'Ban đăng ký', 'Ca học', 'Cơ sở'
    ]

    # Style the header row
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write data
    for row, admission in enumerate(admissions, 2):
        data = [
            admission.full_name,
            admission.gender,
            admission.birthday.strftime('%d/%m/%Y') if admission.birthday else '',
            admission.ethnicity,
            admission.religion,
            admission.email,
            admission.id_number,
            admission.id_issued_date.strftime('%d/%m/%Y') if admission.id_issued_date else '',
            admission.id_issued_place,
            # Địa chỉ thường trú
            f"{admission.cccd_town or ''}, "
            f"{get_ward_name(str(admission.cccd_province) if admission.cccd_province else '', str(admission.cccd_district) if admission.cccd_district else '', str(admission.cccd_ward) if admission.cccd_ward else '')}, "
            f"{get_district_name(str(admission.cccd_province) if admission.cccd_province else '', str(admission.cccd_district) if admission.cccd_district else '')}, "
            f"{get_province_name(str(admission.cccd_province) if admission.cccd_province else '')}",
            # Quê quán
            get_province_name(str(admission.hometown_province) if admission.hometown_province else ''),
            # Nơi đăng ký khai sinh
            f"{admission.birth_reg_town or ''}, "
            f"{get_ward_name(str(admission.birth_reg_province) if admission.birth_reg_province else '', str(admission.birth_reg_district) if admission.birth_reg_district else '', str(admission.birth_reg_ward) if admission.birth_reg_ward else '')}, "
            f"{get_district_name(str(admission.birth_reg_province) if admission.birth_reg_province else '', str(admission.birth_reg_district) if admission.birth_reg_district else '')}, "
            f"{get_province_name(str(admission.birth_reg_province) if admission.birth_reg_province else '')}",
            # Nơi sinh
            f"{admission.birth_place_facility or ''}, "
            f"{get_ward_name(str(admission.birth_place_province) if admission.birth_place_province else '', str(admission.birth_place_district) if admission.birth_place_district else '', str(admission.birth_place_ward) if admission.birth_place_ward else '')}, "
            f"{get_district_name(str(admission.birth_place_province) if admission.birth_place_province else '', str(admission.birth_place_district) if admission.birth_place_district else '')}, "
            f"{get_province_name(str(admission.birth_place_province) if admission.birth_place_province else '')}",
            # Nơi ở hiện tại
            f"{get_ward_name(str(admission.current_province) if admission.current_province else '', str(admission.current_district) if admission.current_district else '', str(admission.current_ward) if admission.current_ward else '')}, "
            f"{get_district_name(str(admission.current_province) if admission.current_province else '', str(admission.current_district) if admission.current_district else '')}, "
            f"{get_province_name(str(admission.current_province) if admission.current_province else '')}",
            # Thông tin cha
            admission.father_name,
            admission.father_job,
            admission.father_birth,
            admission.father_phone,
            # Thông tin mẹ
            admission.mother_name,
            admission.mother_job,
            admission.mother_birth,
            admission.mother_phone,
            # Thông tin học vấn
            admission.graduation_school,
            admission.current_job,
            str(admission.exam_score) if admission.exam_score else '',
            str(admission.avg_score) if admission.avg_score else '',
            str(admission.math_score_6) if admission.math_score_6 else '',
            str(admission.literature_score_6) if admission.literature_score_6 else '',
            str(admission.math_score_7) if admission.math_score_7 else '',
            str(admission.literature_score_7) if admission.literature_score_7 else '',
            str(admission.math_score_8) if admission.math_score_8 else '',
            str(admission.literature_score_8) if admission.literature_score_8 else '',
            str(admission.math_score_9) if admission.math_score_9 else '',
            str(admission.literature_score_9) if admission.literature_score_9 else '',
            admission.conduct,
            admission.graduation_rank,
            admission.graduation_year,
            admission.subject_group.code if admission.subject_group else '',
            admission.shift.name if admission.shift else '',
            admission.campus.name if admission.campus else ''  # Chỉ lấy tên của campus
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 30)

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=danh_sach_hoc_sinh_{datetime.now().strftime("%d%m%Y")}.xlsx'

    # Save the workbook to the response
    wb.save(response)
    return response

def delete_admission(request, admission_id):
    if request.method == 'POST' and request.user.is_authenticated:
        admission = get_object_or_404(AdmissionForm, id=admission_id)
        admission.delete()
        return redirect('adminpage:admission')
    return redirect('adminpage:admission')

@csrf_exempt
def import_cccd_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        cccd_col = None
        start_row = None
        # Tìm dòng tiêu đề chứa 'số CCCD' (không phân biệt hoa thường)
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            for j, cell in enumerate(row):
                if cell and str(cell).strip().lower() in ['số cccd', 'so cccd', 'cccd', 'h2 (số cccd)']:
                    cccd_col = j
                    start_row = i + 1
                    break
            if cccd_col is not None:
                break
        if cccd_col is None:
            return render(request, 'adminpageSIMCODE/admission.html', {'import_result': 'Không tìm thấy cột Số CCCD trong file Excel.'})
        # Lấy danh sách số CCCD từ dòng start_row đến hết
        cccd_list = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            cccd = row[cccd_col]
            if cccd:
                cccd_list.append(str(cccd).strip())
        # Duyệt học viên và gửi email
        from homepage.models import AdmissionForm
        approved_count = 0
        email_count = 0
        for cccd in cccd_list:
            qs = AdmissionForm.objects.filter(id_number=cccd, enable=False)
            for admission in qs:
                admission.enable = True
                admission.save()
                approved_count += 1
                if admission.email:
                    try:
                        html_message = render_to_string('adminpageSIMCODE/email_template.html', {
                            'full_name': admission.full_name,
                            'graduation_year': admission.graduation_year,
                            'campus': admission.campus.name,
                            'shift': admission.shift.name,
                            'subject_group': admission.subject_group.code
                        })
                        plain_message = strip_tags(html_message)
                        send_mail(
                            subject='Thông báo xét duyệt hồ sơ nhập học',
                            message=plain_message,
                            from_email=settings.EMAIL_HOST_USER,
                            recipient_list=[admission.email],
                            html_message=html_message,
                            fail_silently=False,
                        )
                        email_count += 1
                    except Exception as e:
                        print(f"Error sending email to {admission.email}: {e}")
        result_msg = f'Đã duyệt {approved_count} học viên theo danh sách CCCD. Đã gửi email cho {email_count} học viên.'
        return render(request, 'adminpageSIMCODE/admission.html', {'import_result': result_msg})
    return render(request, 'adminpageSIMCODE/admission.html', {'import_result': 'Vui lòng chọn file Excel hợp lệ.'})

@csrf_exempt
def delete_cccd_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        cccd_col = None
        start_row = None
        # Tìm dòng tiêu đề chứa 'số CCCD' (không phân biệt hoa thường)
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            for j, cell in enumerate(row):
                if cell and str(cell).strip().lower() in ['số cccd', 'so cccd', 'cccd', 'h2 (số cccd)']:
                    cccd_col = j
                    start_row = i + 1
                    break
            if cccd_col is not None:
                break
        if cccd_col is None:
            return render(request, 'adminpageSIMCODE/admission.html', {'delete_result': 'Không tìm thấy cột Số CCCD trong file Excel.'})
        # Lấy danh sách số CCCD từ dòng start_row đến hết
        cccd_list = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            cccd = row[cccd_col]
            if cccd:
                cccd_list.append(str(cccd).strip())
        # Xoá học viên
        from homepage.models import AdmissionForm
        deleted_count, _ = AdmissionForm.objects.filter(id_number__in=cccd_list).delete()
        result_msg = f'Đã xoá {deleted_count} biểu mẫu theo danh sách CCCD.'
        return render(request, 'adminpageSIMCODE/admission.html', {'delete_result': result_msg})
    return render(request, 'adminpageSIMCODE/admission.html', {'delete_result': 'Vui lòng chọn file Excel hợp lệ.'})

def export_cccd_for_delete(request):
    from homepage.models import AdmissionForm
    admissions = AdmissionForm.objects.filter(enable=False)
    wb = Workbook()
    ws = wb.active
    ws.title = "CCCD Chua Duyet"
    ws.append(["Số CCCD"])
    for admission in admissions:
        ws.append([admission.id_number])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cccd_chua_duyet.xlsx'
    wb.save(response)
    return response

@csrf_exempt
def update_conduct_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        cccd_col = None
        conduct_col = None
        start_row = None

        # Tìm dòng tiêu đề chứa 'số CCCD' và 'hạnh kiểm'
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            for j, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).strip().lower()
                    if cell_str in ['số cccd', 'so cccd', 'cccd', 'h2 (số cccd)']:
                        cccd_col = j
                    elif cell_str in ['hạnh kiểm', 'hanh kiem', 'conduct', 'hạnh kiểm thpt']:
                        conduct_col = j
            if cccd_col is not None and conduct_col is not None:
                start_row = i + 1
                break

        if cccd_col is None:
            return redirect('adminpage:admission')

        if conduct_col is None:
            return redirect('adminpage:admission')

        # Lấy danh sách CCCD và hạnh kiểm từ dòng start_row đến hết
        update_data = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            cccd = row[cccd_col]
            conduct = row[conduct_col]
            if cccd and conduct:
                update_data.append({
                    'cccd': str(cccd).strip(),
                    'conduct': str(conduct).strip()
                })

        # Cập nhật hạnh kiểm cho học viên
        from homepage.models import AdmissionForm
        updated_count = 0
        not_found_count = 0
        invalid_conduct_count = 0

        valid_conduct_values = ['Tốt', 'Khá', 'Trung bình', 'Yếu']

        for data in update_data:
            cccd = data['cccd']
            conduct = data['conduct']

            # Kiểm tra hạnh kiểm có hợp lệ không
            if conduct not in valid_conduct_values:
                invalid_conduct_count += 1
                continue

            # Tìm học viên theo CCCD
            try:
                admission = AdmissionForm.objects.get(id_number=cccd)
                admission.conduct = conduct
                admission.save()
                updated_count += 1
            except AdmissionForm.DoesNotExist:
                not_found_count += 1

        # Tạo thông báo kết quả
        result_msg = f'Đã cập nhật hạnh kiểm cho {updated_count} học viên.'
        if not_found_count > 0:
            result_msg += f' Không tìm thấy {not_found_count} học viên.'
        if invalid_conduct_count > 0:
            result_msg += f' {invalid_conduct_count} giá trị hạnh kiểm không hợp lệ.'

        # Lưu kết quả vào session để hiển thị
        request.session['conduct_update_result'] = result_msg
        return redirect('adminpage:admission')

    return redirect('adminpage:admission')

def export_conduct_template(request):
    """Export template Excel cho việc cập nhật hạnh kiểm"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Cập nhật Hạnh kiểm"

    # Thêm tiêu đề
    ws.append(["Số CCCD", "Hạnh kiểm"])
    ws.append(["123456789012", "Tốt"])
    ws.append(["987654321098", "Khá"])
    ws.append(["111222333444", "Trung bình"])

    # Định dạng tiêu đề
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=template_cap_nhat_hanh_kiem.xlsx'
    wb.save(response)
    return response

def get_ExamRegistration(request):
    """Quản lý đăng ký thi tốt nghiệp của học viên"""
    if request.user.is_authenticated:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            # Lấy tất cả đăng ký
            registration_list = StudentExamRegistration.objects.select_related('student', 'student__campus', 'student__subject_group').all()

            # Áp dụng các bộ lọc
            student_code_filter = request.GET.get('student_code', '')
            student_name_filter = request.GET.get('student_name', '')
            campus_filter = request.GET.get('campus', '')
            subject_group_filter = request.GET.get('subject_group', '')
            email_filter = request.GET.get('email', '')
            phone_filter = request.GET.get('phone', '')
            class_name_filter = request.GET.get('class_name', '')
            exam_subjects_filter = request.GET.get('exam_subjects', '')  # 'yes' hoặc 'no'

            if student_code_filter:
                registration_list = registration_list.filter(student__student_code__icontains=student_code_filter)

            if student_name_filter:
                registration_list = registration_list.filter(student__full_name__icontains=student_name_filter)

            if campus_filter:
                registration_list = registration_list.filter(student__campus_id=campus_filter)

            if subject_group_filter:
                registration_list = registration_list.filter(student__subject_group_id=subject_group_filter)

            if class_name_filter:
                registration_list = registration_list.filter(student__class_name__icontains=class_name_filter)

            if email_filter:
                registration_list = registration_list.filter(email__icontains=email_filter)

            if phone_filter:
                registration_list = registration_list.filter(phone__icontains=phone_filter)

            # Lọc theo trạng thái đã/chưa chọn môn thi
            # Với JSONField, cần xử lý đặc biệt
            if exam_subjects_filter == 'yes':
                # Đã chọn môn thi (exam_subjects không rỗng)
                # Lọc trong Python sau khi query để xử lý JSONField
                registration_ids = []
                for reg in registration_list:
                    if reg.exam_subjects and len(reg.exam_subjects) > 0:
                        registration_ids.append(reg.id)
                if registration_ids:
                    registration_list = registration_list.filter(id__in=registration_ids)
                else:
                    registration_list = registration_list.none()  # Không có kết quả
            elif exam_subjects_filter == 'no':
                # Chưa chọn môn thi (exam_subjects rỗng hoặc None)
                # Lọc trong Python sau khi query
                registration_ids = []
                for reg in registration_list:
                    if not reg.exam_subjects or len(reg.exam_subjects) == 0:
                        registration_ids.append(reg.id)
                if registration_ids:
                    registration_list = registration_list.filter(id__in=registration_ids)
                else:
                    registration_list = registration_list.none()  # Không có kết quả

            # Sắp xếp theo tên lớp tăng dần (natural sort), sau đó theo ngày đăng ký mới nhất
            # Chuyển QuerySet thành list để sắp xếp tự nhiên
            registration_list = list(registration_list)
            # Sắp xếp tự nhiên theo tên lớp, sau đó theo ngày đăng ký
            registration_list.sort(key=lambda x: (
                natural_sort_key(x.student.class_name if x.student.class_name else ''),
                -(x.created_at.timestamp() if x.created_at else 0)
            ))

            # Phân trang
            paginator = Paginator(registration_list, 15)  # Hiển thị 15 bản ghi mỗi trang
            page = request.GET.get('page', 1)

            try:
                registrations = paginator.page(page)
            except PageNotAnInteger:
                registrations = paginator.page(1)
            except EmptyPage:
                registrations = paginator.page(paginator.num_pages)

            # Lấy dữ liệu cho dropdown filters
            campuses = Campus.objects.all()
            subject_groups = SubjectGroup.objects.all()

            # Thống kê
            total_registrations = StudentExamRegistration.objects.count()
            today_registrations = StudentExamRegistration.objects.filter(created_at__date=timezone.now().date()).count()

            context = {
                'registrations': registrations,
                'total_registrations': total_registrations,
                'today_registrations': today_registrations,
                # Filter values
                'student_code_filter': student_code_filter,
                'student_name_filter': student_name_filter,
                'campus_filter': campus_filter,
                'subject_group_filter': subject_group_filter,
                'class_name_filter': class_name_filter,
                'exam_subjects_filter': exam_subjects_filter,
                'email_filter': email_filter,
                'phone_filter': phone_filter,
                # Dropdown data
                'campuses': campuses,
                'subject_groups': subject_groups,
            }
            return render(request, 'adminpageSIMCODE/exam_registration.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:login')

def view_registration_history(request, registration_id):
    """Xem lịch sử cập nhật của một hồ sơ đăng ký"""
    if request.user.is_authenticated:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            try:
                registration = StudentExamRegistration.objects.select_related('student', 'student__campus', 'student__subject_group').get(pk=registration_id)
                # Lấy tất cả lịch sử cập nhật, sắp xếp theo thời gian mới nhất
                history_list = RegistrationHistory.objects.filter(registration=registration).order_by('-created_at')

                context = {
                    'registration': registration,
                    'history_list': history_list,
                }
                return render(request, 'adminpageSIMCODE/registration_history.html', context)
            except StudentExamRegistration.DoesNotExist:
                messages.error(request, 'Không tìm thấy hồ sơ đăng ký.')
                return redirect('adminpage:exam_registration')
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:login')

def export_exam_registrations(request):
    """Xuất danh sách đăng ký thi tốt nghiệp ra Excel theo bộ lọc"""
    if request.user.is_authenticated:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            # Lấy danh sách đăng ký với các bộ lọc (giống như trong get_ExamRegistration)
            registration_list = StudentExamRegistration.objects.select_related('student', 'student__campus', 'student__subject_group').all()

            # Áp dụng các bộ lọc từ GET parameters
            student_code_filter = request.GET.get('student_code', '')
            student_name_filter = request.GET.get('student_name', '')
            campus_filter = request.GET.get('campus', '')
            subject_group_filter = request.GET.get('subject_group', '')
            class_name_filter = request.GET.get('class_name', '')
            exam_subjects_filter = request.GET.get('exam_subjects', '')
            email_filter = request.GET.get('email', '')
            phone_filter = request.GET.get('phone', '')

            if student_code_filter:
                registration_list = registration_list.filter(student__student_code__icontains=student_code_filter)

            if student_name_filter:
                registration_list = registration_list.filter(student__full_name__icontains=student_name_filter)

            if campus_filter:
                registration_list = registration_list.filter(student__campus_id=campus_filter)

            if subject_group_filter:
                registration_list = registration_list.filter(student__subject_group_id=subject_group_filter)

            if class_name_filter:
                registration_list = registration_list.filter(student__class_name__icontains=class_name_filter)

            if email_filter:
                registration_list = registration_list.filter(email__icontains=email_filter)

            if phone_filter:
                registration_list = registration_list.filter(phone__icontains=phone_filter)

            # Lọc theo trạng thái đã/chưa chọn môn thi
            if exam_subjects_filter == 'yes':
                # Đã chọn môn thi
                registration_ids = []
                for reg in registration_list:
                    if reg.exam_subjects and len(reg.exam_subjects) > 0:
                        registration_ids.append(reg.id)
                if registration_ids:
                    registration_list = registration_list.filter(id__in=registration_ids)
                else:
                    registration_list = registration_list.none()
            elif exam_subjects_filter == 'no':
                # Chưa chọn môn thi
                registration_ids = []
                for reg in registration_list:
                    if not reg.exam_subjects or len(reg.exam_subjects) == 0:
                        registration_ids.append(reg.id)
                if registration_ids:
                    registration_list = registration_list.filter(id__in=registration_ids)
                else:
                    registration_list = registration_list.none()

            # Sắp xếp theo tên lớp tăng dần (natural sort), sau đó theo ngày đăng ký mới nhất
            # Chuyển QuerySet thành list để sắp xếp tự nhiên
            registration_list = list(registration_list)
            # Sắp xếp tự nhiên theo tên lớp, sau đó theo ngày đăng ký
            registration_list.sort(key=lambda x: (
                natural_sort_key(x.student.class_name if x.student.class_name else ''),
                -(x.created_at.timestamp() if x.created_at else 0)
            ))
            registrations = registration_list

            # Tạo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Đăng ký thi tốt nghiệp"

            # Header style
            header_fill = PatternFill(start_color="023EB6", end_color="022468", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Headers
            headers = [
                'STT', 'Mã học viên', 'Họ và tên', 'Lớp', 'Cơ sở', 'Tổ hợp môn',
                'Email', 'Số điện thoại', 'Môn thi 1', 'Môn thi 2', 'Trạng thái chọn môn thi', 'Ngày đăng ký'
            ]
            ws.append(headers)

            # Apply header style
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Data
            for idx, reg in enumerate(registrations, start=1):
                student = reg.student

                # Tách môn thi thành 2 cột riêng
                exam_subject_1 = ''
                exam_subject_2 = ''
                if reg.exam_subjects and isinstance(reg.exam_subjects, list) and len(reg.exam_subjects) > 0:
                    exam_subject_1 = reg.exam_subjects[0] if len(reg.exam_subjects) > 0 else ''
                    exam_subject_2 = reg.exam_subjects[1] if len(reg.exam_subjects) > 1 else ''

                # Xác định trạng thái chọn môn thi
                if reg.exam_subjects and len(reg.exam_subjects) > 0:
                    exam_status = 'Đã chọn'
                else:
                    exam_status = 'Chưa chọn'

                # Ngày đăng ký (ưu tiên registration_date, nếu không có thì dùng created_at)
                registration_date = reg.registration_date if reg.registration_date else reg.created_at

                row = [
                    idx,
                    student.student_code,
                    student.full_name,
                    student.class_name,
                    student.campus.name if student.campus else '',
                    f'Tổ hợp {student.subject_group.code}' if student.subject_group else '',
                    reg.email or '',
                    reg.phone or '',
                    exam_subject_1,
                    exam_subject_2,
                    exam_status,
                    registration_date.strftime('%d/%m/%Y %H:%M') if registration_date else ''
                ]
                ws.append(row)

            # Auto adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'dang_ky_thi_tot_nghiep_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:login')

def import_students_excel(request):
    """Import danh sách học viên từ Excel và hiển thị preview"""
    if request.user.is_authenticated:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST' and request.FILES.get('excel_file'):
                excel_file = request.FILES['excel_file']
                try:
                    wb = load_workbook(excel_file, read_only=True)
                    ws = wb.active

                    # Tìm các cột trong header
                    header_row = None
                    col_mapping = {}

                    # Tìm dòng header
                    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                        row_values = [str(cell).strip().lower() if cell else '' for cell in row]
                        if any(keyword in ' '.join(row_values) for keyword in ['mã hv', 'mã học viên', 'student code', 'ma hv']):
                            header_row = i
                            # Map các cột
                            for j, cell in enumerate(row):
                                if cell:
                                    cell_str = str(cell).strip().lower()
                                    if 'mã hv' in cell_str or 'mã học viên' in cell_str or 'student code' in cell_str:
                                        col_mapping['student_code'] = j
                                    elif 'lớp' in cell_str or 'class' in cell_str:
                                        col_mapping['class_name'] = j
                                    elif 'họ tên' in cell_str or 'họ và tên' in cell_str or 'full name' in cell_str:
                                        col_mapping['full_name'] = j
                                    elif 'ngày sinh' in cell_str or 'date of birth' in cell_str or 'birthday' in cell_str:
                                        col_mapping['birthday'] = j
                                    elif 'số định danh' in cell_str or 'cccd' in cell_str or 'cmnd' in cell_str or 'id number' in cell_str:
                                        col_mapping['id_number'] = j
                                    elif 'nơi sinh' in cell_str or 'birth place' in cell_str or 'tỉnh' in cell_str or 'thành phố' in cell_str:
                                        col_mapping['birth_place'] = j
                                    elif 'dân tộc' in cell_str or 'ethnicity' in cell_str:
                                        col_mapping['ethnicity'] = j
                                    elif 'giới tính' in cell_str or 'gender' in cell_str or 'sex' in cell_str:
                                        col_mapping['gender'] = j
                                    elif 'email' in cell_str:
                                        col_mapping['email'] = j
                                    elif 'số điện thoại' in cell_str or 'phone' in cell_str or 'điện thoại' in cell_str:
                                        col_mapping['phone'] = j
                                    elif 'môn thi' in cell_str or 'exam subject' in cell_str or 'môn' in cell_str:
                                        col_mapping['exam_subjects'] = j
                            break

                    if not header_row or 'student_code' not in col_mapping:
                        return render(request, 'adminpageSIMCODE/import_students.html', {
                            'error': 'Không tìm thấy cột "Mã HV" trong file Excel. Vui lòng kiểm tra lại file.'
                        })

                    # Đọc dữ liệu
                    students_data = []
                    errors = []

                    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                        # Bỏ qua dòng trống
                        if not any(row):
                            continue

                        student_code = str(row[col_mapping['student_code']]).strip() if col_mapping.get('student_code') is not None and row[col_mapping['student_code']] else None

                        if not student_code or len(student_code) != 7 or not student_code.isdigit():
                            errors.append(f"Dòng {i}: Mã học viên không hợp lệ ({student_code})")
                            continue

                        # Parse mã học viên 7 số:
                        # - 2 số đầu: cơ sở (10..17)
                        # - số thứ 3: buổi (0: sáng, 1: tối)
                        # - số thứ 4: tổ hợp môn (1..5)
                        # - 3 số cuối: STT
                        campus_code_num = student_code[:2]
                        shift_digit = student_code[2]
                        subject_group_digit = student_code[3]

                        # Mapping mã cơ sở
                        campus_code_map = {
                            '10': 'AT', '11': 'BS', '12': 'CS', '13': 'ĐS',
                            '14': 'CN', '15': 'VH', '16': 'HT', '17': 'KT'
                        }
                        campus_code = campus_code_map.get(campus_code_num)

                        if not campus_code:
                            errors.append(f"Dòng {i}: Không tìm thấy cơ sở cho mã {campus_code_num}")
                            continue

                        # Lấy campus và subject_group
                        campus = Campus.objects.filter(code=campus_code).first()
                        subject_group = SubjectGroup.objects.filter(code=subject_group_digit).first()

                        if not campus:
                            errors.append(f"Dòng {i}: Không tìm thấy cơ sở với mã {campus_code}")
                            continue

                        if not subject_group:
                            errors.append(f"Dòng {i}: Không tìm thấy tổ hợp môn với mã {subject_group_digit}")
                            continue

                        # Parse ngày sinh
                        birthday = None
                        birthday_display = ''
                        if col_mapping.get('birthday') is not None and row[col_mapping['birthday']]:
                            try:
                                birthday_value = row[col_mapping['birthday']]
                                # Nếu là datetime object từ Excel
                                if isinstance(birthday_value, datetime):
                                    birthday = birthday_value.date()
                                    birthday_display = birthday.strftime('%d/%m/%Y')
                                else:
                                    birthday_str = str(birthday_value).strip()
                                    # Hỗ trợ format DD/MM/YYYY
                                    if '/' in birthday_str:
                                        parts = birthday_str.split('/')
                                        if len(parts) == 3:
                                            from datetime import date
                                            birthday = date(int(parts[2]), int(parts[1]), int(parts[0]))
                                            birthday_display = birthday_str
                                    else:
                                        birthday_display = birthday_str
                            except Exception as e:
                                errors.append(f"Dòng {i}: Lỗi parse ngày sinh: {str(e)}")
                                birthday_display = str(row[col_mapping['birthday']])

                        # Lấy các trường khác
                        student_data = {
                            'row_number': i,
                            'student_code': student_code,
                            'class_name': str(row[col_mapping.get('class_name', 0)]).strip() if col_mapping.get('class_name') is not None and row[col_mapping.get('class_name', 0)] else '',
                            'full_name': str(row[col_mapping.get('full_name', 0)]).strip() if col_mapping.get('full_name') is not None and row[col_mapping.get('full_name', 0)] else '',
                            'birthday': birthday.strftime('%Y-%m-%d') if birthday else None,
                            'birthday_display': birthday_display,
                            'id_number': str(row[col_mapping.get('id_number', 0)]).strip() if col_mapping.get('id_number') is not None and row[col_mapping.get('id_number', 0)] else '',
                            'birth_place': str(row[col_mapping.get('birth_place', 0)]).strip() if col_mapping.get('birth_place') is not None and row[col_mapping.get('birth_place', 0)] else '',
                            'ethnicity': str(row[col_mapping.get('ethnicity', 0)]).strip() if col_mapping.get('ethnicity') is not None and row[col_mapping.get('ethnicity', 0)] else '',
                            'gender': str(row[col_mapping.get('gender', 0)]).strip() if col_mapping.get('gender') is not None and row[col_mapping.get('gender', 0)] else '',
                            'campus_id': campus.id,
                            'campus_name': campus.name,
                            'subject_group_id': subject_group.id,
                            'subject_group_code': subject_group.code,
                            # Thông tin đăng ký thi tốt nghiệp (nếu có)
                            'email': str(row[col_mapping.get('email', 0)]).strip() if col_mapping.get('email') is not None and row[col_mapping.get('email', 0)] else '',
                            'phone': str(row[col_mapping.get('phone', 0)]).strip() if col_mapping.get('phone') is not None and row[col_mapping.get('phone', 0)] else '',
                            'exam_subjects': str(row[col_mapping.get('exam_subjects', 0)]).strip() if col_mapping.get('exam_subjects') is not None and row[col_mapping.get('exam_subjects', 0)] else '',
                        }

                        students_data.append(student_data)

                    # Lưu vào session để import sau
                    request.session['students_import_data'] = students_data
                    request.session['students_import_errors'] = errors

                    context = {
                        'students_data': students_data,
                        'errors': errors,
                        'total_count': len(students_data),
                        'error_count': len(errors)
                    }
                    return render(request, 'adminpageSIMCODE/import_students.html', context)

                except Exception as e:
                    return render(request, 'adminpageSIMCODE/import_students.html', {
                        'error': f'Lỗi khi đọc file Excel: {str(e)}'
                    })

            # GET request - hiển thị form upload
            return render(request, 'adminpageSIMCODE/import_students.html')
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:login')

def save_imported_students(request):
    """Lưu danh sách học viên đã import vào database"""
    if request.user.is_authenticated:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            if request.method == 'POST':
                students_data = request.session.get('students_import_data', [])

                if not students_data:
                    return render(request, 'adminpageSIMCODE/import_students.html', {
                        'error': 'Không có dữ liệu để import. Vui lòng upload file Excel trước.'
                    })

                created_count = 0
                updated_count = 0
                error_count = 0
                errors = []

                for student_data in students_data:
                    try:
                        # Lấy campus và subject_group từ ID
                        campus_id = student_data.get('campus_id')
                        subject_group_id = student_data.get('subject_group_id')

                        if not campus_id or not subject_group_id:
                            error_count += 1
                            errors.append(f"Mã {student_data.get('student_code')}: Thiếu thông tin cơ sở hoặc tổ hợp môn")
                            continue

                        campus = Campus.objects.get(id=campus_id)
                        subject_group = SubjectGroup.objects.get(id=subject_group_id)

                        # Parse birthday
                        birthday = None
                        if student_data.get('birthday'):
                            from datetime import datetime
                            birthday = datetime.strptime(student_data['birthday'], '%Y-%m-%d').date()

                        # Tạo hoặc cập nhật Student
                        student, student_created = Student.objects.update_or_create(
                            student_code=student_data['student_code'],
                            defaults={
                                'campus': campus,
                                'subject_group': subject_group,
                                'class_name': student_data.get('class_name', ''),
                                'full_name': student_data.get('full_name', ''),
                                'birthday': birthday,
                                'id_number': student_data.get('id_number', ''),
                                'birth_place': student_data.get('birth_place', ''),
                                'ethnicity': student_data.get('ethnicity', ''),
                                'gender': student_data.get('gender', ''),
                            }
                        )

                        # Tạo hoặc cập nhật StudentExamRegistration nếu có email hoặc phone
                        email = student_data.get('email', '').strip()
                        phone = student_data.get('phone', '').strip()
                        exam_subjects_str = student_data.get('exam_subjects', '').strip()

                        # Luôn tạo StudentExamRegistration khi import (có thể để trống email/phone để học viên điền sau)
                        # Parse exam_subjects nếu có
                        exam_subjects_list = []
                        if exam_subjects_str:
                            # Hỗ trợ nhiều format: "Môn1, Môn2" hoặc "Môn1; Môn2"
                            exam_subjects_list = [s.strip() for s in exam_subjects_str.replace(';', ',').split(',') if s.strip()]

                        # Tạo hoặc cập nhật StudentExamRegistration
                        registration, reg_created = StudentExamRegistration.objects.update_or_create(
                            student=student,
                            defaults={
                                'email': email if email else None,
                                'phone': phone if phone else None,
                                'exam_subjects': exam_subjects_list if exam_subjects_list else [],
                            }
                        )

                        # Sync sang ExamRoomStudent (dùng riêng cho xếp phòng thi)
                        ExamRoomStudent.objects.update_or_create(
                            student_code=student.student_code,
                            defaults={
                                "campus": student.campus,
                                "subject_group": student.subject_group,
                                "class_name": student.class_name,
                                "full_name": student.full_name,
                            },
                        )

                        # Đếm cả Student và Registration
                        if student_created and reg_created:
                            created_count += 1
                        elif not student_created and not reg_created:
                            updated_count += 1
                        elif student_created:
                            created_count += 1
                        else:
                            updated_count += 1

                    except Exception as e:
                        error_count += 1
                        errors.append(f"Mã {student_data.get('student_code')}: {str(e)}")

                # Xóa session
                request.session.pop('students_import_data', None)
                request.session.pop('students_import_errors', None)

                # Sau khi import xong -> tự động trộn & đánh số báo danh cho toàn bộ ExamRoomStudent
                try:
                    _assign_exam_numbers_for_all_students()
                except Exception as e:
                    # Không chặn import nếu lỗi SBD, chỉ ghi log nhẹ
                    print("Assign exam numbers error:", e)

                context = {
                    'success': True,
                    'created_count': created_count,
                    'updated_count': updated_count,
                    'error_count': error_count,
                    'errors': errors[:20],  # Chỉ hiển thị 20 lỗi đầu tiên
                }
                return render(request, 'adminpageSIMCODE/import_students.html', context)

            return redirect('adminpage:import_students_excel')
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:login')


def exam_room_import_students(request):
    """
    Import danh sách học viên dùng RIÊNG cho xếp phòng thi.
    - Không tạo/đụng tới StudentExamRegistration.
    - Chỉ phục vụ ExamRoomStudent + đánh SBD.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active

            header_row = None
            col_mapping: dict[str, int] = {}

            # Tìm dòng header
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_values = [str(cell).strip().lower() if cell else '' for cell in row]
                if any(keyword in ' '.join(row_values) for keyword in ['mã hv', 'mã học viên', 'student code', 'ma hv']):
                    header_row = i
                    for j, cell in enumerate(row):
                        if not cell:
                            continue
                        cell_str = str(cell).strip().lower()
                        if 'mã hv' in cell_str or 'mã học viên' in cell_str or 'student code' in cell_str:
                            col_mapping['student_code'] = j
                        elif 'lớp' in cell_str or 'class' in cell_str:
                            col_mapping['class_name'] = j
                        elif 'họ tên' in cell_str or 'họ và tên' in cell_str or 'full name' in cell_str:
                            col_mapping['full_name'] = j
                        elif 'ngày sinh' in cell_str or 'date of birth' in cell_str or 'birthday' in cell_str:
                            col_mapping['birthday'] = j
                        elif 'số định danh' in cell_str or 'cccd' in cell_str or 'cmnd' in cell_str or 'id number' in cell_str:
                            col_mapping['id_number'] = j
                        elif 'nơi sinh' in cell_str or 'birth place' in cell_str or 'tỉnh' in cell_str or 'thành phố' in cell_str:
                            col_mapping['birth_place'] = j
                        elif 'dân tộc' in cell_str or 'ethnicity' in cell_str:
                            col_mapping['ethnicity'] = j
                        elif 'giới tính' in cell_str or 'gender' in cell_str or 'sex' in cell_str:
                            col_mapping['gender'] = j
                        elif 'integration' in cell_str or 'inclusion' in cell_str or (
                            ('nhập' in cell_str or 'nhap' in cell_str)
                            and any(h in cell_str for h in ('hòa', 'hoà', 'hoa'))
                        ):
                            col_mapping['integration'] = j
                    break

            if not header_row or 'student_code' not in col_mapping:
                return render(request, 'adminpageSIMCODE/exam_room_import_preview.html', {
                    'students_data': [],
                    'errors': ['Không tìm thấy cột "Mã HV" trong file Excel. Vui lòng kiểm tra lại file.'],
                    'total_count': 0,
                    'error_count': 1,
                })

            students_data: list[dict] = []
            errors: list[str] = []

            for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not any(row):
                    continue

                student_code = str(row[col_mapping['student_code']]).strip() if col_mapping.get('student_code') is not None and row[col_mapping['student_code']] else None
                if not student_code or len(student_code) != 7 or not student_code.isdigit():
                    errors.append(f"Dòng {i}: Mã học viên không hợp lệ ({student_code})")
                    continue

                campus_code_num = student_code[:2]
                subject_group_digit = student_code[3]

                campus_code_map = {
                    '10': 'AT', '11': 'BS', '12': 'CS', '13': 'ĐS',
                    '14': 'CN', '15': 'VH', '16': 'HT', '17': 'KT'
                }
                campus_code = campus_code_map.get(campus_code_num)
                if not campus_code:
                    errors.append(f"Dòng {i}: Không tìm thấy cơ sở cho mã {campus_code_num}")
                    continue

                campus = Campus.objects.filter(code=campus_code).first()
                subject_group = SubjectGroup.objects.filter(code=subject_group_digit).first()
                if not campus:
                    errors.append(f"Dòng {i}: Không tìm thấy cơ sở với mã {campus_code}")
                    continue
                if not subject_group:
                    errors.append(f"Dòng {i}: Không tìm thấy tổ hợp môn với mã {subject_group_digit}")
                    continue

                birthday = None
                birthday_display = ''
                if col_mapping.get('birthday') is not None and row[col_mapping['birthday']]:
                    try:
                        birthday_value = row[col_mapping['birthday']]
                        if isinstance(birthday_value, datetime):
                            birthday = birthday_value.date()
                            birthday_display = birthday.strftime('%d/%m/%Y')
                        else:
                            birthday_str = str(birthday_value).strip()
                            if '/' in birthday_str:
                                parts = birthday_str.split('/')
                                if len(parts) == 3:
                                    from datetime import date
                                    birthday = date(int(parts[2]), int(parts[1]), int(parts[0]))
                                    birthday_display = birthday_str
                            else:
                                birthday_display = birthday_str
                    except Exception as e:  # noqa: BLE001
                        errors.append(f"Dòng {i}: Lỗi parse ngày sinh: {str(e)}")
                        birthday_display = str(row[col_mapping['birthday']])

                integration_cell = None
                if col_mapping.get('integration') is not None:
                    integration_cell = row[col_mapping['integration']]
                is_integration = _exam_room_import_cell_is_integration(integration_cell)

                student_data = {
                    'row_number': i,
                    'student_code': student_code,
                    'class_name': str(row[col_mapping.get('class_name', 0)]).strip() if col_mapping.get('class_name') is not None and row[col_mapping.get('class_name', 0)] else '',
                    'full_name': str(row[col_mapping.get('full_name', 0)]).strip() if col_mapping.get('full_name') is not None and row[col_mapping.get('full_name', 0)] else '',
                    'birthday': birthday.strftime('%Y-%m-%d') if birthday else None,
                    'birthday_display': birthday_display,
                    'id_number': str(row[col_mapping.get('id_number', 0)]).strip() if col_mapping.get('id_number') is not None and row[col_mapping.get('id_number', 0)] else '',
                    'birth_place': str(row[col_mapping.get('birth_place', 0)]).strip() if col_mapping.get('birth_place') is not None and row[col_mapping.get('birth_place', 0)] else '',
                    'ethnicity': str(row[col_mapping.get('ethnicity', 0)]).strip() if col_mapping.get('ethnicity') is not None and row[col_mapping.get('ethnicity', 0)] else '',
                    'gender': str(row[col_mapping.get('gender', 0)]).strip() if col_mapping.get('gender') is not None and row[col_mapping.get('gender', 0)] else '',
                    'campus_id': campus.id,
                    'campus_name': campus.name,
                    'subject_group_id': subject_group.id,
                    'subject_group_code': subject_group.code,
                    'is_integration': is_integration,
                }
                students_data.append(student_data)

            request.session['exam_room_students_import_data'] = students_data
            request.session['exam_room_students_import_errors'] = errors

            context = {
                'students_data': students_data,
                'errors': errors,
                'total_count': len(students_data),
                'error_count': len(errors),
            }
            return render(request, 'adminpageSIMCODE/exam_room_import_preview.html', context)

        except Exception as e:  # noqa: BLE001
            return render(request, 'adminpageSIMCODE/exam_room_import_preview.html', {
                'students_data': [],
                'errors': [f'Lỗi khi đọc file Excel: {str(e)}'],
                'total_count': 0,
                'error_count': 1,
            })

    # GET hoặc không có file -> quay về trang sắp xếp phòng thi
    return redirect('adminpage:exam_room_sort_dashboard')


@transaction.atomic
def exam_room_save_imported_students(request):
    """
    Lưu danh sách học viên import cho xếp phòng thi:
    - Chỉ tạo/cập nhật ExamRoomStudent.
    - Không tạo/đụng StudentExamRegistration.
    - Sau cùng: đánh lại SBD cho toàn bộ ExamRoomStudent.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:exam_room_sort_dashboard')

    students_data = request.session.get('exam_room_students_import_data', [])
    errors_session = request.session.get('exam_room_students_import_errors', [])
    if not students_data:
        messages.error(request, 'Không có dữ liệu để import. Vui lòng upload file Excel trước.')
        return redirect('adminpage:exam_room_sort_dashboard')

    created_count = 0
    updated_count = 0
    error_count = len(errors_session)
    errors: list[str] = list(errors_session)

    for student_data in students_data:
        try:
            campus_id = student_data.get('campus_id')
            subject_group_id = student_data.get('subject_group_id')
            if not campus_id or not subject_group_id:
                error_count += 1
                errors.append(f"Mã {student_data.get('student_code')}: Thiếu thông tin cơ sở hoặc tổ hợp môn")
                continue

            campus = Campus.objects.get(id=campus_id)
            subject_group = SubjectGroup.objects.get(id=subject_group_id)

            exam_student, created = ExamRoomStudent.objects.update_or_create(
                student_code=student_data['student_code'],
                defaults={
                    'campus': campus,
                    'subject_group': subject_group,
                    'class_name': student_data.get('class_name', ''),
                    'full_name': student_data.get('full_name', ''),
                    'is_integration': bool(student_data.get('is_integration', False)),
                },
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        except Exception as e:  # noqa: BLE001
            error_count += 1
            errors.append(f"Mã {student_data.get('student_code')}: {str(e)}")

    # Xoá session tạm
    request.session.pop('exam_room_students_import_data', None)
    request.session.pop('exam_room_students_import_errors', None)

    # Đánh lại SBD cho toàn bộ ExamRoomStudent (an toàn, độc lập với đăng ký tốt nghiệp)
    try:
        _assign_exam_numbers_for_all_students()
    except Exception as e:  # noqa: BLE001
        messages.warning(request, f'Lỗi khi đánh số báo danh tự động: {str(e)}')

    messages.success(
        request,
        f'Import danh sách xếp phòng thi: tạo mới {created_count}, cập nhật {updated_count}, lỗi {error_count}.',
    )
    return redirect('adminpage:exam_room_sort_dashboard')


@transaction.atomic
def exam_room_toggle_integration(request, student_code):
    """Toggle trạng thái học viên hoà nhập / bình thường cho ExamRoomStudent."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:exam_room_sort_dashboard')

    try:
        student = ExamRoomStudent.objects.get(student_code=student_code)
    except ExamRoomStudent.DoesNotExist:
        messages.error(request, 'Không tìm thấy học viên.')
        return redirect('adminpage:exam_room_sort_dashboard')

    student.is_integration = not student.is_integration
    student.save(update_fields=['is_integration'])

    status_label = 'Học viên hoà nhập' if student.is_integration else 'Học viên bình thường'
    messages.success(request, f'Đã cập nhật trạng thái: {student.full_name} -> {status_label}.')
    return redirect('adminpage:exam_room_sort_dashboard')

def _resolve_campus_for_phongthi_import(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.isdigit():
        hv_prefix_to_campus = {
            '10': 'AT',
            '11': 'BS',
            '12': 'CS',
            '13': 'ĐS',
            '14': 'CN',
            '15': 'VH',
            '16': 'HT',
            '17': 'KT',
        }
        if len(s) == 2 and s in hv_prefix_to_campus:
            c = Campus.objects.filter(code__iexact=hv_prefix_to_campus[s]).first()
            if c:
                return c
        return Campus.objects.filter(id=int(s)).first()
    c = Campus.objects.filter(code__iexact=s).first()
    if c:
        return c
    return Campus.objects.filter(name__iexact=s).first()


def _parse_exam_shift_phongthi(val) -> str | None:
    if val is None:
        return None
    kn = _vn_normalize_no_diacritics(str(val).strip())
    if not kn:
        return None
    if kn in ('sang', 'ca sang'):
        return 'sang'
    if kn in ('chieu', 'ca chieu', 'chieu'):
        return 'chieu'
    if kn in ('toi', 'ca toi', 'toi'):
        return 'toi'
    if kn == 'ca':
        return None
    return None


def _parse_hv_buoi_phongthi(val) -> str | None:
    """Cột học viên buổi: 0 = sáng, 1 = tối (theo quy ước mã HV)."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        i = int(val)
        if i == 0:
            return 'sang'
        if i == 1:
            return 'toi'
        return None
    s = str(val).strip().lower()
    if s in ('0', 'sang', 'buoi sang', 'sáng'):
        return 'sang'
    if s in ('1', 'toi', 'buoi toi', 'tối'):
        return 'toi'
    try:
        f = float(s.replace(',', '.'))
        if int(f) == 0:
            return 'sang'
        if int(f) == 1:
            return 'toi'
    except ValueError:
        pass
    return None


def _parse_grade_cell_phongthi(val) -> str | None:
    if val is None or str(val).strip() == '':
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return str(int(val))
    s = str(val).strip()
    try:
        return str(int(float(s.replace(',', '.'))))
    except ValueError:
        m = re.match(r'^(\d{1,2})', s)
        return m.group(1) if m else None


def _parse_group_base_phongthi(val) -> str | None:
    if val is None or str(val).strip() == '':
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        d = int(val)
        if 1 <= d <= 5:
            return str(d)
        return None
    s = str(val).strip()
    for ch in reversed(s):
        if ch.isdigit() and ch in '12345':
            return ch
    return None


def _parse_int_cell_phongthi(val) -> int | None:
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return int(val)
    try:
        return int(float(str(val).strip().replace(',', '.')))
    except ValueError:
        return None


# Import phòng thi: 11 cột A→K đúng thứ tự như file người dùng; dòng 1 luôn là tiêu đề (bỏ qua).
# Cột / Hàng trong Excel = col_count (ghế mỗi hàng) / row_count (số hàng) trong ExamRoom.
PHONGTHI_COLUMN_ORDER: tuple[str, ...] = (
    'campus',
    'room_name',
    'col_count',
    'row_count',
    'shift',
    'hv_buoi',
    'grade',
    'group',
    'limit',
    'extra_subject',
    'integration_limit',
)

PHONGTHI_FORMAT_HINT = (
    'Dòng 1 bắt buộc là tiêu đề (hệ thống không đọc). Từ dòng 2: 11 cột A→K — '
    'Cơ sở · Tên phòng · Cột (số ghế mỗi hàng) · Hàng (số hàng) · Ca thi (sáng/chiều/tối) · '
    'Học viên buổi (0/1) · Khối · Tổ hợp môn (1–5) · Số lượng học viên · '
    'Môn thi (môn thứ tư) · Hoà nhập (số HV hòa nhập cần ưu tiên).'
)


def _phongthi_fixed_col_map() -> dict[str, int]:
    return {name: idx for idx, name in enumerate(PHONGTHI_COLUMN_ORDER)}


def _phongthi_sample_xlsx_path() -> Path:
    """File mẫu trong project: `Sắp xếp phòng thi/phongthi.xlsx` (cùng thư mục với BASE_DIR)."""
    return Path(settings.BASE_DIR) / 'Sắp xếp phòng thi' / 'phongthi.xlsx'


def _phongthi_count_available_students(
    campus,
    shift_exam: str,
    grade: str,
    group_base: str,
    hv_shift: str,
    *,
    integration_only: bool | None = None,
) -> int:
    """Số học viên có thể lấy (cùng logic lọc với _perform_exam_room_initial_assignment, chưa trừ sức chỗ phòng)."""
    assigned_ids = set(
        StudentExamAssignment.objects.filter(shift=shift_exam).values_list('student_id', flat=True)
    )
    expect_hv_digit = '0' if hv_shift == 'sang' else '1'
    campus_students = ExamRoomStudent.objects.filter(campus=campus).select_related('subject_group')
    n = 0
    for s in campus_students:
        if s.pk in assigned_ids:
            continue
        if _exam_hv_shift_digit_from_code(s.student_code or '') != expect_hv_digit:
            continue
        g = _extract_grade_from_class_name(s.class_name)
        if not g or str(g) != str(grade).strip():
            continue
        code = s.subject_group.code if s.subject_group else ''
        if not code or code[-1] != group_base:
            continue
        if integration_only is True and not bool(getattr(s, 'is_integration', False)):
            continue
        if integration_only is False and bool(getattr(s, 'is_integration', False)):
            continue
        n += 1
    return n


def _phongthi_evaluate_data_row(ri: int, row: tuple, col_map: dict[str, int]) -> tuple[dict | None, dict]:
    """Trả về (payload lưu session nếu dòng hợp lệ, dict hiển thị preview)."""

    def cell(key: str):
        j = col_map.get(key)
        if j is None or j >= len(row):
            return None
        return row[j]

    fixed_subjects = ['Toán', 'Văn', 'Sử']
    display: dict = {'row': ri, 'status': 'error', 'note': ''}

    campus = _resolve_campus_for_phongthi_import(cell('campus'))
    if not campus:
        display['room_name'] = str(cell('room_name') or '').strip() or '—'
        display['note'] = 'Không xác định cơ sở (mã AT/BS, 10–17, id…)'
        return None, display
    display['campus'] = campus.name

    room_name = str(cell('room_name') or '').strip()
    display['room_name'] = room_name or '—'
    if not room_name:
        display['note'] = 'Thiếu tên phòng'
        return None, display

    shift = _parse_exam_shift_phongthi(cell('shift'))
    if not shift or shift not in dict(StudentExamAssignment.SHIFT_CHOICES):
        display['note'] = 'Ca thi không hợp lệ (sang / chieu / toi)'
        return None, display
    display['shift'] = shift

    cols_val = _parse_int_cell_phongthi(cell('col_count'))
    rows_val = _parse_int_cell_phongthi(cell('row_count'))
    if not rows_val or not cols_val or rows_val <= 0 or cols_val <= 0:
        display['note'] = 'Cột / Hàng (số ghế mỗi hàng và số hàng) không hợp lệ'
        return None, display
    display['row_count'] = rows_val
    display['col_count'] = cols_val

    grade = _parse_grade_cell_phongthi(cell('grade'))
    group_base = _parse_group_base_phongthi(cell('group'))
    if not grade or not group_base:
        display['note'] = 'Khối hoặc tổ hợp môn không hợp lệ'
        return None, display
    display['grade'] = grade
    display['group'] = group_base

    hv_shift = _parse_hv_buoi_phongthi(cell('hv_buoi'))
    if not hv_shift:
        display['note'] = 'Học viên buổi: 0 = sáng, 1 = tối'
        return None, display
    display['hv_buoi'] = '0 — Sáng' if hv_shift == 'sang' else '1 — Tối'

    limit = _parse_int_cell_phongthi(cell('limit'))
    if not limit or limit <= 0:
        display['note'] = 'Số lượng không hợp lệ'
        return None, display
    display['limit'] = limit

    extra_subject = str(cell('extra_subject') or '').strip()
    if not extra_subject:
        display['note'] = 'Thiếu môn thi (cột môn thứ tư)'
        return None, display
    if extra_subject in fixed_subjects:
        display['note'] = 'Môn còn lại không được trùng Toán / Văn / Sử'
        return None, display
    display['extra_subject'] = extra_subject

    integration_limit = _parse_int_cell_phongthi(cell('integration_limit'))
    if integration_limit is None:
        integration_limit = 0
    if integration_limit < 0:
        display['note'] = 'Hoà nhập không hợp lệ (>= 0)'
        return None, display
    if integration_limit > limit:
        display['note'] = f'Hoà nhập ({integration_limit}) không được lớn hơn số lượng ({limit})'
        return None, display
    display['integration_limit'] = integration_limit

    room = ExamRoom.objects.filter(campus=campus, name=room_name).first()
    if room:
        cfg = ExamRoomShiftConfig.objects.filter(exam_room=room, shift=shift).first()
        if cfg and (cfg.row_count != rows_val or cfg.col_count != cols_val):
            display['note'] = (
                f'Phòng + ca đã tồn tại ({cfg.row_count}×{cfg.col_count}), khác file ({rows_val}×{cols_val})'
            )
            return None, display
        if StudentExamAssignment.objects.filter(exam_room=room, shift=shift).exists():
            display['note'] = 'Phòng này đã có học viên ở ca đó — xoá HV ca đó trước nếu cần import lại'
            return None, display
        remaining = max(
            _get_room_capacity_for_shift(room, shift)
            - StudentExamAssignment.objects.filter(exam_room=room, shift=shift).count(),
            0,
        )
        display['room_action'] = 'Phòng có sẵn'
    else:
        remaining = rows_val * cols_val
        display['room_action'] = 'Tạo phòng mới'

    cand = _phongthi_count_available_students(campus, shift, grade, group_base, hv_shift)
    cand_integration = _phongthi_count_available_students(
        campus,
        shift,
        grade,
        group_base,
        hv_shift,
        integration_only=True,
    )
    display['candidates'] = cand
    display['integration_candidates'] = cand_integration

    if cand < limit:
        display['note'] = f'Chỉ có {cand} HV phù hợp trong pool (cần {limit})'
        return None, display
    if cand_integration < integration_limit:
        display['note'] = f'Chỉ có {cand_integration} HV hoà nhập phù hợp (cần {integration_limit})'
        return None, display
    if remaining < limit:
        display['note'] = f'Phòng chỉ còn {remaining} chỗ trong ca (cần {limit})'
        return None, display

    display['status'] = 'ok'
    display['note'] = (
        f'Sẵn sàng: lấy {limit} HV (ưu tiên {integration_limit} hoà nhập; pool {cand} người, '
        f'hn {cand_integration} người)'
    )

    payload = {
        'row': ri,
        'room_name': room_name,
        'campus_id': campus.id,
        'shift': shift,
        'row_count': rows_val,
        'col_count': cols_val,
        'grade': grade,
        'group_base': group_base,
        'hv_shift': hv_shift,
        'limit': limit,
        'extra_subject': extra_subject,
        'integration_limit': integration_limit,
    }
    return payload, display


def exam_room_phongthi_template(request):
    """Ưu tiên file mẫu trong `Sắp xếp phòng thi/phongthi.xlsx`; không có thì tạo file mẫu tối thiểu."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    sample_path = _phongthi_sample_xlsx_path()
    if sample_path.is_file():
        return FileResponse(
            sample_path.open('rb'),
            as_attachment=True,
            filename='phongthi.xlsx',
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Phong thi'
    headers = [
        'Cơ sở',
        'Tên phòng',
        'Cột',
        'Hàng',
        'Ca thi',
        'Học viên buổi',
        'Khối',
        'Tổ hợp môn',
        'Số lượng học viên',
        'Môn thi',
        'Hoà nhập',
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    ws.append(['AT', '1A', 4, 7, 'sáng', 0, 12, 1, 26, 'Vật lý', 2])
    ws.append(['AT', '2A', 4, 7, 'tối', 1, 12, 1, 26, 'Hoá', 0])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="phongthi_mau.xlsx"'
    return resp


def exam_room_import_phongthi(request):
    """
    Upload Excel → trang xem trước (không ghi DB). Dữ liệu hợp lệ lưu session;
    bấm «Xác nhận import» ở preview mới thực hiện tạo phòng + lấy HV.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST' or not request.FILES.get('excel_file'):
        messages.error(request, 'Vui lòng chọn file Excel.')
        return redirect('adminpage:exam_room_manage')

    excel_file = request.FILES['excel_file']
    format_hint = PHONGTHI_FORMAT_HINT

    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        col_map = _phongthi_fixed_col_map()

        preview_rows: list[dict] = []
        apply_payloads: list[dict | None] = []
        seen_room_dims: dict[tuple[int, str, str], tuple[int, int, int, int]] = {}
        # (campus_id, room_name_lower, shift) -> (row_count, col_count, first_row_excel, first_payload_idx)

        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5000, values_only=True), 1):
            if ri == 1:
                continue
            if not row:
                continue
            if not any(x is not None and str(x).strip() != '' for x in row[:11]):
                continue
            payload, display = _phongthi_evaluate_data_row(ri, row, col_map)
            preview_rows.append(display)
            if payload:
                campus_id = int(payload['campus_id'])
                room_key = (campus_id, str(payload['room_name']).strip().lower(), str(payload['shift']))
                row_count = int(payload['row_count'])
                col_count = int(payload['col_count'])
                conflict = seen_room_dims.get(room_key)
                if conflict and (conflict[0], conflict[1]) != (row_count, col_count):
                    first_row_excel = conflict[2]
                    first_payload_idx = conflict[3]
                    display['status'] = 'error'
                    display['note'] = (
                        f'Phòng trùng tên trong cùng file nhưng khác kích thước với dòng {first_row_excel} '
                        f'({conflict[0]}×{conflict[1]} vs {row_count}×{col_count}).'
                    )
                    if 0 <= first_payload_idx < len(apply_payloads) and apply_payloads[first_payload_idx]:
                        apply_payloads[first_payload_idx] = None
                else:
                    apply_payloads.append(payload)
                    seen_room_dims[room_key] = (row_count, col_count, int(payload['row']), len(apply_payloads) - 1)

        request.session['exam_room_phongthi_apply_rows'] = [p for p in apply_payloads if p]
        ok_count = sum(1 for d in preview_rows if d.get('status') == 'ok')

        error_banner = None
        if not preview_rows:
            request.session.pop('exam_room_phongthi_apply_rows', None)
            error_banner = (
                'Không có dòng dữ liệu từ dòng 2 trở đi (dòng 1 là tiêu đề, không import). '
                + format_hint
            )

        return render(
            request,
            'adminpageSIMCODE/exam_room_phongthi_preview.html',
            {
                'preview_rows': preview_rows,
                'apply_count': len(apply_payloads),
                'total_data_rows': len(preview_rows),
                'ok_count': ok_count,
                'error_row_count': max(len(preview_rows) - ok_count, 0),
                'error_banner': error_banner,
                'format_hint': format_hint,
            },
        )

    except Exception as e:  # noqa: BLE001
        messages.error(request, f'Lỗi đọc file Excel: {e}')
        return redirect('adminpage:exam_room_manage')


def exam_room_save_phongthi_import(request):
    """Thực hiện import sau khi đã xem trước (dữ liệu trong session)."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:exam_room_manage')

    rows = request.session.pop('exam_room_phongthi_apply_rows', None)
    if not rows:
        messages.error(request, 'Không có dữ liệu import. Vui lòng upload file và xem trước lại.')
        return redirect('adminpage:exam_room_manage')

    ok_rows = 0
    log_lines: list[str] = []

    for payload in sorted(rows, key=lambda x: int(x['row'])):
        ri = int(payload['row'])
        try:
            with transaction.atomic():
                campus = Campus.objects.get(id=int(payload['campus_id']))
                room, created = ExamRoom.objects.get_or_create(
                    campus=campus,
                    name=(payload['room_name'] or '').strip(),
                    defaults={
                        'row_count': int(payload['row_count']),
                        'col_count': int(payload['col_count']),
                    },
                )
                shift = payload['shift']
                ExamRoomShiftConfig.objects.update_or_create(
                    exam_room=room,
                    shift=shift,
                    defaults={
                        'row_count': int(payload['row_count']),
                        'col_count': int(payload['col_count']),
                    },
                )
                if not created:
                    if StudentExamAssignment.objects.filter(exam_room=room, shift=shift).exists():
                        raise ValueError('Phòng + ca đã có học viên')

                n, err, picked_initial = _perform_exam_room_initial_assignment(
                    room,
                    shift,
                    grade=str(payload['grade']),
                    group_base=str(payload['group_base']),
                    hv_shift=str(payload['hv_shift']),
                    limit=int(payload['limit']),
                    extra_subject=str(payload['extra_subject']),
                    integration_quota=int(payload.get('integration_limit') or 0),
                )
                if err:
                    # Không rollback việc tạo phòng / cấu hình ca; chỉ ghi log lỗi lấy học viên.
                    log_lines.append(f'Dòng {ri}: Lỗi khi lấy học viên cho phòng {room.name} ca {shift}: {err}')
                else:
                    if n > 0:
                        _sync_exam_sbd_campus_after_room_assignment(campus)
                        log_lines.append(
                            f'Dòng {ri}: nếu có HV hòa nhập, đã đồng bộ SBD {campus.code} (ca → tổ hợp → phòng, ưu tiên tên).',
                        )

            ok_rows += 1
            log_lines.append(
                f'Dòng {ri}: OK — {room.name} ({campus.code}), ca {shift}, đã lấy {n} học viên.',
            )
        except Exception as e:  # noqa: BLE001
            log_lines.append(f'Dòng {ri}: {e}')

    if ok_rows:
        messages.success(request, f'Đã import thành công {ok_rows} dòng (tạo phòng + lấy học viên).')
    else:
        messages.warning(request, 'Không có dòng nào ghi được. Xem chi tiết bên dưới.')

    if log_lines:
        preview = '\n'.join(log_lines[:80])
        if len(log_lines) > 80:
            preview += f'\n… và {len(log_lines) - 80} dòng log khác.'
        messages.info(request, preview)

    return redirect('adminpage:exam_room_manage')

def _redirect_exam_room_manage_preserve_campus(request):
    """Quay lại trang quản lý phòng thi, giữ ?campus_id= nếu form gửi return_campus_id."""
    cid = (request.POST.get('return_campus_id') or '').strip()
    url = reverse('adminpage:exam_room_manage')
    if cid.isdigit():
        return redirect(f'{url}?campus_id={cid}')
    return redirect('adminpage:exam_room_manage')



def exam_room_manage(request):
    """Dashboard quản lý phòng thi: thêm phòng mới + liệt kê + xoá."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    filter_campus = (request.GET.get('campus_id') or '').strip()

    if request.method == 'POST':
        campus_id = request.POST.get('campus_id')
        name = (request.POST.get('name') or '').strip()
        row_count = request.POST.get('row_count')
        col_count = request.POST.get('col_count')

        if not campus_id or not name or not row_count or not col_count:
            messages.error(request, 'Vui lòng nhập đầy đủ thông tin phòng thi.')
        else:
            try:
                campus = Campus.objects.get(id=int(campus_id))
                rows_val = int(row_count)
                cols_val = int(col_count)
                if rows_val <= 0 or cols_val <= 0:
                    raise ValueError('Số hàng/cột phải > 0')
                ExamRoom.objects.create(
                    campus=campus,
                    name=name,
                    row_count=rows_val,
                    col_count=cols_val,
                )
                messages.success(request, f'Đã tạo phòng thi {name} tại {campus.name}.')
                base = reverse('adminpage:exam_room_manage')
                return redirect(f'{base}?campus_id={campus_id}')
            except Campus.DoesNotExist:
                messages.error(request, 'Cơ sở không hợp lệ.')
            except Exception as e:  # noqa: BLE001
                messages.error(request, f'Lỗi khi tạo phòng thi: {str(e)}')

    if request.method == 'POST':
        ret_keep = (request.POST.get('return_campus_id') or '').strip()
        if ret_keep.isdigit():
            filter_campus = ret_keep

    rooms = (
        ExamRoom.objects.select_related('campus')
        .annotate(
            hv_sang=django_models.Count(
                'studentexamassignment',
                filter=django_models.Q(studentexamassignment__shift='sang'),
            ),
            hv_chieu=django_models.Count(
                'studentexamassignment',
                filter=django_models.Q(studentexamassignment__shift='chieu'),
            ),
            hv_toi=django_models.Count(
                'studentexamassignment',
                filter=django_models.Q(studentexamassignment__shift='toi'),
            ),
        )
        .order_by('campus__code', 'name')
    )
    if filter_campus.isdigit() and Campus.objects.filter(id=int(filter_campus)).exists():
        rooms = rooms.filter(campus_id=int(filter_campus))
    elif filter_campus and filter_campus.isdigit():
        filter_campus = ''

    context = {
        'campuses': Campus.objects.all().order_by('code'),
        'rooms': rooms,
        'filter_campus': filter_campus,
    }
    return render(request, 'adminpageSIMCODE/exam_room_manage.html', context)

@transaction.atomic
def exam_room_delete(request, room_id: int):
    """Xoá 1 phòng thi.

    Nếu đã có dữ liệu xếp phòng thì:
    - Xoá toàn bộ StudentExamAssignment của phòng đó.
    - Xoá luôn các bản ghi ExamRoomSubject liên quan.
    Sau đó mới xoá phòng.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    try:
        room = ExamRoom.objects.get(id=room_id)
    except ExamRoom.DoesNotExist:
        messages.error(request, 'Không tìm thấy phòng thi.')
        return _redirect_exam_room_manage_preserve_campus(request)

    # Xoá toàn bộ dữ liệu xếp phòng và môn thi của phòng này trước
    StudentExamAssignment.objects.filter(exam_room=room).delete()
    try:
        from homepage.models import ExamRoomSubject, ExamSubjectSeat  # tránh vòng import nếu có
        ExamRoomSubject.objects.filter(exam_room=room).delete()
        ExamSubjectSeat.objects.filter(exam_room=room).delete()
    except Exception:
        # Nếu vì lý do nào đó không có model/không import được thì bỏ qua, không chặn xoá phòng
        pass

    room.delete()
    messages.success(request, f'Đã xoá phòng thi {room.name} và toàn bộ dữ liệu xếp phòng liên quan.')
    return _redirect_exam_room_manage_preserve_campus(request)

@transaction.atomic
def exam_room_update(request, room_id: int):
    """Sửa phòng thi (popup): cập nhật cơ sở/tên/kích thước."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return _redirect_exam_room_manage_preserve_campus(request)

    try:
        room = ExamRoom.objects.get(id=room_id)
    except ExamRoom.DoesNotExist:
        messages.error(request, 'Không tìm thấy phòng thi.')
        return _redirect_exam_room_manage_preserve_campus(request)

    # Nếu đã có xếp phòng, chặn sửa kích thước/cơ sở để tránh sai lệch
    if StudentExamAssignment.objects.filter(exam_room=room).exists():
        messages.warning(request, f'Không thể sửa phòng {room.name} vì đã có dữ liệu xếp phòng.')
        return _redirect_exam_room_manage_preserve_campus(request)

    campus_id = request.POST.get('campus_id')
    name = (request.POST.get('name') or '').strip()
    row_count = request.POST.get('row_count')
    col_count = request.POST.get('col_count')

    if not campus_id or not name or not row_count or not col_count:
        messages.error(request, 'Vui lòng nhập đầy đủ thông tin phòng thi.')
        return _redirect_exam_room_manage_preserve_campus(request)

    try:
        campus = Campus.objects.get(id=int(campus_id))
        rows_val = int(row_count)
        cols_val = int(col_count)
        if rows_val <= 0 or cols_val <= 0:
            raise ValueError('Số hàng/cột phải > 0')
        room.campus = campus
        room.name = name
        room.row_count = rows_val
        room.col_count = cols_val
        room.save()
        messages.success(request, f'Đã cập nhật phòng thi {room.name}.')
    except Campus.DoesNotExist:
        messages.error(request, 'Cơ sở không hợp lệ.')
    except Exception as e:  # noqa: BLE001
        messages.error(request, f'Lỗi khi cập nhật phòng thi: {str(e)}')

    return _redirect_exam_room_manage_preserve_campus(request)


def exam_room_detail(request, room_id: int):
    """Trang chi tiết phòng thi: chọn ca, môn + layout 2 cột, popup lấy/chọn thêm học viên."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    room = get_object_or_404(ExamRoom, id=room_id)

    def _get_last_param(req, key: str) -> str:
        """Lấy tham số cuối cùng (tránh trường hợp URL có key lặp: shift=sang&shift=chieu)."""
        vals = []
        try:
            vals = req.GET.getlist(key)
        except Exception:
            vals = []
        if not vals:
            try:
                vals = req.POST.getlist(key)
            except Exception:
                v = req.POST.get(key) if hasattr(req.POST, 'get') else ''
                vals = [v] if v else []
        vals = [v for v in vals if v]
        return (vals[-1] if vals else '').strip()

    # Ca thi ưu tiên lấy từ query/form (lấy giá trị cuối nếu bị lặp)
    shift = _get_last_param(request, 'shift') or 'sang'
    # Tên môn hiện tại (nếu có trên URL / form)
    subject = _get_last_param(request, 'subject')

    def _erd_redirect_url(subj=None):
        """Quay lại chi tiết phòng, giữ shift; giữ subject nếu còn (tránh mất môn sau POST)."""
        from urllib.parse import urlencode

        qs = {'shift': shift}
        subj_val = subj if subj is not None else subject
        if subj_val:
            qs['subject'] = subj_val
        return f"{reverse('adminpage:exam_room_detail', args=[room.id])}?{urlencode(qs)}"

    # --------- Xử lý POST: Lấy học viên / Chọn thêm học viên ----------
    if request.method == 'POST':
        # Hỗ trợ cả form POST thường (application/x-www-form-urlencoded) và AJAX JSON
        if request.headers.get('Content-Type', '').startswith('application/json'):
            try:
                payload = json.loads(request.body.decode('utf-8'))
            except Exception:
                payload = {}
            action = payload.get('action')
        else:
            payload = request.POST
            action = payload.get('action')

        # Luôn ưu tiên shift trong payload (đặc biệt khi POST/AJAX không có query ?shift=...)
        payload_shift = (payload.get('shift') or '').strip() if hasattr(payload, 'get') else ''
        if payload_shift in dict(StudentExamAssignment.SHIFT_CHOICES):
            shift = payload_shift

        # Sức chứa còn lại trong ca
        current_count = StudentExamAssignment.objects.filter(exam_room=room, shift=shift).count()
        shift_capacity = _get_room_capacity_for_shift(room, shift)
        remaining_capacity = max(shift_capacity - current_count, 0)

        # Pool student_ids đã có trong ca (bất kể phòng)
        assigned_ids = set(
            StudentExamAssignment.objects.filter(shift=shift).values_list('student_id', flat=True)
        )

        if action == 'initial_pick':
            grade = (payload.get('grade') or '').strip()
            group_base = (payload.get('subject_group') or '').strip()
            hv_shift = (payload.get('hv_shift') or '').strip()
            try:
                limit = int(payload.get('limit') or '0')
            except ValueError:
                limit = 0
            extra_subject = (payload.get('extra_subject') or '').strip()

            n, err, picked_initial = _perform_exam_room_initial_assignment(
                room,
                shift,
                grade=grade,
                group_base=group_base,
                hv_shift=hv_shift,
                limit=limit,
                extra_subject=extra_subject,
            )
            if err:
                if 'đủ' in err and 'chỗ' in err:
                    messages.warning(request, err)
                elif 'Không có học viên phù hợp' in err:
                    messages.warning(request, err)
                else:
                    messages.error(request, err)
                return redirect(_erd_redirect_url())

            if n > 0:
                _sync_exam_sbd_campus_after_room_assignment(room.campus)
            extra_sbd = ''
            if n > 0:
                extra_sbd = (
                    ' Nếu cơ sở có học viên hòa nhập, SBD đã được đánh lại theo ca → tổ hợp môn → phòng (thứ tự phòng: ưu tiên tên trước, HV đại diện mỗi phòng).'
                )
            messages.success(
                request,
                f'Đã lấy {n} học viên và tự động xếp ghế cho phòng {room.name} (ca {shift}).' + extra_sbd,
            )
            return redirect(_erd_redirect_url())

        # Xoá toàn bộ học viên trong phòng + ca hiện tại
        if action == 'delete_all_in_room':
            StudentExamAssignment.objects.filter(exam_room=room, shift=shift).delete()
            # Xoá luôn danh sách môn gắn với phòng + ca này để chọn lại từ đầu nếu cần
            ExamRoomSubject.objects.filter(exam_room=room, shift=shift).delete()
            # Xoá luôn layout ghế theo môn cho phòng + ca này
            ExamSubjectSeat.objects.filter(exam_room=room, shift=shift).delete()
            _sync_exam_sbd_campus_after_room_assignment(room.campus, force=True)
            messages.success(
                request,
                f'Đã xoá toàn bộ học viên khỏi phòng {room.name} (ca {shift}). '
                'Đã đánh lại SBD toàn cơ sở theo ca → tổ hợp môn → phòng.',
            )
            return redirect(_erd_redirect_url())

        # Xoá 1 học viên khỏi phòng + ca hiện tại
        if action == 'delete_one_in_room':
            student_code = (payload.get('student_code') or '').strip()
            if not student_code:
                messages.error(request, 'Thiếu mã học viên cần xoá.')
                return redirect(_erd_redirect_url())

            deleted_count, _ = StudentExamAssignment.objects.filter(
                exam_room=room,
                shift=shift,
                student_id=student_code,
            ).delete()
            # Xoá luôn layout ghế theo môn của học viên này trong phòng + ca
            ExamSubjectSeat.objects.filter(
                exam_room=room,
                shift=shift,
                student_id=student_code,
            ).delete()
            # Sau khi xoá, tự động đánh lại seat_number + regenerate sơ đồ theo môn để khớp dữ liệu thực tế
            with transaction.atomic():
                autos = list(
                    StudentExamAssignment.objects.select_for_update()
                    .select_related('student')
                    .filter(exam_room=room, shift=shift)
                )
                autos.sort(key=_exam_room_seat_order_key)
                seat_no = 1
                for a in autos:
                    a.seat_number = seat_no
                    seat_no += 1
                if autos:
                    StudentExamAssignment.objects.bulk_update(autos, ['seat_number'])
            _regenerate_subject_layouts_for_room_shift(room, shift)
            if deleted_count:
                _sync_exam_sbd_campus_after_room_assignment(room.campus, force=True)
                messages.success(
                    request,
                    f'Đã xoá học viên {student_code} khỏi phòng {room.name} (ca {shift}). '
                    'Đã đánh lại SBD toàn cơ sở theo ca → tổ hợp môn → phòng.',
                )
            else:
                messages.warning(request, 'Không tìm thấy học viên trong phòng/ca này để xoá.')
            return redirect(_erd_redirect_url())

        if action == 'add_more':
            if remaining_capacity <= 0:
                messages.warning(request, f'Phòng {room.name} ca {shift} đã đủ {shift_capacity} chỗ.')
                return redirect(_erd_redirect_url())

            hv_shift = (payload.get('hv_shift') or '').strip()
            if hv_shift not in ('sang', 'toi'):
                messages.error(request, 'Vui lòng chọn học viên buổi sáng hay buổi tối trước khi thêm.')
                return redirect(_erd_redirect_url())
            expect_hv_digit = '0' if hv_shift == 'sang' else '1'

            # Lấy danh sách mã học viên được chọn từ modal "Chọn thêm học viên"
            if hasattr(payload, 'getlist'):
                codes = payload.getlist('student_codes')
            else:
                raw_codes = payload.get('student_codes') or []
                if isinstance(raw_codes, str):
                    codes = [raw_codes]
                else:
                    codes = list(raw_codes)

            codes = [c.strip() for c in codes if c.strip()]
            codes = [c for c in codes if _exam_hv_shift_digit_from_code(c) == expect_hv_digit]
            if not codes:
                messages.warning(
                    request,
                    'Không có học viên được chọn đúng buổi (sáng/tối theo mã HV), hoặc chưa chọn dòng nào.',
                )
                return redirect(_erd_redirect_url())

            codes = codes[:remaining_capacity]

            campus_students = ExamRoomStudent.objects.filter(student_code__in=codes).select_related('subject_group')
            picked: list[ExamRoomStudent] = []
            for s in campus_students:
                if s.pk in assigned_ids:
                    continue
                picked.append(s)

            if not picked:
                messages.warning(request, 'Các học viên chọn thêm đều đã có phòng trong ca này.')
                return redirect(_erd_redirect_url())

            objs = [
                StudentExamAssignment(
                    student=stu,
                    exam_room=room,
                    shift=shift,
                )
                for stu in picked
            ]
            StudentExamAssignment.objects.bulk_create(objs)

            _sync_exam_sbd_campus_after_room_assignment(room.campus, force=True)
            _exam_fill_missing_seats_in_room_shift(room, shift)
            _regenerate_subject_layouts_for_room_shift(room, shift)
            messages.success(
                request,
                f'Đã thêm {len(picked)} học viên vào phòng {room.name} (ca {shift}). '
                'Đã đánh lại SBD toàn cơ sở theo ca → tổ hợp môn → phòng (ưu tiên tên HV đại diện mỗi phòng).',
            )
            return redirect(_erd_redirect_url())

        # Tự động xếp ghế dồn từ trên xuống theo SBD hiện có trong phòng/ca này
        if action == 'auto_layout':
            with transaction.atomic():
                autos = list(
                    StudentExamAssignment.objects.select_for_update()
                    .select_related('student')
                    .filter(exam_room=room, shift=shift)
                )
                autos.sort(key=_exam_room_seat_order_key)
                seat_no = 1
                for a in autos:
                    a.seat_number = seat_no
                    seat_no += 1
                if autos:
                    StudentExamAssignment.objects.bulk_update(autos, ['seat_number'])

            # Sau khi auto layout chung, regenerate lại layout cho tất cả các môn
            _regenerate_subject_layouts_for_room_shift(room, shift)

            messages.success(
                request,
                f'Đã tự động xếp {len(autos)} học viên (thường trước, hòa nhập cuối; trong mỗi nhóm theo SBD) và cập nhật layout môn.',
            )
            return redirect(_erd_redirect_url())

        # Cập nhật ghế khi kéo thả (AJAX)
        if action == 'update_seat':
            student_code = (payload.get('student_code') or '').strip()
            try:
                seat = int(payload.get('seat') or '0')
            except ValueError:
                seat = 0

            # Môn thi đang thao tác (ưu tiên lấy từ payload, fallback về subject hiện tại của view)
            raw_subject = (payload.get('subject') or '').strip()
            subject_name = raw_subject or subject or ''

            if not student_code or seat <= 0 or seat > shift_capacity:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'ok': False, 'error': 'Dữ liệu không hợp lệ.'}, status=400)
                messages.error(request, 'Dữ liệu không hợp lệ khi đổi ghế.')
                return redirect(_erd_redirect_url())

            # Trường hợp 1: CHƯA chọn môn -> giữ nguyên hành vi cũ, cập nhật StudentExamAssignment (layout chung)
            if not subject_name:
                moved_to_other_room = False
                try:
                    with transaction.atomic():
                        assignment = (
                            StudentExamAssignment.objects.select_for_update()
                            .select_related('student')
                            .get(student_id=student_code, shift=shift)
                        )
                        old_room_id = assignment.exam_room_id
                        moved_to_other_room = old_room_id != room.id
                        assignment.exam_room = room
                        old_seat = assignment.seat_number

                        other = (
                            StudentExamAssignment.objects.select_for_update()
                            .filter(exam_room=room, shift=shift, seat_number=seat)
                            .exclude(student_id=student_code)
                            .first()
                        )

                        assignment.seat_number = seat
                        assignment.save()

                        if other:
                            other.seat_number = old_seat
                            other.save()
                except StudentExamAssignment.DoesNotExist:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'ok': False, 'error': 'Không tìm thấy học viên trong ca này.'}, status=404)
                    messages.error(request, 'Không tìm thấy học viên trong ca này.')
                    return redirect(_erd_redirect_url())

                if moved_to_other_room:
                    _sync_exam_sbd_campus_after_room_assignment(room.campus, force=True)

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'ok': True})

                if moved_to_other_room:
                    messages.success(
                        request,
                        'Đã chuyển sang phòng thi khác. '
                        'Đã đánh lại SBD toàn cơ sở theo ca → tổ hợp môn → phòng.',
                    )
                else:
                    messages.success(request, 'Đã cập nhật vị trí ghế.')
                return redirect(_erd_redirect_url())

            # Trường hợp 2: ĐÃ chọn môn -> chỉ cập nhật layout môn đó trong ExamSubjectSeat
            try:
                with transaction.atomic():
                    seat_entry = (
                        ExamSubjectSeat.objects.select_for_update()
                        .select_related('student')
                        .get(
                            exam_room=room,
                            shift=shift,
                            subject_name=subject_name,
                            student_id=student_code,
                        )
                    )
                    old_seat = seat_entry.seat_number

                    other = (
                        ExamSubjectSeat.objects.select_for_update()
                        .filter(
                            exam_room=room,
                            shift=shift,
                            subject_name=subject_name,
                            seat_number=seat,
                        )
                        .exclude(student_id=student_code)
                        .first()
                    )
                    # Để tránh xung đột unique_together (exam_room, shift, subject_name, seat_number),
                    # dùng một ghế "tạm" nằm ngoài dải ghế thật để hoán đổi an toàn.
                    if other:
                        temp_seat = shift_capacity + 1000
                        seat_entry.seat_number = temp_seat
                        seat_entry.save()

                        other.seat_number = old_seat
                        other.save()

                    seat_entry.seat_number = seat
                    seat_entry.save()
            except ExamSubjectSeat.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'ok': False, 'error': 'Không tìm thấy layout môn thi cho học viên này.'}, status=404)
                messages.error(request, 'Không tìm thấy layout môn thi cho học viên này.')
                return redirect(_erd_redirect_url(subject_name))

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'ok': True})

            messages.success(request, 'Đã cập nhật vị trí ghế cho môn thi.')
            return redirect(_erd_redirect_url(subject_name))

    # --------- LOAD DỮ LIỆU HIỂN THỊ ----------
    base_qs = StudentExamAssignment.objects.select_related('student').filter(
        exam_room=room,
        shift=shift,
    )

    assignments = list(base_qs)
    assignments.sort(key=_exam_room_seat_order_key)

    students = [a.student for a in assignments]

    # Chỉ cho chọn khối/tổ hợp môn nếu trong cơ sở phòng đó thực sự có học viên
    campus_students = ExamRoomStudent.objects.filter(campus=room.campus)

    available_grades_set = set()
    grade_to_group_bases: dict[str, set[str]] = {}
    grade_to_group_by_hv: dict[str, dict[str, set[str]]] = {'sang': defaultdict(set), 'toi': defaultdict(set)}
    for s in campus_students.select_related('subject_group'):
        grade_val = _extract_grade_from_class_name(s.class_name)
        if not grade_val:
            continue
        grade_str = str(grade_val)
        available_grades_set.add(grade_str)
        code = s.subject_group.code if s.subject_group else ''
        if not code:
            continue
        base = code[-1]
        if base not in ['1', '2', '3', '4', '5']:
            continue
        grade_to_group_bases.setdefault(grade_str, set()).add(base)

        hv_d = _exam_hv_shift_digit_from_code(s.student_code or '')
        if hv_d == '0':
            grade_to_group_by_hv['sang'][grade_str].add(base)
        elif hv_d == '1':
            grade_to_group_by_hv['toi'][grade_str].add(base)

    available_grades = sorted(available_grades_set, key=int) if available_grades_set else []
    grade_subject_group_bases = {
        g: sorted(bases) for g, bases in grade_to_group_bases.items()
    }
    grade_subject_group_bases_by_hv = {
        'sang': {g: sorted(bases) for g, bases in grade_to_group_by_hv['sang'].items()},
        'toi': {g: sorted(bases) for g, bases in grade_to_group_by_hv['toi'].items()},
    }
    # Danh sách môn thi đã gán cho phòng + ca này (để fill dropdown "Môn thi")
    available_subjects_qs = ExamRoomSubject.objects.filter(
        exam_room=room,
        shift=shift,
    ).values_list("subject_name", flat=True).distinct()
    available_subjects = sorted(list(available_subjects_qs)) if available_subjects_qs else []

    # Sau khi xoá toàn bộ ở sắp xếp phòng thi, danh sách môn có thể rỗng → bỏ chọn môn cũ (tránh hiển thị "Văn" khi không còn dữ liệu)
    if not available_subjects:
        subject = ''
    elif not subject and available_subjects:
        subject = available_subjects[0]
    elif subject and subject not in available_subjects:
        # URL/bookmark còn ?subject=... nhưng môn đó không còn trong ExamRoomSubject → dropdown không có selected,
        # trình duyệt hiện option đầu tiên trong khi tiêu đề sơ đồ vẫn in đúng chuỗi cũ → đồng bộ về môn hợp lệ đầu tiên.
        subject = available_subjects[0]

    # Nếu đã chọn môn, ưu tiên đọc layout từ ExamSubjectSeat (mỗi môn có layout độc lập)
    subject_layout_qs = None
    subject_seat_map = {}
    if subject:
        subject_layout_qs = ExamSubjectSeat.objects.filter(
            exam_room=room,
            shift=shift,
            subject_name=subject,
        )
        subject_seat_map = {s.seat_number: s for s in subject_layout_qs}

    # Map student_id -> bản ghi StudentExamAssignment để template vẫn dùng chung cấu trúc "assignment"
    assignment_by_student_id = {a.student_id: a for a in base_qs}

    # Sinh ma trận ghế dựa trên seat_number (1..capacity) để kéo thả cập nhật được.
    # Nếu có layout theo môn thì bám theo ExamSubjectSeat, ngược lại fallback về seat_number chung.
    seat_map = {a.seat_number: a for a in base_qs if a.seat_number}
    seat_matrix = []
    seat_index = 1
    room_rows, room_cols = _get_room_grid_for_shift(room, shift)
    for _ in range(room_rows):
        row = []
        for _ in range(room_cols):
            # Nếu có layout theo môn thì lấy từ ExamSubjectSeat, rồi map về assignment tương ứng
            assignment = None
            if subject_seat_map:
                subj_seat = subject_seat_map.get(seat_index)
                if subj_seat:
                    assignment = assignment_by_student_id.get(subj_seat.student_id)
            else:
                assignment = seat_map.get(seat_index)

            row.append({
                "index": seat_index,
                "assignment": assignment,
            })
            seat_index += 1
        seat_matrix.append(row)

    can_initial_pick = len(assignments) == 0

    # Ứng viên cho "Chọn thêm học viên": cùng cơ sở, cùng khối + tổ hợp (theo dữ liệu hiện có), chưa có phòng trong ca này
    can_add_more = not can_initial_pick
    add_more_candidates: list[ExamRoomStudent] = []
    if not can_initial_pick:
        first_stu = assignments[0].student
        base_grade = _extract_grade_from_class_name(first_stu.class_name)
        base_group_code = first_stu.subject_group.code if first_stu.subject_group else ''
        base_group_digit = base_group_code[-1] if base_group_code else ''

        if base_grade and base_group_digit:
            assigned_ids_any_shift = set(
                StudentExamAssignment.objects.filter(shift=shift).values_list('student_id', flat=True)
            )
            campus_students_for_add = ExamRoomStudent.objects.filter(campus=room.campus).select_related('subject_group')
            for s in campus_students_for_add:
                if s.pk in assigned_ids_any_shift:
                    continue
                g = _extract_grade_from_class_name(s.class_name)
                if g != base_grade:
                    continue
                code = s.subject_group.code if s.subject_group else ''
                if not code or code[-1] != base_group_digit:
                    continue
                add_more_candidates.append(s)

            # Sắp xếp tăng dần theo SBD (exam_number), sau đó theo họ tên
            add_more_candidates.sort(key=lambda stu: (_exam_number_sort_key(stu.exam_number, stu.student_code), _name_sort_key_vi(stu.full_name or "")))

    campus_rooms = list(
        ExamRoom.objects.filter(campus_id=room.campus_id).order_by('name')
    )

    context = {
        'room': room,
        'shift': shift,
        'subject': subject,
        'campus_rooms': campus_rooms,
        'assignments': assignments,
        'students': students,
        'seat_matrix': seat_matrix,
        'can_initial_pick': can_initial_pick,
        'can_add_more': can_add_more,
        'add_more_candidates': add_more_candidates,
        'SHIFT_CHOICES': StudentExamAssignment.SHIFT_CHOICES,
        'available_subjects': available_subjects,
        'available_grades': available_grades,
        'grade_subject_group_bases': grade_subject_group_bases,
        'grade_subject_group_bases_by_hv': grade_subject_group_bases_by_hv,
    }
    return render(request, 'adminpageSIMCODE/exam_room_detail.html', context)


def _exam_bulk_unique_sheet_name(dest_wb, base: str) -> str:
    """Tên sheet Excel tối đa 31 ký tự, không trùng trong dest_wb."""
    clean = re.sub(r'[^\w\-]', '_', (base or 'sheet').strip())[:28] or 'sheet'
    name = clean[:31]
    i = 1
    existing = set(dest_wb.sheetnames)
    while name in existing:
        suffix = f'_{i}'
        name = (clean[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    return name


def _copy_worksheet_to_workbook(src_ws, dest_wb, sheet_name=None):
    """
    Copy một sheet từ workbook nguồn sang workbook đích, giữ nguyên format (font, border, fill, alignment).
    Trả về worksheet mới trong dest_wb.
    """
    new_title = sheet_name or src_ws.title
    dest_ws = dest_wb.create_sheet(title=new_title[:31])  # Excel sheet name max 31 chars
    # Copy các thiết lập sheet (giữ giống template tối đa)
    try:
        dest_ws.sheet_format = shallow_copy(src_ws.sheet_format)
        dest_ws.sheet_properties = shallow_copy(src_ws.sheet_properties)
        dest_ws.page_margins = shallow_copy(src_ws.page_margins)
        dest_ws.page_setup = shallow_copy(src_ws.page_setup)
        dest_ws.print_options = shallow_copy(src_ws.print_options)
        dest_ws.sheet_view = shallow_copy(src_ws.sheet_view)
        dest_ws.views = shallow_copy(src_ws.views)
        dest_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass

    # AutoFilter
    try:
        if getattr(src_ws.auto_filter, "ref", None):
            dest_ws.auto_filter.ref = src_ws.auto_filter.ref
    except Exception:
        pass

    # Conditional formatting (openpyxl dùng cấu trúc internal)
    try:
        dest_ws.conditional_formatting._cf_rules = shallow_copy(src_ws.conditional_formatting._cf_rules)
    except Exception:
        pass

    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                if cell.font:
                    dest_cell.font = shallow_copy(cell.font)
                if cell.border:
                    dest_cell.border = shallow_copy(cell.border)
                if cell.fill:
                    dest_cell.fill = shallow_copy(cell.fill)
                if cell.alignment:
                    dest_cell.alignment = shallow_copy(cell.alignment)
                if cell.number_format:
                    dest_cell.number_format = cell.number_format
    for merged in list(src_ws.merged_cells.ranges):
        dest_ws.merge_cells(str(merged))
    for col_letter_key, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dest_ws.column_dimensions[col_letter_key].width = dim.width
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dest_ws.row_dimensions[row_num].height = dim.height
    return dest_ws


def _get_merge_origin(ws, row: int, col: int):
    """
    Nếu (row, col) nằm trong một vùng merge thì trả về (min_row, min_col) của vùng đó;
    không thì trả về (row, col). Dùng để ghi value chỉ vào ô gốc, tránh lỗi MergedCell read-only.
    """
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return row, col


def _set_cell_value(ws, row: int, col: int, value):
    """Ghi value vào ô; nếu ô thuộc merged range thì ghi vào ô góc trên-trái của vùng merge."""
    r, c = _get_merge_origin(ws, row, col)
    ws.cell(row=r, column=c, value=value)


def _exam_unmerge_sheet_body(ws1, min_row: int = 9, max_row: int = 500) -> None:
    """
    Gỡ mọi vùng merge **cắt qua** hàng [min_row, max_row] (không giới hạn cột).
    Mẫu thường merge tới cột P trở đi; chỉ gỡ A–O trước đây vẫn để merge dọc → dòng trống / lệch cột.
    """
    refs: list[str] = []
    for rng in list(ws1.merged_cells.ranges):
        if rng.max_row < min_row or rng.min_row > max_row:
            continue
        refs.append(str(rng))
    for r in refs:
        try:
            ws1.unmerge_cells(r)
        except Exception:
            pass


def _exam_snapshot_list_row_styles(ws1, ref_row: int = 9, c1: int = 1, c2: int = 15) -> list[dict]:
    """Chụp font/viền/căn/format từng ô hàng mẫu (trước khi gỡ merge: theo ô gốc merge)."""
    snap: list[dict] = []
    for c in range(c1, c2 + 1):
        r0, c0 = _get_merge_origin(ws1, ref_row, c)
        ref = ws1.cell(row=r0, column=c0)
        snap.append(
            {
                'font': shallow_copy(ref.font) if ref.font else None,
                'border': shallow_copy(ref.border) if ref.border else None,
                'fill': shallow_copy(ref.fill) if ref.fill else None,
                'alignment': shallow_copy(ref.alignment) if ref.alignment else None,
                'number_format': ref.number_format or 'General',
            }
        )
    # Ô nằm trong merge rộng có thể không có font/viền riêng → đồng bộ từ cột đầu có style
    fb_font = next((s['font'] for s in snap if s.get('font')), None)
    fb_border = next((s['border'] for s in snap if s.get('border')), None)
    fb_align = next((s['alignment'] for s in snap if s.get('alignment')), None)
    for s in snap:
        if not s.get('font') and fb_font:
            s['font'] = shallow_copy(fb_font)
        if not s.get('border') and fb_border:
            s['border'] = shallow_copy(fb_border)
        if not s.get('alignment') and fb_align:
            s['alignment'] = shallow_copy(fb_align)
    return snap


def _exam_apply_list_row_styles(ws1, row_num: int, snap: list, c1: int = 1) -> None:
    """Áp dụng style đã chụp cho một hàng dữ liệu (đồng bộ viền/cột, tránh khối trên/dưới lệch)."""
    for i, st in enumerate(snap):
        col = c1 + i
        cell = ws1.cell(row=row_num, column=col)
        if st.get('font'):
            cell.font = st['font']
        if st.get('border'):
            cell.border = st['border']
        if st.get('fill'):
            cell.fill = st['fill']
        if st.get('alignment'):
            cell.alignment = st['alignment']
        nf = st.get('number_format') or 'General'
        cell.number_format = nf


def _get_block_starts_from_template(ws) -> list[int]:
    """
    Suy ra cột bắt đầu của từng khối sơ đồ từ template dựa trên merged cells ở hàng 1.
    Template thường merge tiêu đề từng khối (ví dụ A1:E1, H1:L1, ...).
    """
    starts: list[int] = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row == 1 and rng.max_row == 1 and (rng.max_col - rng.min_col) >= 2:
            starts.append(rng.min_col)
    starts = sorted(set(starts))
    return starts[:4]


def _detect_grid_start_row(ws, block_start_col: int) -> int:
    """
    Tìm dòng bắt đầu của lưới ghế trong 1 khối bằng cách tìm dòng chứa 'Bàn Giáo viên'
    tại cột block_start_col, rồi +2 (1 dòng trống).
    """
    for r in range(1, min(ws.max_row, 80) + 1):
        v = ws.cell(row=r, column=block_start_col).value
        if isinstance(v, str) and v.strip().lower() == 'bàn giáo viên':
            return r + 2
    return 6  # fallback phổ biến của template


def _detect_seat_cells_from_template(ws_values, block_start_col: int, block_width: int) -> list[tuple[int, int]]:
    """
    Dò các ô ghế theo giá trị số đã có trong template (data_only=True).
    Trả về danh sách (row, col) theo thứ tự ghế theo vị trí (row-major) để khớp cách đánh số ghế trên web.
    """
    grid_start = _detect_grid_start_row(ws_values, block_start_col)
    max_scan_rows = min(ws_values.max_row, grid_start + 30)
    candidates: list[tuple[int, int, int]] = []
    for r in range(grid_start, max_scan_rows + 1):
        for c in range(block_start_col, block_start_col + block_width):
            v = ws_values.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v not in (0,):
                try:
                    num = int(v)
                except Exception:
                    continue
                # Tránh nhận nhầm cột lối đi / cột phụ có số nhỏ (ví dụ '3' trong template)
                # SBD thực tế luôn là số đủ lớn.
                if num < 1000:
                    continue
                candidates.append((num, r, c))
    # ưu tiên order theo vị trí vì seat_number trên web tăng theo row-major
    pos = [(r, c) for _, r, c in candidates]
    pos.sort(key=lambda rc: (rc[0], rc[1]))
    return pos


def _detect_seat_cells_with_numbers(ws_values, block_start_col: int, block_width: int) -> list[tuple[int, int, int]]:
    """
    Dò các ô ghế theo giá trị số đã có trong template (data_only=True).
    Trả về danh sách (num, row, col). `num` thường là SBD mẫu trong file.
    """
    grid_start = _detect_grid_start_row(ws_values, block_start_col)
    max_scan_rows = min(ws_values.max_row, grid_start + 30)
    out: list[tuple[int, int, int]] = []
    for r in range(grid_start, max_scan_rows + 1):
        for c in range(block_start_col, block_start_col + block_width):
            v = ws_values.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v not in (0,):
                try:
                    num = int(v)
                except Exception:
                    continue
                if num < 1000:
                    continue
                out.append((num, r, c))
    return out

def _detect_seat_cells_by_template_content(ws_style, block_start_col: int, block_width: int) -> list[tuple[int, int]]:
    """
    Dò ô ghế dựa trên nội dung template (data_only=False):
    - Ô ghế thường có công thức/giá trị sẵn (ví dụ '=A6+5'), còn hành lang để trống (None).
    Trả về danh sách (row, col) theo row-major.
    """
    grid_start = _detect_grid_start_row(ws_style, block_start_col)
    max_scan_rows = min(ws_style.max_row, grid_start + 30)
    col_counts: dict[int, int] = {}
    candidates: list[tuple[int, int]] = []
    for r in range(grid_start, max_scan_rows + 1):
        for c in range(block_start_col, block_start_col + block_width):
            rr, cc = _get_merge_origin(ws_style, r, c)
            if (rr, cc) != (r, c):
                continue
            v = ws_style.cell(row=r, column=c).value
            has_template_value = (v is not None) and (not (isinstance(v, str) and v.strip() == ''))
            if has_template_value:
                candidates.append((r, c))
                col_counts[c] = col_counts.get(c, 0) + 1

    if not candidates:
        return []

    seat_cols = _pick_seat_columns(col_counts, need=4)
    seat_cols_set = set(seat_cols)
    cells = [(r, c) for (r, c) in candidates if c in seat_cols_set]
    cells.sort(key=lambda rc: (rc[0], rc[1]))
    return cells


def _detect_seat_cells_for_export(ws_style, block_start_col: int, block_width: int) -> list[tuple[int, int]]:
    """
    Dò ô ghế để export một cách ổn định:
    - Border có thể xuất hiện ở cột viền/phụ (không phải ghế) -> dễ chọn nhầm.
    - Template của bạn: cột ghế thường có công thức/giá trị; hành lang trống.

    Chiến lược:
    1) Quét vùng lưới lấy candidates theo border.
    2) Tính "content_cols" = cột nào có công thức/giá trị trong template.
       Nếu có >=4 cột content -> ưu tiên chọn 4 cột ghế từ nhóm đó (đảm bảo không chọn cột viền phụ).
    3) Nếu content_cols <4 -> fallback _pick_seat_columns theo border.
    """
    grid_start = _detect_grid_start_row(ws_style, block_start_col)
    max_scan_rows = min(ws_style.max_row, grid_start + 30)

    def _has_border(cell) -> bool:
        b = cell.border
        return bool(
            b
            and ((b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style))
        )

    col_border_counts: dict[int, int] = {}
    col_content_counts: dict[int, int] = {}
    candidates: list[tuple[int, int]] = []

    for r in range(grid_start, max_scan_rows + 1):
        for c in range(block_start_col, block_start_col + block_width):
            rr, cc = _get_merge_origin(ws_style, r, c)
            if (rr, cc) != (r, c):
                continue
            cell = ws_style.cell(row=r, column=c)
            if _has_border(cell):
                candidates.append((r, c))
                col_border_counts[c] = col_border_counts.get(c, 0) + 1
            v = cell.value
            if v is not None and not (isinstance(v, str) and v.strip() == ''):
                col_content_counts[c] = col_content_counts.get(c, 0) + 1

    if not candidates:
        return []

    content_cols = [c for c, cnt in col_content_counts.items() if cnt > 0]
    if len(content_cols) >= 4:
        # Chọn 4 cột ghế trong nhóm content_cols, ưu tiên cột có nhiều border hơn
        seat_cols = [c for c, _ in sorted(((c, col_border_counts.get(c, 0)) for c in content_cols), key=lambda kv: (-kv[1], kv[0]))[:4]]
        seat_cols = sorted(set(seat_cols))
    else:
        seat_cols = _pick_seat_columns(col_border_counts, need=4)

    seat_cols_set = set(seat_cols)
    cells = [(r, c) for (r, c) in candidates if c in seat_cols_set]
    cells.sort(key=lambda rc: (rc[0], rc[1]))
    return cells


def _pick_seat_columns(col_counts: dict[int, int], need: int = 4) -> list[int]:
    """
    Chọn các cột ghế từ thống kê số ô/viền theo cột.
    Template thường có 2 cụm cột ghế (trái) + (phải) ngăn bởi hành lang (có thể 1 hoặc nhiều cột trống).
    Chiến lược:
    - Nếu <= need cột thì lấy hết.
    - Nếu > need: tìm "khoảng cách lớn nhất" giữa các cột -> tách trái/phải,
      rồi lấy top 2 cột mỗi bên theo mật độ (fallback nếu bên thiếu).
    """
    cols = sorted([c for c, cnt in col_counts.items() if cnt > 0])
    if len(cols) <= need:
        return cols

    # tìm điểm tách theo gap lớn nhất
    gaps = [(cols[i + 1] - cols[i], i) for i in range(len(cols) - 1)]
    gap, split_i = max(gaps, key=lambda t: t[0])

    # Nếu các cột gần như liền nhau (gap=1), thường là dạng A,B,|C hành lang|,D,E
    # => lấy 5 cột dày nhất rồi loại cột giữa để còn 4 cột ghế.
    if gap <= 1:
        densest = [c for c, _ in sorted(col_counts.items(), key=lambda kv: (-kv[1], kv[0]))[: max(need + 1, 5)]]
        densest = sorted(set(densest))
        if len(densest) >= 5:
            mid = densest[len(densest) // 2]
            densest = [c for c in densest if c != mid]
        return densest[:need]

    left = cols[: split_i + 1]
    right = cols[split_i + 1 :]

    def top_k(group, k):
        return [c for c, _ in sorted(((c, col_counts.get(c, 0)) for c in group), key=lambda kv: (-kv[1], kv[0]))[:k]]

    picked = top_k(left, 2) + top_k(right, 2)
    picked = sorted(set(picked))

    if len(picked) < need:
        remain = [c for c in cols if c not in picked]
        picked += remain[: (need - len(picked))]

    return picked[:need]


def _fallback_seat_cells_by_border(ws_style, block_start_col: int, block_width: int) -> list[tuple[int, int]]:
    """
    Fallback: dò các ô có border (viền) trong vùng lưới của khối để coi là ghế.
    Thứ tự trả về theo row-major (từ trên xuống, trái qua phải).
    """
    grid_start = _detect_grid_start_row(ws_style, block_start_col)
    max_scan_rows = min(ws_style.max_row, grid_start + 30)
    # Đếm mật độ border theo cột để loại cột "lối đi" / cột phụ
    col_counts: dict[int, int] = {}
    candidates: list[tuple[int, int]] = []
    for r in range(grid_start, max_scan_rows + 1):
        for c in range(block_start_col, block_start_col + block_width):
            rr, cc = _get_merge_origin(ws_style, r, c)
            if (rr, cc) != (r, c):
                continue
            cell = ws_style.cell(row=r, column=c)
            b = cell.border
            has_border = bool(
                b
                and ((b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style))
            )
            if has_border:
                candidates.append((r, c))
                col_counts[c] = col_counts.get(c, 0) + 1

    if not candidates:
        return []

    # Chọn 4 cột ghế (2 trái + 2 phải) theo cấu trúc template
    top_cols = _pick_seat_columns(col_counts, need=4)
    top_cols_set = set(top_cols)
    cells = [(r, c) for (r, c) in candidates if c in top_cols_set]
    cells.sort(key=lambda rc: (rc[0], rc[1]))
    return cells


def _split_ho_ten(full_name):
    """Tách họ và tên: phần cuối là Tên, còn lại là Họ."""
    if not full_name or not full_name.strip():
        return '', ''
    parts = full_name.strip().split()
    if len(parts) <= 1:
        return full_name.strip(), ''
    return ' '.join(parts[:-1]), parts[-1]


def _seat_order_pattern(room, mode: int, shift: str | None = None) -> list[int]:
    """
    Tạo thứ tự seat_number dựa trên lưới phòng (row_count x col_count).
    mode:
      1: ziczac theo chiều dọc (column-major serpentine)
      2: ziczac theo chiều ngang (row-major serpentine)
      3: ziczac theo chiều dọc, đảo ngược (reverse của mode 1)
      4: ziczac theo chiều ngang, đảo ngược (reverse của mode 2)

      5: quét thẳng từ trên xuống, trái qua phải (column-major)
      6: quét thẳng từ trên xuống, phải qua trái (column-major, cột đảo)
      7: quét thẳng từ dưới lên, trái qua phải (column-major, hàng đảo)
      8: quét thẳng từ dưới lên, phải qua trái (column-major, hàng + cột đảo)

      9:  quét thẳng từ trái qua phải, từ trên xuống (row-major)
      10: quét thẳng từ trái qua phải, từ dưới lên (row-major, hàng đảo)
      11: quét thẳng từ phải qua trái, từ trên xuống (row-major, cột đảo)
      12: quét thẳng từ phải qua trái, từ dưới lên (row-major, hàng + cột đảo)
    """
    rows, cols = _get_room_grid_for_shift(room, shift)
    if rows <= 0 or cols <= 0:
        return []

    order: list[int] = []
    if mode in (1, 3):
        # column-major serpentine
        for c in range(cols):
            if c % 2 == 0:
                r_iter = range(rows)
            else:
                r_iter = range(rows - 1, -1, -1)
            for r in r_iter:
                order.append(r * cols + c + 1)
    elif mode in (2, 4):
        # row-major serpentine
        for r in range(rows):
            if r % 2 == 0:
                c_iter = range(cols)
            else:
                c_iter = range(cols - 1, -1, -1)
            for c in c_iter:
                order.append(r * cols + c + 1)
    else:
        # straight scan (no serpentine)
        if mode in (5, 6, 7, 8):
            # column-major with direction controls
            c_iter = range(cols)
            r_iter = range(rows)
            if mode in (6, 8):
                c_iter = range(cols - 1, -1, -1)
            if mode in (7, 8):
                r_iter = range(rows - 1, -1, -1)
            for c in c_iter:
                for r in r_iter:
                    order.append(r * cols + c + 1)
        else:
            # row-major with direction controls
            r_iter = range(rows)
            c_iter = range(cols)
            if mode in (10, 12):
                r_iter = range(rows - 1, -1, -1)
            if mode in (11, 12):
                c_iter = range(cols - 1, -1, -1)
            for r in r_iter:
                for c in c_iter:
                    order.append(r * cols + c + 1)

    if mode in (3, 4):
        order.reverse()
    return order

def _seat_positions_front_filled(room, n_students: int, direction: int, shift: str | None = None) -> list[int]:
    """
    Sinh seat_number cho n_students sao cho:
    - Các hàng ghế phía trên luôn được lấp kín trước (không có ô trống ở các hàng đầu).
    - Chỉ các hàng phía dưới mới được trống.
    - direction (1..4) tạo 4 hướng bắt đầu khác nhau dựa trên "khối ghế đã fill" (không phải toàn phòng).
      Nhờ đó với 4 môn: SBD nhỏ nhất nằm ở 4 hướng khác nhau và SBD lớn nhất cũng nằm ở 4 hướng khác nhau,
      trong khi vẫn đảm bảo các hàng trên không bị trống.

    direction:
      1: bắt đầu TL của khối fill (duyệt hàng: trên->dưới, cột: trái->phải)
      2: bắt đầu TR của khối fill (duyệt hàng: trên->dưới, cột: phải->trái)
      3: bắt đầu BL của khối fill (duyệt hàng: dưới->trên trong khối fill, cột: trái->phải)
      4: bắt đầu BR của khối fill (duyệt hàng: dưới->trên trong khối fill, cột: phải->trái)
    """
    rows, cols = _get_room_grid_for_shift(room, shift)
    if rows <= 0 or cols <= 0 or n_students <= 0:
        return []

    n = min(n_students, rows * cols)
    full_rows = n // cols
    rem = n % cols

    def col_iter():
        return range(cols) if direction in (1, 3) else range(cols - 1, -1, -1)

    order: list[int] = []

    # Xác định các row thuộc khối fill: 0..full_rows-1 (đầy) và full_rows (dư) nếu rem>0
    filled_rows = list(range(full_rows)) + ([full_rows] if rem else [])
    if direction in (3, 4):
        filled_rows = list(reversed(filled_rows))

    for r in filled_rows:
        cols_in = list(col_iter())
        if rem and r == full_rows:
            cols_in = cols_in[:rem]
        for c in cols_in:
            order.append(r * cols + c + 1)

    return order

def _seat_positions_column_balanced(room, n_students: int, direction: int, shift: str | None = None) -> list[int]:
    """
    Chia học sinh theo CỘT trước (balanced) nhưng vẫn đảm bảo hàng trên luôn đầy:
    - Mỗi cột nhận số lượng gần bằng nhau (chênh lệch tối đa 1) => ô trống chỉ rơi xuống hàng dưới.
    - direction điều khiển thứ tự cột và chiều trong cột.

    direction:
      1: trái->phải, trên->dưới
      2: phải->trái, trên->dưới
      3: trái->phải, dưới->lên
      4: phải->trái, dưới->lên
    """
    rows, cols = _get_room_grid_for_shift(room, shift)
    if rows <= 0 or cols <= 0 or n_students <= 0:
        return []

    n = min(n_students, rows * cols)
    base_h = n // cols
    extra = n % cols
    col_heights = [base_h + (1 if i < extra else 0) for i in range(cols)]

    col_indices = list(range(cols - 1, -1, -1)) if direction in (2, 4) else list(range(cols))

    col_seats: dict[int, list[int]] = {}
    for c in range(cols):
        h = col_heights[c]
        if h <= 0:
            col_seats[c] = []
            continue
        if direction in (3, 4):
            rs = list(range(h - 1, -1, -1))  # dưới->lên trong phạm vi khối fill
        else:
            rs = list(range(0, h))  # trên->dưới
        col_seats[c] = [r * cols + c + 1 for r in rs]

    order: list[int] = []
    # Đi THEO CỘT thật sự: fill hết 1 cột rồi mới sang cột tiếp theo.
    # Vì col_heights đã được cân bằng (chênh lệch <= 1) nên các hàng phía trên vẫn đầy,
    # và ghế trống chỉ rơi xuống các hàng dưới cùng.
    for c in col_indices:
        order.extend(col_seats.get(c, []))

    return order

def _exam_export_cell_dash(val):
    """Ô dữ liệu: trống / 0 → dấu '-' (theo mẫu in danh sách)."""
    if val is None:
        return '-'
    if isinstance(val, (int, float)) and not isinstance(val, bool) and val == 0:
        return '-'
    s = str(val).strip()
    if not s:
        return '-'
    # Chuỗi "0" từ Excel/DB vẫn coi là trống (tránh hiện số 0)
    if s in ('0', '0.0', '0.00'):
        return '-'
    return s

# Accounting (0 chữ số, không ký hiệu tiền): số 0 hiển thị thành "-" giống hộp thoại Format Cells
_EXCEL_ACCOUNTING_ZERO_DASH_FMT = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'


def _exam_cell_value_is_blank_or_zero(v) -> bool:
    """Ô rỗng, 0 số, chuỗi '0', Decimal(0) — coi là cần thay bằng '-'."""
    if v is None or v == '':
        return True
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return True
        if s.startswith('='):
            return False
        if s in ('0', '0.0', '0.00'):
            return True
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)) and v == 0:
        return True
    if isinstance(v, Decimal) and v == 0:
        return True
    try:
        if float(v) == 0.0:
            return True
    except (TypeError, ValueError):
        pass
    return False


def _exam_sheet1_apply_dash_cell_style(cell):
    """Đồng bộ định dạng ô '-' với các ô gạch khác (Accounting + bỏ bold)."""
    cell.number_format = _EXCEL_ACCOUNTING_ZERO_DASH_FMT
    try:
        if cell.font:
            nf = shallow_copy(cell.font)
            nf.bold = False
            cell.font = nf
    except Exception:
        pass


def _exam_sheet1_normalize_body_zeros(
    ws,
    row_start: int,
    row_end: int,
    *,
    data_dash_cols: tuple[int, ...] | None = None,
    notes_col: int | None = None,
):
    """
    Dọn vùng thân bảng (STT bắt đầu row_start): các ô nội dung (mặc định C–F) và cột ghi chú (mặc định G)
    nếu là 0/rỗng/công thức → '-'; gán number_format Accounting để Excel hiện gạch đúng kiểu mẫu.
    """
    if row_end < row_start:
        return
    dash_cols = data_dash_cols if data_dash_cols is not None else (3, 4, 5, 6)
    ghi_chu_col = notes_col if notes_col is not None else 7
    for r in range(row_start, row_end + 1):
        for c in dash_cols:
            o_r, o_c = _get_merge_origin(ws, r, c)
            cell = ws.cell(row=o_r, column=o_c)
            v = cell.value
            if isinstance(v, str) and v.strip().startswith('='):
                cell.value = '-'
            elif _exam_cell_value_is_blank_or_zero(v):
                cell.value = '-'
            if cell.value == '-':
                _exam_sheet1_apply_dash_cell_style(cell)
        o_r, o_c = _get_merge_origin(ws, r, ghi_chu_col)
        cell = ws.cell(row=o_r, column=o_c)
        v = cell.value
        vs = str(v).strip() if v is not None else ''
        if vs.upper() == 'HN':
            continue
        if isinstance(v, str) and v.strip().startswith('='):
            cell.value = '-'
        elif _exam_cell_value_is_blank_or_zero(v):
            cell.value = '-'
        if cell.value == '-':
            _exam_sheet1_apply_dash_cell_style(cell)

def _exam_sheet1_footer_start_row(ws, scan_from_row: int = 9) -> int | None:
    """
    Dòng đầu của khối chân mẫu (Tổng cộng / Giám thị / Ký tên).
    Không được ghi đệm/normalize từ dòng này trở xuống — tránh xóa mất footer.
    """
    markers = (
        'tổng cộng',
        'giám thị',
        'ký tên',
        'tong cong',
        'giam thi',
        'ky ten',
    )
    max_r = min(ws.max_row or 0, 500)
    for r in range(max(1, scan_from_row), max_r + 1):
        for c in range(1, 17):
            o_r, o_c = _get_merge_origin(ws, r, c)
            val = ws.cell(row=o_r, column=o_c).value
            if not isinstance(val, str):
                continue
            low = val.lower()
            if any(m in low for m in markers):
                return o_r
    return None



def _build_exam_export_combo_label(assignments):
    """
    Chuỗi hàng 4 cột I–K (một dòng): TỔ HỢP {mã tổ hợp} (các lớp có trong phòng).
    Mã tổ hợp lấy từ SubjectGroup.code của học viên trong phòng/ca.
    """
    codes: list[str] = []
    classes: list[str] = []
    seen_c: set[str] = set()
    seen_cls: set[str] = set()
    for a in assignments:
        stu = getattr(a, 'student', None)
        if not stu:
            continue
        sg = getattr(stu, 'subject_group', None)
        code = (str(sg.code).strip() if sg else '') or ''
        if code and code not in seen_c:
            seen_c.add(code)
            codes.append(code)
        cn = (getattr(stu, 'class_name', None) or '').strip()
        if cn and cn not in seen_cls:
            seen_cls.add(cn)
            classes.append(cn)
    codes.sort(key=lambda x: (len(x), x))
    classes.sort(key=natural_sort_key)
    if not codes and not classes:
        return 'TỔ HỢP — (—)'
    grp = ', '.join(codes)
    cls_part = ', '.join(classes) if classes else '—'
    return f'TỔ HỢP {grp} ({cls_part})'


def _exam_subject_group_th_label(stu) -> str:
    """Nhãn cột Tổ hợp khi export (vd TH1 nếu mã tổ hợp kết thúc bằng 1)."""
    if not stu:
        return '-'
    sg = getattr(stu, 'subject_group', None)
    if not sg:
        return '-'
    code = str(getattr(sg, 'code', '') or '').strip()
    if not code:
        return '-'
    for ch in reversed(code):
        if ch in '12345':
            return f'TH{ch}'
    for ch in reversed(code):
        if ch.isdigit():
            return f'TH{ch}'
    safe = re.sub(r'[^\w\-]', '', code)[:6] or '-'
    return f'TH{safe}'


def _exam_header_cell_norm(v) -> str:
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v).strip().lower())


def _exam_sheet1_guess_stt_header_row(ws1) -> int:
    """Hàng đầu tiên có tiêu đề cột STT (mẫu có thể để tiêu đề bảng ở hàng 6 hoặc 7…)."""
    for r in range(4, 15):
        v = ws1.cell(row=r, column=1).value
        if v is None:
            continue
        t = str(v).strip().lower()
        if t in ('stt', 'stt.', 'stt '):
            return r
    return 6


def _exam_sheet1_template_has_th_column(ws1, header_row: int) -> bool:
    """
    Mẫu có cột TỔ HỢP ở B (chữ «tổ hợp»). Mẫu cũ chỉ có «Phòng thi» ở B — không coi là có cột TH.
    """
    c2 = _exam_header_cell_norm(ws1.cell(row=header_row, column=2).value)
    if not c2:
        return False
    if 'phòng' in c2 and 'thi' in c2.replace(' ', ''):
        return False
    if 'tổ' in c2 or 'to' in c2:
        if 'hợp' in c2 or 'hop' in c2.replace('đ', 'd'):
            return True
    return False


def _exam_sheet1_include_th_layout_kind(ws1, header_row: int) -> str:
    """
    ``wide``: cột C là Phòng thi (mẫu danh sách đầy đủ + MÔN THI).
    ``compact``: cột C là SBD — chỉ A–G (TỔ HỢP, SBD, HỌ VÀ TÊN, Lớp, Ghi chú), không cột Phòng.
    """
    c3 = _exam_header_cell_norm(ws1.cell(row=header_row, column=3).value)
    if not c3:
        return 'wide'
    if 'phòng' in c3 or 'phong' in c3.replace(' ', ''):
        return 'wide'
    if 'sbd' in c3:
        return 'compact'
    return 'wide'


def _exam_sheet1_setup_header_compact_th(ws1, header_top: int) -> None:
    """Mẫu BIÊN BẢN / danh sách 7 cột: Stt | TỔ HỢP | SBD | HỌ VÀ TÊN (D–E) | Lớp | Ghi chú — merge 3 hàng tiêu đề."""
    hr = header_top
    hr1 = hr + 1
    hr2 = hr + 2
    for r in (hr, hr1, hr2):
        for c in range(1, 19):
            ws1.cell(row=r, column=c, value=None)
    ws1.cell(row=hr, column=1, value='Stt')
    ws1.cell(row=hr, column=2, value='Tổ hợp')
    ws1.cell(row=hr, column=3, value='SBD')
    ws1.cell(row=hr, column=4, value='HỌ VÀ TÊN')
    ws1.cell(row=hr, column=6, value='Lớp')
    ws1.cell(row=hr, column=7, value='Ghi chú')
    for m in (
        f'A{hr}:A{hr2}',
        f'B{hr}:B{hr2}',
        f'C{hr}:C{hr2}',
        f'D{hr}:E{hr2}',
        f'F{hr}:F{hr2}',
        f'G{hr}:G{hr2}',
    ):
        ws1.merge_cells(m)


def _exam_sheet1_setup_header_row_with_th_column(
    ws1, subjects_ws1: list[str], header_top: int = 6
) -> None:
    """
    Sau khi gỡ merge 3 hàng tiêu đề: layout ``wide`` — B Tổ hợp, C Phòng thi, **D** SBD (một cột, merge dọc 3 hàng),
    **E:F** merge tiêu đề «Họ và tên» (dữ liệu: họ ở E, tên ở F), **G** Lớp, **H** Ghi chú; MÔN THI **I–P**;
    hàng ``header_top+1`` tên môn, ``header_top+2`` Số tờ / Ký tên / Mã đề.
    """
    hr = header_top
    hr1 = hr + 1
    hr2 = hr + 2
    subjects = list(subjects_ws1[:4])
    while len(subjects) < 4:
        subjects.append('-')
    for r in (hr, hr1, hr2):
        for c in range(1, 19):
            ws1.cell(row=r, column=c, value=None)

    ws1.cell(row=hr, column=1, value='Stt')
    ws1.cell(row=hr, column=2, value='Tổ hợp')
    ws1.cell(row=hr, column=3, value='Phòng\nthi')
    ws1.cell(row=hr, column=4, value='SBD')
    ws1.cell(row=hr, column=5, value='Họ và tên')
    ws1.cell(row=hr, column=7, value='Lớp')
    ws1.cell(row=hr, column=8, value='Ghi \nchú')
    ws1.cell(row=hr, column=9, value='MÔN THI')

    for m in (f'A{hr}:A{hr2}', f'B{hr}:B{hr2}', f'C{hr}:C{hr2}', f'D{hr}:D{hr2}'):
        ws1.merge_cells(m)
    ws1.merge_cells(f'E{hr}:F{hr2}')
    for m in (f'G{hr}:G{hr2}', f'H{hr}:H{hr2}'):
        ws1.merge_cells(m)
    ws1.merge_cells(f'I{hr}:P{hr}')

    for i, subj in enumerate(subjects):
        c0 = 9 + i * 2
        c1 = c0 + 1
        rng = f'{get_column_letter(c0)}{hr1}:{get_column_letter(c1)}{hr1}'
        ws1.merge_cells(rng)
        ws1.cell(row=hr1, column=c0, value=subj)

    row8_labels = ('Số tờ', 'Ký tên', 'Mã đề', 'Ký tên', 'Mã đề', 'Ký tên', 'Mã đề', 'Ký tên')
    for i, lab in enumerate(row8_labels):
        ws1.cell(row=hr2, column=9 + i, value=lab)


def _fill_exam_giao_nhan_merged(
    ws1,
    blocks: list,
    shift: str,
    *,
    global_stt: bool = False,
    include_th_column: bool = False,
    min_data_rows: int | None = None,
) -> None:
    """
    Điền sheet danh sách giao nhận khi gộp nhiều phòng: cùng ca `shift`, mỗi phòng một khối
    STT 1..n theo số HV thực tế (không chèn dòng đệm tới sức chứa — phòng sau nối liền ngay sau phòng trước).
    `blocks`: [(room, assignments_sorted), ...].

    ``global_stt``: True = STT liên tục trên toàn sheet; False = STT 1..n lại từ đầu trong từng phòng (export gộp dùng False).
    ``include_th_column``: cột B = Tổ hợp (TH1…). Bố cục đọc từ mẫu:
    **wide** — C = Phòng thi; **D** = SBD; **E** = họ, **F** = tên (tiêu đề merge E:F «Họ và tên»); **G** = Lớp; **H** = Ghi chú; MÔN THI từ **I**;
    **compact** — mẫu chỉ A–G: C = SBD, D–E = HỌ VÀ TÊN, F = Lớp, G = Ghi chú (không cột Phòng, không khối môn).
    ``min_data_rows``: khi đặt (vd sức chứa phòng), đệm ít nhất từng số hàng dữ liệu (export đơn).
    Hàng tiêu đề STT được đoán (hàng 6 hoặc 7…); dữ liệu bắt đầu ngay dưới 3 hàng tiêu đề.
    """
    all_assignments = [a for _, arr in blocks for a in arr]
    combo_label = _build_exam_export_combo_label(all_assignments)
    names: list[str] = []
    for room, _ in blocks:
        nm = str(room.name or '').strip() or '-'
        if nm not in names:
            names.append(nm)
    room_header = ', '.join(names)
    today_str = timezone.localdate().strftime('%d/%m/%Y')
    first_room = blocks[0][0]
    subjects = list(
        ExamRoomSubject.objects.filter(exam_room=first_room, shift=shift)
        .values_list('subject_name', flat=True)
        .distinct()
    )[:4]
    fixed_subjects = ['Toán', 'Văn', 'Sử']
    extra_subject = ''
    for s in subjects:
        if s and s not in fixed_subjects:
            extra_subject = s
            break
    subjects_ws1: list[str] = []
    for lab in ('Văn', (extra_subject or '').strip() or None, 'Toán', 'Sử'):
        if lab is None:
            subjects_ws1.append('-')
        else:
            subjects_ws1.append(str(lab).strip().upper())

    _set_cell_value(ws1, 4, 1, '')
    _set_cell_value(ws1, 4, 2, f'Ngày: {today_str}')
    _set_cell_value(ws1, 4, 9, combo_label)
    _set_cell_value(ws1, 4, 13, f'PHÒNG THI: {room_header}')
    if _get_merge_origin(ws1, 4, 13) != _get_merge_origin(ws1, 4, 14):
        _set_cell_value(ws1, 4, 14, '')
    _set_cell_value(ws1, 4, 15, '')
    th_kind = 'wide'
    data_start = 9
    if include_th_column:
        header_top = _exam_sheet1_guess_stt_header_row(ws1)
        th_kind = _exam_sheet1_include_th_layout_kind(ws1, header_top)
        _exam_unmerge_sheet_body(ws1, header_top, 500)
        if th_kind == 'compact':
            _exam_sheet1_setup_header_compact_th(ws1, header_top)
            style_c2 = 7
        else:
            _exam_sheet1_setup_header_row_with_th_column(ws1, subjects_ws1, header_top)
            style_c2 = 17
        data_start = header_top + 3
    else:
        for i, subj in enumerate(subjects_ws1[:4]):
            _set_cell_value(ws1, 7, 8 + i * 2, subj)
        _exam_unmerge_sheet_body(ws1)
        style_c2 = 15

    # Chụp style hàng dữ liệu mẫu trước khi gỡ merge; gỡ merge rộng (mọi cột) rồi ghi trực tiếp từng ô
    style_snap = _exam_snapshot_list_row_styles(ws1, data_start, 1, style_c2)
    ref_height = (
        ws1.row_dimensions[data_start].height if data_start in ws1.row_dimensions else None
    )
    col_widths: dict[str, float] = {}
    for c in range(1, style_c2 + 1):
        letter = get_column_letter(c)
        dim = ws1.column_dimensions.get(letter)
        if dim is not None and dim.width is not None:
            col_widths[letter] = dim.width
    for letter, w in col_widths.items():
        ws1.column_dimensions[letter].width = w

    def _write_and_style_row(rn: int) -> None:
        _exam_apply_list_row_styles(ws1, rn, style_snap)
        if ref_height:
            ws1.row_dimensions[rn].height = ref_height

    cur = data_start
    stt_global = 0
    for room, assigns in blocks:
        for idx, a in enumerate(assigns, start=1):
            stu = a.student
            ho, ten = _split_ho_ten((stu.full_name or '') if stu else '')
            sbd_raw = (stu.exam_number or stu.student_code or '') if stu else ''
            sbd = _exam_export_cell_dash(sbd_raw)
            lop = _exam_export_cell_dash((stu.class_name or '') if stu else '')
            ho = _exam_export_cell_dash(ho)
            ten = _exam_export_cell_dash(ten)
            ghi_chu = 'HN' if (stu and getattr(stu, 'is_integration', False)) else '-'
            row_num = cur
            if global_stt:
                stt_global += 1
                stt_val = stt_global
            else:
                stt_val = idx
            ws1.cell(row=row_num, column=1, value=stt_val)
            if include_th_column:
                if th_kind == 'compact':
                    full_nm = _exam_export_cell_dash(
                        ((stu.full_name or '').strip() if stu else '')
                    )
                    if full_nm == '-' and (ho != '-' or ten != '-'):
                        parts = [p for p in (ho, ten) if p and p != '-']
                        full_nm = ' '.join(parts) if parts else '-'
                    ws1.cell(row=row_num, column=2, value=_exam_subject_group_th_label(stu))
                    ws1.cell(row=row_num, column=3, value=sbd)
                    ws1.cell(row=row_num, column=4, value=full_nm)
                    ws1.cell(row=row_num, column=5, value=None)
                    ws1.cell(row=row_num, column=6, value=lop)
                    ws1.cell(row=row_num, column=7, value=ghi_chu)
                    for col in range(8, 18):
                        ws1.cell(row=row_num, column=col, value='')
                else:
                    ws1.cell(row=row_num, column=2, value=_exam_subject_group_th_label(stu))
                    ws1.cell(row=row_num, column=3, value=f'P.{room.name}')
                    ws1.cell(row=row_num, column=4, value=sbd)
                    ws1.cell(row=row_num, column=5, value=ho)
                    ws1.cell(row=row_num, column=6, value=ten)
                    ws1.cell(row=row_num, column=7, value=lop)
                    ws1.cell(row=row_num, column=8, value=ghi_chu)
                    for col in range(9, 18):
                        ws1.cell(row=row_num, column=col, value='')
            else:
                ws1.cell(row=row_num, column=2, value=f'P.{room.name}')
                ws1.cell(row=row_num, column=3, value=sbd)
                ws1.cell(row=row_num, column=4, value=ho)
                ws1.cell(row=row_num, column=5, value=ten)
                ws1.cell(row=row_num, column=6, value=lop)
                ws1.cell(row=row_num, column=7, value=ghi_chu)
                for col in range(8, 16):
                    ws1.cell(row=row_num, column=col, value='')
            _write_and_style_row(row_num)
            cur += 1

    footer_row = _exam_sheet1_footer_start_row(ws1, data_start)
    last_capacity_row = cur - 1
    mr_sheet = ws1.max_row or last_capacity_row
    last_pad_row = min(max(last_capacity_row, mr_sheet), 500)
    if footer_row is not None and footer_row > cur:
        last_pad_row = min(last_pad_row, footer_row - 1)
    if min_data_rows is not None and min_data_rows > 0:
        need_last = data_start + min_data_rows - 1
        last_pad_row = max(last_pad_row, need_last)
        if footer_row is not None:
            last_pad_row = min(last_pad_row, footer_row - 1)
    for row_num in range(cur, last_pad_row + 1):
        if global_stt:
            stt_global += 1
            ws1.cell(row=row_num, column=1, value=stt_global)
        else:
            ws1.cell(row=row_num, column=1, value=row_num - (data_start - 1))
        if include_th_column:
            if th_kind == 'compact':
                ws1.cell(row=row_num, column=2, value='-')
                ws1.cell(row=row_num, column=3, value='-')
                ws1.cell(row=row_num, column=4, value='-')
                ws1.cell(row=row_num, column=5, value=None)
                ws1.cell(row=row_num, column=6, value='-')
                ws1.cell(row=row_num, column=7, value=None)
                for col in range(8, 18):
                    ws1.cell(row=row_num, column=col, value='')
            else:
                ws1.cell(row=row_num, column=2, value='-')
                ws1.cell(row=row_num, column=3, value='-')
                ws1.cell(row=row_num, column=4, value='-')
                ws1.cell(row=row_num, column=5, value='-')
                ws1.cell(row=row_num, column=6, value='-')
                ws1.cell(row=row_num, column=7, value='-')
                ws1.cell(row=row_num, column=8, value=None)
                for col in range(9, 18):
                    ws1.cell(row=row_num, column=col, value='')
        else:
            ws1.cell(row=row_num, column=2, value='-')
            for col in (3, 4, 5, 6):
                ws1.cell(row=row_num, column=col, value='-')
            for col in range(7, 16):
                ws1.cell(row=row_num, column=col, value='')
        _write_and_style_row(row_num)

    row_end = min(last_pad_row, 500)
    if footer_row is not None and footer_row <= row_end + 1:
        row_end = min(row_end, footer_row - 1)
    row_end = max(row_end, data_start)
    if include_th_column:
        if th_kind == 'compact':
            _exam_sheet1_normalize_body_zeros(
                ws1, data_start, row_end, data_dash_cols=(3, 4, 6), notes_col=7
            )
        else:
            _exam_sheet1_normalize_body_zeros(
                ws1, data_start, row_end, data_dash_cols=(4, 5, 6, 7), notes_col=8
            )
    else:
        _exam_sheet1_normalize_body_zeros(ws1, data_start, row_end)


def _exam_export_workbook_paths() -> dict[str, Path]:
    """
    Đường dẫn mẫu Excel xếp phòng thi.
    Mẫu danh sách + sơ đồ chuẩn: ``{BASE_DIR}/templates_excel/EXPORT_TEMPLATE.xlsx``
    (ví dụ: ``.../education-blog/templates_excel/EXPORT_TEMPLATE.xlsx``).
    """
    base_dir = getattr(settings, 'BASE_DIR', None)
    if base_dir is None:
        base_dir = Path(__file__).resolve().parent.parent
    else:
        base_dir = Path(base_dir)
    td = base_dir / 'templates_excel'
    path_so_do = td / 'SO DO CHO NGOI PHONG THI HK I.xlsx'
    if not path_so_do.exists():
        path_so_do = td / 'SƠ ĐỒ CHỖ NGỒI PHÒNG THI HK I.xlsx'
    return {
        'templates_dir': td,
        'export_template': td / 'EXPORT_TEMPLATE.xlsx',
        'ds_thi_tap_trung': td / 'DS.THI TAP TRUNG.xlsx',
        'so_do': path_so_do,
    }


def _build_exam_room_export_workbook_for_shift(room: ExamRoom, shift: str):
    """
    Tạo workbook 2 sheet (danh sách + sơ đồ) cho một phòng và một ca.
    Dùng chung cho export đơn và export gộp nhiều phòng.
    """
    shift_label = dict(StudentExamAssignment.SHIFT_CHOICES).get(shift, shift)

    assignments = list(
        StudentExamAssignment.objects.select_related('student', 'student__campus', 'student__subject_group')
        .filter(exam_room=room, shift=shift)
    )
    assignments.sort(key=_exam_assignment_export_list_order_key)
    subjects = list(
        ExamRoomSubject.objects.filter(exam_room=room, shift=shift)
        .values_list('subject_name', flat=True)
        .distinct()
    )[:4]

    # Sheet "In_GIAO NHAN BAI": thứ tự môn cố định theo mẫu:
    # 1) Văn, 2) Môn tự chọn, 3) Toán, 4) Sử — hiển thị IN HOA; môn tự chọn trống → '-'
    fixed_subjects = ['Toán', 'Văn', 'Sử']
    extra_subject = ''
    for s in subjects:
        if s and s not in fixed_subjects:
            extra_subject = s
            break
    subjects_ws1 = []
    for lab in ('Văn', (extra_subject or '').strip() or None, 'Toán', 'Sử'):
        if lab is None:
            subjects_ws1.append('-')
        else:
            subjects_ws1.append(str(lab).strip().upper())
    so_do_header_label = _build_exam_export_combo_label(assignments)

    tpl = _exam_export_workbook_paths()
    path_export_template = tpl['export_template']
    path_giao_nhan = tpl['ds_thi_tap_trung']
    path_so_do = tpl['so_do']
    # Ưu tiên EXPORT_TEMPLATE.xlsx (2 sheet: In_GIAO NHAN BAI + 4_Sơ đồ mẫu).
    use_templates = path_export_template.exists() or (path_giao_nhan.exists() and path_so_do.exists())

    if use_templates:
        if path_export_template.exists():
            wb = load_workbook(path_export_template, data_only=False)
            if 'In_GIAO NHAN BAI' in wb.sheetnames:
                ws1 = wb['In_GIAO NHAN BAI']
            else:
                ws1 = wb.active
                ws1.title = 'In_GIAO NHAN BAI'
        else:
            wb = load_workbook(path_giao_nhan, data_only=False)
            sheet_names = [s.title for s in wb.worksheets]
            if 'In_GIAO NHAN BAI' in sheet_names:
                ws1 = wb['In_GIAO NHAN BAI']
                for s in reversed(sheet_names):
                    if s != 'In_GIAO NHAN BAI':
                        del wb[s]
            else:
                ws1 = wb.active
                ws1.title = 'In_GIAO NHAN BAI'
                for s in reversed(sheet_names):
                    if s != ws1.title:
                        del wb[s]
        cap = int(getattr(room, 'capacity', 0) or 0)
        hdr_row = _exam_sheet1_guess_stt_header_row(ws1)
        use_th_col = _exam_sheet1_template_has_th_column(ws1, hdr_row)
        _fill_exam_giao_nhan_merged(
            ws1,
            [(room, assignments)],
            shift,
            global_stt=False,
            include_th_column=use_th_col,
            min_data_rows=cap,
        )
        # Sheet 2
        if path_export_template.exists():
            so_do_sheet_name = None
            for name in ('4_Sơ đồ mẫu', '4_So do mau'):
                if name in wb.sheetnames:
                    so_do_sheet_name = name
                    break
            if so_do_sheet_name:
                ws2 = wb[so_do_sheet_name]
            else:
                ws2 = wb.create_sheet('4_Sơ đồ mẫu')
            # dùng chính sheet này làm nguồn values luôn
            ws_src_values = ws2
            ws_src_style = ws2
        else:
            wb_so_do = load_workbook(path_so_do, data_only=False)
            wb_so_do_values = load_workbook(path_so_do, data_only=True)
            so_do_sheet_name = None
            for name in ('4_Sơ đồ mẫu', '4_So do mau'):
                if name in wb_so_do.sheetnames:
                    so_do_sheet_name = name
                    break
            if so_do_sheet_name:
                ws_src_style = wb_so_do[so_do_sheet_name]
                ws_src_values = (
                    wb_so_do_values[so_do_sheet_name]
                    if so_do_sheet_name in wb_so_do_values.sheetnames
                    else wb_so_do_values.active
                )
                ws2 = _copy_worksheet_to_workbook(ws_src_style, wb, sheet_name='4_Sơ đồ mẫu')
            else:
                ws_src_style = None
                ws_src_values = None
                ws2 = wb.create_sheet('4_Sơ đồ mẫu')
            wb_so_do.close()
            wb_so_do_values.close()

        # Suy ra vị trí/độ rộng các khối sơ đồ từ template (tự khớp nếu lớp có layout khác: 2x6, 3x6, 4x5,...)
        if ws_src_style and ws_src_values:
            block_starts = _get_block_starts_from_template(ws_src_style)
            if len(block_starts) >= 2:
                block_w = block_starts[1] - block_starts[0]
            else:
                block_w = 7

            block_seat_cells: dict[int, list[tuple[int, int]]] = {}
            for start in block_starts:
                # Ưu tiên dò theo số mẫu trong template (data_only) để bám đúng "cột ghế" (tránh lối đi).
                seats = _detect_seat_cells_from_template(ws_src_values, start, block_w)
                if not seats:
                    seats = _fallback_seat_cells_by_border(ws2, start, block_w)
                block_seat_cells[start] = seats

            for subj_idx in range(4):
                if subj_idx >= len(block_starts):
                    break
                base_c = block_starts[subj_idx]
                base_r = 1
                seat_cells = block_seat_cells.get(base_c, [])

                if subj_idx < len(subjects):
                    subject_name = subjects[subj_idx]
                    _set_cell_value(ws2, base_r, base_c, so_do_header_label)
                    _set_cell_value(
                        ws2,
                        base_r + 2,
                        base_c,
                        (subject_name or '').strip().upper(),
                    )
                    _set_cell_value(ws2, base_r + 2, base_c + 4, f'P. {room.name}')

                    # Nhận diện ô ghế:
                    # - Ưu tiên dò theo BORDER để bắt cả các ô ghế template đang để trống (không có công thức)
                    # - Bổ sung thêm các ô có công thức/giá trị trong template (nếu có)
                    ordered_cells = _detect_seat_cells_for_export(ws2, base_c, block_w) or seat_cells

                    # Clear tất cả ô ghế trước khi điền
                    for (r, c) in ordered_cells:
                        _set_cell_value(ws2, r, c, '')

                    # Điền theo layout RIÊNG của từng môn (ExamSubjectSeat) để 4 môn có 4 sơ đồ khác nhau như trên web
                    seat_list = list(
                        ExamSubjectSeat.objects.select_related('student')
                        .filter(exam_room=room, shift=shift, subject_name=subject_name)
                        .order_by('seat_number')
                    )
                    for s in seat_list:
                        if not s.student:
                            continue
                        idx = (s.seat_number or 0) - 1
                        if idx < 0 or idx >= len(ordered_cells):
                            continue
                        r, c = ordered_cells[idx]
                        val = s.student.exam_number or s.student.student_code or ''
                        if isinstance(val, float) and val == int(val):
                            val = int(val)
                        _set_cell_value(ws2, r, c, val)

                    # Xoá giá trị số "rác" ở các ô không phải ghế (ví dụ cột lối đi bị hiện số 3 trong template)
                    if ordered_cells:
                        grid_start = min(r for r, _ in ordered_cells)
                        grid_end = max(r for r, _ in ordered_cells)
                        seat_set = set(ordered_cells)
                        for rr in range(grid_start, grid_end + 1):
                            for cc in range(base_c, base_c + block_w):
                                if (rr, cc) in seat_set:
                                    continue
                                # Luôn clear ô không phải ghế trong vùng lưới để hành lang trống.
                                # (Template có thể có công thức; nếu không clear, Excel mở lên sẽ tự tính lại thành '3',...)
                                _set_cell_value(ws2, rr, cc, '')
                else:
                    _set_cell_value(ws2, base_r, base_c, '')
                    _set_cell_value(ws2, base_r + 2, base_c, '')
                    _set_cell_value(ws2, base_r + 2, base_c + 4, '')
                    for (r, c) in seat_cells:
                        _set_cell_value(ws2, r, c, '')
        # Nếu không dò được template (thiếu sheet) thì giữ logic cũ/fallback phía dưới
    else:
        combo_label = _build_exam_export_combo_label(assignments)
        wb = Workbook()
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        thick_blue_side = Side(style='medium', color='0000FF')
        thick_blue_border = Border(
            left=thick_blue_side, right=thick_blue_side,
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        font_title_center = Font(name='Times New Roman', size=13, bold=True)
        font_subtitle = Font(name='Times New Roman', size=11, italic=True)
        font_header_table = Font(name='Times New Roman', size=11, bold=True)
        font_data = Font(name='Times New Roman', size=11)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')
        blue_font_large = Font(name='Times New Roman', size=14, bold=True, color='023EB6')
        black_bold = Font(name='Times New Roman', size=11, bold=True)
        red_font = Font(name='Times New Roman', size=11, bold=True, color='CC0000')
        blue_font_sbd = Font(name='Times New Roman', size=11, color='0000FF')
        red_font_sbd = Font(name='Times New Roman', size=11, color='CC0000')
        font_normal = Font(name='Times New Roman', size=11)
        ws1 = wb.active
        ws1.title = "In_GIAO NHAN BAI"
        ws1.cell(row=1, column=1, value='SỞ GD&ĐT TP HỒ CHÍ MINH')
        ws1.cell(row=1, column=1).font = font_title_center
        ws1.cell(row=1, column=1).alignment = align_center
        ws1.cell(row=1, column=10, value='CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM')
        ws1.cell(row=1, column=10).font = font_title_center
        ws1.cell(row=1, column=10).alignment = align_center
        ws1.cell(row=2, column=1, value='TRUNG TÂM GDNN - GDTX THỦ ĐỨC')
        ws1.cell(row=2, column=1).font = font_subtitle
        ws1.cell(row=2, column=1).alignment = align_center
        ws1.cell(row=2, column=10, value='Độc lập - Tự do - Hạnh phúc')
        ws1.cell(row=2, column=10).font = font_subtitle
        ws1.cell(row=2, column=10).alignment = align_center
        ws1.cell(row=3, column=1, value='DANH SÁCH GIAO NHẬN BÀI KIỂM TRA CUỐI HỌC KÌ I')
        ws1.cell(row=3, column=1).font = Font(name='Times New Roman', size=14, bold=True)
        ws1.cell(row=3, column=1).alignment = align_center
        today_str = timezone.localdate().strftime('%d/%m/%Y')
        ws1.cell(row=4, column=1, value='')
        ws1.cell(row=4, column=1).font = font_data
        ws1.cell(row=4, column=2, value=f'Ngày: {today_str}')
        ws1.cell(row=4, column=2).font = font_data
        ws1.cell(row=4, column=9, value=combo_label)
        ws1.cell(row=4, column=9).font = font_header_table
        room_title_fb = str(room.name or '').strip() or '-'
        ws1.cell(row=4, column=13, value=f'PHÒNG THI: {room_title_fb}')
        ws1.cell(row=4, column=13).font = font_header_table
        ws1.cell(row=4, column=14, value='')
        ws1.cell(row=4, column=14).font = font_header_table
        ws1.cell(row=4, column=15, value='')
        ws1.cell(row=4, column=15).font = font_header_table
        ws1.cell(row=6, column=1, value='Stt')
        ws1.cell(row=6, column=2, value='Phòng\nthi')
        ws1.cell(row=6, column=3, value='SBD')
        ws1.cell(row=6, column=4, value='Họ và')
        ws1.cell(row=6, column=5, value='tên')
        ws1.cell(row=6, column=6, value='Lớp')
        ws1.cell(row=6, column=7, value='Ghi\nchú')
        ws1.cell(row=6, column=8, value='MÔN THI')
        for c in range(1, 16):
            cell = ws1.cell(row=6, column=c)
            cell.font = font_header_table
            cell.alignment = align_center
            cell.border = thin_border
            if c > 8:
                cell.value = cell.value if cell.value else ''
        for i, subj in enumerate(subjects_ws1[:4]):
            c = ws1.cell(row=7, column=8 + i * 2, value=subj)
            c.font = font_data
            c.alignment = align_center
            c.border = thin_border
        for col, val in [(8, 'Số tờ'), (9, 'Ký tên'), (10, 'Mã đề'), (11, 'Ký tên'), (12, 'Mã đề'), (13, 'Ký tên'), (14, 'Mã đề'), (15, 'Ký tên')]:
            c = ws1.cell(row=8, column=col, value=val)
            c.font = font_data
            c.alignment = align_center
            c.border = thin_border
        cap_fb = int(getattr(room, 'capacity', 0) or 0)
        for idx, a in enumerate(assignments, start=1):
            stu = a.student
            ho, ten = _split_ho_ten((stu.full_name or '') if stu else '')
            sbd_raw = (stu.exam_number or stu.student_code or '') if stu else ''
            sbd = _exam_export_cell_dash(sbd_raw)
            lop = _exam_export_cell_dash((stu.class_name or '') if stu else '')
            ho = _exam_export_cell_dash(ho)
            ten = _exam_export_cell_dash(ten)
            ghi_chu = 'HN' if (stu and getattr(stu, 'is_integration', False)) else '-'
            row_num = 9 + idx - 1
            for col, val, align in [
                (1, idx, align_center), (2, f'P.{room.name}', align_center), (3, sbd, align_center),
                (4, ho, align_left), (5, ten, align_left), (6, lop, align_center), (7, ghi_chu, align_center),
            ]:
                cell = ws1.cell(row=row_num, column=col, value=val)
                cell.font = font_data
                cell.alignment = align
                cell.border = thin_border
            for col in range(8, 16):
                cell = ws1.cell(row=row_num, column=col, value='')
                cell.font = font_data
                cell.border = thin_border
        n_fb = len(assignments)
        for row_num in range(9 + n_fb, 9 + cap_fb):
            cell = ws1.cell(row=row_num, column=1, value=row_num - 8)
            cell.font = font_data
            cell.border = thin_border
            cell = ws1.cell(row=row_num, column=2, value='-')
            cell.font = font_data
            cell.border = thin_border
            for col in (3, 4, 5, 6):
                cell = ws1.cell(row=row_num, column=col, value='-')
                cell.font = font_data
                cell.border = thin_border
            for col in range(7, 16):
                cell = ws1.cell(row=row_num, column=col, value='')
                cell.font = font_data
                cell.border = thin_border
        row_end_fb = max(9 + n_fb - 1, 8 + cap_fb, 9)
        mr_fb = ws1.max_row or row_end_fb
        row_end_fb = min(max(row_end_fb, mr_fb), 250)
        _exam_sheet1_normalize_body_zeros(ws1, 9, row_end_fb)
        for c in range(1, 16):
            ws1.column_dimensions[get_column_letter(c)].width = 12
        ws2 = wb.create_sheet("4_Sơ đồ mẫu", 1)
        block_w = 8
        data_cols = [0, 1, 3, 4]
        grid_rows = 6
        seat_to_cell = []
        for c in range(4):
            if c == 0:
                for r in range(5):
                    seat_to_cell.append((r, 0))
            elif c == 1:
                for r in range(6):
                    seat_to_cell.append((r, 1))
            elif c == 2:
                for r in range(6):
                    seat_to_cell.append((r, 2))
            else:
                for r in range(5):
                    seat_to_cell.append((r, 3))
        for subj_idx, subject_name in enumerate(subjects):
            base_c = 1 + subj_idx * block_w
            base_r = 1
            ws2.cell(row=base_r, column=base_c, value=so_do_header_label)
            ws2.cell(row=base_r, column=base_c).font = blue_font_large
            ws2.cell(row=base_r, column=base_c).alignment = align_left
            ws2.cell(row=base_r + 1, column=base_c, value='Sơ đồ chỗ ngồi')
            ws2.cell(row=base_r + 1, column=base_c).font = black_bold
            ws2.cell(row=base_r + 1, column=base_c).alignment = align_left
            ws2.cell(row=base_r + 2, column=base_c, value=(subject_name or '').strip().upper())
            ws2.cell(row=base_r + 2, column=base_c).font = blue_font_large
            ws2.cell(row=base_r + 2, column=base_c).alignment = align_left
            ws2.cell(row=base_r + 2, column=base_c + 4, value=f'P. {room.name}')
            ws2.cell(row=base_r + 2, column=base_c + 4).font = red_font
            ws2.cell(row=base_r + 2, column=base_c + 4).alignment = align_right
            ws2.cell(row=base_r + 3, column=base_c, value='Bàn Giáo viên')
            ws2.cell(row=base_r + 3, column=base_c).font = font_normal
            ws2.cell(row=base_r + 3, column=base_c).alignment = align_left
            seat_list = list(
                ExamSubjectSeat.objects.select_related('student')
                .filter(exam_room=room, shift=shift, subject_name=subject_name)
                .order_by('seat_number')
            )
            seat_to_sbd = {}
            for s in seat_list:
                if s.student:
                    val = s.student.exam_number or s.student.student_code or ''
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    seat_to_sbd[s.seat_number] = val
            total_seats = len(seat_to_cell)
            for seat_num in range(1, total_seats + 1):
                if seat_num > total_seats:
                    break
                r0, c0 = seat_to_cell[seat_num - 1]
                cell_r = base_r + 5 + r0
                cell_c = base_c + data_cols[c0]
                sbd = seat_to_sbd.get(seat_num, '')
                if isinstance(sbd, float) and sbd == int(sbd):
                    sbd = int(sbd)
                cell = ws2.cell(row=cell_r, column=cell_c, value=sbd)
                cell.alignment = align_center
                cell.border = thin_border
                cell.font = font_normal
                if seat_num == 1:
                    cell.font = red_font_sbd
                elif seat_num == total_seats:
                    cell.font = blue_font_sbd
            thin_side = Side(style='thin')
            for r in range(base_r, base_r + 5 + grid_rows):
                cell_left = ws2.cell(row=r, column=base_c)
                cell_left.border = Border(left=thick_blue_side, right=thin_side, top=thin_side, bottom=thin_side)
                cell_right = ws2.cell(row=r, column=base_c + block_w - 1)
                cell_right.border = Border(left=thin_side, right=thick_blue_side, top=thin_side, bottom=thin_side)
            for c in range(block_w):
                ws2.column_dimensions[get_column_letter(base_c + c)].width = 10
        if not subjects:
            ws2.cell(row=1, column=1, value=f'Phòng {room.name} - Ca: {shift_label}')
            ws2.cell(row=2, column=1, value='Chưa có môn thi nào. Chọn môn trong chi tiết phòng thi.')
    return wb


def _exam_merged_list_prepare_sheet(wb_out: Workbook, blocks: list, shift: str, suggested_base: str):
    """
    Thêm vào wb_out một sheet mẫu In_GIAO NHAN BAI (ưu tiên copy từ ``templates_excel/EXPORT_TEMPLATE.xlsx``).
    Trả về worksheet vừa tạo (để gọi _fill_exam_giao_nhan_merged).
    """
    first_room, _ = blocks[0]
    sheet_name = _exam_bulk_unique_sheet_name(wb_out, suggested_base)
    tpl_paths = _exam_export_workbook_paths()
    path_export_template = tpl_paths['export_template']
    path_giao_nhan = tpl_paths['ds_thi_tap_trung']

    if path_export_template.exists():
        tpl = load_workbook(path_export_template, data_only=False)
        try:
            if 'In_GIAO NHAN BAI' in tpl.sheetnames:
                ws_src = tpl['In_GIAO NHAN BAI']
            else:
                ws_src = tpl.active
            _copy_worksheet_to_workbook(ws_src, wb_out, sheet_name=sheet_name)
        finally:
            tpl.close()
        return wb_out[sheet_name]

    if path_giao_nhan.exists():
        tpl = load_workbook(path_giao_nhan, data_only=False)
        try:
            sn_list = [s.title for s in tpl.worksheets]
            if 'In_GIAO NHAN BAI' in sn_list:
                ws_src = tpl['In_GIAO NHAN BAI']
            else:
                ws_src = tpl.active
            _copy_worksheet_to_workbook(ws_src, wb_out, sheet_name=sheet_name)
        finally:
            tpl.close()
        return wb_out[sheet_name]

    wb_seed = _build_exam_room_export_workbook_for_shift(first_room, shift)
    ws_src = None
    for t in list(wb_seed.sheetnames):
        tit_up = (t or '').upper()
        if 'GIAO' in tit_up or (t or '').startswith('In_'):
            ws_src = wb_seed[t]
            break
    if ws_src is None:
        ws_src = wb_seed.active
    _copy_worksheet_to_workbook(ws_src, wb_out, sheet_name=sheet_name)
    return wb_out[sheet_name]


def _build_bulk_multi_room_workbook(rooms: list[ExamRoom]) -> Workbook:
    """
    Một workbook gộp khi export nhiều phòng:
    - Sheet danh sách: mỗi (cơ sở, ca) một sheet — định dạng mẫu từ ``templates_excel/EXPORT_TEMPLATE.xlsx``
      (sheet In_GIAO NHAN BAI), gộp tất cả phòng của cơ sở trong ca; STT 1..n theo từng phòng; mỗi phòng sắp theo SBD tăng dần; cột TỔ HỢP (TH1…) ở cột B (Phòng → C, SBD → D, …).
    - Sheet sơ đồ: mỗi (phòng, ca) một sheet (4 môn trên cùng sheet như export đơn).
    """
    shift_order = [c[0] for c in StudentExamAssignment.SHIFT_CHOICES]
    assignments_index: dict[tuple[int, str], list] = {}

    for room in rooms:
        for shift in shift_order:
            assigns = list(
                StudentExamAssignment.objects.select_related('student', 'student__subject_group')
                .filter(exam_room_id=room.id, shift=shift)
            )
            if not assigns:
                continue
            assigns.sort(key=_exam_assignment_export_list_order_key)
            assignments_index[(room.id, shift)] = assigns

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for shift in shift_order:
        active_rooms = [r for r in rooms if (r.id, shift) in assignments_index]
        if not active_rooms:
            continue
        by_campus: dict[int, list[ExamRoom]] = defaultdict(list)
        for room in active_rooms:
            by_campus[room.campus_id].append(room)

        def _campus_key(cid: int) -> tuple:
            r0 = next(r for r in active_rooms if r.campus_id == cid)
            code = (r0.campus.code or '') if r0.campus else ''
            return natural_sort_key(code)

        for campus_id in sorted(by_campus.keys(), key=_campus_key):
            room_list = by_campus[campus_id]
            room_list.sort(key=lambda r: natural_sort_key(r.name or ''))
            campus = room_list[0].campus
            cc = re.sub(r'[^\w\-]', '', (getattr(campus, 'code', None) or '') or '')[:8] or str(campus.pk)
            blocks = [(r, assignments_index[(r.id, shift)]) for r in room_list]
            base_sn = f'DS_{cc}_{shift}'
            ws1 = _exam_merged_list_prepare_sheet(wb_out, blocks, shift, base_sn)
            _fill_exam_giao_nhan_merged(
                ws1,
                blocks,
                shift,
                global_stt=False,
                include_th_column=True,
            )

    for room in sorted(
        rooms,
        key=lambda r: (natural_sort_key((r.campus.code or '') if r.campus else ''), natural_sort_key(r.name or '')),
    ):
        for shift in shift_order:
            if (room.id, shift) not in assignments_index:
                continue
            wb_part = _build_exam_room_export_workbook_for_shift(room, shift)
            for title in list(wb_part.sheetnames):
                tit_up = (title or '').upper()
                if 'GIAO' in tit_up or (title or '').startswith('In_'):
                    continue
                ws_src = wb_part[title]
                safe_room = re.sub(r'[^\w\-]', '_', (room.name or str(room.id)).strip())[:10]
                sn = _exam_bulk_unique_sheet_name(wb_out, f'SD_{safe_room}_{shift}')
                _copy_worksheet_to_workbook(ws_src, wb_out, sheet_name=sn)

    return wb_out


def exam_room_detail_export_excel(request, room_id: int):
    """
    Export Excel chi tiết phòng thi: 2 sheet.
    - Sheet 1: Chỉ cột A-O, bố cục/định dạng giống sheet "In_GIAO NHAN BAI" trong ``templates_excel/EXPORT_TEMPLATE.xlsx``, chỉ thay dữ liệu.
    - Sheet 2: Bố cục sheet "4_Sơ đồ mẫu" (SƠ ĐỒ CHỖ NGỒI PHÒNG THI HK I) - 4 sơ đồ y chang (TỔ HỢP, Sơ đồ chỗ ngồi, Môn, P.xx, Bàn GV, lưới ghế 2 khối), chỉ thay dữ liệu.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Homepage')

    room = get_object_or_404(ExamRoom, id=room_id)
    shift = request.GET.get('shift', 'sang')
    wb = _build_exam_room_export_workbook_for_shift(room, shift)
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_name = re.sub(r'[^\w\s-]', '', room.name).strip()[:30]
    resp['Content-Disposition'] = f'attachment; filename="phong_{room_id}_{safe_name}_{shift}.xlsx"'
    wb.save(resp)
    return resp


def _exam_room_bulk_xlsx_filename(room: ExamRoom) -> str:
    """Tên file .xlsx theo cơ sở + tên phòng (tránh ký tự không hợp lệ)."""
    code = re.sub(r'[^\w\-]', '', (getattr(room.campus, 'code', None) or '') or '')[:12]
    name = re.sub(r'[^\w\-\.]', '_', (room.name or str(room.id)).strip())[:40]
    base = '_'.join(p for p in (code, name) if p).strip('_') or f'phong_{room.id}'
    return f'{base}.xlsx'


def exam_room_bulk_export_excel(request):
    """
    POST: chọn một hoặc nhiều phòng thi.
    - Một phòng: một file .xlsx (mỗi ca có HV: 2 sheet danh sách + sơ đồ), tải trực tiếp.
    - Nhiều phòng: một file .xlsx gộp — mỗi (cơ sở, ca) một sheet danh sách theo mẫu In_GIAO NHAN BAI
      (STT theo từng phòng 1..n, cột TỔ HỢP TH1… ở cột B); sơ đồ mỗi (phòng, ca) một sheet.
    """
    if not request.user.is_authenticated:
        return redirect('homepage:login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Homepage')

    cid = (request.POST.get('return_campus_id') or '').strip()
    manage_url = reverse('adminpage:exam_room_manage')
    if cid.isdigit():
        redirect_manage = redirect(f'{manage_url}?campus_id={cid}')
    else:
        redirect_manage = redirect('adminpage:exam_room_manage')

    if request.method != 'POST':
        messages.warning(request, 'Vui lòng chọn phòng và bấm Export từ trang Quản lý phòng thi.')
        return redirect_manage

    raw_ids = request.POST.getlist('room_ids')
    id_set: set[int] = set()
    for x in raw_ids:
        s = str(x).strip()
        if s.isdigit():
            id_set.add(int(s))
    if not id_set:
        messages.error(request, 'Chưa chọn phòng thi nào.')
        return redirect_manage

    rooms = list(
        ExamRoom.objects.filter(id__in=id_set).select_related('campus').order_by('campus__code', 'name')
    )
    if not rooms:
        messages.error(request, 'Không tìm thấy phòng thi hợp lệ.')
        return redirect_manage

    shift_order = [c[0] for c in StudentExamAssignment.SHIFT_CHOICES]

    if len(rooms) > 1:
        has_any = StudentExamAssignment.objects.filter(exam_room_id__in=[r.id for r in rooms]).exists()
        if not has_any:
            messages.warning(
                request,
                'Không có phòng nào trong lựa chọn có học viên ở bất kỳ ca nào — không export được file.',
            )
            return redirect_manage
        wb = _build_bulk_multi_room_workbook(rooms)
        buf = BytesIO()
        wb.save(buf)
        ts = timezone.localdate().strftime('%Y%m%d')
        fname = f'phong_thi_gop_{ts}.xlsx'
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    # (tên file trong zip / response, nội dung bytes)
    outputs: list[tuple[str, bytes]] = []
    used_zip_names: set[str] = set()

    for room in rooms:
        shifts_with_students = [
            sh
            for sh in shift_order
            if StudentExamAssignment.objects.filter(exam_room=room, shift=sh).exists()
        ]
        if not shifts_with_students:
            continue

        wb_room = Workbook()
        wb_room.remove(wb_room.active)
        for shift in shifts_with_students:
            wb_part = _build_exam_room_export_workbook_for_shift(room, shift)
            for title in list(wb_part.sheetnames):
                ws_src = wb_part[title]
                kind = 'DS' if ('GIAO' in (title or '').upper() or title.startswith('In_')) else 'SD'
                base = f'{room.id}_{shift}_{kind}'
                sn = _exam_bulk_unique_sheet_name(wb_room, base)
                _copy_worksheet_to_workbook(ws_src, wb_room, sheet_name=sn)

        buf = BytesIO()
        wb_room.save(buf)
        xlsx_bytes = buf.getvalue()

        arc = _exam_room_bulk_xlsx_filename(room)
        if arc in used_zip_names:
            stem = arc[:-5] if arc.lower().endswith('.xlsx') else arc
            arc = f'{stem}_{room.id}.xlsx'
        used_zip_names.add(arc)
        outputs.append((arc, xlsx_bytes))

    if not outputs:
        messages.warning(
            request,
            'Không có phòng nào trong lựa chọn có học viên ở bất kỳ ca nào — không export được file.',
        )
        return redirect_manage

    if len(outputs) == 1:
        arc, xlsx_bytes = outputs[0]
        resp = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{arc}"'
        return resp

    ts = timezone.localdate().strftime('%Y%m%d')
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for arc, xlsx_bytes in outputs:
            zf.writestr(arc, xlsx_bytes)
    zip_name = f'phong_thi_{ts}.zip'
    resp = HttpResponse(zip_buf.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    return resp


def _perform_exam_room_initial_assignment(
    room,
    shift: str,
    *,
    grade: str,
    group_base: str,
    hv_shift: str,
    limit: int,
    extra_subject: str,
    integration_quota: int = 0,
) -> tuple[int, str | None, list]:
    """
    Cùng logic với action 'initial_pick' (exam_room_detail): tạo StudentExamAssignment,
    ExamRoomSubject (Toán/Văn/Sử + môn thêm), đánh ghế, regenerate layout theo môn.
    Kiểm tra môn thêm trước khi gán học viên (tránh gán rồi mới báo lỗi).
    Trả về (số HV đã gán, None, danh_sách_ExamRoomStudent) hoặc (0, lỗi, []).
    """
    fixed_subjects = ['Toán', 'Văn', 'Sử']
    extra_subject = (extra_subject or '').strip()
    if hv_shift not in ('sang', 'toi'):
        return 0, 'Vui lòng chọn học viên buổi sáng hay buổi tối (theo mã HV: 0 = sáng, 1 = tối).', []
    if not grade or not group_base:
        return 0, 'Vui lòng chọn đủ khối và tổ hợp môn.', []
    if limit <= 0:
        return 0, 'Vui lòng nhập số lượng học viên hợp lệ (> 0).', []
    if integration_quota < 0:
        return 0, 'Số lượng hoà nhập không hợp lệ (>= 0).', []
    if integration_quota > limit:
        return 0, f'Số lượng hoà nhập ({integration_quota}) không được lớn hơn số lượng học viên ({limit}).', []
    if not extra_subject:
        return 0, 'Vui lòng chọn thêm 1 môn còn lại (ngoài Toán, Văn, Sử).', []
    if extra_subject in fixed_subjects:
        return 0, 'Môn chọn thêm không được trùng Toán/Văn/Sử.', []

    assigned_ids = set(
        StudentExamAssignment.objects.filter(shift=shift).values_list('student_id', flat=True)
    )
    current_count = StudentExamAssignment.objects.filter(exam_room=room, shift=shift).count()
    shift_capacity = _get_room_capacity_for_shift(room, shift)
    remaining_capacity = max(shift_capacity - current_count, 0)
    if remaining_capacity <= 0:
        return 0, f'Phòng {room.name} ca {shift} đã đủ {shift_capacity} chỗ.', []

    limit = min(limit, remaining_capacity)
    expect_hv_digit = '0' if hv_shift == 'sang' else '1'
    campus_students = ExamRoomStudent.objects.filter(campus=room.campus).select_related('subject_group')
    candidates: list[ExamRoomStudent] = []
    for s in campus_students:
        if s.pk in assigned_ids:
            continue
        if _exam_hv_shift_digit_from_code(s.student_code or '') != expect_hv_digit:
            continue
        g = _extract_grade_from_class_name(s.class_name)
        if not g or str(g) != str(grade).strip():
            continue
        code = s.subject_group.code if s.subject_group else ''
        if not code or code[-1] != group_base:
            continue
        candidates.append(s)

    candidates.sort(key=lambda stu: (stu.exam_number or '', _name_sort_key_vi(stu.full_name or "")))
    integration_candidates = [s for s in candidates if bool(getattr(s, 'is_integration', False))]
    normal_candidates = [s for s in candidates if not bool(getattr(s, 'is_integration', False))]

    integration_need = min(integration_quota, limit)
    if len(integration_candidates) < integration_need:
        return (
            0,
            (
                f'Không đủ học viên hoà nhập theo yêu cầu: cần {integration_need}, '
                f'chỉ có {len(integration_candidates)}.'
            ),
            [],
        )

    picked: list[ExamRoomStudent] = []
    picked.extend(integration_candidates[:integration_need])
    remaining = limit - len(picked)
    if remaining > 0:
        picked.extend(normal_candidates[:remaining])
        remaining = limit - len(picked)
    if remaining > 0:
        # Nếu thiếu HV thường thì cho phép lấy thêm HN để vẫn đủ chỉ tiêu.
        picked.extend(integration_candidates[integration_need:integration_need + remaining])

    if len(picked) < limit:
        return (
            0,
            'Không có học viên phù hợp (khối, tổ hợp, buổi sáng/tối theo mã HV) hoặc đều đã có phòng trong ca.',
            [],
        )

    objs = [
        StudentExamAssignment(student=stu, exam_room=room, shift=shift)
        for stu in picked
    ]
    StudentExamAssignment.objects.bulk_create(objs)

    subjects = fixed_subjects + [extra_subject]
    ExamRoomSubject.objects.filter(exam_room=room, shift=shift).delete()
    for subj_name in subjects:
        ExamRoomSubject.objects.get_or_create(
            exam_room=room,
            shift=shift,
            subject_name=subj_name,
        )

    with transaction.atomic():
        autos = list(
            StudentExamAssignment.objects.select_for_update()
            .select_related('student')
            .filter(exam_room=room, shift=shift)
        )
        autos.sort(key=_exam_room_seat_order_key)
        seat_no = 1
        for a in autos:
            a.seat_number = seat_no
            seat_no += 1
        if autos:
            StudentExamAssignment.objects.bulk_update(autos, ['seat_number'])

    _regenerate_subject_layouts_for_room_shift(room, shift)
    return len(picked), None, picked


def _regenerate_subject_layouts_for_room_shift(room, shift: str):
    """
    Sinh lại layout ghế cho toàn bộ môn thi của một phòng + ca:
    - Danh sách học viên lấy từ StudentExamAssignment (phòng, ca).
    - Mỗi môn có một thứ tự ngẫu nhiên riêng (xáo trộn).
    - Ghi vào ExamSubjectSeat (exam_room, shift, subject_name, student, seat_number).
    """
    # Lấy toàn bộ học viên đang thuộc phòng/ca này
    assignments = list(
        StudentExamAssignment.objects.select_related('student')
        .filter(exam_room=room, shift=shift)
    )
    assignments.sort(key=_exam_room_seat_order_key)
    students = [a.student for a in assignments]

    # Nếu không có học viên thì xoá sạch layout cũ và thoát
    ExamSubjectSeat.objects.filter(exam_room=room, shift=shift).delete()
    if not students:
        return

    # Danh sách môn thi của phòng/ca này
    subjects = list(
        ExamRoomSubject.objects.filter(exam_room=room, shift=shift)
        .values_list('subject_name', flat=True)
        .distinct()
    )
    if not subjects:
        return

    # Giữ thứ tự thường trước / hòa nhập cuối (không sort thuần SBD — tránh HN nằm giữa danh sách).
    sorted_students = sorted(students, key=_exam_student_room_order_tuple)
    seat_orders_cache: dict[int, list[int]] = {}

    _room_rows, room_cols = _get_room_grid_for_shift(room, shift)

    def _seat_to_rc(seat_no: int) -> tuple[int, int]:
        cols = int(room_cols or 0)
        if cols <= 0:
            return (0, 0)
        idx0 = max(0, int(seat_no) - 1)
        return (idx0 // cols, idx0 % cols)

    def _direction_for_seat_in_block(seat_no: int, seat_order: list[int]) -> str:
        """
        Trả về 1 trong 4 hướng: TL/TR/BL/BR, tính theo "khối ghế đã fill" của seat_order.
        Nếu ghế không đúng ở góc tuyệt đối, phân loại theo top/bottom + left/right trong bounding box.
        """
        if not seat_order:
            return "TL"
        coords = [_seat_to_rc(s) for s in seat_order]
        rs = [r for (r, _c) in coords]
        cs = [c for (_r, c) in coords]
        min_r, max_r = min(rs), max(rs)
        min_c, max_c = min(cs), max(cs)
        r, c = _seat_to_rc(seat_no)
        # Exact corners first (hi-signal)
        if (r, c) == (min_r, min_c):
            return "TL"
        if (r, c) == (min_r, max_c):
            return "TR"
        if (r, c) == (max_r, min_c):
            return "BL"
        if (r, c) == (max_r, max_c):
            return "BR"
        # Otherwise classify by quadrant within bbox
        mid_r = (min_r + max_r) / 2.0
        mid_c = (min_c + max_c) / 2.0
        top = r <= mid_r
        left = c <= mid_c
        if top and left:
            return "TL"
        if top and not left:
            return "TR"
        if not top and left:
            return "BL"
        return "BR"

    def _get_seat_order_for_pattern(pattern_id: int, n_students: int) -> list[int]:
        if pattern_id in (1, 2, 3, 4):
            return _seat_positions_column_balanced(room, n_students, pattern_id, shift=shift)
        ff_dir_map = {5: 1, 6: 3, 7: 2, 8: 4}
        return _seat_positions_front_filled(room, n_students, ff_dir_map.get(pattern_id, 1), shift=shift)

    # Quy luật 4 môn:
    # - In riêng từng môn 1 sơ đồ (ExamSubjectSeat) đúng như web.
    # - Mỗi lần regenerate sẽ lấy NGẪU NHIÊN 4 kiểu trong 8 kiểu, với điều kiện:
    #   - 2 kiểu thuộc nhóm 1..4 (chia theo CỘT trước - balanced)
    #   - 2 kiểu thuộc nhóm 5..8 (lấp kín theo HÀNG trước - front-filled)
    #
    # Nhóm 1..4 (column-balanced):
    #   1: trái->phải, trên->dưới
    #   2: phải->trái, trên->dưới
    #   3: trái->phải, dưới->lên
    #   4: phải->trái, dưới->lên
    #
    # Nhóm 5..8 (front-filled):
    #   5: trên->dưới, trái->phải
    #   6: dưới->trên, trái->phải
    #   7: trên->dưới, phải->trái
    #   8: dưới->trên, phải->trái
    group_a = [1, 2, 3, 4]
    group_b = [5, 6, 7, 8]
    # Chọn 4 kiểu sao cho:
    # - 2 kiểu thuộc nhóm 1..4 và 2 kiểu thuộc nhóm 5..8
    # - SBD nhỏ nhất (seat_order[0]) của 4 môn nằm ở 4 hướng khác nhau (TL/TR/BL/BR)
    # - SBD lớn nhất (seat_order[n-1]) của 4 môn cũng nằm ở 4 hướng khác nhau
    # Nếu không thể (phòng quá nhỏ / quá ít học viên), fallback về random 2+2.
    directions: list[int] = []

    n_students = len(sorted_students)
    if n_students > 0:
        import itertools

        candidates_a = list(itertools.combinations(group_a, 2))
        candidates_b = list(itertools.combinations(group_b, 2))
        subjects4 = subjects[:4]

        found = None
        for a2 in candidates_a:
            for b2 in candidates_b:
                patt4 = list(a2) + list(b2)
                # Thử mọi cách gán 4 kiểu cho 4 môn
                for perm in itertools.permutations(patt4, len(subjects4)):
                    start_dirs: set[str] = set()
                    end_dirs: set[str] = set()
                    ok = True
                    for patt in perm:
                        so = _get_seat_order_for_pattern(patt, n_students)
                        if not so:
                            ok = False
                            break
                        start_dirs.add(_direction_for_seat_in_block(so[0], so))
                        end_dirs.add(_direction_for_seat_in_block(so[min(n_students, len(so)) - 1], so))
                    if ok and len(start_dirs) == len(subjects4) and len(end_dirs) == len(subjects4):
                        found = list(perm)
                        break
                if found:
                    break
            if found:
                break
        if found:
            directions = found

    if not directions:
        directions = random.sample(group_a, 2) + random.sample(group_b, 2)
        random.shuffle(directions)

    for idx, subj in enumerate(subjects[:4]):
        direction = directions[idx % len(directions)]
        if direction not in seat_orders_cache:
            seat_orders_cache[direction] = _get_seat_order_for_pattern(direction, n_students)
        seat_order = seat_orders_cache[direction]
        if not seat_order or n_students <= 0:
            continue

        objs = []
        for i, stu in enumerate(sorted_students):
            if i >= len(seat_order):
                break
            objs.append(
                ExamSubjectSeat(
                    exam_room=room,
                    shift=shift,
                    subject_name=subj,
                    student=stu,
                    seat_number=seat_order[i],
                )
            )
        ExamSubjectSeat.objects.bulk_create(objs)


def _exam_fill_missing_seats_in_room_shift(room, shift: str) -> None:
    """Gán seat_number cho bản ghi thiếu ghế trong đúng phòng+ca (không xáo trộn ghế đã có)."""
    with transaction.atomic():
        autos = list(
            StudentExamAssignment.objects.select_for_update()
            .filter(exam_room=room, shift=shift)
        )
        used = {a.seat_number for a in autos if a.seat_number is not None}
        missing = [a for a in autos if a.seat_number is None]
        if not missing:
            return
        n = 1
        cap = int(getattr(room, "capacity", 0) or 0)
        to_update: list = []
        for a in sorted(missing, key=lambda x: x.student_id):
            while n <= cap and n in used:
                n += 1
            if n > cap:
                break
            a.seat_number = n
            to_update.append(a)
            used.add(n)
            n += 1
        if to_update:
            StudentExamAssignment.objects.bulk_update(to_update, ["seat_number"])


def _sync_exam_sbd_campus_after_room_assignment(campus, *, force: bool = False) -> None:
    """
    Đồng bộ SBD cả cơ sở (ca → tổ hợp môn → phòng; chỉ ``exam_number``).
    ``force=True``: luôn đánh lại (sau xoá/thêm HV trong phòng). Mặc định: chỉ khi có HV hòa nhập.
    """
    _assign_exam_numbers_for_campus_sync_room_display(campus, force=force)


# ---------- Chấm công tự động ----------
THU_LABELS = ['', 'CN', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']  # index 2->8 tương ứng Thứ 2 -> CN


def _date_to_day_of_week(d):
    """Chuyển date sang day_of_week (2=Thứ 2, ..., 8=Chủ nhật)."""
    return d.weekday() + 2  # Monday=0 -> 2, Sunday=6 -> 8


def _week_ranges_in_month(year, month):
    """
    Trả về dict tuần -> (day_start, day_end) theo chuẩn tuần Thứ 2 -> Chủ nhật.
    Các ngày được cắt theo phạm vi trong tháng.
    """
    first_weekday, days_in_month = calendar.monthrange(year, month)  # Monday=0 ... Sunday=6
    max_week = ((days_in_month + first_weekday - 1) // 7) + 1
    ranges = {}
    for week_no in range(1, max_week + 1):
        day_start = max(1, (7 * (week_no - 1)) - first_weekday + 1)
        day_end = min(days_in_month, (7 * week_no) - first_weekday)
        ranges[week_no] = (day_start, day_end)
    return ranges


def _week_of_month(year, month, day):
    """Tuần trong tháng theo chuẩn tuần Thứ 2 -> Chủ nhật (có thể từ 4 đến 6 tuần/tháng)."""
    first_weekday, _ = calendar.monthrange(year, month)  # Monday=0 ... Sunday=6
    return ((day + first_weekday - 1) // 7) + 1


def _resolve_version_for_day(year, month, day, version_cache):
    """
    Áp dụng TKB theo tuần trong cùng tháng: nếu tuần hiện tại chưa nhập thì dùng TKB tuần trước (cascade).
    version_cache: dict (year, month, week) -> ScheduleVersion (hoặc None).
    """
    w = _week_of_month(year, month, day)
    candidates = [(year, month, wk) for wk in range(w, 0, -1)]  # (y, m, w), (y, m, w-1), ..., (y, m, 1)
    for (y, m, wk) in candidates:
        if wk < 1:
            continue
        v = version_cache.get((y, m, wk))
        if v is not None:
            return v
    return None


def attendance_pivot(request):
    """Giao diện bảng pivot chấm công: cột STT, Họ tên, Môn; hàng ngang ngày 1-31; ô = tiết quy đổi."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')

    now = timezone.now().date()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))
    location_id = request.GET.get('location_id', '')
    grade = request.GET.get('grade', '')
    size_group = request.GET.get('size_group', '')
    version_id = request.GET.get('version_id', '')

    pivot_rows, calendar_days, days_in_month, excluded_days = _build_attendance_pivot_data(
        year, month, location_id, grade, size_group, version_id
    )
    campuses = Campus.objects.exclude(code='AS').order_by('code')
    versions = ScheduleVersion.objects.all().order_by('-year', '-month', '-week')[:30]

    context = {
        'year': year,
        'month': month,
        'days_in_month': days_in_month,
        'calendar_days': calendar_days,
        'campuses': campuses,
        'versions': versions,
        'pivot_rows': pivot_rows,
        'excluded_days': sorted(excluded_days),
        'location_id': location_id,
        'grade': grade,
        'size_group': size_group,
        'version_id': version_id,
    }
    return render(request, 'adminpageSIMCODE/attendance_pivot.html', context)


def _build_attendance_pivot_data(year, month, location_id, grade, size_group, version_id):
    """Xây dựng pivot_rows và calendar_days cho bảng chấm công (dùng chung cho view và export)."""
    excluded_days = set(
        AttendanceExcludedDay.objects.filter(year=year, month=month).values_list('day', flat=True)
    )
    _, days_in_month = calendar.monthrange(year, month)
    calendar_days = []
    for d in range(1, days_in_month + 1):
        dt = datetime(year, month, d).date()
        calendar_days.append({
            'day': d,
            'weekday': _date_to_day_of_week(dt),
            'label': THU_LABELS[_date_to_day_of_week(dt)] if _date_to_day_of_week(dt) < len(THU_LABELS) else '',
        })
    version_cache = {}
    for v in ScheduleVersion.objects.filter(year__isnull=False, month__isnull=False, week__isnull=False):
        version_cache[(v.year, v.month, v.week)] = v
    base_qs = Schedule.objects.select_related('teacher', 'classroom', 'classroom__managing_campus')
    base_qs = base_qs.filter(classroom__managing_campus_id__isnull=False)
    if location_id:
        base_qs = base_qs.filter(classroom__managing_campus_id=location_id)
    if grade:
        base_qs = base_qs.filter(classroom__grade=int(grade))
    if size_group == 'gte47':
        base_qs = base_qs.filter(classroom__class_size__gte=47)
    elif size_group == 'lt47':
        base_qs = base_qs.filter(classroom__class_size__lt=47)
    use_single_version = bool(version_id)
    if use_single_version:
        schedules_qs = base_qs.filter(version_id=version_id)
    else:
        version_ids_in_month = set()
        for d in range(1, days_in_month + 1):
            v = _resolve_version_for_day(year, month, d, version_cache)
            if v:
                version_ids_in_month.add(v.id)
        schedules_qs = base_qs.filter(version_id__in=version_ids_in_month) if version_ids_in_month else base_qs.none()
    teacher_ids = schedules_qs.values_list('teacher_id', flat=True).distinct()
    teachers = Teacher.objects.filter(id__in=teacher_ids).order_by('full_name')
    by_teacher_dow_version = defaultdict(list)
    for s in schedules_qs:
        if s.version_id:
            by_teacher_dow_version[(s.teacher_id, s.day_of_week, s.version_id)].append(s)
    pivot_cells = defaultdict(lambda: defaultdict(int))
    sh_extra_by_teacher = defaultdict(int)
    for teacher in teachers:
        for day_num in range(1, days_in_month + 1):
            dt = datetime(year, month, day_num).date()
            dow = _date_to_day_of_week(dt)
            if day_num in excluded_days:
                continue
            vid = int(version_id) if use_single_version and version_id else (_resolve_version_for_day(year, month, day_num, version_cache).id if _resolve_version_for_day(year, month, day_num, version_cache) else None)
            if not vid:
                continue
            for sch in by_teacher_dow_version.get((teacher.id, dow, vid), []):
                pivot_cells[teacher.id][day_num] += get_daily_periods_for_schedule(sch)
                sh_extra_by_teacher[teacher.id] += get_hoso_extra_for_schedule(sch)
    overrides = {(o.teacher_id, o.day): o.value for o in AttendanceOverride.objects.filter(
        teacher_id__in=[t.id for t in teachers], year=year, month=month
    )}
    for teacher in teachers:
        for day_num in range(1, days_in_month + 1):
            key = (teacher.id, day_num)
            if day_num in excluded_days:
                continue
            if key in overrides:
                pivot_cells[teacher.id][day_num] = overrides[key]
    teacher_subject_from_tkb = {}
    for teacher in teachers:
        subs = list(schedules_qs.filter(teacher=teacher).values_list('subject_name', flat=True).distinct())
        teacher_subject_from_tkb[teacher.id] = subs[0] if subs else ''
    pivot_rows = []
    for stt, teacher in enumerate(teachers, 1):
        row_total = 0
        days_list = []
        for day_num in range(1, days_in_month + 1):
            val = pivot_cells[teacher.id].get(day_num, 0)
            days_list.append({'day': day_num, 'val': val, 'excluded': day_num in excluded_days})
            row_total += val
        so_tiet_lam_ho_so = sh_extra_by_teacher.get(teacher.id, 0)
        tong_cong = row_total + so_tiet_lam_ho_so
        subject_display = (teacher.display_subject or '').strip() or teacher_subject_from_tkb.get(teacher.id, '')
        pivot_rows.append({
            'stt': stt, 'teacher': teacher, 'subject_display': subject_display,
            'days_list': days_list, 'so_tiet_lam_ho_so': so_tiet_lam_ho_so, 'tong_cong': tong_cong,
        })
    return pivot_rows, calendar_days, days_in_month, excluded_days



def _copy_cell_style(source_cell, target_cell):
    """Sao chép định dạng (font, border, fill, alignment...) từ ô nguồn sang ô đích."""
    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = shallow_copy(source_cell.font)
        if source_cell.border:
            target_cell.border = shallow_copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = shallow_copy(source_cell.fill)
        if source_cell.number_format:
            target_cell.number_format = shallow_copy(source_cell.number_format)
        if source_cell.alignment:
            target_cell.alignment = shallow_copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = shallow_copy(source_cell.protection)


def _split_vn_name(full_name):
    """Tách họ đệm và tên từ họ tên đầy đủ (tách tại khoảng trắng cuối)."""
    if not full_name or not full_name.strip():
        return '', ''
    s = full_name.strip()
    i = s.rfind(' ')
    if i <= 0:
        return '', s
    return s[:i + 1].rstrip() or '', s[i + 1:].lstrip() or s


def export_attendance_excel(request):
    """Xuất bảng chấm công ra Excel theo template CDCN.xlsx và bộ lọc đang áp dụng."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')
    now = timezone.now().date()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))
    location_id = request.GET.get('location_id', '')
    grade = request.GET.get('grade', '')
    size_group = request.GET.get('size_group', '')
    version_id = request.GET.get('version_id', '')
    pivot_rows, calendar_days, days_in_month, excluded_days = _build_attendance_pivot_data(
        year, month, location_id, grade, size_group, version_id
    )
    template_path = os.path.join(settings.BASE_DIR, 'chấm công tự động', 'CDCN.xlsx')
    if os.path.isfile(template_path):
        try:
            wb = load_workbook(template_path)
            if 'T11.25' in wb.sheetnames:
                ws = wb['T11.25']
            else:
                ws = wb.active
            ws.title = f"T{month}.{year}"
            loc_label = ''
            if location_id:
                try:
                    c = Campus.objects.get(id=location_id)
                    loc_label = f' {c.code} - {c.name}'
                except Campus.DoesNotExist:
                    pass
            grade_label = f' KHỐI {grade}' if grade else ''
            size_label = ''
            if size_group == 'gte47':
                size_label = ' SĨ SỐ ≥ 47'
            elif size_group == 'lt47':
                size_label = ' SĨ SỐ < 47'
            title = f'BẢNG TỔNG HỢP GIỜ DẠY{loc_label}{grade_label}{size_label} THÁNG {month:02d}/{year}'
            ws.cell(3, 1, title)
            day_start_col = 5
            day_end_col = day_start_col + days_in_month - 1
            to_unmerge = [m for m in list(ws.merged_cells.ranges) if m.min_row >= 4 and m.max_row <= 6]
            for m in to_unmerge:
                ws.unmerge_cells(str(m))
            ws.merge_cells(start_row=4, start_column=1, end_row=6, end_column=1)
            ws.merge_cells(start_row=4, start_column=2, end_row=6, end_column=3)
            ws.merge_cells(start_row=4, start_column=4, end_row=6, end_column=4)
            ws.merge_cells(start_row=4, start_column=day_start_col, end_row=4, end_column=day_end_col)
            ws.cell(4, 1, 'STT')
            ws.cell(4, 2, 'HỌ VÀ TÊN')
            ws.cell(4, 4, 'Môn')
            # Dòng tiêu đề phụ (vùng cột ngày) cập nhật theo đúng bộ lọc đang áp dụng.
            subtitle_grade = f'KHỐI {grade}' if grade else 'CÁC KHỐI'
            if size_group == 'gte47':
                subtitle_size = 'SĨ SỐ HV TRÊN 47'
            elif size_group == 'lt47':
                subtitle_size = 'SĨ SỐ HV DƯỚI 47'
            else:
                subtitle_size = 'TẤT CẢ SĨ SỐ'

            base_qs = Schedule.objects.select_related('classroom').filter(classroom__managing_campus_id__isnull=False)
            if location_id:
                base_qs = base_qs.filter(classroom__managing_campus_id=location_id)
            if grade:
                try:
                    base_qs = base_qs.filter(classroom__grade=int(grade))
                except (TypeError, ValueError):
                    pass
            if size_group == 'gte47':
                base_qs = base_qs.filter(classroom__class_size__gte=47)
            elif size_group == 'lt47':
                base_qs = base_qs.filter(classroom__class_size__lt=47)

            if version_id:
                filtered_qs = base_qs.filter(version_id=version_id)
            else:
                version_cache = {}
                for v in ScheduleVersion.objects.filter(year__isnull=False, month__isnull=False, week__isnull=False):
                    version_cache[(v.year, v.month, v.week)] = v
                version_ids_in_month = set()
                for d in range(1, days_in_month + 1):
                    resolved_version = _resolve_version_for_day(year, month, d, version_cache)
                    if resolved_version:
                        version_ids_in_month.add(resolved_version.id)
                filtered_qs = (
                    base_qs.filter(version_id__in=version_ids_in_month)
                    if version_ids_in_month
                    else base_qs.none()
                )

            raw_class_names = list(filtered_qs.values_list('classroom__name', flat=True))
            # Khử trùng lặp ở tầng Python (phòng trường hợp DISTINCT SQL vẫn trả về trùng do dữ liệu/order_by).
            unique_class_map = {}
            for class_name in raw_class_names:
                normalized = (class_name or '').strip().upper()
                if not normalized:
                    continue
                if normalized not in unique_class_map:
                    unique_class_map[normalized] = (class_name or '').strip()
            class_names = sorted(list(unique_class_map.values()), key=natural_sort_key)
            if class_names:
                if len(class_names) > 8:
                    class_text = ', '.join(class_names[:8]) + f', ... ({len(class_names)} lớp)'
                else:
                    class_text = ', '.join(class_names)
                subtitle = f'SỐ TIẾT DẠY {subtitle_grade} {subtitle_size} (LỚP {class_text})'
            else:
                subtitle = f'SỐ TIẾT DẠY {subtitle_grade} {subtitle_size}'
            ws.cell(4, day_start_col, subtitle)
            ws.cell(4, day_end_col + 1, 'Số tiết\nlàm hồ sơ CN')
            ws.cell(4, day_end_col + 2, 'Tổng\ncộng')
            ws.merge_cells(start_row=4, start_column=day_end_col + 1, end_row=6, end_column=day_end_col + 1)
            ws.merge_cells(start_row=4, start_column=day_end_col + 2, end_row=6, end_column=day_end_col + 2)
            for c, cd in enumerate(calendar_days, start=day_start_col):
                ws.cell(5, c, cd['day'])
                ws.cell(6, c, cd['label'])
            # Sao chép định dạng cho phần tiêu đề ngày/thứ nếu vượt cột mẫu
            header_base = ws.cell(5, day_start_col)
            weekday_base = ws.cell(6, day_start_col)
            for c in range(day_start_col + 1, day_end_col + 1):
                _copy_cell_style(header_base, ws.cell(5, c))
                _copy_cell_style(weekday_base, ws.cell(6, c))
            for row_idx in range(7, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None:
                        cell.value = None
            for idx, row in enumerate(pivot_rows, start=7):
                ho_dem, ten = _split_vn_name(row['teacher'].full_name)
                ws.cell(idx, 1, row['stt'])
                ws.cell(idx, 2, ho_dem)
                ws.cell(idx, 3, ten)
                ws.cell(idx, 4, row['subject_display'])
                for c, item in enumerate(row['days_list'], start=day_start_col):
                    ws.cell(idx, c, item['val'] if item['val'] else None)
                ws.cell(idx, day_end_col + 1, row['so_tiet_lam_ho_so'])
                ws.cell(idx, day_end_col + 2, row['tong_cong'])
            # Sao chép định dạng từ dòng mẫu (7) sang tất cả dòng dữ liệu, kể cả dòng vượt quá mẫu
            num_cols = day_end_col + 2
            for idx in range(7, 7 + len(pivot_rows)):
                for col in range(1, num_cols + 1):
                    src = ws.cell(7, col)
                    if not src.has_style and col >= day_start_col:
                        src = ws.cell(7, day_start_col)
                    tgt = ws.cell(idx, col)
                    _copy_cell_style(src, tgt)
        except Exception:
            wb = None
    else:
        wb = None
    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Cham_cong_T{month}_{year}"
        header = ['STT', 'Họ và Tên', 'Môn'] + [f"{cd['day']}({cd['label']})" for cd in calendar_days] + ['Số tiết làm hồ sơ', 'Tổng cộng']
        ws.append(header)
        for row in pivot_rows:
            r = [row['stt'], row['teacher'].full_name, row['subject_display']]
            r += [item['val'] for item in row['days_list']]
            r += [row['so_tiet_lam_ho_so'], row['tong_cong']]
            ws.append(r)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cham_cong_T{month}_{year}.xlsx"'
    wb.save(response)
    return response


def import_tkb_excel(request):
    """Import TKB từ Excel theo phiên bản thời gian (tuần/tháng)."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        year_import = request.POST.get('year')
        month_import = request.POST.get('month')
        week_import = request.POST.get('week')
        session = request.POST.get('session', 'sang')
        if session not in ('sang', 'chieu', 'toi'):
            session = 'sang'
        if not year_import or not month_import or not week_import:
            messages.warning(request, 'Vui lòng chọn Năm, Tháng và Tuần.')
            return redirect('adminpage:import_tkb_excel')
        try:
            year_import = int(year_import)
            month_import = int(month_import)
            week_import = int(week_import)
            week_ranges = _week_ranges_in_month(year_import, month_import)
            if week_import not in week_ranges:
                raise ValueError(f'Tuần phải nằm trong khoảng 1-{len(week_ranges)} của tháng đã chọn.')
        except (ValueError, TypeError):
            messages.warning(request, 'Năm/Tháng/Tuần không hợp lệ.')
            return redirect('adminpage:import_tkb_excel')

        day_start, day_end = week_ranges[week_import]
        effective_from = datetime(year_import, month_import, day_start).date()
        effective_to = datetime(year_import, month_import, day_end).date()
        version_name = f"T{month_import}/{year_import} Tuần {week_import}"

        # Cùng tuần (năm, tháng, tuần) chỉ dùng 1 phiên bản — Sáng/Chiều/Tối cộng dồn vào 1 phiên
        version, version_created = ScheduleVersion.objects.get_or_create(
            year=year_import,
            month=month_import,
            week=week_import,
            defaults={
                'name': version_name,
                'effective_from': effective_from,
                'effective_to': effective_to,
            },
        )
        excel_file = request.FILES['excel_file']
        try:
            n = import_tkb_from_excel(excel_file, version, session=session)
            session_label = dict(Schedule.SESSION_CHOICES).get(session, session)
            if version_created:
                messages.success(request, f'Đã tạo phiên bản "{version.name}" và import TKB buổi {session_label}: {n} tiết.')
            else:
                messages.success(request, f'Đã thêm TKB buổi {session_label} vào phiên bản "{version.name}": {n} tiết (cộng dồn với Sáng/Chiều/Tối).')
        except Exception as e:
            if version_created:
                version.delete()
            messages.error(request, f'Lỗi import: {str(e)}')
        return redirect('adminpage:attendance_pivot')
    selected_year = int(request.GET.get('year', timezone.now().year))
    selected_month = int(request.GET.get('month', timezone.now().month))
    week_ranges = _week_ranges_in_month(selected_year, selected_month)
    week_options = [
        {
            'value': week_no,
            'start_day': day_start,
            'end_day': day_end,
        }
        for week_no, (day_start, day_end) in week_ranges.items()
    ]
    return render(
        request,
        'adminpageSIMCODE/import_tkb.html',
        {
            'selected_year': selected_year,
            'selected_month': selected_month,
            'week_options': week_options,
        },
    )


def import_dsgv_excel(request):
    """Import riêng file DSGV (danh sách giáo viên). Cột: STT, Mã GV, Họ tên [, Môn]."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')

    if request.method == 'POST' and request.FILES.get('dsgv_file'):
        f = request.FILES['dsgv_file']
        try:
            n = import_dsgv_from_excel(f)
            messages.success(request, f'Đã import DSGV: {n} giáo viên mới (đã cập nhật họ tên và môn nếu có).')
        except Exception as e:
            messages.error(request, f'Lỗi import DSGV: {str(e)}')
        return redirect('adminpage:attendance_pivot')
    return render(request, 'adminpageSIMCODE/import_dsgv.html', {})


def import_dsl_excel(request):
    """Import danh sách lớp (DSL) cập nhật sĩ số ClassRoom. Cột: Lớp, Số HS."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')

    if request.method == 'POST' and request.FILES.get('dsl_file'):
        f = request.FILES['dsl_file']
        try:
            n = import_dsl_from_excel(f)
            messages.success(request, f'Đã import DSL: cập nhật sĩ số cho {n} lớp.')
        except Exception as e:
            messages.error(request, f'Lỗi import DSL: {str(e)}')
        return redirect('adminpage:attendance_pivot')
    return render(request, 'adminpageSIMCODE/import_dsl.html', {})


def import_dsl_from_excel(excel_file):
    """Đọc file DSL: cột Lớp, Số HS. Cập nhật ClassRoom.class_size theo tên lớp."""
    try:
        import pandas as pd
    except ImportError:
        raise ValueError('Cần cài đặt pandas: pip install pandas xlrd openpyxl')
    df = pd.read_excel(excel_file, header=None)
    rows = df.values.tolist()
    updated = 0
    for i, row in enumerate(rows):
        if i == 0:
            continue
        if len(row) < 2:
            continue
        try:
            class_name = str(row[0]).strip() if pd.notna(row[0]) else None
            so_hs = int(float(row[1])) if pd.notna(row[1]) else None
        except (ValueError, TypeError):
            continue
        if not class_name or so_hs is None or class_name.upper() == 'LỚP':
            continue
        n = ClassRoom.objects.filter(name=class_name).update(class_size=so_hs)
        if n:
            updated += 1
    return updated


def clear_attendance_data(request):
    """Xóa hết dữ liệu DSGV (Teacher) và TKB (Schedule, ScheduleVersion, ClassRoom). Chỉ admin."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')

    if request.method == 'POST':
        # Thứ tự xóa: Schedule (FK) -> ScheduleVersion, ClassRoom -> Teacher
        n_schedule = Schedule.objects.count()
        n_version = ScheduleVersion.objects.count()
        n_classroom = ClassRoom.objects.count()
        n_teacher = Teacher.objects.count()
        Schedule.objects.all().delete()
        ScheduleVersion.objects.all().delete()
        ClassRoom.objects.all().delete()
        Teacher.objects.all().delete()
        messages.success(
            request,
            f'Đã xóa hết: {n_schedule} tiết TKB, {n_version} phiên bản TKB, {n_classroom} lớp, {n_teacher} giáo viên.'
        )
        return redirect('adminpage:attendance_pivot')

    return render(request, 'adminpageSIMCODE/clear_attendance_data.html', {})

def delete_schedule_version(request):
    """Xóa một phiên bản TKB (kèm toàn bộ tiết thuộc phiên đó). Chỉ admin."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')

    if request.method != 'POST':
        return redirect('adminpage:attendance_pivot')

    version_id = request.POST.get('version_id')
    if not version_id:
        messages.warning(request, 'Vui lòng chọn phiên bản TKB cần xóa.')
        return redirect('adminpage:attendance_pivot')

    version = ScheduleVersion.objects.filter(id=version_id).first()
    if not version:
        messages.error(request, 'Không tìm thấy phiên bản TKB cần xóa.')
        return redirect('adminpage:attendance_pivot')

    schedule_count = Schedule.objects.filter(version=version).count()
    version_name = version.name
    version.delete()
    messages.success(
        request,
        f'Đã xóa phiên bản "{version_name}" và {schedule_count} tiết TKB thuộc phiên này.'
    )
    return redirect('adminpage:attendance_pivot')


def _ensure_campuses_and_linked_points():
    """Tất cả (cơ sở + liên kết) lấy từ homepage.Campus. Bỏ AS, chỉ dùng AT."""
    # Cơ sở: AT, BS, CS. Điểm liên kết: CN, ĐS, KT, HT, VH — đều trong Campus
    codes_names = [
        ('AT', 'Trụ sở chính'), ('BS', 'Cơ sở 1'), ('CS', 'Cơ sở 2'),
        ('CN', 'Trường Cao đẳng Công nghệ TPHCM'), ('ĐS', 'Trường Trung cấp Đông Sài Gòn'),
        ('KT', 'Trường Cao đẳng Kinh tế Kỹ thuật Thủ Đức'), ('HT', 'Trung tâm Huấn luyện Thể thao Quốc gia'),
        ('VH', 'Trường Cao đẳng Kỹ Nghệ II'),
    ]
    for code, name in codes_names:
        Campus.objects.get_or_create(code=code, defaults={'name': name, 'address': ''})


def _loc_code_to_campus_code(loc_code):
    """AS và AT đều quy về Campus AT; BS, CS, CN, ĐS, KT, HT, VH giữ nguyên (đều trong Campus)."""
    if loc_code == 'AS' or loc_code == 'AT':
        return 'AT'
    if loc_code in ('BS', 'CS', 'CN', 'ĐS', 'KT', 'HT', 'VH'):
        return loc_code
    return None


def _parse_class_name(class_name):
    """Từ tên lớp như 10AS1, 12CN3 trả về (grade, location_code). Ví dụ: 10AS1 -> (10, 'AS')."""
    s = (class_name or '').strip()
    m = re.match(r'^(\d+)\s*([A-Za-zĐ]+)\s*\d*$', s)
    if m:
        grade = int(m.group(1))
        loc_code = m.group(2).upper()
        if 'Đ' in loc_code or 'D' in loc_code:
            loc_code = loc_code.replace('D', 'Đ')
        return (grade, loc_code)
    return (None, None)

def _is_yellow_excel_cell(cell):
    """True nếu ô Excel có tô nền vàng (các mã vàng thường gặp)."""
    fill = getattr(cell, 'fill', None)
    if not fill or fill.patternType != 'solid':
        return False
    color_candidates = [getattr(fill, 'fgColor', None), getattr(fill, 'start_color', None)]
    yellow_rgbs = {'FFFF00', 'FFFFFF00', 'FFEB9C', 'FFFFEB9C', 'FFE699', 'FFFFE699', 'FFF2CC', 'FFFFF2CC'}
    yellow_indexed = {5, 6, 27, 36}
    for color in color_candidates:
        if not color:
            continue
        color_type = getattr(color, 'type', None)
        if color_type == 'rgb':
            rgb = (getattr(color, 'rgb', '') or '').upper()
            if rgb in yellow_rgbs:
                return True
        elif color_type == 'indexed':
            if getattr(color, 'indexed', None) in yellow_indexed:
                return True
    return False


def import_dsgv_from_excel(excel_file):
    """
    Đọc file DSGV.xls: dòng 6+ có cột STT, Mã GV, Họ tên [, Môn].
    Tạo/cập nhật Teacher (teacher_code, full_name, display_subject từ cột Môn nếu có).
    """
    try:
        import pandas as pd
    except ImportError:
        raise ValueError('Cần cài đặt pandas: pip install pandas xlrd openpyxl')
    df = pd.read_excel(excel_file, header=None)
    rows = df.values.tolist()
    created = 0
    updated = 0
    for i, row in enumerate(rows):
        if i < 5:
            continue
        if len(row) < 3:
            continue
        try:
            code = str(row[1]).strip() if pd.notna(row[1]) else None
            name = str(row[2]).strip() if pd.notna(row[2]) else None
            mon = str(row[3]).strip() if len(row) > 3 and pd.notna(row[3]) else ''
            if mon and (mon.upper() == 'MÔN' or mon == 'Môn'):
                mon = ''
        except (ValueError, TypeError):
            continue
        if not code or not name or code.upper() == 'MÃ GV':
            continue
        teacher, created_this = Teacher.objects.get_or_create(
            teacher_code=code,
            defaults={'full_name': name, 'display_subject': mon or ''},
        )
        if created_this:
            created += 1
        else:
            teacher.full_name = name
            if len(row) > 3:
                teacher.display_subject = mon or ''
            teacher.save(update_fields=['full_name', 'display_subject'])
            updated += 1
    return created


def import_tkb_from_excel(excel_file, version, session='sang'):
    """
    Đọc file TKB.xls (ma trận): dòng header có "Tiết" và tên lớp (10AS1, 10AS2, ...);
    mỗi ô dữ liệu dạng "MÔN - MãGV" (vd: SH - C.Quyên, TOAN - T.T.Bảo).
    Cột 0 = Ngày (Thứ 2, Thứ 3...), cột 1 = Tiết; từ cột 2 = từng lớp.
    """
    try:
        import pandas as pd
    except ImportError:
        raise ValueError('Cần cài đặt pandas: pip install pandas xlrd openpyxl')
    _ensure_campuses_and_linked_points()
    yellow_cells = set()
    try:
        excel_file.seek(0)
        wb_style = load_workbook(excel_file, data_only=True)
        ws_style = wb_style.active
        for excel_row in ws_style.iter_rows(min_row=1, max_row=ws_style.max_row, min_col=1, max_col=ws_style.max_column):
            for excel_cell in excel_row:
                if _is_yellow_excel_cell(excel_cell):
                    yellow_cells.add((excel_cell.row - 1, excel_cell.column - 1))
    except Exception:
        # Fallback: nếu không đọc được style (vd file .xls), vẫn import như cũ.
        yellow_cells = set()

    excel_file.seek(0)
    df = pd.read_excel(excel_file, header=None)
    rows = df.values.tolist()
    # Tìm dòng header: có "Tiết" ở cột 1 và các cột sau là tên lớp (số+chữ+số)
    header_row_idx = None
    for i, row in enumerate(rows):
        if len(row) < 3:
            continue
        c1 = str(row[1]).strip() if pd.notna(row[1]) else ''
        if 'Tiết' in c1 or (c1.isdigit() and i > 5):
            # Cột 2 có thể là tên lớp kiểu 10AS1
            c2 = str(row[2]).strip() if pd.notna(row[2]) else ''
            if c2 and re.match(r'^\d+[A-ZĐ]', c2, re.IGNORECASE):
                header_row_idx = i
                break
    if header_row_idx is None:
        for i in range(min(15, len(rows))):
            row = rows[i]
            if len(row) >= 10:
                c1 = str(row[1]).strip() if pd.notna(row[1]) else ''
                c2 = str(row[2]).strip() if pd.notna(row[2]) else ''
                if 'Tiết' in c1 or (c1 == '1' and c2):
                    header_row_idx = i
                    break
    if header_row_idx is None:
        header_row_idx = 7
    header_row = rows[header_row_idx]
    class_names = []
    for j in range(2, len(header_row)):
        cn = str(header_row[j]).strip() if pd.notna(header_row[j]) else ''
        if cn and (re.match(r'^\d+[A-ZĐ]', cn, re.IGNORECASE) or cn.isdigit() is False):
            class_names.append((j, cn))
    # Map Thứ N -> day_of_week
    def day_label_to_dow(label):
        s = (str(label or '').strip()).lower()
        if 'thứ 2' in s or 'thu 2' in s:
            return 2
        if 'thứ 3' in s or 'thu 3' in s:
            return 3
        if 'thứ 4' in s or 'thu 4' in s:
            return 4
        if 'thứ 5' in s or 'thu 5' in s:
            return 5
        if 'thứ 6' in s or 'thu 6' in s:
            return 6
        if 'thứ 7' in s or 'thu 7' in s:
            return 7
        if 'chủ nhật' in s or 'cn' == s or 'chu nhat' in s:
            return 8
        return None
    current_day = None
    created_schedules = 0
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if len(row) < 3:
            continue
        day_label = row[0]
        period_val = row[1]
        if pd.notna(day_label) and str(day_label).strip():
            current_day = day_label_to_dow(day_label)
        try:
            period = int(float(period_val)) if pd.notna(period_val) else None
        except (ValueError, TypeError):
            period = None
        if current_day is None or period is None:
            continue
        for col_idx, class_name in class_names:
            if col_idx >= len(row):
                continue
            # Ô TKB tô vàng được xem là "không tính tiết" => bỏ qua khi import.
            if (i, col_idx) in yellow_cells:
                continue
            cell = row[col_idx]
            if pd.isna(cell) or not str(cell).strip():
                continue
            cell_str = str(cell).strip()
            if ' - ' in cell_str:
                parts = cell_str.split(' - ', 1)
                subject_name = (parts[0] or '').strip()
                teacher_code = (parts[1] or '').strip()
            else:
                continue
            if not teacher_code or not subject_name:
                continue
            teacher = Teacher.objects.filter(teacher_code=teacher_code).first()
            if not teacher:
                teacher = Teacher.objects.create(teacher_code=teacher_code, full_name=teacher_code)
            grade, loc_code = _parse_class_name(class_name)
            campus_code = _loc_code_to_campus_code(loc_code) if loc_code else None
            managing_campus = None
            if campus_code:
                managing_campus = Campus.objects.filter(code=campus_code).first()
                if not managing_campus:
                    managing_campus = Campus.objects.create(code=campus_code, name=campus_code, address='')
            classroom, _ = ClassRoom.objects.get_or_create(
                name=class_name,
                defaults={
                    'grade': grade,
                    'class_size': 0,
                    'managing_campus': managing_campus,
                },
            )
            if managing_campus and classroom.managing_campus_id != managing_campus.id:
                classroom.managing_campus = managing_campus
                classroom.save(update_fields=['managing_campus'])
            _, created = Schedule.objects.get_or_create(
                teacher=teacher,
                classroom=classroom,
                day_of_week=current_day,
                period=period,
                version=version,
                defaults={
                    'subject_name': subject_name,
                    'session': session,
                    'effective_date': version.effective_from,
                },
            )
            if created:
                created_schedules += 1
    return created_schedules

# ---------- Sổ đầu bài số (theo nhóm bộ môn) ----------

def _require_journal_admin(request):
    """Kiểm tra quyền admin cho sổ đầu bài."""
    if not request.user.is_authenticated:
        return redirect('homepage:Login')
    try:
        account = Account.objects.get(user=request.user)
        accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
        if accounttype.accounttype_role != 'admin':
            return redirect('homepage:Homepage')
    except (Account.DoesNotExist, AccountType.DoesNotExist):
        return redirect('homepage:Login')
    return None


def _create_13_weeks(subject_journal):
    """Tạo 13 tuần liền kề từ week1_start_date."""
    from datetime import timedelta
    start = subject_journal.week1_start_date
    if not start:
        return
    for w in range(1, 14):
        week_start = start + timedelta(days=(w - 1) * 7)
        week_end = week_start + timedelta(days=6)
        JournalWeek.objects.update_or_create(
            subject_journal=subject_journal, week_number=w,
            defaults={'start_date': week_start, 'end_date': week_end, 'is_locked': False}
        )


def _add_next_week_for_journal(subject_journal):
    """
    Thêm 1 tuần mới (tuần kế tiếp) cho một sổ đầu bài.
    Trả về số tuần mới nếu tạo thành công, None nếu không tạo được.
    """
    from datetime import timedelta

    last_week = (
        JournalWeek.objects.filter(subject_journal=subject_journal)
        .order_by('-week_number')
        .first()
    )
    if not last_week:
        if not subject_journal.week1_start_date:
            return None
        week_start = subject_journal.week1_start_date
        week_end = week_start + timedelta(days=6)
        JournalWeek.objects.create(
            subject_journal=subject_journal,
            week_number=1,
            start_date=week_start,
            end_date=week_end,
            is_locked=False,
            allow_late_edit=False,
        )
        return 1

    new_num = last_week.week_number + 1
    if JournalWeek.objects.filter(subject_journal=subject_journal, week_number=new_num).exists():
        return None

    week_start = last_week.end_date + timedelta(days=1)
    week_end = week_start + timedelta(days=6)
    JournalWeek.objects.create(
        subject_journal=subject_journal,
        week_number=new_num,
        start_date=week_start,
        end_date=week_end,
        is_locked=False,
        allow_late_edit=False,
    )
    return new_num


def _add_next_week_for_all_journals(year=None):
    """Thêm 1 tuần cho mọi sổ đầu bài (lọc theo năm nếu có). Trả về (created, skipped_no_weeks)."""
    qs = SubjectJournal.objects.all()
    if year is not None:
        qs = qs.filter(year=year)
    created = 0
    skipped = 0
    for sj in qs:
        if not JournalWeek.objects.filter(subject_journal=sj).exists():
            skipped += 1
            continue
        if _add_next_week_for_journal(sj) is not None:
            created += 1
    return created, skipped



def journal_manager_dashboard(request):
    """Quản lý sổ đầu bài theo nhóm bộ môn: tạo sổ (môn+năm), set tuần 1, import DSGV, import DSL."""
    err = _require_journal_admin(request)
    if err:
        return err
    from datetime import date
    current_year = date.today().year
    journals = SubjectJournal.objects.prefetch_related('weeks', 'rows').order_by('-year', 'subject')
    subject_choices = SUBJECT_CHOICES

    if request.method == 'POST':
        action = request.POST.get('journal_action')
        if action == 'delete_selected':
            selected_ids = request.POST.getlist('selected_journal_ids')
            if not selected_ids:
                messages.error(request, 'Vui lòng chọn ít nhất 1 sổ đầu bài để xoá.')
                return redirect('adminpage:journal_manager_dashboard')
            deleted_count = SubjectJournal.objects.filter(id__in=selected_ids).count()
            SubjectJournal.objects.filter(id__in=selected_ids).delete()
            messages.success(request, f'Đã xoá {deleted_count} sổ đầu bài đã chọn.')
            return redirect('adminpage:journal_manager_dashboard')
        if action == 'create_journal':
            subject = (request.POST.get('subject') or '').strip()
            year = request.POST.get('year', current_year)
            try:
                year = int(year)
            except ValueError:
                year = current_year
            if subject:
                sj, created = SubjectJournal.objects.get_or_create(subject=subject, year=year)
                messages.success(request, f'Đã tạo sổ {subject} năm {year}.')
            else:
                messages.error(request, 'Chọn môn.')
            return redirect('adminpage:journal_manager_dashboard')
        if action == 'set_week1':
            journal_id = request.POST.get('journal_id')
            week1_str = request.POST.get('week1_start_date', '').strip()
            try:
                sj = SubjectJournal.objects.get(id=journal_id)
                from datetime import datetime
                d = datetime.strptime(week1_str, '%Y-%m-%d').date()
                sj.week1_start_date = d
                sj.save()
                _create_13_weeks(sj)
                messages.success(request, f'Đã set tuần 1 và tạo 13 tuần cho {sj.subject} {sj.year}.')
            except (SubjectJournal.DoesNotExist, ValueError) as e:
                messages.error(request, 'Ngày không hợp lệ hoặc sổ không tồn tại.')
            return redirect('adminpage:journal_manager_dashboard')
        if action == 'toggle_week_lock':
            week_id = request.POST.get('week_id')
            journal_id = request.POST.get('journal_id')
            try:
                w = JournalWeek.objects.get(id=week_id)
                is_expired = w.end_date < date.today()
                effective_locked = w.is_locked or (is_expired and not w.allow_late_edit)
                if effective_locked:
                    # Mở lại: nếu quá hạn thì bật cờ cho phép sửa quá hạn
                    w.is_locked = False
                    if is_expired:
                        w.allow_late_edit = True
                else:
                    # Khóa lại: tuần quá hạn sẽ quay về cơ chế tự khóa
                    w.is_locked = True
                    w.allow_late_edit = False
                w.save()
                is_now_locked = w.is_locked or (is_expired and not w.allow_late_edit)
                messages.success(request, f'Tuần {w.week_number} đã {"khóa" if is_now_locked else "mở lại"}.')
            except JournalWeek.DoesNotExist:
                messages.error(request, 'Tuần không tồn tại.')
            if journal_id:
                return redirect('adminpage:journal_subject_detail', journal_id=journal_id)
            return redirect('adminpage:journal_manager_dashboard')
        if action == 'add_week_all':
            try:
                year = int(request.POST.get('year', current_year))
            except (TypeError, ValueError):
                year = current_year
            created, skipped = _add_next_week_for_all_journals(year=year)
            messages.success(
                request,
                f'Đã thêm tuần mới cho {created} sổ đầu bài năm {year}.'
                + (f' ({skipped} sổ chưa có tuần nào — cần set ngày tuần 1 trước.)' if skipped else '')
            )
            return redirect('adminpage:journal_manager_dashboard')
        if action == 'add_week':
            journal_id = request.POST.get('journal_id')
            try:
                sj = SubjectJournal.objects.get(id=journal_id)
                new_num = _add_next_week_for_journal(sj)
                if new_num:
                    messages.success(request, f'Đã thêm tuần {new_num} cho sổ {sj.get_subject_display() if hasattr(sj, "get_subject_display") else sj.subject} {sj.year}.')
                else:
                    messages.error(request, 'Không thêm được tuần (chưa có tuần trước hoặc tuần đã tồn tại).')
            except SubjectJournal.DoesNotExist:
                messages.error(request, 'Sổ đầu bài không tồn tại.')
            return redirect('adminpage:journal_subject_detail', journal_id=journal_id)

    from django.db.models import Max
    max_week_number = JournalWeek.objects.aggregate(m=Max('week_number'))['m'] or 13
    max_week_number = max(max_week_number, 13)

    context = {
        'journals': journals,
        'subject_choices': subject_choices,
        'current_year': current_year,
        'max_week_number': max_week_number,
    }
    return render(request, 'adminpageSIMCODE/journal_manager_dashboard.html', context)


def journal_subject_detail(request, journal_id):
    """Chi tiết sổ đầu bài: danh sách hàng, tuần, khóa/mở tuần."""
    err = _require_journal_admin(request)
    if err:
        return err
    from datetime import date
    journal = get_object_or_404(SubjectJournal, id=journal_id)
    rows = JournalRow.objects.filter(subject_journal=journal).order_by('row_order')
    weeks = JournalWeek.objects.filter(subject_journal=journal).order_by('week_number')

    if request.method == 'POST' and request.POST.get('journal_action') == 'add_week':
        new_num = _add_next_week_for_journal(journal)
        if new_num:
            messages.success(request, f'Đã thêm tuần {new_num} ({journal.get_subject_display() if hasattr(journal, "get_subject_display") else journal.subject} {journal.year}).')
        else:
            messages.error(request, 'Không thêm được tuần. Kiểm tra đã có tuần trước đó chưa.')
        return redirect('adminpage:journal_subject_detail', journal_id=journal_id)

    context = {'journal': journal, 'rows': rows, 'weeks': weeks, 'today': date.today()}
    return render(request, 'adminpageSIMCODE/journal_subject_detail.html', context)

def journal_thong_ke_export(request):
    """Export Excel thống kê GV ghi sổ đầu bài theo mẫu thong-ke.xlsx: Môn | Họ tên GV | Ngày dạy | Các tiết."""
    err = _require_journal_admin(request)
    if err:
        return err
    from io import BytesIO
    from datetime import date

    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        week_number = int(request.GET.get('week', 1))
    except (TypeError, ValueError):
        week_number = 1

    # Tất cả entry trong tuần này (năm đã chọn)
    entries = JournalEntry.objects.filter(
        journal_row__subject_journal__year=year,
        week_number=week_number,
    ).select_related('journal_row__teacher', 'journal_row__subject_journal').order_by(
        'journal_row__subject_journal__subject',
        'journal_row__teacher__full_name',
        'lesson_date',
        'period',
    )

    # Gom theo (môn, giáo viên, ngày) -> danh sách tiết
    group = defaultdict(list)
    for e in entries:
        sj = e.journal_row.subject_journal
        teacher = e.journal_row.teacher
        key = (sj.id, sj.get_subject_display(), teacher.id, teacher.full_name, e.lesson_date)
        group[key].append(e.period)

    # Mỗi dòng: (subject_display, teacher_name, lesson_date, periods_str) — sort tiết
    rows_data = []
    for (sj_id, subject_label, _tid, teacher_name, lesson_date), periods in group.items():
        periods_str = ','.join(str(p) for p in sorted(set(periods)))
        rows_data.append((subject_label, teacher_name, lesson_date, periods_str))
    rows_data.sort(key=lambda r: (r[0], r[1], r[2]))

    wb = Workbook()
    ws = wb.active
    ws.title = f'Tuan {week_number}'
    title = f'THỐNG KÊ GIÁO VIÊN GHI SỔ ĐẦU BÀI TUẦN {week_number} NĂM HỌC {year}'
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=2, column=1, value=None)
    prev_subject = None
    for idx, (subject_label, teacher_name, lesson_date, periods_str) in enumerate(rows_data, start=3):
        cell_subject = subject_label if subject_label != prev_subject else None
        if subject_label is not None:
            prev_subject = subject_label
        ws.cell(row=idx, column=1, value=cell_subject)
        ws.cell(row=idx, column=2, value=teacher_name or '')
        ws.cell(row=idx, column=3, value=lesson_date)
        ws.cell(row=idx, column=4, value=periods_str)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'Thong-ke-so-dau-bai-tuan-{week_number}-{year}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def journal_week_summary(request, journal_id, week_number):
    """Tổng hợp sổ đầu bài theo tuần: tất cả tiết đã nhập, sắp xếp theo ngày dạy + tiết."""
    err = _require_journal_admin(request)
    if err:
        return err
    journal = get_object_or_404(SubjectJournal, id=journal_id)
    week_obj = JournalWeek.objects.filter(
        subject_journal=journal, week_number=week_number
    ).first()
    if not week_obj:
        messages.error(request, f'Không tìm thấy tuần {week_number}.')
        return redirect('adminpage:journal_subject_detail', journal_id=journal_id)

    if request.method == 'POST' and request.POST.get('journal_action') == 'set_teacher_over_limit':
        teacher_id = request.POST.get('teacher_id')
        allow_over_limit = request.POST.get('allow_over_limit') == '1'
        teacher_obj = get_object_or_404(JournalTeacher, id=teacher_id)
        JournalTeacherWeekLimitOverride.objects.update_or_create(
            journal_week=week_obj,
            teacher=teacher_obj,
            defaults={'allow_over_limit': allow_over_limit},
        )
        messages.success(
            request,
            f'Đã cập nhật quyền nhập vượt giới hạn cho GV {teacher_obj.full_name} (tuần {week_number}).',
        )
        return redirect('adminpage:journal_week_summary', journal_id=journal_id, week_number=week_number)

    rows = JournalRow.objects.filter(subject_journal=journal).select_related('teacher').order_by('row_order')
    entries = []
    for row in rows:
        ents = JournalEntry.objects.filter(
            journal_row=row, week_number=week_number
        ).order_by('lesson_date', 'period')
        for e in ents:
            entries.append({'entry': e, 'teacher': row.teacher})
    entries.sort(key=lambda x: (x['teacher'].access_code, x['entry'].lesson_date, x['entry'].period))

    # Dữ liệu giới hạn theo tuần cho từng giáo viên (để admin bật/tắt "mở giới hạn tiết")
    required_qs = JournalRow.objects.filter(subject_journal=journal).values(
        'teacher_id',
        'teacher__access_code',
        'teacher__full_name',
    ).annotate(required=django_models.Count('id')).order_by('teacher__access_code')

    actual_qs = JournalEntry.objects.filter(
        journal_row__subject_journal=journal,
        week_number=week_number,
    ).values('journal_row__teacher_id').annotate(actual=django_models.Count('id'))

    actual_map = {row['journal_row__teacher_id']: row['actual'] for row in actual_qs}

    overrides = JournalTeacherWeekLimitOverride.objects.filter(journal_week=week_obj).values('teacher_id', 'allow_over_limit')
    override_map = {o['teacher_id']: o['allow_over_limit'] for o in overrides}

    teacher_limits = []
    for rec in required_qs:
        tid = rec['teacher_id']
        teacher_limits.append({
            'teacher_id': tid,
            'teacher_access_code': rec.get('teacher__access_code', ''),
            'teacher_full_name': rec.get('teacher__full_name', ''),
            'required': rec.get('required', 0),
            'actual': actual_map.get(tid, 0),
            'allow_over_limit': bool(override_map.get(tid, False)),
        })

    context = {
        'journal': journal,
        'week_obj': week_obj,
        'entries': entries,
        'teacher_limits': teacher_limits,
    }
    return render(request, 'adminpageSIMCODE/journal_week_summary.html', context)

def journal_week_missing_export(request, journal_id, week_number):
    """Export Excel: thống kê giáo viên chưa ghi đủ sổ đầu bài trong tuần (còn bao nhiêu tiết)."""
    err = _require_journal_admin(request)
    if err:
        return err
    from io import BytesIO

    journal = get_object_or_404(SubjectJournal, id=journal_id)
    week_obj = JournalWeek.objects.filter(
        subject_journal=journal, week_number=week_number
    ).first()
    if not week_obj:
        messages.error(request, f'Không tìm thấy tuần {week_number}.')
        return redirect('adminpage:journal_subject_detail', journal_id=journal_id)

    # Số hàng (dòng) mỗi giáo viên trong sổ này
    rows_qs = JournalRow.objects.filter(subject_journal=journal).select_related('teacher')
    required = rows_qs.values(
        'teacher_id',
        'teacher__full_name',
        'teacher__access_code',
        'teacher__subject',
    ).annotate(total_rows=django_models.Count('id')).order_by('teacher__access_code')

    # Số tiết đã ghi trong tuần này của mỗi giáo viên
    actual_qs = JournalEntry.objects.filter(
        journal_row__subject_journal=journal,
        week_number=week_number,
    ).values('journal_row__teacher_id').annotate(total_entries=django_models.Count('id'))
    actual_map = {row['journal_row__teacher_id']: row['total_entries'] for row in actual_qs}

    missing_list = []
    for rec in required:
        tid = rec['teacher_id']
        total_required = rec['total_rows']
        total_actual = actual_map.get(tid, 0)
        missing = max(total_required - total_actual, 0)
        if missing > 0:
            missing_list.append({
                'teacher_code': rec['teacher__access_code'],
                'teacher_name': rec['teacher__full_name'],
                'teacher_subject': rec['teacher__subject'],
                'required': total_required,
                'actual': total_actual,
                'missing': missing,
            })

    wb = Workbook()
    ws = wb.active
    ws.title = f'Tuan {week_number}'
    header_font = Font(bold=True)
    headers = [
        'STT', 'Mã GV', 'Họ và tên giáo viên', 'Môn',
        'Số hàng (cần)', 'Số tiết đã ghi', 'Số tiết còn thiếu',
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    for idx, row in enumerate(missing_list, start=1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=row['teacher_code'])
        ws.cell(row=idx + 1, column=3, value=row['teacher_name'])
        ws.cell(row=idx + 1, column=4, value=row['teacher_subject'])
        ws.cell(row=idx + 1, column=5, value=row['required'])
        ws.cell(row=idx + 1, column=6, value=row['actual'])
        ws.cell(row=idx + 1, column=7, value=row['missing'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    subject_label = journal.get_subject_display() if hasattr(journal, 'get_subject_display') and callable(getattr(journal, 'get_subject_display')) else journal.subject
    safe_name = re.sub(r'[^\w\s-]', '', str(subject_label)).strip() or journal.subject
    filename = f'Thong-ke-chua-ghi-{safe_name}-tuan-{week_number}-{journal.year}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


def journal_missing_all_subjects_week_export(request):
    """Export Excel: thống kê GV còn thiếu tiết ở TẤT CẢ môn trong một tuần (theo năm + tuần)."""
    err = _require_journal_admin(request)
    if err:
        return err
    from datetime import date
    from io import BytesIO

    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        week_number = int(request.GET.get('week', 1))
    except (TypeError, ValueError):
        week_number = 1

    journals = SubjectJournal.objects.filter(year=year).prefetch_related('weeks')
    rows_global = []

    for journal in journals:
        week_obj = next((w for w in journal.weeks.all() if w.week_number == week_number), None)
        if not week_obj:
            continue

        rows_qs = JournalRow.objects.filter(subject_journal=journal).select_related('teacher')
        if not rows_qs.exists():
            continue

        required = rows_qs.values(
            'teacher_id',
            'teacher__full_name',
            'teacher__access_code',
            'teacher__subject',
        ).annotate(total_rows=django_models.Count('id')).order_by('teacher__access_code')

        actual_qs = JournalEntry.objects.filter(
            journal_row__subject_journal=journal,
            week_number=week_number,
        ).values('journal_row__teacher_id').annotate(total_entries=django_models.Count('id'))
        actual_map = {row['journal_row__teacher_id']: row['total_entries'] for row in actual_qs}

        subject_label = journal.get_subject_display() if hasattr(journal, 'get_subject_display') and callable(getattr(journal, 'get_subject_display')) else journal.subject

        for rec in required:
            tid = rec['teacher_id']
            total_required = rec['total_rows']
            total_actual = actual_map.get(tid, 0)
            missing = max(total_required - total_actual, 0)
            if missing > 0:
                rows_global.append({
                    'subject': subject_label,
                    'teacher_code': rec['teacher__access_code'],
                    'teacher_name': rec['teacher__full_name'],
                    'teacher_subject': rec['teacher__subject'],
                    'required': total_required,
                    'actual': total_actual,
                    'missing': missing,
                })

    rows_global.sort(key=lambda r: (str(r['subject']), str(r['teacher_code'])))

    wb = Workbook()
    ws = wb.active
    ws.title = f'Tuan {week_number}'
    header_font = Font(bold=True)
    headers = [
        'STT', 'Môn', 'Mã GV', 'Họ và tên giáo viên',
        'Môn (từ DSGV)', 'Số hàng (cần)', 'Số tiết đã ghi', 'Số tiết còn thiếu',
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    for idx, row in enumerate(rows_global, start=1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=row['subject'])
        ws.cell(row=idx + 1, column=3, value=row['teacher_code'])
        ws.cell(row=idx + 1, column=4, value=row['teacher_name'])
        ws.cell(row=idx + 1, column=5, value=row['teacher_subject'])
        ws.cell(row=idx + 1, column=6, value=row['required'])
        ws.cell(row=idx + 1, column=7, value=row['actual'])
        ws.cell(row=idx + 1, column=8, value=row['missing'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'Thong-ke-chua-ghi-tuan-{week_number}-{year}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


def journal_export_excel(request, journal_id):
    """Export sổ đầu bài 1 môn ra Excel: mỗi tuần một sheet. Không STT, tối ưu độ rộng cột để in 2 trang ngang."""
    err = _require_journal_admin(request)
    if err:
        return err
    from io import BytesIO

    journal = get_object_or_404(SubjectJournal, id=journal_id)
    weeks = JournalWeek.objects.filter(subject_journal=journal).order_by('week_number')
    if not weeks.exists():
        messages.error(request, 'Chưa có tuần nào. Set ngày bắt đầu tuần 1 trước.')
        return redirect('adminpage:journal_subject_detail', journal_id=journal_id)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    # Theo bố cục in thực tế (như ảnh): không cần STT
    header_row4 = [
        'Họ và tên giáo viên', 'Ngày dạy', 'Lớp dạy', 'Tiết', 'Sĩ số',
        'Học viên vắng', 'Tên bài giảng', 'NHẬN XÉT CỦA GIÁO VIÊN SAU TIẾT DẠY'
    ]
    row6_vals = [1, 2, 3, 4, 5, 6, 7, 8]
    rows_by_journal = JournalRow.objects.filter(subject_journal=journal).select_related('teacher').order_by('row_order')

    for week_obj in weeks:
        sheet_name = f"Tuần {week_obj.week_number}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_journal_sheet_one_week(
            ws, week_obj, rows_by_journal,
            thin_border, header_font, header_fill, header_row4, row6_vals,
            fit_two_pages=True
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    subject_label = journal.get_subject_display() if hasattr(journal, 'get_subject_display') and callable(getattr(journal, 'get_subject_display')) else journal.subject
    safe_name = re.sub(r'[^\w\s-]', '', str(subject_label)).strip() or journal.subject
    filename = f"So-dau-bai-{safe_name}-{journal.year}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _write_journal_sheet_one_week(ws, week_obj, rows_by_journal, thin_border, header_font, header_fill, header_row4, row6_vals, wrap_lop_vang=False, fit_two_pages=True):
    """Ghi một sheet sổ đầu bài cho một tuần: header + dữ liệu. Không STT, tối ưu độ rộng cột để phủ hết 2 trang ngang."""
    # Cột A-H: A=Họ tên, B=Ngày, C=Lớp, D=Tiết, E=Sĩ số, F=Học viên vắng, G=Tên bài, H=Nhận xét
    # Style giống dashboard (journal_week_summary.html): font 13, border 1px, header nền xanh nhạt.
    # Chỉ cần đảm bảo cột gọn 1 trang ngang; 2 cột cuối rộng nhất.
    col_widths = [22, 12, 14, 6, 6, 18, 28, 36]
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Lấy tên môn hiển thị để ghi tiêu đề lớn
    subject_label = ''
    if rows_by_journal:
        sj = rows_by_journal[0].subject_journal
        if hasattr(sj, 'get_subject_display') and callable(getattr(sj, 'get_subject_display')):
            subject_label = sj.get_subject_display()
        else:
            subject_label = str(sj.subject)

    # Tiêu đề tuần (hàng 1) trải ngang toàn bảng để không bị cụt
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(
        row=1,
        column=1,
        value=f"Tuần: {week_obj.week_number}   Từ ngày: {week_obj.start_date.strftime('%d/%m/%Y')}   Đến ngày: {week_obj.end_date.strftime('%d/%m/%Y')}"
    )
    title_cell.border = thin_border
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Hàng 2-3: tiêu đề lớn ÔN THI TỐT NGHIỆP...
    ws.merge_cells('A2:H3')
    big_title = f"ÔN THI TỐT NGHIỆP THPT NĂM 2026 - MÔN {subject_label.upper() if subject_label else ''}"
    big_title_cell = ws.cell(row=2, column=1, value=big_title)
    big_title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    big_title_cell.font = Font(name='Times New Roman', bold=True, size=14)

    for col, h in enumerate(header_row4, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Header căn giữa
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 18
    for col_idx, val in enumerate(row6_vals, start=1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Hàng 6 (số thứ tự cột) căn giữa
        cell.alignment = Alignment(horizontal='center', vertical='center')

    entries = []
    for row in rows_by_journal:
        for e in JournalEntry.objects.filter(journal_row=row, week_number=week_obj.week_number).order_by('lesson_date', 'period'):
            entries.append({'entry': e, 'teacher': row.teacher})
    entries.sort(key=lambda x: (x['teacher'].access_code, x['entry'].lesson_date, x['entry'].period))

    for offset, item in enumerate(entries):
        e, t = item['entry'], item['teacher']
        row_num = 7 + offset
        classes_val = (e.classes_taught or '').strip()
        absent_val = (e.absent_students or '').strip()
        if absent_val in ('0', '00'):
            absent_val = ''

        data_cols = [
            t.full_name or '',
            e.lesson_date.strftime('%d/%m/%Y') if e.lesson_date else '',
            classes_val,
            e.period or '',
            e.student_count if e.student_count is not None else '',
            absent_val,
            e.lesson_title or '',
            e.comment or ''
        ]
        for col, val in enumerate(data_cols, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            # Canh giống dashboard: Ngày/Tiết/Sĩ số center; còn lại left.
            if col in (2, 4, 5):  # Ngày, Tiết, Sĩ số
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            else:
                # Giống dashboard: nội dung dài tự xuống dòng trong ô
                wrap = col in (3, 6, 7, 8)  # Lớp dạy, HV vắng, Tên bài, Nhận xét
                # Riêng cột Lớp (col=3) căn giữa theo chiều dọc
                v_align = 'center' if col == 3 else ('top' if wrap else 'center')
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical=v_align,
                    wrap_text=wrap
                )

    if fit_two_pages:
        # Vùng in + lặp lại header cho trang 2
        last_row = max(ws.max_row, 7)
        ws.print_area = f"A1:H{last_row}"
        ws.print_title_rows = "1:6"

        # In: cột gọn trong 1 trang ngang, hàng tối đa 2 trang dọc
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 2

        # Lề
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.print_options.horizontalCentered = False


def journal_export_excel_all_week(request):
    """Export sổ đầu bài: 1 file Excel, 1 sheet = 1 môn, theo tuần đã chọn. Autofit, in vừa 2 trang, Lớp/Học viên vắng xuống dòng."""
    err = _require_journal_admin(request)
    if err:
        return err
    from io import BytesIO
    from datetime import date

    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        week_number = int(request.GET.get('week', 1))
    except (TypeError, ValueError):
        week_number = 1

    # Tất cả sổ đầu bài (môn) trong năm có tuần này
    journals = SubjectJournal.objects.filter(year=year).prefetch_related('weeks')
    journals_with_week = [j for j in journals if any(w.week_number == week_number for w in j.weeks.all())]
    if not journals_with_week:
        messages.error(request, f'Năm {year} không có sổ đầu bài nào có tuần {week_number}.')
        return redirect('adminpage:journal_manager_dashboard')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    header_row4 = [
        'Họ và tên giáo viên', 'Ngày dạy', 'Lớp dạy', 'Tiết', 'Sĩ số',
        'Học viên vắng', 'Tên bài giảng', 'NHẬN XÉT CỦA GIÁO VIÊN SAU TIẾT DẠY'
    ]
    row6_vals = [1, 2, 3, 4, 5, 6, 7, 8]

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    for journal in journals_with_week:
        week_obj = next((w for w in journal.weeks.all() if w.week_number == week_number), None)
        if not week_obj:
            continue
        subject_label = journal.get_subject_display() if hasattr(journal, 'get_subject_display') and callable(getattr(journal, 'get_subject_display')) else journal.subject
        sheet_name = re.sub(r'[^\w\s-]', '', str(subject_label)).strip() or journal.subject
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        rows_by_journal = JournalRow.objects.filter(subject_journal=journal).select_related('teacher').order_by('row_order')
        _write_journal_sheet_one_week(
            ws, week_obj, rows_by_journal,
            thin_border, header_font, header_fill, header_row4, row6_vals,
            fit_two_pages=True
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"So-dau-bai-tat-ca-mon-tuan-{week_number}-{year}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def import_journal_dsgv(request):
    """Import DSGV: Mã GV, Họ tên, Môn, Số lớp. Tạo sổ môn nếu chưa có, tạo GV + hàng (số lớp × 2)."""
    err = _require_journal_admin(request)
    if err:
        return err
    from datetime import date
    current_year = date.today().year
    if request.method == 'POST' and request.FILES.get('dsgv_file'):
        try:
            import pandas as pd
            df = pd.read_excel(request.FILES['dsgv_file'], header=0)
            if df.shape[1] < 4:
                messages.error(request, 'File cần ít nhất 4 cột: Mã GV, Họ và tên, Môn, Số lớp.')
                return redirect('adminpage:import_journal_dsgv')
            df.columns = [str(c).strip() for c in df.columns]
            col_map = {}
            for i, c in enumerate(df.columns):
                c_lower = str(c).lower()
                if 'mã' in c_lower or 'ma' in c_lower:
                    col_map['code'] = i
                elif 'họ' in c_lower or 'tên' in c_lower or 'tên' in c_lower:
                    col_map['name'] = i
                elif 'môn' in c_lower or 'mon' in c_lower:
                    col_map['subject'] = i
                elif 'số lớp' in c_lower or 'so lop' in c_lower or 'lớp' in c_lower:
                    col_map['num_classes'] = i
            if 'code' not in col_map or 'name' not in col_map or 'subject' not in col_map:
                messages.error(request, 'File cần cột Mã GV, Họ và tên, Môn.')
                return redirect('adminpage:import_journal_dsgv')
            num_created = 0
            num_rows_created = 0
            for _, row in df.iterrows():
                code = str(row.iloc[col_map['code']]).strip()
                name = str(row.iloc[col_map['name']]).strip()
                subject_raw = str(row.iloc[col_map['subject']]).strip()
                subject_code = normalize_subject_code(subject_raw)
                num_classes = 1
                if 'num_classes' in col_map:
                    try:
                        num_classes = max(1, int(float(row.iloc[col_map['num_classes']])))
                    except (ValueError, TypeError):
                        pass
                if not code or not name or not subject_raw:
                    continue
                teacher, created = JournalTeacher.objects.update_or_create(
                    access_code=code, defaults={'full_name': name, 'subject': subject_raw, 'num_classes': num_classes}
                )
                if created:
                    num_created += 1
                sj, _ = SubjectJournal.objects.get_or_create(subject=subject_code, year=current_year)
                existing_rows = JournalRow.objects.filter(subject_journal=sj, teacher=teacher).count()
                if existing_rows == 0:
                    max_order = JournalRow.objects.filter(subject_journal=sj).aggregate(
                        m=django_models.Max('row_order')
                    )['m'] or 0
                    for r in range(teacher.num_rows()):
                        JournalRow.objects.create(subject_journal=sj, teacher=teacher, row_order=max_order + r + 1)
                        num_rows_created += 1
            messages.success(request, f'Đã import: {num_created} GV mới, {num_rows_created} hàng.')
        except Exception as e:
            messages.error(request, f'Lỗi: {str(e)}')
        return redirect('adminpage:journal_manager_dashboard')
    return render(request, 'adminpageSIMCODE/import_journal_dsgv.html', {})


def import_journal_dsl(request):
    """Import DSL: cột Lớp -> JournalClass."""
    err = _require_journal_admin(request)
    if err:
        return err
    if request.method == 'POST' and request.FILES.get('dsl_file'):
        try:
            import pandas as pd
            df = pd.read_excel(request.FILES['dsl_file'], header=0)
            col_class = None
            for i, c in enumerate(df.columns):
                if 'lớp' in str(c).lower() and 'chủ nhiệm' not in str(c).lower():
                    col_class = i
                    break
            if col_class is None:
                messages.error(request, 'Không tìm thấy cột Lớp.')
                return redirect('adminpage:import_journal_dsl')
            num = 0
            for _, row in df.iterrows():
                val = str(row.iloc[col_class]).strip()
                if val and len(val) <= 50:
                    _, c = JournalClass.objects.get_or_create(name=val)
                    if c:
                        num += 1
            messages.success(request, f'Đã thêm {num} lớp mới.')
        except Exception as e:
            messages.error(request, f'Lỗi: {str(e)}')
        return redirect('adminpage:journal_manager_dashboard')
    return render(request, 'adminpageSIMCODE/import_journal_dsl.html', {})


