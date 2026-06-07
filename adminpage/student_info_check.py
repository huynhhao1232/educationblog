"""
Quản lý & import Excel — Kiểm tra thông tin học viên (THPT).
"""
from __future__ import annotations

import re
import unicodedata
from datetime import datetime

from django.contrib import messages
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from homepage.models import (
    Account,
    AccountType,
    Campus,
    Student,
    StudentInfoRecord,
    StudentInfoVerification,
)


def _get_filtered_student_info_qs(request):
    """Dùng chung cho bảng quản lý và export Excel (áp dụng cùng bộ lọc)."""
    answered_q = _verification_answered_q()

    qs = StudentInfoRecord.objects.all().order_by("class_name", "full_name")
    filter_campus = request.GET.get("campus_id", "").strip()
    filter_class = request.GET.get("class_name", "").strip()
    filter_name = request.GET.get("full_name", "").strip()
    filter_status = request.GET.get("status", "").strip()

    if filter_campus.isdigit():
        codes = Student.objects.filter(campus_id=int(filter_campus)).values_list("student_code", flat=True)
        qs = qs.filter(student_code__in=list(codes))
    if filter_class:
        qs = qs.filter(class_name__iexact=filter_class)
    if filter_name:
        qs = qs.filter(full_name__icontains=filter_name)
    if filter_status == "done":
        qs = qs.filter(answered_q)
    elif filter_status == "pending":
        qs = qs.exclude(answered_q)

    return qs


def _vn_normalize_no_diacritics(text):
    if not text:
        return ""
    text = str(text).strip()
    nfkd = unicodedata.normalize("NFKD", text)
    without_diacritics = "".join(c for c in nfkd if not unicodedata.combining(c))
    return without_diacritics.lower().strip()


def _normalize_si_header(value):
    if value is None:
        return ""
    txt = str(value).strip().lower()
    txt = txt.replace("\n", " ").replace("\r", " ")
    txt = re.sub(r"\s+", " ", txt)
    return txt


def _parse_excel_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


STUDENT_INFO_DEFAULT_HIGHSCHOOL_PLACE = "TT GDNN-GDTX TP Thủ Đức"


def _default_highschool_place_if_empty(value):
    if value is None:
        return STUDENT_INFO_DEFAULT_HIGHSCHOOL_PLACE
    t = str(value).strip()
    if not t:
        return STUDENT_INFO_DEFAULT_HIGHSCHOOL_PLACE
    return t


def _resolve_student_code_from_graduation_data(*, id_number, class_name, full_name, birthday):
    """Đối chiếu bảng Student: ưu tiên CCCD, sau đó lớp + họ tên + ngày sinh."""
    id_num = (id_number or "").strip()
    if id_num:
        s = Student.objects.filter(id_number=id_num).first()
        if s:
            return s.student_code
    fn = (full_name or "").strip()
    if not fn:
        return None
    fn_norm = _vn_normalize_no_diacritics(re.sub(r"\s+", " ", fn))
    qs = Student.objects.all()
    c = (class_name or "").strip()
    if c:
        qs = qs.filter(class_name__iexact=c)
    if birthday:
        qs = qs.filter(birthday=birthday)
    for s in qs:
        if _vn_normalize_no_diacritics(re.sub(r"\s+", " ", s.full_name or "")) == fn_norm:
            return s.student_code
    return None


def _verification_instance_has_answer(v: StudentInfoVerification | None) -> bool:
    if not v:
        return False
    for fn in (
        "class_name_status",
        "full_name_status",
        "birthday_status",
        "birth_place_status",
        "gender_status",
        "ethnicity_status",
        "id_number_status",
        "contact_address_status",
        "email_status",
        "phone_status",
        "highschool_10_status",
        "highschool_11_status",
        "exam_subjects_status",
    ):
        if getattr(v, fn, "") in ("D", "S"):
            return True
    return False


def _verification_answered_q():
    from django.db.models import Q

    return (
        Q(verification__class_name_status__in=["D", "S"])
        | Q(verification__full_name_status__in=["D", "S"])
        | Q(verification__birthday_status__in=["D", "S"])
        | Q(verification__birth_place_status__in=["D", "S"])
        | Q(verification__gender_status__in=["D", "S"])
        | Q(verification__ethnicity_status__in=["D", "S"])
        | Q(verification__id_number_status__in=["D", "S"])
        | Q(verification__contact_address_status__in=["D", "S"])
        | Q(verification__email_status__in=["D", "S"])
        | Q(verification__phone_status__in=["D", "S"])
        | Q(verification__highschool_10_status__in=["D", "S"])
        | Q(verification__highschool_11_status__in=["D", "S"])
        | Q(verification__exam_subjects_status__in=["D", "S"])
    )


def student_info_manage(request):
    if not request.user.is_authenticated:
        return redirect("homepage:login")
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return redirect("homepage:Homepage")

    answered_q = _verification_answered_q()
    qs = _get_filtered_student_info_qs(request)
    filter_campus = request.GET.get("campus_id", "").strip()
    filter_class = request.GET.get("class_name", "").strip()
    filter_name = request.GET.get("full_name", "").strip()
    filter_status = request.GET.get("status", "").strip()

    # Thống kê 3 ô = đúng bộ dữ liệu đang lọc (cùng điều kiện với bảng + phân trang)
    total_count = qs.count()
    if filter_status == "done":
        verified_count = total_count
        pending_count = 0
    elif filter_status == "pending":
        verified_count = 0
        pending_count = total_count
    else:
        verified_count = qs.filter(answered_q).count()
        pending_count = max(0, total_count - verified_count)

    paginator = Paginator(qs, 50)
    page = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    codes = [r.student_code for r in page_obj.object_list]
    campus_map = {}
    if codes:
        for s in Student.objects.filter(student_code__in=codes).select_related("campus"):
            campus_map[s.student_code] = s.campus.name if s.campus else ""

    rows = []
    for r in page_obj.object_list:
        try:
            v = r.verification
        except StudentInfoVerification.DoesNotExist:
            v = None
        rows.append(
            {
                "student_code": r.student_code,
                "campus_name": campus_map.get(r.student_code, ""),
                "class_name": r.class_name,
                "full_name": r.full_name,
                "birthday": r.birthday,
                "id_number": r.id_number,
                "phone": r.phone,
                "exam_subjects": r.exam_subjects or [],
                "verification": v,
                "has_answer": _verification_instance_has_answer(v),
            }
        )

    import_preview_count = len(request.session.get("student_info_import_data", []))
    import_errors = request.session.get("student_info_import_errors", [])[:10]

    return render(
        request,
        "adminpageSIMCODE/student_info_manage.html",
        {
            "page_obj": page_obj,
            "rows": rows,
            "total_count": total_count,
            "verified_count": verified_count,
            "pending_count": pending_count,
            "campuses": Campus.objects.order_by("name"),
            "filter_campus": filter_campus,
            "filter_class": filter_class,
            "filter_name": filter_name,
            "filter_status": filter_status,
            "import_preview_count": import_preview_count,
            "import_errors": import_errors,
        },
    )


def student_info_export_excel(request):
    """Admin: export Excel danh sách kiểm tra; các ô học viên tick 'Sai' sẽ được tô đỏ."""
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415

    if not request.user.is_authenticated:
        return redirect("homepage:login")
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return redirect("homepage:Homepage")

    qs = _get_filtered_student_info_qs(request).select_related("verification")
    records = list(qs)

    codes = [r.student_code for r in records if r.student_code]
    campus_map: dict[str, str] = {}
    if codes:
        for s in Student.objects.filter(student_code__in=codes).select_related("campus"):
            campus_map[s.student_code] = s.campus.name if s.campus else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "KIEM_TRA_THONG_TIN"

    header_fill = PatternFill(start_color="023EB6", end_color="023EB6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrong_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    wrong_font = Font(color="B71C1C", bold=True)

    status_fields = (
        "class_name_status",
        "full_name_status",
        "birthday_status",
        "birth_place_status",
        "gender_status",
        "ethnicity_status",
        "id_number_status",
        "contact_address_status",
        "email_status",
        "phone_status",
        "highschool_10_status",
        "highschool_11_status",
        "exam_subjects_status",
    )

    columns = [
        ("Cơ sở", "campus_name", None),
        ("Mã học viên", "student_code", None),
        ("Lớp", "class_name", "class_name_status"),
        ("Họ và tên", "full_name", "full_name_status"),
        ("Ngày sinh", "birthday", "birthday_status"),
        ("Nơi sinh", "birth_place", "birth_place_status"),
        ("Giới tính", "gender", "gender_status"),
        ("Dân tộc", "ethnicity", "ethnicity_status"),
        ("Số CCCD", "id_number", "id_number_status"),
        ("Địa chỉ liên hệ", "contact_address", "contact_address_status"),
        ("Gmail", "email", "email_status"),
        ("Số điện thoại", "phone", "phone_status"),
        ("Nơi học THPT lớp 10", "highschool_10", "highschool_10_status"),
        ("Nơi học THPT lớp 11", "highschool_11", "highschool_11_status"),
        ("Môn thi TN THPT", "exam_subjects", "exam_subjects_status"),
        ("Trạng thái", "overall_status", None),
        ("Ghi chú", "note", None),
        ("Số lần đã gửi", "submit_count", None),
    ]

    ws.append([c[0] for c in columns])
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in records:
        v = getattr(r, "verification", None)
        submit_count = getattr(v, "submit_count", 0) if v else 0
        note = (getattr(v, "note", "") or "").strip() if v else ""

        overall_status = "Chưa gửi"
        has_wrong = False
        if v and submit_count > 0:
            has_wrong = any(getattr(v, fn, "") == "S" for fn in status_fields)
            overall_status = "Sai" if has_wrong else "Đúng"

        row_values = []
        for _, key, _status_attr in columns:
            if key == "campus_name":
                row_values.append(campus_map.get(r.student_code, ""))
            elif key == "exam_subjects":
                row_values.append(", ".join(r.exam_subjects or []))
            elif key == "birthday":
                row_values.append(r.birthday.strftime("%d/%m/%Y") if r.birthday else "")
            elif key == "overall_status":
                row_values.append(overall_status)
            elif key == "note":
                row_values.append(note)
            elif key == "submit_count":
                row_values.append(submit_count)
            else:
                row_values.append(getattr(r, key, "") or "")

        ws.append(row_values)

        # Tô đỏ ô dữ liệu nào học viên xác nhận sai (status = "S")
        row_idx = ws.max_row
        for col_idx, (_label, _key, status_attr) in enumerate(columns, start=1):
            if not status_attr or not v:
                continue
            if getattr(v, status_attr, "") == "S":
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = wrong_fill
                c.font = wrong_font

    # Auto width đơn giản
    for col_idx, (label, *_rest) in enumerate(columns, start=1):
        width = max(12, min(40, len(label) + 2))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
    resp["Content-Disposition"] = f'attachment; filename="kiem_tra_thong_tin_hoc_vien_{ts}.xlsx"'
    wb.save(resp)
    return resp


STUDENT_INFO_HISTORY_PAYLOAD_LABELS = (
    ("class_name_status", "Lớp"),
    ("full_name_status", "Họ và tên"),
    ("birthday_status", "Ngày sinh"),
    ("birth_place_status", "Nơi sinh"),
    ("gender_status", "Giới tính"),
    ("ethnicity_status", "Dân tộc"),
    ("id_number_status", "Số CCCD"),
    ("contact_address_status", "Địa chỉ liên hệ"),
    ("email_status", "Gmail"),
    ("phone_status", "Số điện thoại"),
    ("highschool_10_status", "Nơi học THPT lớp 10"),
    ("highschool_11_status", "Nơi học THPT lớp 11"),
    ("exam_subjects_status", "Môn thi TN THPT"),
)


def _history_payload_display_rows(payload: dict):
    rows = []
    for key, label in STUDENT_INFO_HISTORY_PAYLOAD_LABELS:
        val = (payload or {}).get(key, "") or ""
        disp = {"D": "Đúng", "S": "Sai"}.get(val, "—")
        rows.append({"label": label, "display": disp})
    return rows


def student_info_verification_history(request, student_code):
    """Admin: xem lịch sử các lần học viên gửi xác nhận (tối đa 2)."""
    if not request.user.is_authenticated:
        return redirect("homepage:login")
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return redirect("homepage:Homepage")

    record = StudentInfoRecord.objects.filter(student_code=student_code).first()
    if not record:
        messages.error(request, "Không tìm thấy học viên trong danh sách kiểm tra.")
        return redirect("adminpage:student_info_manage")

    verification = StudentInfoVerification.objects.filter(student=record).first()
    history_blocks = []
    if verification:
        for h in verification.history_entries.order_by("-created_at"):
            pl = h.payload or {}
            history_blocks.append(
                {
                    "submit_index": h.submit_index,
                    "created_at": timezone.localtime(h.created_at),
                    "rows": _history_payload_display_rows(pl),
                    "note": pl.get("note") or "",
                    "wrong_field_labels": pl.get("wrong_field_labels") or [],
                    "all_correct": pl.get("all_correct"),
                }
            )

    return render(
        request,
        "adminpageSIMCODE/student_info_verification_history.html",
        {
            "record": record,
            "verification": verification,
            "history_blocks": history_blocks,
            "submit_count": verification.submit_count if verification else 0,
        },
    )


def student_info_verification_json(request, student_code):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return JsonResponse({"error": "Forbidden"}, status=403)

    record = StudentInfoRecord.objects.filter(student_code=student_code).first()
    if not record:
        return JsonResponse({"error": "Không tìm thấy học viên."}, status=404)

    v = StudentInfoVerification.objects.filter(student=record).first()

    def _disp(code):
        if code == "D":
            return "Đúng"
        if code == "S":
            return "Sai"
        return "Chưa chọn"

    if not v:
        return JsonResponse(
            {
                "has_verification": False,
                "student_code": record.student_code,
                "full_name": record.full_name,
                "message": "Học viên chưa gửi xác nhận.",
            }
        )

    submitted = timezone.localtime(v.submitted_at).strftime("%d/%m/%Y %H:%M") if v.submitted_at else ""
    fields = [
        {"label": "Lớp", "status": v.class_name_status, "display": _disp(v.class_name_status)},
        {"label": "Họ và tên", "status": v.full_name_status, "display": _disp(v.full_name_status)},
        {"label": "Ngày sinh", "status": v.birthday_status, "display": _disp(v.birthday_status)},
        {"label": "Nơi sinh", "status": v.birth_place_status, "display": _disp(v.birth_place_status)},
        {"label": "Giới tính", "status": v.gender_status, "display": _disp(v.gender_status)},
        {"label": "Dân tộc", "status": v.ethnicity_status, "display": _disp(v.ethnicity_status)},
        {"label": "Số CCCD", "status": v.id_number_status, "display": _disp(v.id_number_status)},
        {"label": "Địa chỉ liên hệ", "status": v.contact_address_status, "display": _disp(v.contact_address_status)},
        {"label": "Gmail", "status": v.email_status, "display": _disp(v.email_status)},
        {"label": "Số điện thoại", "status": v.phone_status, "display": _disp(v.phone_status)},
        {"label": "Nơi học THPT lớp 10", "status": v.highschool_10_status, "display": _disp(v.highschool_10_status)},
        {"label": "Nơi học THPT lớp 11", "status": v.highschool_11_status, "display": _disp(v.highschool_11_status)},
        {"label": "Môn thi TN THPT", "status": v.exam_subjects_status, "display": _disp(v.exam_subjects_status)},
    ]

    return JsonResponse(
        {
            "has_verification": True,
            "student_code": record.student_code,
            "full_name": record.full_name,
            "class_name": record.class_name,
            "fields": fields,
            "note": v.note or "",
            "submitted_at": submitted,
        }
    )


def student_info_import_preview(request):
    if not request.user.is_authenticated:
        return redirect("homepage:login")
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return redirect("homepage:Homepage")

    rows = request.session.get("student_info_import_data", [])
    errors = request.session.get("student_info_import_errors", [])
    if not rows and not errors:
        messages.info(request, "Chưa có dữ liệu import tạm để xem trước.")
        return redirect("adminpage:student_info_manage")

    return render(
        request,
        "adminpageSIMCODE/student_info_import_preview.html",
        {
            "students_data": rows,
            "errors": errors[:30],
            "total_count": len(rows),
            "error_count": len(errors),
        },
    )


def student_info_import_excel(request):
    from openpyxl import load_workbook  # noqa: PLC0415

    if not request.user.is_authenticated:
        return redirect("homepage:login")
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return redirect("homepage:Homepage")

    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "Vui lòng chọn file Excel để import.")
        return redirect("adminpage:student_info_manage")

    excel_file = request.FILES["excel_file"]
    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Lỗi đọc file Excel: {exc}")
        return redirect("adminpage:student_info_manage")

    row6 = [cell for cell in next(ws.iter_rows(min_row=6, max_row=6, values_only=True))]
    row7 = [cell for cell in next(ws.iter_rows(min_row=7, max_row=7, values_only=True))]

    col_mapping: dict[str, int] = {}
    for idx in range(max(len(row6), len(row7))):
        v6 = _normalize_si_header(row6[idx] if idx < len(row6) else "")
        v7 = _normalize_si_header(row7[idx] if idx < len(row7) else "")
        combined = f"{v6} {v7}".strip()
        if "s t t" in combined or combined == "stt":
            col_mapping["stt"] = idx
        elif "lớp 10" in combined:
            col_mapping["highschool_10"] = idx
        elif "lớp 11" in combined:
            col_mapping["highschool_11"] = idx
        elif "lớp" in combined:
            col_mapping["class_name"] = idx
        elif combined == "họ và":
            col_mapping["last_name"] = idx
        elif combined == "tên hv":
            col_mapping["first_name"] = idx
        elif "ngày sinh" in combined:
            col_mapping["birthday"] = idx
        elif "nơi sinh" in combined:
            col_mapping["birth_place"] = idx
        elif "giới tính" in combined:
            col_mapping["gender"] = idx
        elif "dân tộc" in combined:
            col_mapping["ethnicity"] = idx
        elif "số cccd" in combined:
            col_mapping["id_number"] = idx
        elif "địa chỉ liên hệ" in combined:
            col_mapping["contact_address"] = idx
        elif "gmail" in combined:
            col_mapping["email"] = idx
        elif "số điện thoại" in combined:
            col_mapping["phone"] = idx
        elif combined.endswith("lý"):
            col_mapping["ly"] = idx
        elif combined.endswith("hóa") or combined.endswith("hoá"):
            col_mapping["hoa"] = idx
        elif combined.endswith("sinh"):
            col_mapping["sinh"] = idx
        elif combined.endswith("sử"):
            col_mapping["su"] = idx
        elif combined.endswith("địa"):
            col_mapping["dia"] = idx
        elif combined.endswith("ktpl"):
            col_mapping["ktpl"] = idx
        elif combined.endswith("tin") and "thông tin" not in combined:
            col_mapping["tin"] = idx
        elif "tiếng anh" in combined or "tieng anh" in _vn_normalize_no_diacritics(combined):
            col_mapping["tieng_anh"] = idx

    if col_mapping.get("tieng_anh") is None and col_mapping.get("tin") is not None:
        nxt = col_mapping["tin"] + 1
        max_h = max(len(row6), len(row7))
        if nxt < max_h:
            v6n = _normalize_si_header(row6[nxt] if nxt < len(row6) else "")
            v7n = _normalize_si_header(row7[nxt] if nxt < len(row7) else "")
            combined_next = f"{v6n} {v7n}".strip()
            if (
                not combined_next
                or "tiếng anh" in combined_next
                or "tieng anh" in _vn_normalize_no_diacritics(combined_next)
            ):
                col_mapping["tieng_anh"] = nxt

    required = ["stt", "class_name", "last_name", "first_name"]
    if any(key not in col_mapping for key in required):
        messages.error(
            request,
            "Không đọc được header file mẫu. Vui lòng kiểm tra đúng mẫu 12BS_12CN.DS T.TIN HV...",
        )
        return redirect("adminpage:student_info_manage")

    subject_keys = ["ly", "hoa", "sinh", "su", "dia", "ktpl", "tin", "tieng_anh"]
    subject_labels = {
        "ly": "LÝ",
        "hoa": "HÓA",
        "sinh": "SINH",
        "su": "SỬ",
        "dia": "ĐỊA",
        "ktpl": "KTPL",
        "tin": "TIN",
        "tieng_anh": "TIẾNG ANH",
    }

    data: list[dict] = []
    errors: list[str] = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8):
        if not any(row):
            continue
        stt_value = row[col_mapping["stt"]] if col_mapping["stt"] < len(row) else None
        class_name = (
            str(row[col_mapping["class_name"]]).strip()
            if col_mapping["class_name"] < len(row) and row[col_mapping["class_name"]]
            else ""
        )
        last_name = (
            str(row[col_mapping["last_name"]]).strip()
            if col_mapping["last_name"] < len(row) and row[col_mapping["last_name"]]
            else ""
        )
        first_name = (
            str(row[col_mapping["first_name"]]).strip()
            if col_mapping["first_name"] < len(row) and row[col_mapping["first_name"]]
            else ""
        )

        if not class_name or not first_name:
            errors.append(f"Dòng {excel_row}: thiếu lớp hoặc tên học viên.")
            continue

        birthday_value = (
            _parse_excel_date(row[col_mapping["birthday"]])
            if col_mapping.get("birthday") is not None
            and col_mapping["birthday"] < len(row)
            else None
        )
        full_name = f"{last_name} {first_name}".strip()
        id_number = (
            str(row[col_mapping["id_number"]]).strip()
            if col_mapping.get("id_number") is not None
            and col_mapping["id_number"] < len(row)
            and row[col_mapping["id_number"]]
            else ""
        )
        student_code = _resolve_student_code_from_graduation_data(
            id_number=id_number,
            class_name=class_name,
            full_name=full_name,
            birthday=birthday_value,
        )
        if not student_code:
            errors.append(
                f"Dòng {excel_row}: không tìm thấy mã học viên chuẩn đăng ký tốt nghiệp cho '{full_name}' "
                f"(CCCD: {id_number or 'trống'})."
            )
            continue
        if len(str(student_code)) != 7 or not str(student_code).isdigit():
            errors.append(f"Dòng {excel_row}: mã học viên '{student_code}' không đúng chuẩn 7 số.")
            continue
        chosen_subjects = []
        for key in subject_keys:
            sidx = col_mapping.get(key)
            if sidx is None or sidx >= len(row):
                continue
            raw = row[sidx]
            if raw is None:
                continue
            text = str(raw).strip().upper()
            if text.startswith("="):
                continue
            if text and text not in ("0", ".", "X"):
                chosen_subjects.append(subject_labels[key])

        entry = {
            "student_code": student_code,
            "stt": int(stt_value) if str(stt_value).strip().isdigit() else None,
            "class_name": class_name,
            "full_name": full_name,
            "birthday": birthday_value.strftime("%Y-%m-%d") if birthday_value else None,
            "birth_place": str(row[col_mapping["birth_place"]]).strip()
            if col_mapping.get("birth_place") is not None
            and col_mapping["birth_place"] < len(row)
            and row[col_mapping["birth_place"]]
            else "",
            "gender": str(row[col_mapping["gender"]]).strip()
            if col_mapping.get("gender") is not None
            and col_mapping["gender"] < len(row)
            and row[col_mapping["gender"]]
            else "",
            "ethnicity": str(row[col_mapping["ethnicity"]]).strip()
            if col_mapping.get("ethnicity") is not None
            and col_mapping["ethnicity"] < len(row)
            and row[col_mapping["ethnicity"]]
            else "",
            "id_number": id_number,
            "contact_address": str(row[col_mapping["contact_address"]]).strip()
            if col_mapping.get("contact_address") is not None
            and col_mapping["contact_address"] < len(row)
            and row[col_mapping["contact_address"]]
            else "",
            "email": str(row[col_mapping["email"]]).strip()
            if col_mapping.get("email") is not None
            and col_mapping["email"] < len(row)
            and row[col_mapping["email"]]
            else "",
            "phone": str(row[col_mapping["phone"]]).strip()
            if col_mapping.get("phone") is not None
            and col_mapping["phone"] < len(row)
            and row[col_mapping["phone"]]
            else "",
            "highschool_10": _default_highschool_place_if_empty(
                str(row[col_mapping["highschool_10"]]).strip()
                if col_mapping.get("highschool_10") is not None
                and col_mapping["highschool_10"] < len(row)
                and row[col_mapping["highschool_10"]]
                else ""
            ),
            "highschool_11": _default_highschool_place_if_empty(
                str(row[col_mapping["highschool_11"]]).strip()
                if col_mapping.get("highschool_11") is not None
                and col_mapping["highschool_11"] < len(row)
                and row[col_mapping["highschool_11"]]
                else ""
            ),
            "exam_subjects": chosen_subjects,
        }
        data.append(entry)

    request.session["student_info_import_data"] = data
    request.session["student_info_import_errors"] = errors
    return redirect("adminpage:student_info_import_preview")


def student_info_save_imported(request):
    if not request.user.is_authenticated:
        return redirect("homepage:login")
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != "admin":
        return redirect("homepage:Homepage")
    if request.method != "POST":
        return redirect("adminpage:student_info_manage")

    rows = request.session.get("student_info_import_data", [])
    if not rows:
        messages.error(request, "Không có dữ liệu import tạm. Vui lòng upload file trước.")
        return redirect("adminpage:student_info_manage")

    created_count = 0
    updated_count = 0
    for row in rows:
        birthday_raw = row.get("birthday")
        birthday_obj = None
        if birthday_raw:
            try:
                birthday_obj = datetime.strptime(str(birthday_raw), "%Y-%m-%d").date()
            except ValueError:
                birthday_obj = None
        row_to_save = dict(row)
        row_to_save["birthday"] = birthday_obj
        row_to_save["highschool_10"] = _default_highschool_place_if_empty(row.get("highschool_10"))
        row_to_save["highschool_11"] = _default_highschool_place_if_empty(row.get("highschool_11"))

        defaults = {k: v for k, v in row_to_save.items() if k != "student_code"}
        _, created = StudentInfoRecord.objects.update_or_create(
            student_code=row["student_code"],
            defaults=defaults,
        )
        if created:
            created_count += 1
        else:
            updated_count += 1

    request.session.pop("student_info_import_data", None)
    request.session.pop("student_info_import_errors", None)
    messages.success(
        request,
        f"Đã lưu danh sách kiểm tra thông tin: tạo mới {created_count}, cập nhật {updated_count}.",
    )
    return redirect("adminpage:student_info_manage")
