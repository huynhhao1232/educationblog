import json
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.core.mail import send_mail
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Count
from django.db.utils import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.html import strip_tags
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from homepage.models import (
    Account,
    AccountType,
    AdmissionForm,
    Campus,
    CampusShiftGroup,
    CampusVocationalLink,
    Shift,
    SubjectGroup,
    VocationalCampus,
    VocationalTrade,
)
from homepage.admission_validation import (
    parse_graduation_scores,
    validate_admission_conduct,
    validate_admission_email,
    validate_admission_image,
)
from homepage.validators import validate_vn_cccd, validate_vn_phone

def get_Admission(request):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            # Get all admission records
            admission_list = AdmissionForm.objects.all()

            # Apply filters
            name_filter = request.GET.get('name', '').strip()
            id_number_filter = request.GET.get('id_number', '').strip()
            gender_filter = request.GET.get('gender', '')
            status_filter = request.GET.get('status', '')
            campus_filter = request.GET.get('campus', '')
            shift_filter = request.GET.get('shift', '')
            subject_group_filter = request.GET.get('subject_group', '')
            exam_score_order = request.GET.get('exam_score', '')
            avg_score_order = request.GET.get('avg_score', '')
            graduation_year_filter = request.GET.get('graduation_year', '')
            conduct_filter = request.GET.get('conduct', '')
            from_date_filter = request.GET.get('from_date', '')
            to_date_filter = request.GET.get('to_date', '')

            if name_filter:
                admission_list = admission_list.filter(
                    full_name__contains=name_filter.upper()
                )

            if id_number_filter:
                admission_list = admission_list.filter(id_number__icontains=id_number_filter)

            if gender_filter:
                admission_list = admission_list.filter(gender=gender_filter)

            if status_filter:
                status_bool = status_filter == '1'
                admission_list = admission_list.filter(enable=status_bool)

            if campus_filter:
                admission_list = admission_list.filter(campus_id=campus_filter)

            if shift_filter:
                admission_list = admission_list.filter(shift_id=shift_filter)

            if subject_group_filter:
                admission_list = admission_list.filter(subject_group_id=subject_group_filter)

            if exam_score_order:
                order_by = '-exam_score' if exam_score_order == 'desc' else 'exam_score'
                admission_list = admission_list.order_by(order_by)

            if avg_score_order:
                order_by = '-avg_score' if avg_score_order == 'desc' else 'avg_score'
                admission_list = admission_list.order_by(order_by)

            if graduation_year_filter:
                admission_list = admission_list.filter(graduation_year=graduation_year_filter)

            if conduct_filter:
                admission_list = admission_list.filter(conduct=conduct_filter)
            if from_date_filter:
                admission_list = admission_list.filter(created_at__date__gte=from_date_filter)
            if to_date_filter:
                admission_list = admission_list.filter(created_at__date__lte=to_date_filter)

            # Set up pagination
            paginator = Paginator(admission_list, 10)  # Show 10 records per page
            page = request.GET.get('page', 1)

            try:
                ad = paginator.page(page)
            except PageNotAnInteger:
                ad = paginator.page(1)
            except EmptyPage:
                ad = paginator.page(paginator.num_pages)

            # Get data for filter dropdowns
            from homepage.models import Campus, Shift, SubjectGroup
            campuses = Campus.objects.all()
            shifts = Shift.objects.all()
            subject_groups = SubjectGroup.objects.all()
            # Graduation years for dropdown
            graduation_years = AdmissionForm.objects.order_by('graduation_year').values_list('graduation_year', flat=True).distinct()

            # Lấy thông báo từ session nếu có
            conduct_update_result = request.session.pop('conduct_update_result', None)

            context = {
                'ad': ad,
                'new_users': admission_list.count(),
                # Pass filter values back to template
                'name_filter': name_filter,
                'id_number_filter': id_number_filter,
                'graduation_year_filter': graduation_year_filter,
                'status_filter': status_filter,
                'campus_filter': campus_filter,
                'shift_filter': shift_filter,
                'subject_group_filter': subject_group_filter,
                'exam_score_order': exam_score_order,
                'avg_score_order': avg_score_order,
                # Pass data for filter dropdowns
                'campuses': campuses,
                'shifts': shifts,
                'subject_groups': subject_groups,
                'graduation_years': graduation_years,
                'conduct_filter': conduct_filter,
                'from_date_filter': from_date_filter,
                'to_date_filter': to_date_filter,
                'conduct_update_result': conduct_update_result,
                'total_admissions': AdmissionForm.objects.count(),
            }
            return render(request, 'adminpageSIMCODE/admission.html', context)
        else:
            return redirect('homepage:Homepage')
    return redirect('homepage:login')


def _calc_admission_capacity(num_classes):
    """50 học viên/lớp + 20% dự phòng."""
    return int(num_classes * 50 * 1.2)


@transaction.atomic
def delete_all_admissions(request):
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method != 'POST':
        return redirect('adminpage:admission')

    confirm_text = request.POST.get('confirm_text', '').strip().upper()
    if confirm_text != 'XOA TAT CA':
        messages.error(request, 'Vui lòng nhập đúng "XOA TAT CA" để xác nhận xoá.')
        return redirect('adminpage:admission')

    from django.db.models import Count

    combos = (
        AdmissionForm.objects.values('campus_id', 'shift_id', 'subject_group_id')
        .annotate(cnt=Count('id'))
    )
    for row in combos:
        try:
            csg = CampusShiftGroup.objects.get(
                campus_id=row['campus_id'],
                shift_id=row['shift_id'],
                subject_group_id=row['subject_group_id'],
            )
            csg.registration_count += row['cnt']
            csg.save(update_fields=['registration_count'])
        except CampusShiftGroup.DoesNotExist:
            pass

    deleted_count = AdmissionForm.objects.count()
    AdmissionForm.objects.all().delete()
    messages.success(
        request,
        f'Đã xoá {deleted_count} học viên đăng ký tuyển sinh và hoàn trả chỗ đăng ký.',
    )
    return redirect('adminpage:admission')


def campus_shift_group_manage(request):
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == '0':
            campus_id = request.POST.get('campus_id')
            shift_id = request.POST.get('shift_id')
            subject_group_id = request.POST.get('subject_group_id')
            try:
                number_of_classes = max(0, int(request.POST.get('number_of_classes', 0)))
            except (TypeError, ValueError):
                number_of_classes = 0
            auto_calc = request.POST.get('auto_calc') == '1'
            if auto_calc:
                registration_count = _calc_admission_capacity(number_of_classes)
            else:
                try:
                    registration_count = max(0, int(request.POST.get('registration_count', 0)))
                except (TypeError, ValueError):
                    registration_count = 0

            if CampusShiftGroup.objects.filter(
                campus_id=campus_id,
                shift_id=shift_id,
                subject_group_id=subject_group_id,
            ).exists():
                messages.error(request, 'Cơ sở + ca học + tổ hợp môn này đã tồn tại.')
            else:
                CampusShiftGroup.objects.create(
                    campus_id=campus_id,
                    shift_id=shift_id,
                    subject_group_id=subject_group_id,
                    number_of_classes=number_of_classes,
                    registration_count=registration_count,
                )
                messages.success(request, 'Đã thêm cấu hình tuyển sinh.')
        elif action == '1':
            csg = get_object_or_404(CampusShiftGroup, id=request.POST.get('csg_id'))
            try:
                number_of_classes = max(0, int(request.POST.get('number_of_classes', csg.number_of_classes)))
            except (TypeError, ValueError):
                number_of_classes = csg.number_of_classes
            auto_calc = request.POST.get('auto_calc') == '1'
            if auto_calc:
                registration_count = _calc_admission_capacity(number_of_classes)
            else:
                try:
                    registration_count = max(0, int(request.POST.get('registration_count', csg.registration_count)))
                except (TypeError, ValueError):
                    registration_count = csg.registration_count
            csg.number_of_classes = number_of_classes
            csg.registration_count = registration_count
            csg.save()
            messages.success(request, 'Đã cập nhật cấu hình tuyển sinh.')
        elif action == '2':
            csg = get_object_or_404(CampusShiftGroup, id=request.POST.get('csg_id'))
            if AdmissionForm.objects.filter(
                campus=csg.campus,
                shift=csg.shift,
                subject_group=csg.subject_group,
            ).exists():
                messages.error(request, 'Không thể xoá vì đã có học viên đăng ký.')
            else:
                csg.delete()
                messages.success(request, 'Đã xoá cấu hình tuyển sinh.')
        return redirect('adminpage:campus_shift_group_manage')

    campus_filter = request.GET.get('campus', '')
    shift_filter = request.GET.get('shift', '')

    queryset = CampusShiftGroup.objects.select_related(
        'campus', 'shift', 'subject_group'
    ).order_by('campus__name', 'shift__name', 'subject_group__code')

    if campus_filter:
        queryset = queryset.filter(campus_id=campus_filter)
    if shift_filter:
        queryset = queryset.filter(shift_id=shift_filter)

    rows = []
    for csg in queryset:
        registered = AdmissionForm.objects.filter(
            campus=csg.campus,
            shift=csg.shift,
            subject_group=csg.subject_group,
        ).count()
        rows.append({
            'obj': csg,
            'registered': registered,
            'max_capacity': _calc_admission_capacity(csg.number_of_classes),
        })

    context = {
        'rows': rows,
        'campuses': Campus.objects.all(),
        'shifts': Shift.objects.all(),
        'subject_groups': SubjectGroup.objects.all(),
        'campus_filter': campus_filter,
        'shift_filter': shift_filter,
    }
    return render(request, 'adminpageSIMCODE/campus_shift_group.html', context)


def admission_vocational_manage(request):
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_vc':
            name = (request.POST.get('vc_name') or '').strip()
            address = (request.POST.get('vc_address') or '').strip()
            code = (request.POST.get('vc_code') or '').strip()
            if not name:
                messages.error(request, 'Tên cơ sở dạy nghề không được để trống.')
            elif code and VocationalCampus.objects.filter(code=code).exists():
                messages.error(request, 'Mã cơ sở dạy nghề đã tồn tại.')
            else:
                vc = VocationalCampus.objects.create(name=name, address=address, code=code or '')
                if not vc.code:
                    vc.code = f'VN{vc.id}'
                    vc.save(update_fields=['code'])
                messages.success(request, 'Đã thêm cơ sở dạy nghề.')

        elif action == 'edit_vc':
            vc = get_object_or_404(VocationalCampus, id=request.POST.get('vc_id'))
            name = (request.POST.get('vc_name') or '').strip()
            address = (request.POST.get('vc_address') or '').strip()
            code = (request.POST.get('vc_code') or '').strip()
            if not name:
                messages.error(request, 'Tên cơ sở dạy nghề không được để trống.')
            elif code and VocationalCampus.objects.filter(code=code).exclude(id=vc.id).exists():
                messages.error(request, 'Mã cơ sở dạy nghề đã tồn tại.')
            else:
                vc.name = name
                vc.address = address
                if code:
                    vc.code = code
                vc.save()
                messages.success(request, 'Đã cập nhật cơ sở dạy nghề.')

        elif action == 'delete_vc':
            vc = get_object_or_404(VocationalCampus, id=request.POST.get('vc_id'))
            if AdmissionForm.objects.filter(vocational_campus=vc).exists():
                messages.error(request, 'Không thể xóa vì đã có học viên đăng ký nghề tại cơ sở này.')
            else:
                vc.delete()
                messages.success(request, 'Đã xóa cơ sở dạy nghề.')

        elif action == 'add_trade':
            vc = get_object_or_404(VocationalCampus, id=request.POST.get('vc_id'))
            trade_name = (request.POST.get('trade_name') or '').strip()
            if not trade_name:
                messages.error(request, 'Tên nghề không được để trống.')
            elif VocationalTrade.objects.filter(vocational_campus=vc, name=trade_name).exists():
                messages.error(request, 'Nghề này đã tồn tại tại cơ sở dạy nghề.')
            else:
                VocationalTrade.objects.create(vocational_campus=vc, name=trade_name)
                messages.success(request, 'Đã thêm nghề.')

        elif action == 'delete_trade':
            trade = get_object_or_404(VocationalTrade, id=request.POST.get('trade_id'))
            if AdmissionForm.objects.filter(vocational_trade=trade).exists():
                messages.error(request, 'Không thể xóa vì đã có học viên đăng ký nghề này.')
            else:
                trade.delete()
                messages.success(request, 'Đã xóa nghề.')

        elif action == 'link':
            admission_campus_id = request.POST.get('admission_campus_id')
            vocational_campus_id = request.POST.get('vocational_campus_id')
            if not admission_campus_id or not vocational_campus_id:
                messages.error(request, 'Vui lòng chọn đủ cơ sở tuyển sinh và cơ sở dạy nghề.')
            elif CampusVocationalLink.objects.filter(
                admission_campus_id=admission_campus_id,
                vocational_campus_id=vocational_campus_id,
            ).exists():
                messages.error(request, 'Liên kết này đã tồn tại.')
            else:
                CampusVocationalLink.objects.create(
                    admission_campus_id=admission_campus_id,
                    vocational_campus_id=vocational_campus_id,
                )
                messages.success(request, 'Đã liên kết cơ sở dạy nghề.')

        elif action == 'unlink':
            link = get_object_or_404(CampusVocationalLink, id=request.POST.get('link_id'))
            link.delete()
            messages.success(request, 'Đã gỡ liên kết cơ sở dạy nghề.')

        campus_filter = request.POST.get('campus_filter', '')
        url = reverse('adminpage:admission_vocational_manage')
        if campus_filter:
            url += f'?campus={campus_filter}'
        return redirect(url)

    campus_filter = request.GET.get('campus', '')
    vocational_campuses = VocationalCampus.objects.prefetch_related('trades').order_by('name')

    links = CampusVocationalLink.objects.select_related(
        'admission_campus', 'vocational_campus'
    ).order_by('admission_campus__name', 'vocational_campus__name')
    if campus_filter:
        links = links.filter(admission_campus_id=campus_filter)

    linked_vc_ids = set()
    if campus_filter:
        linked_vc_ids = set(
            CampusVocationalLink.objects.filter(
                admission_campus_id=campus_filter
            ).values_list('vocational_campus_id', flat=True)
        )

    context = {
        'campuses': Campus.objects.all().order_by('name'),
        'vocational_campuses': vocational_campuses,
        'links': links,
        'campus_filter': campus_filter,
        'linked_vc_ids': linked_vc_ids,
    }
    return render(request, 'adminpageSIMCODE/admission_vocational.html', context)


def _parse_optional_decimal(raw):
    if raw is None:
        return None
    text = str(raw).strip().replace(',', '.')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _apply_admission_scores(admission, post_data):
    graduation_year = post_data.get('graduation_year') or admission.graduation_year
    scores, err = parse_graduation_scores(post_data, graduation_year)
    if err:
        return err
    for field, value in scores.items():
        setattr(admission, field, value)
    return None


def _score_for_input(value):
    """Chuẩn hoá điểm cho input HTML: rỗng nếu null, dấu chấm thập phân."""
    if value is None:
        return ''
    text = str(value).strip()
    if not text or text.lower() == 'none':
        return ''
    return text.replace(',', '.')


def _admission_score_values(admission):
    fields = (
        'exam_score', 'avg_score', 'math_score', 'literature_score',
        'math_score_6', 'literature_score_6',
        'math_score_7', 'literature_score_7',
        'math_score_8', 'literature_score_8',
        'math_score_9', 'literature_score_9',
    )
    return {field: _score_for_input(getattr(admission, field)) for field in fields}


def _restore_admission_slot(admission):
    try:
        with transaction.atomic():
            csg = CampusShiftGroup.objects.select_for_update().get(
                campus=admission.campus,
                shift=admission.shift,
                subject_group=admission.subject_group,
            )
            csg.registration_count += 1
            csg.save(update_fields=['registration_count'])
    except CampusShiftGroup.DoesNotExist:
        pass


def _take_admission_slot(campus, shift, subject_group):
    try:
        csg = CampusShiftGroup.objects.select_for_update().get(
            campus=campus,
            shift=shift,
            subject_group=subject_group,
        )
    except CampusShiftGroup.DoesNotExist:
        return False, 'Không tìm thấy cấu hình tuyển sinh'
    if csg.registration_count <= 0:
        return False, 'Tổ hợp môn này đã hết chỗ'
    csg.registration_count -= 1
    csg.save(update_fields=['registration_count'])
    return True, None


def _count_registered(campus, shift, subject_group):
    return AdmissionForm.objects.filter(
        campus=campus,
        shift=shift,
        subject_group=subject_group,
    ).count()


def download_admission_docx(request, admission_id):
    if not request.user.is_authenticated:
        return redirect('homepage:login')
    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    admission = get_object_or_404(
        AdmissionForm.objects.select_related('campus', 'shift', 'subject_group'),
        id=admission_id,
    )
    try:
        from .admission_docx import admission_docx_filename, build_admission_docx
        buffer = build_admission_docx(admission)
        filename = admission_docx_filename(admission)
    except FileNotFoundError as exc:
        messages.error(request, str(exc))
        return redirect('adminpage:letter', admission_id=admission_id)

    from urllib.parse import quote
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


def get_Letter(request, admission_id):
    if request.user.is_authenticated:
        account = Account.objects.get(user = request.user)
        accounttype = AccountType.objects.get(accounttype_id = account.accounttype.accounttype_id)
        if accounttype.accounttype_role == 'admin':
            # Get admission record
            admission = get_object_or_404(AdmissionForm, id=admission_id)

            if request.method == 'POST':
                action = request.POST.get('action', 'update')
                if action == 'delete':
                    with transaction.atomic():
                        _restore_admission_slot(admission)
                        admission.delete()
                    return JsonResponse({
                        'status': 'success',
                        'redirect': reverse('adminpage:admission'),
                    })
                try:
                    # Kiểm tra trạng thái trước khi cập nhật
                    old_status = admission.enable
                    new_status = request.POST.get('enable') == '1'

                    # Nếu đã duyệt rồi thì không cho phép thay đổi trạng thái
                    if old_status:
                        new_status = True

                    # Update admission record
                    admission.full_name = request.POST.get('full_name')
                    admission.gender = request.POST.get('gender')
                    admission.birthday = request.POST.get('birthday')
                    admission.ethnicity = request.POST.get('ethnicity') or ''
                    admission.email = (request.POST.get('email') or '').strip()
                    email_err = validate_admission_email(admission.email)
                    if email_err:
                        return JsonResponse({'status': 'error', 'message': email_err})

                    phone = (request.POST.get('phone') or '').strip()
                    phone_err = validate_vn_phone(phone)
                    if phone_err:
                        return JsonResponse({'status': 'error', 'message': phone_err})

                    id_number = (request.POST.get('id_number') or '').strip()
                    cccd_err = validate_vn_cccd(id_number)
                    if cccd_err:
                        return JsonResponse({'status': 'error', 'message': cccd_err})

                    if AdmissionForm.objects.filter(id_number=id_number).exclude(pk=admission.pk).exists():
                        return JsonResponse({
                            'status': 'error',
                            'message': 'Số CCCD này đã được sử dụng bởi hồ sơ khác',
                        })

                    father_phone = (request.POST.get('father_phone') or '').strip()
                    father_phone_err = validate_vn_phone(father_phone, 'Số điện thoại của cha')
                    if father_phone_err:
                        return JsonResponse({'status': 'error', 'message': father_phone_err})

                    mother_phone = (request.POST.get('mother_phone') or '').strip()
                    mother_phone_err = validate_vn_phone(mother_phone, 'Số điện thoại của mẹ')
                    if mother_phone_err:
                        return JsonResponse({'status': 'error', 'message': mother_phone_err})

                    admission.id_number = id_number
                    admission.id_issued_date = request.POST.get('id_issued_date')

                    admission.phone = phone

                    cccd_image = request.FILES.get('cccd_image')
                    if cccd_image:
                        img_err = validate_admission_image(cccd_image, 'Ảnh CCCD')
                        if img_err:
                            return JsonResponse({'status': 'error', 'message': img_err})
                        if admission.cccd_image:
                            try:
                                os.remove(os.path.join(settings.MEDIA_ROOT, str(admission.cccd_image)))
                            except OSError:
                                pass
                        admission.cccd_image = cccd_image

                    school_record_image = request.FILES.get('school_record_image')
                    if school_record_image:
                        img_err = validate_admission_image(school_record_image, 'Ảnh học bạ')
                        if img_err:
                            return JsonResponse({'status': 'error', 'message': img_err})
                        if admission.school_record_image:
                            try:
                                os.remove(os.path.join(settings.MEDIA_ROOT, str(admission.school_record_image)))
                            except OSError:
                                pass
                        admission.school_record_image = school_record_image

                    admission.cccd_province = request.POST.get('cccd_province', '')
                    admission.cccd_district = request.POST.get('cccd_district', '')
                    admission.cccd_ward = request.POST.get('cccd_ward', '')
                    admission.cccd_town = request.POST.get('cccd_town', '')

                    admission.birth_reg_town = request.POST.get('birth_reg_town', '')
                    admission.current_district = request.POST.get('current_district', '')
                    admission.current_ward = request.POST.get('current_ward', '')
                    admission.current_province = request.POST.get('current_province', '')

                    admission.birth_place_facility = request.POST.get('birth_place_facility', '')

                    admission.graduation_school = request.POST.get('graduation_school')
                    conduct_err = validate_admission_conduct(
                        request.POST.get('conduct_6', ''),
                        request.POST.get('conduct_7', ''),
                        request.POST.get('conduct_8', ''),
                        request.POST.get('conduct_9', ''),
                    )
                    if conduct_err:
                        return JsonResponse({'status': 'error', 'message': conduct_err})
                    for field in ('conduct_6', 'conduct_7', 'conduct_8', 'conduct_9'):
                        value = request.POST.get(field, '')
                        setattr(admission, field, value or None)
                    admission.conduct = admission.conduct_9 or admission.conduct

                    study_vocational = request.POST.get('study_vocational', 'no')
                    admission.study_vocational = study_vocational
                    if study_vocational == 'yes':
                        vc_id = request.POST.get('vocational_campus_id')
                        vt_id = request.POST.get('vocational_trade_id')
                        if not vc_id or not vt_id:
                            return JsonResponse({'status': 'error', 'message': 'Vui lòng chọn cơ sở và nghề dạy nghề'})
                        if not CampusVocationalLink.objects.filter(
                            admission_campus=admission.campus,
                            vocational_campus_id=vc_id,
                        ).exists():
                            return JsonResponse({'status': 'error', 'message': 'Cơ sở dạy nghề không hợp lệ'})
                        try:
                            trade = VocationalTrade.objects.get(id=vt_id, vocational_campus_id=vc_id)
                            admission.vocational_campus = trade.vocational_campus
                            admission.vocational_trade = trade
                        except VocationalTrade.DoesNotExist:
                            return JsonResponse({'status': 'error', 'message': 'Nghề đăng ký không hợp lệ'})
                    else:
                        admission.vocational_campus = None
                        admission.vocational_trade = None

                    # Thông tin phụ huynh
                    admission.father_name = request.POST.get('father_name')
                    admission.father_job = request.POST.get('father_job')
                    admission.father_birth = request.POST.get('father_birth')
                    admission.father_phone = father_phone

                    admission.mother_name = request.POST.get('mother_name')
                    admission.mother_job = request.POST.get('mother_job')
                    admission.mother_birth = request.POST.get('mother_birth')
                    admission.mother_phone = mother_phone

                    score_err = _apply_admission_scores(admission, request.POST)
                    if score_err:
                        return JsonResponse({'status': 'error', 'message': score_err})

                    admission.enable = new_status

                    shift_id = request.POST.get('shift')
                    subject_group_id = request.POST.get('subject_group')
                    new_shift_id = int(shift_id) if shift_id else admission.shift_id
                    new_sg_id = int(subject_group_id) if subject_group_id else admission.subject_group_id
                    slot_changed = (
                        new_shift_id != admission.shift_id
                        or new_sg_id != admission.subject_group_id
                    )

                    with transaction.atomic():
                        if slot_changed:
                            _restore_admission_slot(admission)
                            try:
                                new_shift = Shift.objects.get(id=new_shift_id)
                                new_sg = SubjectGroup.objects.get(id=new_sg_id)
                            except (Shift.DoesNotExist, SubjectGroup.DoesNotExist):
                                return JsonResponse({
                                    'status': 'error',
                                    'message': 'Ca học hoặc tổ hợp môn không hợp lệ',
                                })
                            ok, slot_err = _take_admission_slot(
                                admission.campus, new_shift, new_sg
                            )
                            if not ok:
                                return JsonResponse({'status': 'error', 'message': slot_err})
                            admission.shift = new_shift
                            admission.subject_group = new_sg

                        admission.save()

                    # Nếu trạng thái thay đổi từ chưa duyệt sang đã duyệt và có email
                    if not old_status and new_status and admission.email:
                        try:
                            # Chuẩn bị nội dung email
                            html_message = render_to_string('adminpageSIMCODE/email_template.html', {
                                'full_name': admission.full_name,
                                'graduation_year': admission.graduation_year,
                                'campus': admission.campus.name,
                                'shift': admission.shift.name,
                                'subject_group': admission.subject_group.code
                            })
                            plain_message = strip_tags(html_message)

                            # Gửi email
                            send_mail(
                                subject='Thông báo xét duyệt hồ sơ nhập học',
                                message=plain_message,
                                from_email=settings.EMAIL_HOST_USER,
                                recipient_list=[admission.email],
                                html_message=html_message,
                                fail_silently=False,
                            )
                        except Exception as e:
                            print(f"Error sending email: {e}")
                            # Continue with success response even if email fails

                    return JsonResponse({'status': 'success'})
                except Exception as e:
                    print(f"Error updating admission: {e}")
                    return JsonResponse({'status': 'error', 'message': str(e)})

            import json
            vocational_campuses = VocationalCampus.objects.filter(
                admission_links__admission_campus=admission.campus
            ).distinct().order_by('name')
            vc_trades = {}
            for vc in vocational_campuses.prefetch_related('trades'):
                vc_trades[str(vc.id)] = [
                    {'id': t.id, 'name': t.name} for t in vc.trades.all()
                ]

            context = {
                'admission': admission,
                'scores': _admission_score_values(admission),
                'subject_groups': SubjectGroup.objects.all(),
                'shifts': Shift.objects.all(),
                'vocational_campuses': vocational_campuses,
                'vocational_trades_json': json.dumps(vc_trades),
            }
            return render(request, 'adminpageSIMCODE/letter.html', context)
    return redirect('adminpage:login')

def load_vn_address_data():
    try:
        data_path = os.path.join(settings.BASE_DIR, 'adminpage', 'static', 'data.json')
        with open(data_path, encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading VN address data: {e}")
        return []

VN_ADDRESS_DATA = load_vn_address_data()

def get_province_name(code):
    try:
        if not code or code == 'None' or code == '':
            return ''
        for province in VN_ADDRESS_DATA:
            if province.get('Id') == code:
                return province.get('Name', code)
        return code
    except Exception as e:
        print(f"Error in get_province_name: {e}")
        return code if code else ''

def get_district_name(province_code, district_code):
    try:
        if not province_code or province_code == 'None' or province_code == '' or not district_code or district_code == 'None' or district_code == '':
            return ''
        for province in VN_ADDRESS_DATA:
            if province.get('Id') == province_code:
                for district in province.get('Districts', []):
                    if district.get('Id') == district_code:
                        return district.get('Name', district_code)
        return district_code
    except Exception as e:
        print(f"Error in get_district_name: {e}")
        return district_code if district_code else ''

def get_ward_name(province_code, district_code, ward_code):
    try:
        if not province_code or province_code == 'None' or province_code == '' or not district_code or district_code == 'None' or district_code == '' or not ward_code or ward_code == 'None' or ward_code == '':
            return ''
        for province in VN_ADDRESS_DATA:
            if province.get('Id') == province_code:
                for district in province.get('Districts', []):
                    if district.get('Id') == district_code:
                        for ward in district.get('Wards', []):
                            if ward.get('Id') == ward_code:
                                return ward.get('Name', ward_code)
        return ward_code
    except Exception as e:
        print(f"Error in get_ward_name: {e}")
        return ward_code if ward_code else ''

def export_approved_admissions(request):
    """Export approved admissions to Excel file."""
    if not request.user.is_authenticated:
        return redirect('homepage:login')

    account = Account.objects.get(user=request.user)
    accounttype = AccountType.objects.get(accounttype_id=account.accounttype.accounttype_id)
    if accounttype.accounttype_role != 'admin':
        return redirect('homepage:Homepage')

    # Get all admissions (not just approved ones, apply filters)
    admissions = AdmissionForm.objects.all()

    # Apply filters from GET parameters
    name_filter = request.GET.get('name', '').strip()
    id_number_filter = request.GET.get('id_number', '').strip()
    gender_filter = request.GET.get('gender', '')
    status_filter = request.GET.get('status', '')
    campus_filter = request.GET.get('campus', '')
    shift_filter = request.GET.get('shift', '')
    subject_group_filter = request.GET.get('subject_group', '')
    exam_score_order = request.GET.get('exam_score', '')
    avg_score_order = request.GET.get('avg_score', '')
    graduation_year_filter = request.GET.get('graduation_year', '')
    conduct_filter = request.GET.get('conduct', '')
    from_date_filter = request.GET.get('from_date', '')
    to_date_filter = request.GET.get('to_date', '')

    if name_filter:
        admissions = admissions.filter(
            full_name__contains=name_filter.upper()
        )

    if id_number_filter:
        admissions = admissions.filter(id_number__icontains=id_number_filter)

    if gender_filter:
        admissions = admissions.filter(gender=gender_filter)

    if status_filter:
        status_bool = status_filter == '1'
        admissions = admissions.filter(enable=status_bool)

    if campus_filter:
        admissions = admissions.filter(campus_id=campus_filter)

    if shift_filter:
        admissions = admissions.filter(shift_id=shift_filter)

    if subject_group_filter:
        admissions = admissions.filter(subject_group_id=subject_group_filter)

    if exam_score_order:
        order_by = '-exam_score' if exam_score_order == 'desc' else 'exam_score'
        admissions = admissions.order_by(order_by)

    if avg_score_order:
        order_by = '-avg_score' if avg_score_order == 'desc' else 'avg_score'
        admissions = admissions.order_by(order_by)

    if graduation_year_filter:
        admissions = admissions.filter(graduation_year=graduation_year_filter)

    if conduct_filter:
        admissions = admissions.filter(conduct=conduct_filter)
    if from_date_filter:
        admissions = admissions.filter(created_at__date__gte=from_date_filter)
    if to_date_filter:
        admissions = admissions.filter(created_at__date__lte=to_date_filter)

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Set title based on filters
    title = "Danh sách học sinh"
    if status_filter == '1':
        title += " đã duyệt"
    elif status_filter == '0':
        title += " chưa duyệt"
    else:
        title += " (tất cả)"

    ws.title = title

    # Define headers
    headers = [
        'Họ và tên', 'Giới tính', 'Ngày sinh', 'Dân tộc', 'Tôn giáo',
        'Email', 'CCCD/CMND', 'Ngày cấp', 'Nơi cấp',
        'Địa chỉ thường trú', 'Quê quán',
        'Nơi đăng ký khai sinh', 'Nơi sinh', 'Nơi ở hiện tại',
        'Họ tên cha', 'Nghề nghiệp cha', 'Năm sinh cha', 'SĐT cha',
        'Họ tên mẹ', 'Nghề nghiệp mẹ', 'Năm sinh mẹ', 'SĐT mẹ',
        'Trường đang học', 'Công việc hiện tại',
        'Điểm thi chuyển cấp', 'Điểm TB Toán, Văn',
        'Điểm Toán lớp 6', 'Điểm Văn lớp 6',
        'Điểm Toán lớp 7', 'Điểm Văn lớp 7',
        'Điểm Toán lớp 8', 'Điểm Văn lớp 8',
        'Điểm Toán lớp 9', 'Điểm Văn lớp 9',
        'Hạnh kiểm', 'Học lực', 'Năm tốt nghiệp',
        'Ban đăng ký', 'Ca học', 'Cơ sở'
    ]

    # Style the header row
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write data
    for row, admission in enumerate(admissions, 2):
        data = [
            admission.full_name,
            admission.gender,
            admission.birthday.strftime('%d/%m/%Y') if admission.birthday else '',
            admission.ethnicity,
            admission.religion,
            admission.email,
            admission.id_number,
            admission.id_issued_date.strftime('%d/%m/%Y') if admission.id_issued_date else '',
            admission.id_issued_place,
            # Địa chỉ thường trú
            f"{admission.cccd_town or ''}, "
            f"{get_ward_name(str(admission.cccd_province) if admission.cccd_province else '', str(admission.cccd_district) if admission.cccd_district else '', str(admission.cccd_ward) if admission.cccd_ward else '')}, "
            f"{get_district_name(str(admission.cccd_province) if admission.cccd_province else '', str(admission.cccd_district) if admission.cccd_district else '')}, "
            f"{get_province_name(str(admission.cccd_province) if admission.cccd_province else '')}",
            # Quê quán
            get_province_name(str(admission.hometown_province) if admission.hometown_province else ''),
            # Nơi đăng ký khai sinh
            f"{admission.birth_reg_town or ''}, "
            f"{get_ward_name(str(admission.birth_reg_province) if admission.birth_reg_province else '', str(admission.birth_reg_district) if admission.birth_reg_district else '', str(admission.birth_reg_ward) if admission.birth_reg_ward else '')}, "
            f"{get_district_name(str(admission.birth_reg_province) if admission.birth_reg_province else '', str(admission.birth_reg_district) if admission.birth_reg_district else '')}, "
            f"{get_province_name(str(admission.birth_reg_province) if admission.birth_reg_province else '')}",
            # Nơi sinh
            f"{admission.birth_place_facility or ''}, "
            f"{get_ward_name(str(admission.birth_place_province) if admission.birth_place_province else '', str(admission.birth_place_district) if admission.birth_place_district else '', str(admission.birth_place_ward) if admission.birth_place_ward else '')}, "
            f"{get_district_name(str(admission.birth_place_province) if admission.birth_place_province else '', str(admission.birth_place_district) if admission.birth_place_district else '')}, "
            f"{get_province_name(str(admission.birth_place_province) if admission.birth_place_province else '')}",
            # Nơi ở hiện tại
            f"{get_ward_name(str(admission.current_province) if admission.current_province else '', str(admission.current_district) if admission.current_district else '', str(admission.current_ward) if admission.current_ward else '')}, "
            f"{get_district_name(str(admission.current_province) if admission.current_province else '', str(admission.current_district) if admission.current_district else '')}, "
            f"{get_province_name(str(admission.current_province) if admission.current_province else '')}",
            # Thông tin cha
            admission.father_name,
            admission.father_job,
            admission.father_birth,
            admission.father_phone,
            # Thông tin mẹ
            admission.mother_name,
            admission.mother_job,
            admission.mother_birth,
            admission.mother_phone,
            # Thông tin học vấn
            admission.graduation_school,
            admission.current_job,
            str(admission.exam_score) if admission.exam_score else '',
            str(admission.avg_score) if admission.avg_score else '',
            str(admission.math_score_6) if admission.math_score_6 else '',
            str(admission.literature_score_6) if admission.literature_score_6 else '',
            str(admission.math_score_7) if admission.math_score_7 else '',
            str(admission.literature_score_7) if admission.literature_score_7 else '',
            str(admission.math_score_8) if admission.math_score_8 else '',
            str(admission.literature_score_8) if admission.literature_score_8 else '',
            str(admission.math_score_9) if admission.math_score_9 else '',
            str(admission.literature_score_9) if admission.literature_score_9 else '',
            admission.conduct,
            admission.graduation_rank,
            admission.graduation_year,
            admission.subject_group.code if admission.subject_group else '',
            admission.shift.name if admission.shift else '',
            admission.campus.name if admission.campus else ''  # Chỉ lấy tên của campus
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 30)

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=danh_sach_hoc_sinh_{datetime.now().strftime("%d%m%Y")}.xlsx'

    # Save the workbook to the response
    wb.save(response)
    return response

def delete_admission(request, admission_id):
    if request.method == 'POST' and request.user.is_authenticated:
        admission = get_object_or_404(AdmissionForm, id=admission_id)
        with transaction.atomic():
            _restore_admission_slot(admission)
            admission.delete()
        return redirect('adminpage:admission')
    return redirect('adminpage:admission')

@csrf_exempt
def import_cccd_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        cccd_col = None
        start_row = None
        # Tìm dòng tiêu đề chứa 'số CCCD' (không phân biệt hoa thường)
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            for j, cell in enumerate(row):
                if cell and str(cell).strip().lower() in ['số cccd', 'so cccd', 'cccd', 'h2 (số cccd)']:
                    cccd_col = j
                    start_row = i + 1
                    break
            if cccd_col is not None:
                break
        if cccd_col is None:
            return render(request, 'adminpageSIMCODE/admission.html', {'import_result': 'Không tìm thấy cột Số CCCD trong file Excel.'})
        # Lấy danh sách số CCCD từ dòng start_row đến hết
        cccd_list = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            cccd = row[cccd_col]
            if cccd:
                cccd_list.append(str(cccd).strip())
        # Duyệt học viên và gửi email
        from homepage.models import AdmissionForm
        approved_count = 0
        email_count = 0
        for cccd in cccd_list:
            qs = AdmissionForm.objects.filter(id_number=cccd, enable=False)
            for admission in qs:
                admission.enable = True
                admission.save()
                approved_count += 1
                if admission.email:
                    try:
                        html_message = render_to_string('adminpageSIMCODE/email_template.html', {
                            'full_name': admission.full_name,
                            'graduation_year': admission.graduation_year,
                            'campus': admission.campus.name,
                            'shift': admission.shift.name,
                            'subject_group': admission.subject_group.code
                        })
                        plain_message = strip_tags(html_message)
                        send_mail(
                            subject='Thông báo xét duyệt hồ sơ nhập học',
                            message=plain_message,
                            from_email=settings.EMAIL_HOST_USER,
                            recipient_list=[admission.email],
                            html_message=html_message,
                            fail_silently=False,
                        )
                        email_count += 1
                    except Exception as e:
                        print(f"Error sending email to {admission.email}: {e}")
        result_msg = f'Đã duyệt {approved_count} học viên theo danh sách CCCD. Đã gửi email cho {email_count} học viên.'
        return render(request, 'adminpageSIMCODE/admission.html', {'import_result': result_msg})
    return render(request, 'adminpageSIMCODE/admission.html', {'import_result': 'Vui lòng chọn file Excel hợp lệ.'})

@csrf_exempt
def delete_cccd_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        cccd_col = None
        start_row = None
        # Tìm dòng tiêu đề chứa 'số CCCD' (không phân biệt hoa thường)
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            for j, cell in enumerate(row):
                if cell and str(cell).strip().lower() in ['số cccd', 'so cccd', 'cccd', 'h2 (số cccd)']:
                    cccd_col = j
                    start_row = i + 1
                    break
            if cccd_col is not None:
                break
        if cccd_col is None:
            return render(request, 'adminpageSIMCODE/admission.html', {'delete_result': 'Không tìm thấy cột Số CCCD trong file Excel.'})
        # Lấy danh sách số CCCD từ dòng start_row đến hết
        cccd_list = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            cccd = row[cccd_col]
            if cccd:
                cccd_list.append(str(cccd).strip())
        # Xoá học viên
        from homepage.models import AdmissionForm
        deleted_count, _ = AdmissionForm.objects.filter(id_number__in=cccd_list).delete()
        result_msg = f'Đã xoá {deleted_count} biểu mẫu theo danh sách CCCD.'
        return render(request, 'adminpageSIMCODE/admission.html', {'delete_result': result_msg})
    return render(request, 'adminpageSIMCODE/admission.html', {'delete_result': 'Vui lòng chọn file Excel hợp lệ.'})

def export_cccd_for_delete(request):
    from homepage.models import AdmissionForm
    admissions = AdmissionForm.objects.filter(enable=False)
    wb = Workbook()
    ws = wb.active
    ws.title = "CCCD Chua Duyet"
    ws.append(["Số CCCD"])
    for admission in admissions:
        ws.append([admission.id_number])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cccd_chua_duyet.xlsx'
    wb.save(response)
    return response

@csrf_exempt
def update_conduct_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        cccd_col = None
        conduct_col = None
        start_row = None

        # Tìm dòng tiêu đề chứa 'số CCCD' và 'hạnh kiểm'
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            for j, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).strip().lower()
                    if cell_str in ['số cccd', 'so cccd', 'cccd', 'h2 (số cccd)']:
                        cccd_col = j
                    elif cell_str in ['hạnh kiểm', 'hanh kiem', 'conduct', 'hạnh kiểm thpt']:
                        conduct_col = j
            if cccd_col is not None and conduct_col is not None:
                start_row = i + 1
                break

        if cccd_col is None:
            return redirect('adminpage:admission')

        if conduct_col is None:
            return redirect('adminpage:admission')

        # Lấy danh sách CCCD và hạnh kiểm từ dòng start_row đến hết
        update_data = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            cccd = row[cccd_col]
            conduct = row[conduct_col]
            if cccd and conduct:
                update_data.append({
                    'cccd': str(cccd).strip(),
                    'conduct': str(conduct).strip()
                })

        # Cập nhật hạnh kiểm cho học viên
        from homepage.models import AdmissionForm
        updated_count = 0
        not_found_count = 0
        invalid_conduct_count = 0

        valid_conduct_values = ['Tốt', 'Khá', 'Trung bình', 'Yếu']

        for data in update_data:
            cccd = data['cccd']
            conduct = data['conduct']

            # Kiểm tra hạnh kiểm có hợp lệ không
            if conduct not in valid_conduct_values:
                invalid_conduct_count += 1
                continue

            # Tìm học viên theo CCCD
            try:
                admission = AdmissionForm.objects.get(id_number=cccd)
                admission.conduct = conduct
                admission.save()
                updated_count += 1
            except AdmissionForm.DoesNotExist:
                not_found_count += 1

        # Tạo thông báo kết quả
        result_msg = f'Đã cập nhật hạnh kiểm cho {updated_count} học viên.'
        if not_found_count > 0:
            result_msg += f' Không tìm thấy {not_found_count} học viên.'
        if invalid_conduct_count > 0:
            result_msg += f' {invalid_conduct_count} giá trị hạnh kiểm không hợp lệ.'

        # Lưu kết quả vào session để hiển thị
        request.session['conduct_update_result'] = result_msg
        return redirect('adminpage:admission')

    return redirect('adminpage:admission')

def export_conduct_template(request):
    """Export template Excel cho việc cập nhật hạnh kiểm"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Cập nhật Hạnh kiểm"

    # Thêm tiêu đề
    ws.append(["Số CCCD", "Hạnh kiểm"])
    ws.append(["123456789012", "Tốt"])
    ws.append(["987654321098", "Khá"])
    ws.append(["111222333444", "Trung bình"])

    # Định dạng tiêu đề
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=template_cap_nhat_hanh_kiem.xlsx'
    wb.save(response)
    return response
