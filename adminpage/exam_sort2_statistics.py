"""
Xuất thống kê phòng thi SXPT II theo mẫu Excel (mỗi điểm thi một sheet).
"""
from __future__ import annotations

import re
from collections import OrderedDict
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from adminpage.exam_sort2_logic import (
    _roster_pair_context,
    elective_pair_label,
    get_room_assignments_ordered,
    roster_pair_group_key,
)
from homepage.models import ExamSort2Room, ExamSort2SeatAssignment, ExamSort2Venue

STATS_TITLE = 'THỐNG KÊ SỐ LƯỢNG PHÒNG THI THỬ KỲ THI TN THPT NĂM 2026'
WED_SCHEDULE = " - SÁNG: Ngữ văn (120')\n - CHIỀU: Toán (90')"
THU_MORNING = '- SÁNG:'
EXAM_DATE_WED = '(27/5/2026)'
EXAM_DATE_THU = '(28/5/2026)'

# Tên hiển thị môn thi Thứ Năm (theo file mẫu)
ELECTIVE_STATS_DISPLAY = {
    'Lý': 'Lý', 'Hóa': 'Hóa', 'Sinh': 'Sinh', 'Sử': 'Lịch sử', 'Địa': 'Địa lý',
    'KTPL': 'KTPL', 'Tin': 'Tin', 'T.Anh': 'Tiếng Anh', 'T.Nhật': 'Tiếng Nhật',
}
ELECTIVE_EXAM_MINUTES = {
    'Lý': 50, 'Hóa': 50, 'Sinh': 50, 'Sử': 50, 'Địa': 50,
    'KTPL': 50, 'Tin': 50, 'T.Anh': 50, 'T.Nhật': 50,
}

_THIN = Side(style='thin')
_DASHED = Side(style='dashed')
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_COLOR_ROOM = 'FFC00000'
_COLOR_SESSION = 'FF000066'
_FONT_TITLE = Font(bold=True, size=14)
_FONT_SUBTITLE = Font(bold=True, size=12)
_FONT_HEADER = Font(bold=True)
_FONT_ROOM = Font(bold=True, color=_COLOR_ROOM)
_FONT_COMBO = Font(bold=True, color=_COLOR_ROOM)
_FONT_SESSION = Font(color=_COLOR_SESSION)
_FONT_WED = Font(color=_COLOR_SESSION)
_ALIGN_CENTER = Alignment(vertical='center', horizontal='center')
_ALIGN_WRAP = Alignment(wrap_text=True, vertical='center', horizontal='center')
_ALIGN_LEFT_WRAP = Alignment(wrap_text=True, vertical='center', horizontal='left')
_ALIGN_TOP = Alignment(wrap_text=True, vertical='top', horizontal='left')
_ROW_H_SINGLE = 40
_ROW_H_COMBO = 28


def exam_sort2_statistics_template_path() -> Path:
    return (
        Path(settings.BASE_DIR)
        / 'Sắp xếp phòng thi II'
        / 'THỐNG KÊ PHÒNG THI THỬ TN THPT NĂM 2026.xlsx'
    )


def _sanitize_sheet_title(name: str) -> str:
    s = re.sub(r'[\[\]:*?/\\]', '_', (name or 'Sheet').strip())[:31]
    return s or 'Sheet'


def _pick_prototype_sheet(wb, venue_code: str):
    code = (venue_code or '').strip().casefold()
    for title in wb.sheetnames:
        if title.casefold() == code:
            return wb[title]
    for title in wb.sheetnames:
        if code and code in title.casefold():
            return wb[title]
    return wb[wb.sheetnames[0]]


def _combo_short_label(pair_key: tuple[str, str]) -> str:
    """Nhãn tổ hợp theo thứ tự hiển thị trong phòng (không sort alphabet môn)."""
    parts: list[str] = []
    for subj in pair_key:
        if subj:
            parts.append(ELECTIVE_STATS_DISPLAY.get(subj, subj))
    return ' + '.join(parts) if parts else '(chưa có môn TC)'


def _format_elective_cell_oriented(pair_key: tuple[str, str]) -> str | None:
    """Chi tiết Thứ Năm theo đúng thứ tự cặp (môn 1 / môn 2) như danh sách phòng."""
    parts: list[str] = []
    for subj in pair_key:
        if not subj:
            continue
        disp = ELECTIVE_STATS_DISPLAY.get(subj, subj)
        mins = ELECTIVE_EXAM_MINUTES.get(subj, 50)
        parts.append(f"- {disp} ({mins}')")
    if not parts:
        return None
    return ' \n'.join(parts)


def _room_combo_groups(room: ExamSort2Room) -> list[dict[str, Any]]:
    """Nhóm thí sinh theo tổ hợp — khớp thứ tự danh sách trong phòng thi."""
    from collections import Counter

    assignments = get_room_assignments_ordered(room)
    if not assignments:
        return [{
            'key': ('', ''),
            'label': None,
            'detail': None,
            'count': 0,
        }]

    candidates = [a.venue_candidate.candidate for a in assignments]
    subject_slots, subject_weights = _roster_pair_context(candidates)
    counter: Counter[tuple[str, str]] = Counter()
    for a in assignments:
        gk = roster_pair_group_key(
            a.venue_candidate.candidate,
            subject_slots=subject_slots,
            subject_weights=subject_weights,
        )
        counter[gk] += 1

    groups = []
    for key, cnt in sorted(
        counter.items(),
        key=lambda item: (
            -item[1],
            tuple(s.casefold() for s in item[0]),
        ),
    ):
        groups.append({
            'key': key,
            'label': _combo_short_label(key),
            'detail': _format_elective_cell_oriented(key),
            'count': cnt,
        })
    return groups


def _venue_blocks(venue: ExamSort2Venue) -> list[dict[str, Any]]:
    rooms = list(
        ExamSort2Room.objects.filter(venue=venue)
        .select_related('target_campus')
        .order_by('target_campus__code', 'sort_order', 'name')
    )
    grouped: OrderedDict[int, dict] = OrderedDict()
    for room in rooms:
        cid = room.target_campus_id
        if cid not in grouped:
            grouped[cid] = {
                'campus': room.target_campus,
                'rooms': [],
            }
        grouped[cid]['rooms'].append(room)
    return list(grouped.values())


def _campus_block_label(campus_code: str) -> str:
    return f'KHỐI 12{campus_code}'


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _copy_sheet_header(proto_ws, ws) -> None:
    """Sao chép hàng tiêu đề (1–4), độ rộng cột và merge tiêu đề (không merge hàng 4–5 bảng)."""
    for col_letter, dim in proto_ws.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[col_letter].width = dim.width
    for row in range(1, 5):
        ws.row_dimensions[row].height = proto_ws.row_dimensions[row].height
        for col in range(1, 8):
            src = proto_ws.cell(row, col)
            dst = ws.cell(row, col, value=src.value)
            _copy_cell_style(src, dst)
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    ws.merge_cells('C4:E4')


def _cell_border(bottom_style: str = 'thin') -> Border:
    bottom = _DASHED if bottom_style == 'dashed' else _THIN
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=bottom)


def _style_room_cell(ws, row: int, col: int, *, bottom: str = 'thin') -> None:
    c = ws.cell(row, col)
    c.border = _cell_border(bottom)
    if col == 1:
        c.alignment = _ALIGN_WRAP
    elif col == 2:
        c.font = _FONT_ROOM
        c.alignment = _ALIGN_WRAP
    elif col == 3:
        c.font = _FONT_WED
        c.alignment = _ALIGN_LEFT_WRAP
    elif col == 4:
        c.font = _FONT_SESSION
        c.alignment = _ALIGN_TOP
    elif col == 5:
        c.font = _FONT_COMBO
        c.alignment = _ALIGN_WRAP
    elif col == 6:
        c.alignment = _ALIGN_WRAP
    else:
        c.alignment = _ALIGN_LEFT_WRAP


def _merge_vertical(ws, col: int, start_row: int, end_row: int) -> None:
    if end_row > start_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=col,
            end_row=end_row,
            end_column=col,
        )


def _write_room_block(
    ws,
    start_row: int,
    stt: int,
    room_name: str,
    combo_groups: list[dict[str, Any]],
) -> int:
    """
    Ghi một phòng thi.
    - 1 tổ hợp: 2 hàng (dòng trống spacer như mẫu).
    - Nhiều tổ hợp: mỗi tổ hợp 1 hàng, STT/tên/lịch Tư gộp dọc, cột E = KTPL + Hóa, F = số HV.
    """
    multi = len(combo_groups) > 1

    if multi:
        end_row = start_row + len(combo_groups) - 1
        ws.cell(start_row, 1, stt)
        ws.cell(start_row, 2, room_name)
        ws.cell(start_row, 3, WED_SCHEDULE)
        ws.cell(start_row, 4, THU_MORNING)
        _merge_vertical(ws, 1, start_row, end_row)
        _merge_vertical(ws, 2, start_row, end_row)
        _merge_vertical(ws, 3, start_row, end_row)
        _merge_vertical(ws, 4, start_row, end_row)

        for i, grp in enumerate(combo_groups):
            r = start_row + i
            is_last = i == len(combo_groups) - 1
            bottom = 'thin' if is_last else 'dashed'
            ws.cell(r, 5, grp['label'])
            ws.cell(r, 6, grp['count'] if grp['count'] else None)
            ws.row_dimensions[r].height = _ROW_H_COMBO
            for col in range(1, 8):
                if col in (1, 2, 3, 4) and r != start_row:
                    _style_room_cell(ws, r, col, bottom=bottom)
                else:
                    _style_room_cell(ws, r, col, bottom=bottom)
        return end_row + 1

    grp = combo_groups[0]
    r = start_row
    ws.cell(r, 1, stt)
    ws.cell(r, 2, room_name)
    ws.cell(r, 3, WED_SCHEDULE)
    elective_detail = grp.get('detail')
    ws.cell(r, 4, THU_MORNING if elective_detail else None)
    ws.cell(r, 5, elective_detail)
    ws.cell(r, 6, grp['count'] if grp['count'] else None)

    for col in range(1, 8):
        _style_room_cell(ws, r, col)

    for col in (1, 2, 3, 4, 5, 6, 7):
        ws.merge_cells(start_row=r, start_column=col, end_row=r + 1, end_column=col)
    ws.row_dimensions[r].height = _ROW_H_SINGLE
    ws.row_dimensions[r + 1].height = _ROW_H_SINGLE
    _style_room_cell(ws, r + 1, 1)
    _style_room_cell(ws, r + 1, 2)
    _style_room_cell(ws, r + 1, 3)
    _style_room_cell(ws, r + 1, 4)
    _style_room_cell(ws, r + 1, 5)
    _style_room_cell(ws, r + 1, 6)
    _style_room_cell(ws, r + 1, 7)
    return r + 2


def fill_venue_statistics_sheet(ws, venue: ExamSort2Venue, proto_ws=None) -> None:
    """Điền dữ liệu thống kê cho một sheet (đã có header 1–4)."""
    if proto_ws is not None:
        _copy_sheet_header(proto_ws, ws)
    else:
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        ws.merge_cells('C4:E4')
        for col, label in enumerate(
            ['STT', 'Tên\n phòng thi', 'Ngày thi', None, None, 'Số học viên', 'Lưu ý'],
            start=1,
        ):
            if label:
                c = ws.cell(4, col, label)
                c.font = _FONT_HEADER
                c.alignment = _ALIGN_WRAP

    ws['A1'] = STATS_TITLE
    ws['A1'].font = _FONT_TITLE
    subtitle = venue.name.strip()
    if not subtitle.startswith('ĐIỂM') and not subtitle.startswith('Điểm'):
        subtitle = f' {subtitle}' if subtitle else f' {venue.code}'
    else:
        subtitle = f' {subtitle}' if not subtitle.startswith(' ') else subtitle
    ws['A2'] = subtitle
    ws['A2'].font = _FONT_SUBTITLE

    # Xóa nội dung cũ từ hàng 5 (giữ header)
    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)

    blocks = _venue_blocks(venue)
    row = 5
    grand_total_row = row if len(blocks) > 1 else None

    if grand_total_row:
        codes = ' + '.join(f'12{b["campus"].code}' for b in blocks)
        ws.cell(row, 2, f'TỔNG CỘNG {codes}:')
        ws.cell(row, 2).font = _FONT_HEADER
        row += 1

    stt = 0
    block_sum_rows: list[int] = []

    for block in blocks:
        campus = block['campus']
        rooms = block['rooms']
        block_header_row = row

        ws.cell(row, 1, _campus_block_label(campus.code))
        ws.cell(row, 1).font = _FONT_HEADER
        ws.cell(row, 3, 'Thứ TƯ')
        ws.cell(row, 4, 'Thứ NĂM')
        row += 1
        ws.cell(row, 3, EXAM_DATE_WED)
        ws.cell(row, 4, EXAM_DATE_THU)
        row += 1

        first_data_row = row
        for room in rooms:
            stt += 1
            combo_groups = _room_combo_groups(room)
            row = _write_room_block(ws, row, stt, room.name, combo_groups)

        last_data_row = row - 1
        if first_data_row <= last_data_row:
            ws.cell(block_header_row, 6).value = (
                f'=SUM(F{first_data_row}:F{last_data_row})'
            )
            block_sum_rows.append(block_header_row)

    if grand_total_row and block_sum_rows:
        ws.cell(grand_total_row, 6).value = (
            '=' + '+'.join(f'F{r}' for r in block_sum_rows)
        )

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1


def build_venue_statistics_workbook(venue: ExamSort2Venue) -> BytesIO:
    """Một điểm thi → workbook một sheet."""
    return build_all_statistics_workbook([venue])


def build_all_statistics_workbook(venues: list[ExamSort2Venue]) -> BytesIO:
    template_path = exam_sort2_statistics_template_path()
    if not template_path.is_file():
        raise FileNotFoundError(f'Không tìm thấy file mẫu thống kê: {template_path}')

    tpl_wb = load_workbook(template_path)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    if not venues:
        ws = out_wb.create_sheet('Trống')
        ws['A1'] = 'Chưa có điểm thi.'
        buf = BytesIO()
        out_wb.save(buf)
        buf.seek(0)
        return buf

    for venue in venues:
        proto = _pick_prototype_sheet(tpl_wb, venue.code)
        title = _sanitize_sheet_title(venue.code)
        base = 1
        while title in out_wb.sheetnames:
            title = _sanitize_sheet_title(f'{venue.code}_{base}')
            base += 1
        ws = out_wb.create_sheet(title)
        fill_venue_statistics_sheet(ws, venue, proto_ws=proto)

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf


# --- Thống kê số lượng môn tự chọn theo vị trí môn 1 / môn 2 (mỗi môn một dòng / phòng) ---

ELECTIVE_COUNT_DATE = '28/5/26'
ELECTIVE_COUNT_SESSION = 'chiều'
ELECTIVE_SLOT_COUNT_TITLE = (
    'THỐNG KÊ SỐ LƯỢNG MÔN TỰ CHỌN — THỬ KỲ THI TN THPT NĂM 2026'
)

ELECTIVE_SLOT_SHORT = {
    'Lý': 'LÝ',
    'Hóa': 'HÓA',
    'Sinh': 'SINH',
    'Sử': 'SỬ',
    'Địa': 'ĐỊA',
    'KTPL': 'KTPL',
    'Tin': 'TIN',
    'T.Anh': 'T.Anh',
    'T.Nhật': 'T.Nhật',
}

_FONT_SUBJECT_COUNT = Font(bold=True, color=_COLOR_ROOM)
_ALIGN_CENTER_LEFT = Alignment(vertical='center', horizontal='left')


def _slot_subject_display(subj: str) -> str:
    s = (subj or '').strip()
    if not s:
        return '—'
    return ELECTIVE_SLOT_SHORT.get(s, s.upper())


def _format_slot_count_cell(subj: str, count: int) -> str:
    return f'{_slot_subject_display(subj)}({count})'


def _room_elective_slot_subject_counts(
    room: ExamSort2Room,
    slot: int,
) -> list[tuple[str, int]]:
    """
    Đếm số thí sinh theo môn ở vị trí môn 1 hoặc môn 2 trong phòng
    (thứ tự cặp khớp danh sách / sơ đồ phòng).
    """
    from collections import Counter

    assignments = list(
        ExamSort2SeatAssignment.objects.filter(room=room)
        .select_related('venue_candidate__candidate'),
    )
    if not assignments:
        return []

    candidates = [a.venue_candidate.candidate for a in assignments]
    subject_slots, subject_weights = _roster_pair_context(candidates)
    idx = 0 if slot == 1 else 1
    counter: Counter[str] = Counter()
    for a in assignments:
        cand = a.venue_candidate.candidate
        key = roster_pair_group_key(
            cand,
            subject_slots=subject_slots,
            subject_weights=subject_weights,
        )
        subj = (key[idx] or '').strip()
        if subj:
            counter[subj] += 1

    return sorted(counter.items(), key=lambda x: (-x[1], x[0].casefold()))


def _collect_elective_slot_count_rows(
    venues: list[ExamSort2Venue],
    slot: int,
) -> list[dict[str, Any]]:
    """Mỗi (phòng, môn ở vị trí slot) → một dòng."""
    multi_venue = len(venues) > 1
    rows: list[dict[str, Any]] = []
    for venue in sorted(venues, key=lambda v: (v.sort_order, (v.code or '').casefold())):
        room_ids = (
            ExamSort2SeatAssignment.objects.filter(room__venue=venue)
            .values_list('room_id', flat=True)
            .distinct()
        )
        rooms = list(
            ExamSort2Room.objects.filter(pk__in=room_ids)
            .select_related('target_campus')
            .order_by('target_campus__code', 'sort_order', 'name'),
        )
        for room in rooms:
            counts = _room_elective_slot_subject_counts(room, slot)
            if not counts:
                continue
            for subj, cnt in counts:
                rows.append({
                    'date': ELECTIVE_COUNT_DATE,
                    'session': ELECTIVE_COUNT_SESSION,
                    'slot_label': f'Môn {slot}:',
                    'subject_cell': _format_slot_count_cell(subj, cnt),
                    'room_name': room.name,
                    'venue_code': venue.code if multi_venue else '',
                    'venue_name': venue.name if multi_venue else '',
                })
    return rows


def _write_elective_slot_count_sheet(
    ws,
    rows: list[dict[str, Any]],
    *,
    slot: int,
    subtitle: str = '',
    show_venue: bool = False,
) -> None:
    headers = ['Ngày', 'Buổi', '', 'môn', 'phòng']
    if show_venue:
        headers.insert(0, 'Điểm thi')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(1, 1, ELECTIVE_SLOT_COUNT_TITLE)
    t.font = _FONT_TITLE
    t.alignment = _ALIGN_CENTER

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        s = ws.cell(2, 1, subtitle)
        s.font = _FONT_SUBTITLE
        s.alignment = _ALIGN_CENTER

    hdr_row = 4 if subtitle else 3
    for col, label in enumerate(headers, start=1):
        c = ws.cell(hdr_row, col, label)
        c.font = _FONT_HEADER
        c.alignment = _ALIGN_WRAP if label != 'môn' else _ALIGN_CENTER

    data_start = hdr_row + 1
    for i, row in enumerate(rows):
        r = data_start + i
        col = 1
        if show_venue:
            ws.cell(r, col, row.get('venue_code') or '')
            col += 1
        ws.cell(r, col, row['date'])
        ws.cell(r, col + 1, row['session'])
        ws.cell(r, col + 2, row['slot_label'])
        subj_cell = ws.cell(r, col + 3, row['subject_cell'])
        subj_cell.font = _FONT_SUBJECT_COUNT
        room_cell = ws.cell(r, col + 4, row['room_name'])
        room_cell.font = _FONT_ROOM
        room_cell.alignment = _ALIGN_CENTER_LEFT
        ws.row_dimensions[r].height = 22

    col_widths = [10, 12, 10, 8, 18, 12] if show_venue else [12, 10, 10, 8, 18, 12]
    for i, w in enumerate(col_widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1


def build_elective_slot_counts_workbook(
    venues: list[ExamSort2Venue],
) -> BytesIO:
    """
    Workbook 2 sheet: Môn 1 và Môn 2.
    Mỗi phòng — mỗi môn xuất hiện ở vị trí đó → một dòng (vd. 1A: LÝ(13), HÓA(8)).
    """
    venues = list(venues)
    show_venue = len(venues) > 1
    if len(venues) == 1:
        v = venues[0]
        subtitle = f'Điểm thi: {v.code} — {v.name}'
    elif venues:
        codes = ', '.join(v.code for v in venues[:6])
        if len(venues) > 6:
            codes += ', …'
        subtitle = f'Tất cả điểm thi ({len(venues)}): {codes}'
    else:
        subtitle = ''

    rows1 = _collect_elective_slot_count_rows(venues, 1)
    rows2 = _collect_elective_slot_count_rows(venues, 2)

    out_wb = Workbook()
    ws1 = out_wb.active
    ws1.title = _sanitize_sheet_title('Môn 1')
    _write_elective_slot_count_sheet(
        ws1, rows1, slot=1, subtitle=subtitle, show_venue=show_venue,
    )

    ws2 = out_wb.create_sheet(_sanitize_sheet_title('Môn 2'))
    _write_elective_slot_count_sheet(
        ws2, rows2, slot=2, subtitle=subtitle, show_venue=show_venue,
    )

    if not rows1 and not rows2:
        ws1.cell(5, 1, 'Chưa có phòng nào có thí sinh được xếp ghế.')

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf


def build_venue_elective_slot_counts_workbook(venue: ExamSort2Venue) -> BytesIO:
    return build_elective_slot_counts_workbook([venue])
