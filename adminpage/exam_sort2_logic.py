"""
Logic nghiệp vụ Sắp xếp phòng thi II: import thí sinh, xếp ghế theo cột×hàng.
"""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from django.conf import settings
from django.db import transaction

from homepage.models import (
    EXAM_SORT2_CAMPUS_CODE_BY_INT,
    Campus,
    ExamSort2Candidate,
    ExamSort2Room,
    ExamSort2SeatAssignment,
    ExamSort2Venue,
    ExamSort2VenueCandidate,
)


def exam_sort2_campus_from_import_value(raw) -> Campus | None:
    """Chuyển mã cơ sở (số nguyên 1–9 hoặc mã chữ AS, ĐS, …) sang bản ghi Campus."""
    if raw is None or str(raw).strip() == '':
        return None
    s = str(raw).strip()
    code = None
    try:
        n = int(float(s))
        code = EXAM_SORT2_CAMPUS_CODE_BY_INT.get(n)
    except (TypeError, ValueError):
        code = s.upper()
        if code == 'DS':
            code = 'ĐS'
    if not code:
        return None
    campus = Campus.objects.filter(code__iexact=code).first()
    if campus:
        return campus
    return Campus.objects.create(code=code, name=code)


def _normalize_header(cell: str) -> str:
    return (cell or '').strip().lower()


def _combine_full_name(parts: list[str]) -> str:
    return ' '.join(p.strip() for p in parts if p and str(p).strip())


# Cột môn tự chọn trong file Excel (ngoài VĂN, TOÁN bắt buộc)
EXAM_SORT2_ELECTIVE_HEADER_KEYS = (
    'lý', 'ly', 'hóa', 'hoa', 'sinh', 'sử', 'su', 'địa', 'dia',
    'ktpl', 'tin', 't.anh', 't anh', 'tiếng anh', 'tieng anh',
    't.nhật', 't.nhat', 't nhật', 't nhat', 'tiếng nhật', 'tieng nhat',
    'nhật', 'nhat',
)

EXAM_SORT2_SUBJECT_DISPLAY = {
    'VĂN': 'Văn', 'VAN': 'Văn',
    'TOÁN': 'Toán', 'TOAN': 'Toán',
    'LÝ': 'Lý', 'LY': 'Lý',
    'HÓA': 'Hóa', 'HOA': 'Hóa',
    'SINH': 'Sinh',
    'SỬ': 'Sử', 'SU': 'Sử',
    'ĐỊA': 'Địa', 'DIA': 'Địa',
    'KTPL': 'KTPL',
    'TIN': 'Tin',
    'T.ANH': 'T.Anh', 'T ANH': 'T.Anh', 'TIẾNG ANH': 'T.Anh',
    'T.NHẬT': 'T.Nhật', 'T.NHAT': 'T.Nhật', 'T NHẬT': 'T.Nhật',
    'TIẾNG NHẬT': 'T.Nhật', 'TIENG NHAT': 'T.Nhật', 'NHẬT': 'T.Nhật', 'NHAT': 'T.Nhật',
}


def _normalize_subject_key(text: str) -> str:
    import unicodedata
    s = unicodedata.normalize('NFD', str(text or '').strip().upper())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').replace('.', '').replace(' ', '')


def _elective_normalized_map() -> dict[str, str]:
    skip = {_normalize_subject_key('VĂN'), _normalize_subject_key('TOÁN')}
    return {
        _normalize_subject_key(k): v
        for k, v in EXAM_SORT2_SUBJECT_DISPLAY.items()
        if _normalize_subject_key(k) not in skip
    }


def exam_sort2_subject_display_name(raw_header: str) -> str:
    key = _normalize_subject_key(raw_header)
    for k, label in EXAM_SORT2_SUBJECT_DISPLAY.items():
        if _normalize_subject_key(k) == key:
            return label
    norm = _elective_normalized_map()
    if key in norm:
        return norm[key]
    return str(raw_header or '').strip().title() or raw_header


def _elective_from_cell(cell_val, header_label: str) -> str | None:
    """
    Nhận môn tự chọn từ ô Excel.
    Chấp nhận: ô trùng tiêu đề cột (Lý trong cột LÝ) HOẶC ô ghi tên môn khác
    (VD: T.Nhật trong cột T.ANH khi thí sinh chọn Tiếng Nhật thay Tiếng Anh).
    """
    if cell_val is None:
        return None
    if isinstance(cell_val, float):
        if cell_val == 0.0:
            return None
        s = str(int(cell_val)) if cell_val == int(cell_val) else str(cell_val)
    else:
        s = str(cell_val).strip()
    if not s or s in ('-', '—'):
        return None
    try:
        if float(s) == 0:
            return None
    except ValueError:
        pass

    cell_key = _normalize_subject_key(s)
    norm = _elective_normalized_map()
    if cell_key in norm:
        return norm[cell_key]
    if cell_key == _normalize_subject_key(header_label):
        return exam_sort2_subject_display_name(header_label)
    return None


def _extract_elective_subjects(row: list, elective_cols: list[tuple[int, str]]) -> tuple[str, str, list[str]]:
    """
    elective_cols: [(col_index, header_name), ...]
    Trả về (môn1, môn2, cảnh báo phụ).
    """
    picked: list[str] = []
    for col_idx, header in elective_cols:
        if col_idx >= len(row):
            continue
        subj = _elective_from_cell(row[col_idx], header)
        if subj and subj not in picked:
            picked.append(subj)
    warnings: list[str] = []
    if len(picked) < 2:
        warnings.append(f'chỉ có {len(picked)} môn tự chọn ({", ".join(picked) or "không có"})')
    elif len(picked) > 2:
        warnings.append(f'có {len(picked)} môn tự chọn: {", ".join(picked)} — lấy 2 môn đầu')
        picked = picked[:2]
    s1 = picked[0] if len(picked) > 0 else ''
    s2 = picked[1] if len(picked) > 1 else ''
    return s1, s2, warnings


def parse_exam_sort2_import_rows(file_obj, filename: str) -> tuple[list[dict], list[str]]:
    """
    Đọc Excel (.xls/.xlsx) hoặc CSV.
    Cột đầu: mã cơ sở (số); cột lớp; họ tên (có thể tách 2 cột như file test.xls).
    """
    errors: list[str] = []
    rows_out: list[dict] = []
    ext = (filename or '').rsplit('.', 1)[-1].lower()

    if ext == 'csv':
        text = file_obj.read()
        if isinstance(text, bytes):
            text = text.decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(text))
        matrix = list(reader)
    elif ext == 'xls':
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_obj.read())
        sh = wb.sheet_by_index(0)
        matrix = []
        for r in range(sh.nrows):
            matrix.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]

    header_row_idx = None
    col_coso = col_lop = col_ho = col_ten = None
    elective_cols: list[tuple[int, str]] = []
    header_cells_raw: list[str] = []

    for i, row in enumerate(matrix):
        cells = [_normalize_header(str(c) if c is not None else '') for c in row]
        joined = ' '.join(cells)
        if 'cơ sở' in joined or 'co so' in joined:
            header_row_idx = i
            header_cells_raw = [str(c).strip() if c is not None else '' for c in row]
            for j, h in enumerate(cells):
                raw_h = header_cells_raw[j] if j < len(header_cells_raw) else ''
                if 'cơ sở' in h or h == 'co so':
                    col_coso = j
                elif h == 'lớp' or h == 'lop' or 'class' in h:
                    col_lop = j
                elif 'họ và' in h or 'họ tên' in h or 'ho va' in h:
                    col_ho = j
                elif h == 'tên' or h == 'ten':
                    col_ten = j
                elif h in ('văn', 'van', 'toán', 'toan'):
                    continue
                elif h in EXAM_SORT2_ELECTIVE_HEADER_KEYS or _normalize_subject_key(raw_h) in (
                    _normalize_subject_key(k) for k in (
                        'LÝ', 'HÓA', 'SINH', 'SỬ', 'ĐỊA', 'KTPL', 'TIN', 'T.ANH',
                    )
                ):
                    elective_cols.append((j, raw_h or h.upper()))
            if col_coso is None:
                col_coso = 0
            if col_lop is None:
                col_lop = 1
            if col_ho is None:
                for j, h in enumerate(cells):
                    if 'tên' in h and 'họ' not in h:
                        col_ten = j
                    elif 'họ' in h:
                        col_ho = j
            break

    if header_row_idx is None:
        errors.append('Không tìm thấy dòng tiêu đề có cột "Cơ Sở".')
        return rows_out, errors

    data_start = header_row_idx + 1
    if data_start < len(matrix):
        next_row = matrix[data_start]
        if next_row and col_ho is not None and col_ho < len(next_row):
            sub = _normalize_header(str(next_row[col_ho]))
            if sub in ('tên', 'ten') and col_ten is None and col_ho + 1 < len(next_row):
                col_ten = col_ho + 1
                data_start += 1

    for i in range(data_start, len(matrix)):
        row = matrix[i]
        if not row or not any(row):
            continue
        line_no = i + 1
        raw_coso = row[col_coso] if col_coso is not None and col_coso < len(row) else None
        if raw_coso is None or str(raw_coso).strip() == '':
            continue
        campus = exam_sort2_campus_from_import_value(raw_coso)
        if not campus:
            errors.append(f'Dòng {line_no}: Mã cơ sở không hợp lệ ({raw_coso}).')
            continue

        class_name = ''
        if col_lop is not None and col_lop < len(row) and row[col_lop] is not None:
            class_name = str(row[col_lop]).strip()
            if class_name.endswith('.0') and class_name[:-2].isdigit():
                class_name = class_name[:-2]

        name_parts = []
        if col_ho is not None and col_ho < len(row):
            v = row[col_ho]
            if v is not None and str(v).strip():
                name_parts.append(str(v).strip())
        if col_ten is not None and col_ten < len(row):
            v = row[col_ten]
            if v is not None and str(v).strip():
                name_parts.append(str(v).strip())
        if not name_parts:
            for j in range(2, min(len(row), 6)):
                if j in (col_coso, col_lop):
                    continue
                v = row[j]
                if v is not None and str(v).strip() and not str(v).replace('.', '').isdigit():
                    name_parts.append(str(v).strip())
                    if len(name_parts) >= 2:
                        break

        full_name = _combine_full_name(name_parts)
        if not full_name:
            errors.append(f'Dòng {line_no}: Thiếu họ tên.')
            continue
        if not class_name:
            errors.append(f'Dòng {line_no}: Thiếu lớp.')
            continue

        elective_1, elective_2, elective_warns = _extract_elective_subjects(row, elective_cols)
        for w in elective_warns:
            errors.append(f'Dòng {line_no}: {w}.')

        rows_out.append({
            'row_number': line_no,
            'full_name': full_name,
            'class_name': class_name,
            'campus_id': campus.id,
            'campus_code': campus.code,
            'campus_name': campus.name,
            'elective_subject_1': elective_1,
            'elective_subject_2': elective_2,
            'exam_subjects_display': f'Văn, Toán' + (
                f', {elective_1}, {elective_2}' if elective_1 and elective_2
                else (f', {elective_1}' if elective_1 else '')
            ),
        })

    return rows_out, errors


def parse_exam_sort2_import_rooms(file_obj, filename: str) -> tuple[list[dict], list[str]]:
    """
    Import phòng thi từ Excel/CSV (mẫu phongthi.xlsx):
    Tên Phòng | Số cột | Số hàng | Đối tượng | STT (mã cơ sở: AS, AT, … hoặc số 1–9).
    """
    errors: list[str] = []
    rows_out: list[dict] = []
    ext = (filename or '').rsplit('.', 1)[-1].lower()

    if ext == 'csv':
        text = file_obj.read()
        if isinstance(text, bytes):
            text = text.decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(text))
        matrix = list(reader)
    elif ext == 'xls':
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_obj.read())
        sh = wb.sheet_by_index(0)
        matrix = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]

    header_row_idx = None
    col_name = col_col = col_row = col_target = col_stt = None

    for i, row in enumerate(matrix):
        cells = [_normalize_header(str(c) if c is not None else '') for c in row]
        joined = ' '.join(cells)
        if 'tên phòng' in joined or 'ten phong' in joined or (
            'phòng' in joined and 'tên' in joined
        ):
            header_row_idx = i
            for j, h in enumerate(cells):
                if 'tên' in h and 'phòng' in h or h == 'ten phong':
                    col_name = j
                elif 'cột' in h or h == 'cot' or 'col' in h:
                    col_col = j
                elif 'hàng' in h or h == 'hang' or 'row' in h:
                    col_row = j
                elif 'đối tượng' in h or 'doi tuong' in h or 'cơ sở' in h or 'co so' in h:
                    col_target = j
                elif h == 'stt' or h.startswith('stt '):
                    col_stt = j
            if col_name is None:
                col_name = 0
            if col_col is None:
                col_col = 1
            if col_row is None:
                col_row = 2
            if col_target is None:
                col_target = 3
            if col_stt is None and max(col_name, col_col, col_row, col_target) < 4:
                col_stt = 4
            break

    if header_row_idx is None:
        errors.append(
            'Không tìm thấy dòng tiêu đề (Tên Phòng, Số cột, Số hàng, Đối tượng, STT).',
        )
        return rows_out, errors

    seen_names: set[str] = set()
    for i in range(header_row_idx + 1, len(matrix)):
        row = matrix[i]
        if not row or not any(row):
            continue
        line_no = i + 1

        def cell(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name_raw = cell(col_name)
        name = str(name_raw).strip() if name_raw is not None else ''
        if name.endswith('.0') and name[:-2].replace(' ', '').isalnum():
            try:
                name = str(int(float(name)))
            except ValueError:
                pass
        if not name:
            errors.append(f'Dòng {line_no}: Thiếu tên phòng.')
            continue

        try:
            col_count = max(1, int(float(cell(col_col) or 0)))
            row_count = max(1, int(float(cell(col_row) or 0)))
        except (TypeError, ValueError):
            errors.append(f'Dòng {line_no}: Số cột/hàng không hợp lệ.')
            continue

        campus = exam_sort2_campus_from_import_value(cell(col_target))
        if not campus:
            errors.append(f'Dòng {line_no}: Đối tượng cơ sở không hợp lệ ({cell(col_target)}).')
            continue

        name_key = name.casefold()
        if name_key in seen_names:
            errors.append(f'Dòng {line_no}: Trùng tên phòng "{name}" trong file.')
            continue
        seen_names.add(name_key)

        sort_order = 0
        if col_stt is not None:
            stt_raw = cell(col_stt)
            if stt_raw is not None and str(stt_raw).strip() != '':
                try:
                    sort_order = max(0, int(float(stt_raw)))
                except (TypeError, ValueError):
                    errors.append(f'Dòng {line_no}: STT không hợp lệ ({stt_raw}).')
                    continue

        rows_out.append({
            'row_number': line_no,
            'name': name,
            'col_count': col_count,
            'row_count': row_count,
            'capacity': col_count * row_count,
            'target_campus_id': campus.id,
            'target_campus_code': campus.code,
            'target_campus_name': campus.name,
            'sort_order': sort_order,
        })

    return rows_out, errors


def seat_number_to_grid(seat_number: int, col_count: int) -> tuple[int, int]:
    """Ghế 1..n → (hàng 1-based, cột 1-based), điền theo hàng: hàng 1 cột 1..col, hàng 2..."""
    if seat_number < 1 or col_count < 1:
        return 0, 0
    idx = seat_number - 1
    row = idx // col_count + 1
    col = idx % col_count + 1
    return row, col


def grid_to_seat_number(row: int, col: int, col_count: int) -> int:
    return (row - 1) * col_count + col


def _venue_assigned_vc_ids(venue: ExamSort2Venue) -> set[int]:
    return set(
        ExamSort2SeatAssignment.objects.filter(room__venue=venue)
        .values_list('venue_candidate_id', flat=True)
    )


def elective_pair_key(candidate: ExamSort2Candidate) -> tuple[str, str]:
    """Khóa tổ hợp 2 môn tự chọn (thứ tự không quan trọng)."""
    e1 = (candidate.elective_subject_1 or '').strip()
    e2 = (candidate.elective_subject_2 or '').strip()
    return tuple(sorted([e1, e2], key=lambda s: s.casefold()))


def elective_pair_label(key: tuple[str, str]) -> str:
    parts = [p for p in key if p]
    return ' + '.join(parts) if parts else '(chưa có môn TC)'


def _room_roster_anchor_subject(candidates: list[ExamSort2Candidate]) -> str:
    """
    Môn neo để gom cặp trong phòng — môn xuất hiện nhiều nhất (vd. Sử trong Sử+Địa và Lý+Sử).
    """
    from collections import Counter

    counts: Counter[str] = Counter()
    for cand in candidates:
        e1 = (cand.elective_subject_1 or '').strip()
        e2 = (cand.elective_subject_2 or '').strip()
        if e1:
            counts[e1] += 1
        if e2:
            counts[e2] += 1
    if not counts:
        return ''
    return max(counts.keys(), key=lambda s: (counts[s], s.casefold()))


def _subject_slot_preferences(
    candidates: list[ExamSort2Candidate],
) -> dict[str, int]:
    """
    Mỗi môn: vị trí ưu tiên trong nhãn cặp (0 = môn 1, 1 = môn 2) theo đa số trong phòng.
    vd. Lý→môn1, Hóa→môn2 (Lý+Hóa), Sử/Sinh→môn2.
    """
    from collections import Counter

    as_first: Counter[str] = Counter()
    as_second: Counter[str] = Counter()
    for cand in candidates:
        e1 = (cand.elective_subject_1 or '').strip()
        e2 = (cand.elective_subject_2 or '').strip()
        if e1:
            as_first[e1] += 1
        if e2:
            as_second[e2] += 1
    prefs: dict[str, int] = {}
    for subj in set(as_first) | set(as_second):
        prefs[subj] = 1 if as_second[subj] >= as_first[subj] else 0
    return prefs


def _subject_presence_weights(
    candidates: list[ExamSort2Candidate],
) -> dict[str, int]:
    """Số lần môn xuất hiện trong phòng (để xử lý hai môn cùng muốn đứng môn 2)."""
    from collections import Counter

    weights: Counter[str] = Counter()
    for cand in candidates:
        e1 = (cand.elective_subject_1 or '').strip()
        e2 = (cand.elective_subject_2 or '').strip()
        if e1:
            weights[e1] += 1
        if e2:
            weights[e2] += 1
    return dict(weights)


def roster_pair_group_key(
    candidate: ExamSort2Candidate,
    anchor_subject: str = '',
    *,
    anchor_slot: int | None = None,
    subject_slots: dict[str, int] | None = None,
    subject_weights: dict[str, int] | None = None,
) -> tuple[str, str]:
    """
    Khóa cặp môn trong phòng: mỗi môn giữ vị trí theo thói quen của phòng
    (vd. Lý+Sử, Lý+Hóa, Sinh+Hóa — Hóa cùng ở môn 2 dù nhóm neo là Lý).
    """
    e1 = (candidate.elective_subject_1 or '').strip()
    e2 = (candidate.elective_subject_2 or '').strip()
    if not e2:
        return (e1, e2)
    slots = subject_slots
    if slots is None and anchor_subject:
        anchor = anchor_subject.strip()
        slot = 0 if anchor_slot is None else anchor_slot
        if anchor == e1:
            return (e1, e2) if slot == 0 else (e2, e1)
        if anchor == e2:
            return (e2, e1) if slot == 0 else (e1, e2)
        return (e1, e2)
    if not slots:
        return (e1, e2)

    s1, s2 = e1, e2
    p1, p2 = slots.get(s1, 0), slots.get(s2, 0)
    if p1 == 0 and p2 == 1:
        return (s1, s2)
    if p1 == 1 and p2 == 0:
        return (s2, s1)
    if p1 == 1 and p2 == 1:
        w = subject_weights or {}
        if w.get(s1, 0) >= w.get(s2, 0):
            return (s2, s1)
        return (s1, s2)
    return (e1, e2)


def roster_pair_display_label(
    candidate: ExamSort2Candidate,
    anchor_subject: str = '',
    *,
    anchor_slot: int | None = None,
    subject_slots: dict[str, int] | None = None,
    subject_weights: dict[str, int] | None = None,
) -> str:
    return elective_pair_label(
        roster_pair_group_key(
            candidate,
            anchor_subject,
            anchor_slot=anchor_slot,
            subject_slots=subject_slots,
            subject_weights=subject_weights,
        ),
    )


def _roster_pair_context(
    candidates: list[ExamSort2Candidate],
) -> tuple[dict[str, int], dict[str, int]]:
    """(vị trí ưu tiên từng môn, trọng số từng môn) cho sắp danh sách phòng."""
    return (
        _subject_slot_preferences(candidates),
        _subject_presence_weights(candidates),
    )


def _pair_subjects(key: tuple[str, str]) -> set[str]:
    return {s for s in key if s}


def _shared_subject_count(a: tuple[str, str], b: tuple[str, str]) -> int:
    return len(_pair_subjects(a) & _pair_subjects(b))


def _room_subject_overlap(
    pair_key: tuple[str, str],
    elective_keys: set[tuple[str, str]],
) -> int:
    """Số môn chung tối đa giữa pair_key và một tổ hợp đã có trong phòng (0–2)."""
    if not elective_keys:
        return 0
    return max(_shared_subject_count(pair_key, k) for k in elective_keys)


def _room_accepts_shared_combo(rs: dict, pair_key: tuple[str, str]) -> bool:
    """
    Tổ hợp có thể ghép vào phòng khi trùng môn với cụm neo hoặc tổ hợp khác đã có
    (vd. Tin+T.Anh vào phòng đã có Tin+Địa).
    """
    if pair_key in rs['elective_keys']:
        return True
    if _room_is_anchor_locked(rs) and pair_key == rs.get('anchor_pair_key'):
        return False
    anchor = rs.get('anchor_pair_key')
    if anchor and _pair_shares_subject_with_anchor(pair_key, anchor):
        return True
    return _room_subject_overlap(pair_key, rs['elective_keys']) > 0


def _vc_sort_key(vc: ExamSort2VenueCandidate) -> tuple:
    """Thứ tự họ tên (tiếng Việt) — dùng khi xếp ghế / đánh SBD."""
    return _candidate_list_order_key(vc.candidate) + (vc.pk,)


def _candidate_name_sort_key_vi(full_name: str) -> tuple:
    """
    Họ và tên theo bảng chữ cái tiếng Việt (cùng quy tắc module xếp phòng cũ):
    Tên (từ cuối) → tên đệm → Họ — có phân biệt ă/â, d/đ, dấu thanh.
    """
    from adminpage.views import _name_sort_key_vi

    return _name_sort_key_vi(full_name)


def _candidate_list_order_key(candidate: ExamSort2Candidate) -> tuple:
    """Tên → tên đệm → Họ (bảng chữ cái tiếng Việt)."""
    return (
        _candidate_name_sort_key_vi(candidate.full_name),
        candidate.pk,
    )


def sort_venue_candidates_by_vi_name(
    candidates: list[ExamSort2VenueCandidate],
) -> list[ExamSort2VenueCandidate]:
    """Danh sách thí sinh tại điểm thi — tăng dần theo bảng chữ cái TV (dùng khi xếp phòng)."""
    return sorted(candidates, key=_vc_sort_key)


def sort_venue_candidates_for_display(
    candidates: list[ExamSort2VenueCandidate],
) -> list[ExamSort2VenueCandidate]:
    """
    Hiển thị danh sách điểm thi: SBD tăng dần nếu đã đánh; chưa có SBD xếp cuối (theo tên TV).
    """
    from adminpage.views import _exam_number_sort_key

    def display_key(vc: ExamSort2VenueCandidate) -> tuple:
        sbd = (vc.candidate.exam_number or '').strip()
        if sbd:
            return (0, _exam_number_sort_key(sbd), _vc_sort_key(vc))
        return (1, _vc_sort_key(vc))

    return sorted(candidates, key=display_key)


def sort_exam_sort2_candidates_by_vi_name(
    candidates: list[ExamSort2Candidate],
) -> list[ExamSort2Candidate]:
    return sorted(candidates, key=_candidate_list_order_key)


def next_room_sort_order(venue: ExamSort2Venue) -> int:
    from django.db.models import Max

    current = (
        ExamSort2Room.objects.filter(venue=venue)
        .aggregate(m=Max('sort_order'))['m']
    )
    return int(current or 0) + 1


def _sort_assignments_by_roster(
    assignments: list[ExamSort2SeatAssignment],
    *,
    anchor_subject: str | None = None,  # noqa: ARG001 — giữ tương thích chữ ký cũ
) -> list[ExamSort2SeatAssignment]:
    """
    Danh sách trong phòng: cặp môn (mỗi môn đúng vị trí trong phòng) → tên TV.
    """
    if not assignments:
        return []
    from collections import Counter

    candidates = [a.venue_candidate.candidate for a in assignments]
    subject_slots, subject_weights = _roster_pair_context(candidates)

    def group_key(a: ExamSort2SeatAssignment) -> tuple[str, str]:
        return roster_pair_group_key(
            a.venue_candidate.candidate,
            subject_slots=subject_slots,
            subject_weights=subject_weights,
        )

    group_sizes: Counter[tuple[str, str]] = Counter()
    for a in assignments:
        group_sizes[group_key(a)] += 1
    return sorted(
        assignments,
        key=lambda a: (
            -group_sizes[group_key(a)],
            tuple(s.casefold() for s in group_key(a)),
            _candidate_list_order_key(a.venue_candidate.candidate),
        ),
    )


def reorder_room_seats_by_candidate_order(room: ExamSort2Room) -> None:
    """Gán lại số ghế 1..n: cặp môn (gom môn neo) → Tên → tên đệm → Họ (tiếng Việt)."""
    assignments = list(
        ExamSort2SeatAssignment.objects.filter(room=room)
        .select_related('venue_candidate__candidate')
    )
    if not assignments:
        return
    assignments = _sort_assignments_by_roster(assignments)
    temp_base = max(room.capacity, len(assignments)) + 1000
    for idx, a in enumerate(assignments):
        a.seat_number = temp_base + idx
    ExamSort2SeatAssignment.objects.bulk_update(assignments, ['seat_number'])
    for seat_no, a in enumerate(assignments, start=1):
        a.seat_number = seat_no
    ExamSort2SeatAssignment.objects.bulk_update(assignments, ['seat_number'])


def reorder_venue_rooms_by_roster(venue: ExamSort2Venue) -> int:
    """Sắp lại STT ghế 1..n trong mỗi phòng: cặp môn → tên TV."""
    room_ids = (
        ExamSort2SeatAssignment.objects.filter(room__venue=venue)
        .values_list('room_id', flat=True)
        .distinct()
    )
    count = 0
    for room in ExamSort2Room.objects.filter(pk__in=room_ids).order_by(
        'sort_order', 'name', 'pk',
    ):
        reorder_room_seats_by_candidate_order(room)
        count += 1
    return count


def reorder_all_seated_rooms_by_roster() -> int:
    """Bước 1 trước khi đánh SBD: mọi phòng có thí sinh — sắp xếp ghế theo họ tên TV."""
    room_ids = (
        ExamSort2SeatAssignment.objects.values_list('room_id', flat=True).distinct()
    )
    count = 0
    for room in ExamSort2Room.objects.filter(pk__in=room_ids).order_by(
        'venue__sort_order', 'venue__code', 'sort_order', 'name',
    ):
        reorder_room_seats_by_candidate_order(room)
        count += 1
    return count


def _collect_assignments_for_sbd() -> list[ExamSort2SeatAssignment]:
    """
    Thứ tự đánh SBD (sau khi đã sắp tên trong từng phòng):
    STT điểm thi → STT phòng → Tên → tên đệm → Họ (TV).
    Xong toàn bộ phòng ở điểm thi STT 1 rồi mới sang điểm thi STT 2.
    """
    ordered: list[ExamSort2SeatAssignment] = []
    for venue in ExamSort2Venue.objects.order_by('sort_order', 'code', 'pk'):
        rooms = ExamSort2Room.objects.filter(venue=venue).order_by(
            'sort_order', 'name', 'pk',
        )
        for room in rooms:
            assignments = list(
                ExamSort2SeatAssignment.objects.filter(room=room)
                .select_related(
                    'venue_candidate__candidate',
                    'venue_candidate__candidate__campus',
                ),
            )
            if assignments:
                ordered.extend(_sort_assignments_by_roster(assignments))
    return ordered


def assign_exam_numbers_sort2(
    *,
    start_serial: int | None = None,
    city_prefix: str | None = None,
) -> dict[str, Any]:
    """
    Đánh số báo danh SXPT II (thí sinh đã xếp ghế):
    1. Sắp xếp từng phòng: Tên → tên đệm → Họ (bảng chữ cái tiếng Việt).
    2. Đánh SBD tăng dần: STT điểm thi → STT phòng → thứ tự đó trong phòng.
    Định dạng: {mã TP 2 số}{STT 6 số}, VD 79000001.
    """
    prefix = (city_prefix or getattr(settings, 'EXAM_SORT2_SBD_CITY_PREFIX', '79')).strip()
    try:
        serial_start = int(
            start_serial if start_serial is not None
            else getattr(settings, 'EXAM_SORT2_SBD_START_SERIAL', 1)
        )
    except (TypeError, ValueError):
        serial_start = 1
    if serial_start < 1:
        serial_start = 1
    if len(prefix) != 2 or not prefix.isdigit():
        return {
            'assigned': 0,
            'last_serial': serial_start - 1,
            'errors': [f'Mã thành phố SBD phải đúng 2 chữ số (hiện: "{prefix}").'],
        }

    to_update: list[ExamSort2Candidate] = []
    serial = serial_start
    errors: list[str] = []
    rooms_reordered = 0

    with transaction.atomic():
        rooms_reordered = reorder_all_seated_rooms_by_roster()
        roster = _collect_assignments_for_sbd()

        if roster:
            ExamSort2Candidate.objects.filter(
                pk__in={a.venue_candidate.candidate_id for a in roster},
            ).update(exam_number='')

        for a in roster:
            cand = a.venue_candidate.candidate
            cand.exam_number = f'{prefix}{serial:06d}'
            to_update.append(cand)
            serial += 1

        if to_update:
            ExamSort2Candidate.objects.bulk_update(to_update, ['exam_number'])

    if serial > 1_000_000:
        errors.append('Vượt quá 6 chữ số STT (tối đa 999999 trong mã thành phố).')

    return {
        'assigned': len(to_update),
        'rooms_reordered': rooms_reordered,
        'first_sbd': f'{prefix}{serial_start:06d}' if to_update else '',
        'last_sbd': f'{prefix}{serial - 1:06d}' if to_update else '',
        'errors': errors,
    }


def _room_existing_assignments(room: ExamSort2Room) -> tuple[dict[int, ExamSort2SeatAssignment], set[int]]:
    existing = {
        a.seat_number: a
        for a in ExamSort2SeatAssignment.objects.filter(room=room)
        .select_related('venue_candidate__candidate')
    }
    used = set(existing.keys())
    return existing, used


ASSIGN_FLEX = 1
# Sàn cứng: mỗi phòng (đủ sức chứa) phải có ít nhất N thí sinh sau khi xếp
MIN_STUDENTS_PER_ROOM = 24
MAX_COMBOS_PER_ROOM = 2
MAX_COMBOS_PER_ROOM_UNDER_BAND = 3
# Trần cứng — không phòng nào quá 3 tổ hợm môn (kể cả bước khẩn cấp)
MAX_COMBOS_HARD_CAP = 3
# Tổ hợp chỉ 1–2 thí sinh: dùng “thêm/bớt 1” so với trung bình để ghép làm tổ hợp phụ
TINY_COMBO_MAX = 2


def _count_band_from_avg(target_avg: float) -> tuple[int, int, int]:
    """Biên số thí sinh/phòng: TB làm tròn; cho phép TB−1 … TB+1 (ưu tiên ≤ TB)."""
    mid = round(target_avg)
    return mid, mid - ASSIGN_FLEX, mid + ASSIGN_FLEX


def _room_count_bands_for_capacity(
    capacity: int,
    count_mid: int,
    band_lo: int,
    band_hi: int,
) -> tuple[int, int, int]:
    """Áp dụng sàn MIN_STUDENTS_PER_ROOM khi phòng đủ sức chứa."""
    mid = count_mid
    lo = band_lo
    hi = band_hi
    if capacity >= MIN_STUDENTS_PER_ROOM:
        lo = max(lo, MIN_STUDENTS_PER_ROOM)
        hi = max(hi, lo)
    lo = min(lo, capacity)
    hi = min(hi, capacity)
    mid = min(mid, capacity)
    return mid, lo, hi


def _room_is_anchor_locked(rs: dict) -> bool:
    return bool(rs.get('anchor_locked'))


def _pair_shares_subject_with_anchor(
    pair_key: tuple[str, str],
    anchor_key: tuple[str, str] | None,
) -> bool:
    if not anchor_key:
        return False
    return _shared_subject_count(pair_key, anchor_key) > 0


def _room_may_accept_pair(
    rs: dict,
    pair_key: tuple[str, str],
    room_states: list[dict] | None = None,
) -> bool:
    """Phòng có nhận thêm tổ hợp này không (tôn trọng neo khóa + trần band_hi)."""
    if rs['assigned_count'] >= _room_band_hi(rs):
        return False
    if _room_is_anchor_locked(rs):
        if pair_key == rs.get('anchor_pair_key'):
            return False
        if pair_key in rs['elective_keys']:
            return True
        if not _room_accepts_shared_combo(rs, pair_key):
            return False
        return len(rs['elective_keys']) < MAX_COMBOS_HARD_CAP
    if pair_key in rs['elective_keys']:
        return True
    states = room_states if room_states is not None else [rs]
    if not _can_add_combo(rs, pair_key):
        return False
    return _room_may_open_new_combo(rs, states)


def _room_band_hi(rs: dict) -> int:
    return rs['band_hi']


def _room_band_lo(rs: dict) -> int:
    return rs['band_lo']


def _room_count_mid(rs: dict) -> int:
    return rs['count_mid']


def _rooms_below_mid(room_states: list[dict]) -> bool:
    """Còn phòng dưới mức trung bình (làm tròn) và còn ghế."""
    return any(
        rs['assigned_count'] < _room_count_mid(rs) and rs['remaining'] > 0
        for rs in room_states
    )


def _room_may_exceed_avg(rs: dict, room_states: list[dict]) -> bool:
    """
    Chỉ nhận thêm quá TB (+1 tối đa) khi không còn phòng nào dưới TB cần lấp.
    """
    if rs['assigned_count'] < _room_count_mid(rs):
        return False
    if rs['assigned_count'] >= _room_band_hi(rs):
        return False
    return not _rooms_below_mid(room_states)


def _assign_ceiling(rs: dict, room_states: list[dict] | None) -> int:
    """Trần nhận thêm TS: ưu tiên ≤ TB; chỉ tới TB+1 khi được phép vượt TB."""
    states = room_states if room_states is not None else [rs]
    if _room_may_exceed_avg(rs, states):
        return _room_band_hi(rs)
    return _room_count_mid(rs)


def _cap_take_to_band(rs: dict, take: int) -> int:
    """Không cho phòng vượt quá band_hi (TB+1) sau khi nhận thêm."""
    if take <= 0:
        return 0
    return min(take, max(0, _room_band_hi(rs) - rs['assigned_count']))


def _quota_ceiling(rs: dict, room_states: list[dict] | None) -> int:
    """Trần gán: khi còn phòng dưới sàn, không vượt chỉ tiêu chia đều (ideal_target)."""
    ceiling = _assign_ceiling(rs, room_states)
    if room_states and _rooms_below_band(room_states):
        ideal = rs.get('ideal_target')
        if ideal is not None:
            return min(ceiling, ideal)
    return ceiling


def _cap_take_for_assign(
    rs: dict,
    take: int,
    room_states: list[dict] | None = None,
) -> int:
    """Ưu tiên ≤ TB; chỉ tới TB+1 khi mọi phòng đã đạt TB."""
    if take <= 0:
        return 0
    ceiling = _quota_ceiling(rs, room_states)
    return min(take, max(0, ceiling - rs['assigned_count']))


def _rooms_below_band(room_states: list[dict]) -> bool:
    return any(rs['assigned_count'] < _room_band_lo(rs) for rs in room_states)


def _rooms_need_fill(room_states: list[dict]) -> bool:
    """Còn phòng dưới TB (làm tròn) và còn ghế trống."""
    return any(
        rs['assigned_count'] < _room_count_mid(rs) and rs['remaining'] > 0
        for rs in room_states
    )


def _room_may_open_new_combo(rs: dict, room_states: list[dict]) -> bool:
    """
    Mở tổ hợp mới: được nếu phòng chưa đạt TB (count_mid).
    Phòng đã ≥ TB mà phòng khác còn dưới TB → không mở tổ hợp lẻ trên phòng đầy.
    """
    if _room_is_anchor_locked(rs):
        return False
    if rs['assigned_count'] < _room_count_mid(rs):
        return True
    if rs['assigned_count'] >= _room_count_mid(rs) and _rooms_below_mid(room_states):
        return False
    return True


def _max_combos_for_room(rs: dict) -> int:
    """
    Giới hạn tổ hợm môn/phòng (tối đa MAX_COMBOS_HARD_CAP = 3):
    - 2: phòng đã đạt ≥ trung bình −1
    - 3: phòng đang thiếu so với trung bình −1
    """
    if rs['assigned_count'] < _room_count_mid(rs):
        return min(MAX_COMBOS_PER_ROOM_UNDER_BAND, MAX_COMBOS_HARD_CAP)
    return min(MAX_COMBOS_PER_ROOM, MAX_COMBOS_HARD_CAP)


def _combo_limit(rs: dict) -> int:
    return _max_combos_for_room(rs)


def _can_add_combo(rs: dict, pair_key: tuple[str, str]) -> bool:
    if _room_is_anchor_locked(rs):
        if pair_key == rs.get('anchor_pair_key'):
            return False
        if pair_key in rs['elective_keys']:
            return True
        if not _room_accepts_shared_combo(rs, pair_key):
            return False
        return len(rs['elective_keys']) < MAX_COMBOS_HARD_CAP
    if pair_key in rs['elective_keys']:
        return True
    return len(rs['elective_keys']) < _max_combos_for_room(rs)


def _can_add_combo_emergency(rs: dict, pair_key: tuple[str, str]) -> bool:
    """Bước cuối: vẫn tối đa MAX_COMBOS_HARD_CAP tổ hợp/phòng."""
    if rs['assigned_count'] >= _room_band_hi(rs):
        return False
    if _room_is_anchor_locked(rs):
        if pair_key == rs.get('anchor_pair_key'):
            return False
        if pair_key in rs['elective_keys']:
            return True
        if not _room_accepts_shared_combo(rs, pair_key):
            return False
        return len(rs['elective_keys']) < MAX_COMBOS_HARD_CAP
    if pair_key in rs['elective_keys']:
        return True
    return len(rs['elective_keys']) < MAX_COMBOS_HARD_CAP


def _room_can_absorb_count(
    rs: dict,
    add_count: int,
    room_states: list[dict] | None = None,
) -> bool:
    """Sau khi thêm vẫn trong biên: ≥ TB−1, ưu tiên ≤ TB (chỉ TB+1 khi được phép)."""
    if add_count <= 0 or rs['remaining'] < add_count:
        return False
    after = rs['assigned_count'] + add_count
    states = room_states if room_states is not None else [rs]
    hi = _quota_ceiling(rs, states)
    return _room_band_lo(rs) <= after <= hi


def _pick_room_for_tiny_combo(
    room_states: list[dict],
    pair_key: tuple[str, str],
    group_size: int,
    rr_index: int,
) -> tuple[dict | None, int]:
    """
    Ghép tổ hợp 1–2 thí sinh vào phòng đã có 1 tổ hợp chính (tổ hợp phụ thứ 2).
    Dùng biên ±1: ví dụ phòng 25–26 TS có thể +1 hoặc +2 để nhận nhóm nhỏ.
    """
    candidates: list[dict] = []
    for rs in room_states:
        if _room_is_anchor_locked(rs) and not _room_accepts_shared_combo(rs, pair_key):
            continue
        if not _room_can_absorb_count(rs, group_size, room_states):
            continue
        if pair_key in rs['elective_keys']:
            candidates.append(rs)
            continue
        if not _room_may_open_new_combo(rs, room_states):
            continue
        if len(rs['elective_keys']) != 1:
            continue
        if not _can_add_combo(rs, pair_key):
            continue
        candidates.append(rs)

    if not candidates:
        for rs in room_states:
            if _room_is_anchor_locked(rs) and not _room_accepts_shared_combo(rs, pair_key):
                continue
            if not _room_can_absorb_count(rs, group_size, room_states):
                continue
            if pair_key in rs['elective_keys']:
                candidates.append(rs)
            elif _can_add_combo(rs, pair_key) and _room_may_open_new_combo(rs, room_states):
                candidates.append(rs)

    if not candidates:
        return None, 0

    def sort_key(r: dict) -> tuple:
        after = r['assigned_count'] + group_size
        has = pair_key in r['elective_keys']
        overlap = _room_subject_overlap(pair_key, r['elective_keys'])
        at_mid = r['assigned_count'] == _room_count_mid(r)
        return (
            len(r['elective_keys']) + (0 if has else 1),
            0 if has else 1,
            -overlap,
            0 if at_mid else 1,
            r['assigned_count'],
            abs(after - r['count_mid']),
            r['room'].name,
        )

    candidates.sort(key=sort_key)
    best_score = sort_key(candidates[0])
    tier = [c for c in candidates if sort_key(c) == best_score]
    picked = tier[rr_index % len(tier)]
    take = min(group_size, picked['remaining'])
    return picked, take


def _can_add_to_room_under_global(
    rs: dict,
    pair_key: tuple[str, str],
    *,
    has_pair: bool,
    n_combos: int,
    room_states: list[dict],
    chunk_size: int = 0,
) -> bool:
    """Không nhồi phòng đã đủ band_hi khi còn phòng dưới band_lo (trừ phòng đang thiếu)."""
    if not _rooms_below_band(room_states):
        return True
    if has_pair:
        return True
    if rs['assigned_count'] < _room_count_mid(rs):
        return True
    if rs['assigned_count'] >= _assign_ceiling(rs, room_states):
        return False
    if (
        rs['assigned_count'] >= rs['count_mid']
        and _rooms_below_mid(room_states)
        and not has_pair
        and chunk_size > TINY_COMBO_MAX
    ):
        return False
    return True


def _fair_split_quotas(total: int, n_parts: int) -> list[int]:
    """Chia đều total vào n_parts phòng (chênh lệch tối đa 1) — mức lý tưởng mỗi phòng."""
    if n_parts <= 0:
        return []
    base = total // n_parts
    rem = total % n_parts
    return [base + (1 if i < rem else 0) for i in range(n_parts)]


def _best_pair_key_for_room(
    rs: dict,
    queue: list[ExamSort2VenueCandidate],
    room_states: list[dict] | None = None,
) -> tuple[tuple[str, str], int] | None:
    """Tổ hợp kế tiếp: phòng thiếu → lấp nhiều TS; đã đủ TB → ưu tiên ít tổ hợp mới."""
    if not queue or rs['remaining'] <= 0:
        return None
    if rs['assigned_count'] >= _room_band_hi(rs):
        return None

    states = room_states if room_states is not None else [rs]
    ceiling = _quota_ceiling(rs, states)
    need = min(rs['remaining'], ceiling - rs['assigned_count'])
    n_existing = len(rs['elective_keys'])
    under_lo = rs['assigned_count'] < _room_band_lo(rs)
    under_mid = rs['assigned_count'] < _room_count_mid(rs)
    sort_by = 'size_desc' if (under_lo or under_mid) else 'venue'
    best: tuple[tuple, tuple[str, str], int] | None = None

    for pair_key, n in _sorted_combo_keys_from_queue(queue, sort_by=sort_by):
        if not _can_add_combo(rs, pair_key):
            continue
        has = pair_key in rs['elective_keys']
        if not has and not _room_may_open_new_combo(rs, states):
            continue
        overlap = _room_subject_overlap(pair_key, rs['elective_keys'])
        take = min(n, need, rs['remaining'])
        take = _cap_take_for_assign(rs, take, states)
        if take <= 0:
            continue
        after = rs['assigned_count'] + take
        over_mid = max(0, after - rs['count_mid'])
        after_combos = n_existing if has else n_existing + 1
        first_i = next(
            (
                i for i, vc in enumerate(queue)
                if elective_pair_key(vc.candidate) == pair_key
            ),
            999999,
        )
        if under_lo or under_mid:
            # Đã có 2 tổ hợp mà vẫn dưới TB → ưu tiên mở tổ hợp thứ 3 (vd. 6B–9B).
            need_third = (
                not has
                and n_existing >= MAX_COMBOS_PER_ROOM
                and rs['assigned_count'] + take < _room_count_mid(rs)
            )
            score = (
                over_mid * 100,
                -take,
                0 if has or need_third else 1,
                after_combos - (1 if need_third else 0),
                abs(after - rs['count_mid']),
                first_i,
                elective_pair_label(pair_key).casefold(),
            )
        else:
            score = (
                over_mid * 100,
                0 if has else 1,
                after_combos,
                -overlap if not has else 0,
                abs(after - rs['count_mid']),
                -take if has else 0,
                first_i,
                elective_pair_label(pair_key).casefold(),
            )
        if best is None or score < best[0]:
            best = (score, pair_key, take)
    if best is None:
        return None
    return best[1], best[2]


def _anchor_take_headroom(rs: dict, pair_key: tuple[str, str], n_available: int) -> int:
    """Số TS có thể gán trên phòng neo (tối đa band_hi)."""
    if not _room_may_accept_pair(rs, pair_key):
        return 0
    return min(
        n_available,
        _room_band_hi(rs) - rs['assigned_count'],
        rs['remaining'],
    )


def _anchor_fill_band_hi_with_tiny_shared_combos(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Phòng neo (vd. Sử+Địa 26 / 25): lấp tới TB+1 bằng tổ hợp lẻ KHÁC trùng môn
    (vd. Tin+Địa +1, Sử+Lý +2) — không thêm cùng tổ hợp neo.
    """
    placed = 0
    for _ in range(len(room_states) * 30):
        targets = sorted(
            [
                rs for rs in room_states
                if _room_is_anchor_locked(rs)
                and rs['assigned_count'] < _room_band_hi(rs)
            ],
            key=lambda r: (
                _room_band_hi(r) - r['assigned_count'],
                r['assigned_count'],
                r['room'].sort_order,
            ),
            reverse=True,
        )
        if not targets:
            break
        progressed = False
        for rs in targets:
            anchor = rs.get('anchor_pair_key')
            need = _room_band_hi(rs) - rs['assigned_count']
            if need <= 0:
                continue
            best: tuple[tuple, tuple[str, str], int] | None = None
            for pk, n in _sorted_combo_keys_from_queue(
                queue, min_size=1, max_size=TINY_COMBO_MAX, sort_by='size_asc',
            ):
                if pk == anchor:
                    continue
                if not _pair_shares_subject_with_anchor(pk, anchor):
                    continue
                if not _room_may_accept_pair(rs, pk, room_states):
                    continue
                take = min(n, need)
                if take <= 0:
                    continue
                score = (
                    abs(n - need),
                    n,
                    -_shared_subject_count(pk, anchor or ('', '')),
                    elective_pair_label(pk).casefold(),
                )
                if best is None or score < best[0]:
                    best = (score, pk, take)
            if best is None:
                continue
            pk, take = best[1], best[2]
            chunk = _pop_from_venue_queue(queue, pk, take)
            if chunk:
                _assign_chunk_to_room(rs, pk, chunk)
                placed += len(chunk)
                progressed = True
                break
        if not progressed:
            break
    return placed


def _anchor_large_combo_clusters(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Neo cụm lớn (trước round-robin):
    - Phòng trống: chia đều cụm (fair split) vào số phòng cần thiết (vd. 51→26+25).
    - Khóa phòng; lấp tới TB+1 bằng tổ hợp lẻ KHÁC trùng môn (không thêm cùng TC).
    """
    if not queue or not room_states:
        return 0
    count_mid = _room_count_mid(room_states[0])
    placed = 0

    def empty_rooms() -> list[dict]:
        return sorted(
            [
                rs for rs in room_states
                if not rs['elective_keys']
                and rs['assigned_count'] == 0
                and rs['remaining'] > 0
                and not _room_is_anchor_locked(rs)
            ],
            key=lambda r: (r['room'].sort_order, r['room'].name),
        )

    for pair_key, _n in _sorted_combo_keys_from_queue(queue, sort_by='size_desc'):
        while _venue_queue_count(queue, pair_key) > 0:
            empties = empty_rooms()
            if not empties:
                break
            rs = empties[0]
            n_left = _venue_queue_count(queue, pair_key)
            if count_mid > 0:
                rooms_for_cluster = min(
                    len(empties),
                    max(1, (n_left + count_mid - 1) // count_mid),
                )
            else:
                rooms_for_cluster = 1
            take = _fair_split_quotas(n_left, rooms_for_cluster)[0]
            take = _anchor_take_headroom(rs, pair_key, take)
            if take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)
            rs['anchor_locked'] = True
            rs['anchor_pair_key'] = pair_key
            placed += len(chunk)

    placed += _anchor_fill_band_hi_with_tiny_shared_combos(room_states, queue)
    return placed


def _distribute_combos_round_robin(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Chia từng tổ hợp lên phòng chưa đủ mức ưu tiên (≤ TB) — vòng tròn theo số TS thấp nhất.
    Tránh dồn hết một cụm vào vài phòng đầu khi phòng sau còn trống.
    """
    placed = 0
    for _ in range(len(room_states) * 30):
        if not queue:
            break
        items = _sorted_combo_keys_from_queue(queue, sort_by='size_desc')
        progressed = False
        for pair_key, _n in items:
            if _venue_queue_count(queue, pair_key) <= 0:
                continue
            receivers = [
                rs for rs in room_states
                if rs['remaining'] > 0
                and rs['assigned_count'] < _quota_ceiling(rs, room_states)
                and _room_may_accept_pair(rs, pair_key, room_states)
            ]
            if not receivers:
                continue
            receivers.sort(
                key=lambda r: (
                    pair_key not in r['elective_keys'],
                    r['assigned_count'],
                    len(r['elective_keys']),
                    r['room'].sort_order,
                ),
            )
            for rs in receivers:
                if _venue_queue_count(queue, pair_key) <= 0:
                    break
                need = _quota_ceiling(rs, room_states) - rs['assigned_count']
                take = min(
                    _venue_queue_count(queue, pair_key),
                    need,
                    rs['remaining'],
                )
                take = _cap_take_for_assign(rs, take, room_states)
                if take <= 0:
                    continue
                chunk = _pop_from_venue_queue(queue, pair_key, take)
                if chunk:
                    _assign_chunk_to_room(rs, pair_key, chunk)
                    placed += len(chunk)
                    progressed = True
        if not progressed:
            break
    return placed


def _assign_rooms_balance_min_combos(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Xếp theo phòng (thiếu nhất trước): giữ TB ±1, tối thiểu số tổ hợm/phòng.
    Không gom toàn bộ một cụm tổ hợp lớn trước khi các phòng khác đủ người.
    """
    placed = 0
    max_pass = max(len(room_states) * 4, 8)
    for _ in range(max_pass):
        if not queue:
            break
        progressed = False
        ordered = sorted(
            room_states,
            key=lambda r: (
                r['assigned_count'] >= _quota_ceiling(r, room_states),
                r['assigned_count'],
                len(r['elective_keys']),
                r['room'].sort_order,
                r['room'].name,
            ),
        )
        for rs in ordered:
            if rs['remaining'] <= 0:
                continue
            while (
                queue
                and rs['remaining'] > 0
                and rs['assigned_count'] < _quota_ceiling(rs, room_states)
            ):
                pick = _best_pair_key_for_room(rs, queue, room_states)
                if pick is None:
                    break
                pair_key, take = pick
                chunk = _pop_from_venue_queue(queue, pair_key, take)
                if not chunk:
                    break
                _assign_chunk_to_room(rs, pair_key, chunk)
                placed += len(chunk)
                progressed = True
        if not progressed:
            break
    return placed


def _pick_room_for_assign(
    room_states: list[dict],
    pair_key: tuple[str, str],
    chunk_size: int,
) -> tuple[dict | None, int]:
    """
    Chọn phòng khi còn gán theo tổ hợp (bước phụ).

    Ưu tiên: phòng thiếu / gần TB, ít tổ hợm môn; không bắt gom cùng tổ hợp trước.
    """
    best_state = None
    best_take = 0
    best_score: tuple | None = None

    for rs in room_states:
        if rs['remaining'] <= 0:
            continue

        has_pair = pair_key in rs['elective_keys']
        n_combos = len(rs['elective_keys'])

        if not _can_add_combo(rs, pair_key):
            continue
        if not has_pair and not _room_may_open_new_combo(rs, room_states):
            continue
        if not _can_add_to_room_under_global(
            rs,
            pair_key,
            has_pair=has_pair,
            n_combos=n_combos,
            room_states=room_states,
            chunk_size=chunk_size,
        ):
            continue

        take = min(chunk_size, rs['remaining'])
        take = _cap_take_for_assign(rs, take, room_states)
        if take <= 0:
            continue

        after = rs['assigned_count'] + take
        band_hi = _room_band_hi(rs)
        band_lo = _room_band_lo(rs)
        mid = _room_count_mid(rs)
        ceiling = _quota_ceiling(rs, room_states)

        if has_pair:
            combo_rank = n_combos
        elif n_combos == 0:
            combo_rank = 1
        else:
            combo_rank = 10 + n_combos

        overlap = _room_subject_overlap(pair_key, rs['elective_keys']) if not has_pair else 0

        if after > ceiling:
            continue
        if after < band_lo:
            band_penalty = abs(after - band_lo) * 5
        else:
            band_penalty = abs(after - mid) * 2

        over_mid = max(0, after - mid) * 80

        score = (
            over_mid,
            combo_rank,
            band_penalty,
            rs['assigned_count'],
            -overlap,
            -take if rs['assigned_count'] < band_lo else 0,
            rs['room'].sort_order,
            rs['room'].pk,
        )
        if best_score is None or score < best_score:
            best_score = score
            best_state = rs
            best_take = take

    return best_state, best_take


def _assign_chunk_to_room(rs: dict, pair_key: tuple[str, str], chunk: list) -> None:
    rs['new_assignments'].extend(chunk)
    rs['elective_keys'].add(pair_key)
    rs['combo_counts'][pair_key] = rs['combo_counts'].get(pair_key, 0) + len(chunk)
    rs['assigned_count'] += len(chunk)
    rs['remaining'] -= len(chunk)


def _venue_queue_count(
    queue: list[ExamSort2VenueCandidate],
    pair_key: tuple[str, str],
) -> int:
    return sum(
        1 for vc in queue if elective_pair_key(vc.candidate) == pair_key
    )


def _pop_from_venue_queue(
    queue: list[ExamSort2VenueCandidate],
    pair_key: tuple[str, str],
    take: int,
) -> list[ExamSort2VenueCandidate]:
    """Lấy tối đa ``take`` thí sinh đầu tiên (trên→dưới) trong queue đúng tổ hợp môn."""
    if take <= 0:
        return []
    out: list[ExamSort2VenueCandidate] = []
    i = 0
    while i < len(queue) and len(out) < take:
        vc = queue[i]
        if elective_pair_key(vc.candidate) == pair_key:
            out.append(vc)
            queue.pop(i)
        else:
            i += 1
    return out


def _sorted_combo_keys_from_queue(
    queue: list[ExamSort2VenueCandidate],
    *,
    min_size: int = 1,
    max_size: int | None = None,
    sort_by: str = 'size_desc',
) -> list[tuple[tuple[str, str], int]]:
    """
    Liệt kê tổ hợp môn còn trong queue (chưa xếp phòng).

    sort_by — thứ tự xử lý phân bổ (logic cũ), khác với thứ tự lấy HV cụ thể (luôn trên→dưới queue):
    - size_desc: nhóm đông trước (cụm lớn)
    - size_asc: nhóm 1–2 TS trước (tiny)
    - venue: tổ hợp xuất hiện sớm trên danh sách điểm thi trước
    """
    first_rank: dict[tuple[str, str], int] = {}
    counts: dict[tuple[str, str], int] = {}
    for i, vc in enumerate(queue):
        k = elective_pair_key(vc.candidate)
        counts[k] = counts.get(k, 0) + 1
        if k not in first_rank:
            first_rank[k] = i

    items: list[tuple[tuple[str, str], int]] = []
    for k, n in counts.items():
        if n < min_size:
            continue
        if max_size is not None and n > max_size:
            continue
        items.append((k, n))

    label_key = lambda x: elective_pair_label(x[0]).casefold()

    if sort_by == 'venue':
        return sorted(items, key=lambda x: (first_rank[x[0]], -x[1], label_key(x)))
    if sort_by == 'size_asc':
        return sorted(items, key=lambda x: (x[1], label_key(x)))
    # size_desc — mặc định logic cũ (cụm lớn trước)
    return sorted(items, key=lambda x: (-x[1], first_rank[x[0]], label_key(x)))


def _fill_rooms_to_band(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> None:
    """Lấp phòng chưa đạt band_lo (trung bình −1), không vượt band_hi."""
    for rs in sorted(room_states, key=lambda r: r['assigned_count']):
        while rs['assigned_count'] < _room_band_lo(rs) and rs['remaining'] > 0:
            candidates: list[tuple[int, tuple[str, str], int]] = []
            if rs['elective_keys']:
                for pair_key in rs['elective_keys']:
                    n = _venue_queue_count(queue, pair_key)
                    if n:
                        first_i = next(
                            (
                                i for i, vc in enumerate(queue)
                                if elective_pair_key(vc.candidate) == pair_key
                            ),
                            999999,
                        )
                        candidates.append((n, pair_key, first_i))
            if not candidates:
                for pair_key, n in _sorted_combo_keys_from_queue(queue, sort_by='size_desc'):
                    if not _can_add_combo(rs, pair_key):
                        continue
                    if (
                        pair_key not in rs['elective_keys']
                        and not _room_may_open_new_combo(rs, room_states)
                    ):
                        continue
                    first_i = next(
                        (
                            i for i, vc in enumerate(queue)
                            if elective_pair_key(vc.candidate) == pair_key
                        ),
                        999999,
                    )
                    candidates.append((n, pair_key, first_i))
            if not candidates:
                break

            def fill_sort_key(item: tuple[int, tuple[str, str], int]) -> tuple:
                size, pk, _first_i = item
                has = pk in rs['elective_keys']
                need = _room_band_lo(rs) - rs['assigned_count']
                return (
                    0 if has else 1,
                    len(rs['elective_keys']) + (0 if has else 1),
                    -_room_subject_overlap(pk, rs['elective_keys']),
                    abs((rs['assigned_count'] + min(size, need)) - rs['count_mid']),
                    -size,
                    elective_pair_label(pk).casefold(),
                )

            candidates.sort(key=fill_sort_key)
            pair_key = candidates[0][1]
            need = _room_band_lo(rs) - rs['assigned_count']
            take = min(
                _venue_queue_count(queue, pair_key),
                rs['remaining'],
                need,
                _room_band_hi(rs) - rs['assigned_count'],
            )
            take = _cap_take_for_assign(rs, take, room_states)
            if take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)


def _fill_rooms_to_mid(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> None:
    """Lấp các phòng chưa đạt mức trung bình (làm tròn), vẫn ưu tiên ≤ TB."""
    for rs in sorted(room_states, key=lambda r: r['assigned_count']):
        while (
            rs['assigned_count'] < _room_count_mid(rs)
            and rs['remaining'] > 0
            and queue
        ):
            pick = _best_pair_key_for_room(rs, queue, room_states)
            if pick is None:
                break
            pair_key, take = pick
            take = _cap_take_for_assign(rs, take, room_states)
            if take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)


def _fill_third_combo_for_under_mid_rooms(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Phòng đã có 2 tổ hợp nhưng vẫn dưới TB → mở tổ hợp thứ 3 (vd. 6B–9B: 22–24 → ~26).
    """
    placed = 0
    targets = sorted(
        [
            rs for rs in room_states
            if rs['assigned_count'] < _room_count_mid(rs)
            and len(rs['elective_keys']) >= MAX_COMBOS_PER_ROOM
            and rs['remaining'] > 0
            and len(rs['elective_keys']) < _max_combos_for_room(rs)
        ],
        key=lambda r: (r['assigned_count'], r['room'].sort_order, r['room'].name),
    )
    for rs in targets:
        while (
            rs['assigned_count'] < _room_count_mid(rs)
            and rs['remaining'] > 0
            and queue
            and len(rs['elective_keys']) < _max_combos_for_room(rs)
        ):
            best: tuple[tuple, tuple[str, str], int] | None = None
            for pair_key, n in _sorted_combo_keys_from_queue(queue, sort_by='size_desc'):
                if pair_key in rs['elective_keys']:
                    continue
                if not _can_add_combo(rs, pair_key):
                    continue
                if not _room_may_open_new_combo(rs, room_states):
                    continue
                take = min(
                    n,
                    rs['remaining'],
                    _room_count_mid(rs) - rs['assigned_count'],
                )
                take = _cap_take_for_assign(rs, take, room_states)
                if take <= 0:
                    continue
                overlap = _room_subject_overlap(pair_key, rs['elective_keys'])
                score = (
                    -take,
                    -overlap,
                    elective_pair_label(pair_key).casefold(),
                )
                if best is None or score < best[0]:
                    best = (score, pair_key, take)
            if best is None:
                break
            pair_key, take = best[1], best[2]
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)
            placed += len(chunk)
    return placed


def _place_tiny_combos(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> None:
    """Tổ hợp 1–2 thí sinh → ghép vào phòng đã có 1 tổ hợp (±1 TS so với trung bình), rải vòng."""
    items = _sorted_combo_keys_from_queue(
        queue, min_size=1, max_size=TINY_COMBO_MAX, sort_by='size_asc',
    )
    rr = 0
    for pair_key, _n in items:
        while _venue_queue_count(queue, pair_key) > 0:
            size = _venue_queue_count(queue, pair_key)
            rs, take = _pick_room_for_tiny_combo(room_states, pair_key, size, rr)
            rr += 1
            if rs is None or take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)


def _spread_small_groups(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> None:
    """Rải nhóm 3 TS (và phần tổ hợp 1–2 còn sót), tối đa 2 tổ hợp/phòng."""
    items = _sorted_combo_keys_from_queue(queue, sort_by='size_asc')
    rr = 0
    for pair_key, _n in items:
        while _venue_queue_count(queue, pair_key) > 0:
            receivers = [rs for rs in room_states if rs['remaining'] > 0]
            if not receivers:
                break
            picked = None
            same_rooms = [rs for rs in receivers if pair_key in rs['elective_keys']]
            if same_rooms:
                same_rooms.sort(key=lambda r: r['assigned_count'])
                picked = same_rooms[0]
            else:
                candidates = [
                    rs for rs in receivers
                    if _can_add_combo(rs, pair_key)
                    and rs['assigned_count'] < _quota_ceiling(rs, room_states)
                    and _room_may_open_new_combo(rs, room_states)
                ]
                if not candidates:
                    break
                candidates.sort(
                    key=lambda r: (
                        len(r['elective_keys']),
                        -_room_subject_overlap(pair_key, r['elective_keys']),
                        r['assigned_count'],
                        r['room'].name,
                    ),
                )
                picked = candidates[rr % len(candidates)]
                rr += 1
            take = min(_venue_queue_count(queue, pair_key), picked['remaining'])
            take = _cap_take_for_assign(picked, take, room_states)
            if take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(picked, pair_key, chunk)


def _top_up_existing_combos(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> None:
    """Lấp thêm vào phòng đã có tổ hợp (trong biên ±FLEX) trước khi mở tổ hợp mới."""
    for rs in sorted(room_states, key=lambda r: r['assigned_count']):
        if not rs['elective_keys'] or rs['remaining'] <= 0:
            continue
        cap_band = _quota_ceiling(rs, room_states)

        def key_rank(pk: tuple[str, str]) -> int:
            for i, vc in enumerate(queue):
                if elective_pair_key(vc.candidate) == pk:
                    return i
            return 999999

        for pair_key in sorted(rs['elective_keys'], key=key_rank):
            while (
                _venue_queue_count(queue, pair_key) > 0
                and rs['remaining'] > 0
                and rs['assigned_count'] < cap_band
            ):
                take = min(
                    _venue_queue_count(queue, pair_key),
                    rs['remaining'],
                    cap_band - rs['assigned_count'],
                )
                take = _cap_take_for_assign(rs, take, room_states)
                if take <= 0:
                    break
                chunk = _pop_from_venue_queue(queue, pair_key, take)
                if not chunk:
                    break
                _assign_chunk_to_room(rs, pair_key, chunk)


def _transfer_vc_between_rooms(
    don: dict,
    rec: dict,
    vc: ExamSort2VenueCandidate,
    *,
    from_index: int | None = None,
) -> bool:
    """Chuyển một thí sinh (đã xếp tạm) giữa hai phòng. Trả về False nếu không tìm thấy trong phòng nguồn."""
    lst = don['new_assignments']
    if from_index is not None:
        if from_index < 0 or from_index >= len(lst):
            return False
        popped = lst.pop(from_index)
        if popped.pk != vc.pk:
            lst.insert(from_index, popped)
            return False
    else:
        try:
            idx = lst.index(vc)
        except ValueError:
            idx = next((i for i, x in enumerate(lst) if x.pk == vc.pk), -1)
            if idx < 0:
                return False
        lst.pop(idx)

    k = elective_pair_key(vc.candidate)
    don['assigned_count'] -= 1
    don['remaining'] += 1
    don['combo_counts'][k] = don['combo_counts'].get(k, 1) - 1
    if don['combo_counts'][k] <= 0:
        don['combo_counts'].pop(k, None)
        don['elective_keys'].discard(k)
    rec['new_assignments'].append(vc)
    rec['elective_keys'].add(k)
    rec['combo_counts'][k] = rec['combo_counts'].get(k, 0) + 1
    rec['assigned_count'] += 1
    rec['remaining'] -= 1
    return True


def _spread_tiny_combos_from_crowded_rooms(room_states: list[dict]) -> int:
    """
    Phòng có quá nhiều tổ hợp (>2): chuyển nhóm 1–2 TS sang phòng có **môn trùng**,
    ưu tiên phòng đang 26 TS → 27 (TB+1). VD: Hóa+KTPL từ 9B → 2B/3B (Hóa+Lý).
    """
    from collections import defaultdict

    moved = 0
    for _ in range(len(room_states) * 50):
        progressed = False
        crowded = sorted(
            [
                rs for rs in room_states
                if len(rs['elective_keys']) > MAX_COMBOS_PER_ROOM
                and rs['new_assignments']
            ],
            key=lambda r: (-len(r['elective_keys']), -r['assigned_count']),
        )
        for don in crowded:
            by_combo: dict[tuple[str, str], list[ExamSort2VenueCandidate]] = defaultdict(list)
            for vc in don['new_assignments']:
                by_combo[elective_pair_key(vc.candidate)].append(vc)

            for pair_key in sorted(by_combo.keys(), key=lambda k: len(by_combo[k])):
                if len(by_combo[pair_key]) > TINY_COMBO_MAX:
                    continue
                for vc in list(by_combo[pair_key]):
                    k = pair_key
                    receivers = [
                        rs for rs in room_states
                        if rs is not don
                        and rs['remaining'] > 0
                        and rs['assigned_count'] < _room_band_hi(rs)
                        and _room_subject_overlap(k, rs['elective_keys']) > 0
                        and _can_add_combo(rs, k)
                    ]
                    if not receivers:
                        receivers = [
                            rs for rs in room_states
                            if rs is not don
                            and rs['remaining'] > 0
                            and rs['assigned_count'] < _room_band_hi(rs)
                            and _room_subject_overlap(k, rs['elective_keys']) > 0
                            and _can_add_combo_emergency(rs, k)
                        ]
                    if not receivers:
                        continue

                    receivers.sort(
                        key=lambda r: (
                            k not in r['elective_keys'],
                            len(r['elective_keys']),
                            -_room_subject_overlap(k, r['elective_keys']),
                            0 if r['assigned_count'] == _room_count_mid(r) else 1,
                            r['assigned_count'],
                            r['room'].sort_order,
                            r['room'].name,
                        ),
                    )
                    if _transfer_vc_between_rooms(don, receivers[0], vc):
                        moved += 1
                        progressed = True
                    break
                if progressed:
                    break
            if progressed:
                break
        if not progressed:
            break
    return moved


def _pull_from_full_rooms(room_states: list[dict]) -> None:
    """Chuyển thí sinh từ phòng > band_lo sang phòng < band_lo (có thể khác tổ hợp nếu phòng thiếu còn slot)."""
    for _ in range(len(room_states) * 20):
        receivers = [
            rs for rs in room_states
            if rs['assigned_count'] < _room_band_lo(rs) and rs['remaining'] > 0
        ]
        if not receivers:
            break
        receivers.sort(
            key=lambda r: (_room_band_lo(r) - r['assigned_count']),
            reverse=True,
        )
        moved = False
        for rec in receivers:
            donors = sorted(
                [
                    d for d in room_states
                    if d['assigned_count'] > _room_band_lo(d) and d['new_assignments']
                ],
                key=lambda d: -d['assigned_count'],
            )
            for don in donors:
                if don is rec:
                    continue
                best_idx = None
                best_prio = 999
                for idx, vc in enumerate(don['new_assignments']):
                    k = elective_pair_key(vc.candidate)
                    if not _can_add_combo(rec, k):
                        continue
                    if k in rec['elective_keys']:
                        prio = 0
                    elif _room_subject_overlap(k, rec['elective_keys']) > 0:
                        prio = 1
                    else:
                        prio = 2
                    if prio < best_prio:
                        best_prio = prio
                        best_idx = idx
                if best_idx is None:
                    continue
                vc = don['new_assignments'][best_idx]
                if _transfer_vc_between_rooms(don, rec, vc, from_index=best_idx):
                    moved = True
                break
            if moved:
                break
        if not moved:
            break


def _balance_count_band(room_states: list[dict]) -> None:
    """Chuyển thí sinh từ phòng đông → phòng thiếu (ưu tiên cùng tổ hợp, tối đa 2 tổ hợp)."""
    for _ in range(len(room_states) * 12):
        receivers = [
            rs for rs in room_states
            if rs['assigned_count'] < _room_band_lo(rs) and rs['remaining'] > 0
        ]
        if not receivers:
            break
        donors = [
            rs for rs in room_states
            if rs['assigned_count'] > _room_band_lo(rs) and rs['new_assignments']
        ]
        if not donors:
            break
        receivers.sort(
            key=lambda r: (_room_band_lo(r) - r['assigned_count']),
            reverse=True,
        )
        donors.sort(key=lambda r: -r['assigned_count'])
        moved = False
        for rec in receivers:
            for don in donors:
                if don is rec:
                    continue
                for idx, vc in enumerate(list(don['new_assignments'])):
                    k = elective_pair_key(vc.candidate)
                    if not _can_add_combo(rec, k):
                        continue
                    if rec['assigned_count'] >= _room_band_hi(rec):
                        break
                    if don['assigned_count'] <= _room_band_lo(don):
                        continue
                    vc = don['new_assignments'][idx]
                    if _transfer_vc_between_rooms(don, rec, vc, from_index=idx):
                        moved = True
                    break
                if moved:
                    break
            if moved:
                break
        if not moved:
            break


def _campus_cannot_fill_min_per_room(n_students: int, room_states: list[dict]) -> bool:
    """Tổng TS cơ sở < số phòng × sàn 24 → không thể mọi phòng đạt 24."""
    cap_rooms = [rs for rs in room_states if rs['capacity'] >= MIN_STUDENTS_PER_ROOM]
    return bool(cap_rooms) and n_students < MIN_STUDENTS_PER_ROOM * len(cap_rooms)


def _max_rooms_for_min_students(n_students: int) -> int:
    return max(0, n_students // MIN_STUDENTS_PER_ROOM)


def _donor_floor_for_rebalance(
    don: dict,
    rec: dict,
    *,
    campus_shortage: bool = False,
) -> int:
    """Sàn số TS trên phòng cho sau khi chuyển đi."""
    ideal_rec = rec.get('ideal_target', _room_count_mid(rec))
    if rec['assigned_count'] < ideal_rec:
        if campus_shortage:
            # TB làm tròn có thể = 24: cho phép 24→23 để lấp phòng dưới chỉ tiêu
            return max(_room_count_mid(don) - 1, ideal_rec)
        if rec['assigned_count'] < _room_band_lo(rec):
            return _room_count_mid(don)
        if rec['assigned_count'] < _room_count_mid(rec):
            return _room_count_mid(don)
    return _room_band_lo(don)


def _donor_has_transfer_for_rec(don: dict, rec: dict) -> bool:
    """Phòng cho còn TS có thể chuyển sang phòng nhận (cùng tổ hợp / còn slot tổ hợp)."""
    if not don['new_assignments']:
        return False
    rec_keys = rec['elective_keys']
    max_c = _max_combos_for_room(rec)
    for vc in don['new_assignments']:
        k = elective_pair_key(vc.candidate)
        if k in rec_keys:
            return True
        if len(rec_keys) < max_c:
            if not rec_keys or _room_subject_overlap(k, rec_keys) > 0:
                return True
    return False


def _transfer_one_to_receiver(
    don: dict,
    rec: dict,
    *,
    campus_shortage: bool = False,
) -> bool:
    """Chuyển 1 TS phù hợp nhất từ don → rec."""
    if not _donor_has_transfer_for_rec(don, rec):
        return False
    if don['assigned_count'] <= _donor_floor_for_rebalance(
        don, rec, campus_shortage=campus_shortage,
    ):
        return False
    if rec['assigned_count'] >= rec.get('ideal_target', _room_count_mid(rec)):
        return False
    if rec['assigned_count'] >= _room_band_hi(rec):
        return False
    best_idx: int | None = None
    best_prio = 999
    for idx, vc in enumerate(don['new_assignments']):
        k = elective_pair_key(vc.candidate)
        can = k in rec['elective_keys'] or _can_add_combo(rec, k)
        if not can and rec['assigned_count'] < rec.get('ideal_target', _room_count_mid(rec)):
            can = _can_add_combo_emergency(rec, k)
        if not can:
            continue
        if k in rec['elective_keys']:
            prio = 0
        elif _room_subject_overlap(k, rec['elective_keys']) > 0:
            prio = 1
        elif len(rec['elective_keys']) < _max_combos_for_room(rec):
            prio = 2
        else:
            continue
        if prio < best_prio:
            best_prio = prio
            best_idx = idx
    if best_idx is None:
        return False
    return _transfer_vc_between_rooms(
        don, rec, don['new_assignments'][best_idx], from_index=best_idx,
    )


def _rebalance_surplus_to_deficit(
    room_states: list[dict],
    n_campus_students: int = 0,
) -> int:
    """
    Kéo TS từ phòng vượt chỉ tiêu / trên TB sang phòng thiếu (dưới sàn hoặc dưới TB).
    Cho phép mở tổ hợp thứ 3 trên phòng nhận (6B–9B).
    """
    campus_shortage = _campus_cannot_fill_min_per_room(n_campus_students, room_states)
    moved = 0
    for _ in range(len(room_states) * 60):
        receivers = [
            rs for rs in room_states
            if rs['remaining'] > 0
            and rs['assigned_count'] < rs.get('ideal_target', _room_count_mid(rs))
        ]
        if not receivers:
            receivers = [
                rs for rs in room_states
                if rs['remaining'] > 0
                and rs['assigned_count'] < _room_count_mid(rs)
            ]
        if not receivers:
            break
        receivers.sort(
            key=lambda r: (
                r.get('ideal_target', _room_count_mid(r)) - r['assigned_count'],
                0 if r['assigned_count'] < _room_band_lo(r) else 1,
                r['room'].sort_order,
            ),
            reverse=True,
        )
        donors = [
            rs for rs in room_states
            if rs['new_assignments']
            and rs['assigned_count'] > rs.get('ideal_target', _room_count_mid(rs))
        ]
        if campus_shortage and not donors:
            donors = [
                rs for rs in room_states
                if rs['new_assignments']
                and rs['assigned_count'] > max(
                    _room_count_mid(rs) - 1,
                    rs.get('ideal_target', _room_count_mid(rs)),
                )
            ]
        if not donors:
            break
        donors.sort(
            key=lambda d: (
                d['assigned_count'] - d.get('ideal_target', _room_count_mid(d)),
                d['assigned_count'],
            ),
            reverse=True,
        )
        progressed = False
        for rec in receivers:
            goal = rec.get('ideal_target', _room_count_mid(rec))
            if rec['assigned_count'] >= goal:
                continue
            compatible = [
                d for d in donors
                if d is not rec and _donor_has_transfer_for_rec(d, rec)
            ]
            for don in compatible or donors:
                if don is rec:
                    continue
                if not compatible and not _donor_has_transfer_for_rec(don, rec):
                    continue
                if _transfer_one_to_receiver(
                    don, rec, campus_shortage=campus_shortage,
                ):
                    moved += 1
                    progressed = True
                    break
            if progressed:
                break
        if not progressed:
            break
    return moved


def _enforce_ideal_quotas(
    room_states: list[dict],
    n_campus_students: int,
) -> int:
    """Cân theo chỉ tiêu chia đều — tránh 1 phòng (vd. 13) gánh hết phần thiếu."""
    return _rebalance_surplus_to_deficit(room_states, n_campus_students)


def _pick_room_emergency_third(
    room_states: list[dict],
    pair_key: tuple[str, str],
    chunk_size: int,
) -> tuple[dict | None, int]:
    """Tổ hợp thứ 3 hoặc lấp ghế khi phòng còn chỗ nhưng đã đủ 2 tổ hợp / band_hi."""
    best_rs = None
    best_take = 0
    best_key: tuple | None = None
    for rs in sorted(
        room_states,
        key=lambda r: (r['assigned_count'], len(r['elective_keys'])),
    ):
        if rs['remaining'] <= 0:
            continue
        if pair_key in rs['elective_keys']:
            score = (0, rs['assigned_count'], -rs['remaining'])
        elif len(rs['elective_keys']) >= _max_combos_for_room(rs):
            continue
        elif not _room_may_open_new_combo(rs, room_states):
            continue
        else:
            score = (
                1,
                rs['assigned_count'],
                -_room_subject_overlap(pair_key, rs['elective_keys']),
                -rs['remaining'],
            )
        take = min(chunk_size, rs['remaining'])
        if take <= 0:
            continue
        if best_key is None or score < best_key:
            best_key = score
            best_rs = rs
            best_take = take
    return best_rs, best_take


def _finish_remaining_assignments(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
    errors: list[str],
    campus_code: str,
) -> None:
    """Gán nốt; tối đa 2 tổ hợp/phòng, chỉ mở tổ hợp 3 khi bắt buộc."""
    items = _sorted_combo_keys_from_queue(queue, sort_by='size_desc')
    for pair_key, _n in items:
        while _venue_queue_count(queue, pair_key) > 0:
            pending_n = _venue_queue_count(queue, pair_key)
            if pending_n <= TINY_COMBO_MAX:
                rs, take = _pick_room_for_tiny_combo(
                    room_states, pair_key, pending_n, 0,
                )
            else:
                rs, take = _pick_room_for_assign(room_states, pair_key, pending_n)
            if rs is None or take <= 0:
                rs, take = _pick_room_emergency_third(
                    room_states, pair_key, pending_n,
                )
            if rs is None or take <= 0:
                label = elective_pair_label(pair_key)
                errors.append(
                    f'Không đủ chỗ cho {pending_n} thí sinh tổ hợp {label} (cơ sở {campus_code}).',
                )
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)


def _boost_underfilled_rooms_from_queue(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Kéo phòng đang thiếu (dưới band_lo / count_mid) lên gần mức trung bình:
    ưu tiên bổ sung tổ hợp đã có trong phòng, rồi mới mở tổ hợp mới (tối đa 3 khi còn thiếu).
    """
    placed = 0
    target_rooms = sorted(
        room_states,
        key=lambda rs: (rs['assigned_count'], rs['room'].sort_order, rs['room'].name),
    )
    for rs in target_rooms:
        if rs['assigned_count'] < _room_count_mid(rs):
            goal = _room_count_mid(rs)
        elif _room_may_exceed_avg(rs, room_states):
            goal = _room_band_hi(rs)
        else:
            goal = _room_count_mid(rs)
        while rs['assigned_count'] < goal and rs['remaining'] > 0 and queue:
            moved = False
            for pair_key in list(rs['elective_keys']):
                n = _venue_queue_count(queue, pair_key)
                if n <= 0:
                    continue
                take = min(
                    n,
                    rs['remaining'],
                    goal - rs['assigned_count'],
                )
                take = _cap_take_for_assign(rs, take, room_states)
                if take <= 0:
                    continue
                chunk = _pop_from_venue_queue(queue, pair_key, take)
                if chunk:
                    _assign_chunk_to_room(rs, pair_key, chunk)
                    placed += len(chunk)
                    moved = True
                    break
            if moved:
                continue
            pick = _best_pair_key_for_room(rs, queue, room_states)
            if pick is None:
                break
            pair_key, take = pick
            take = min(take, goal - rs['assigned_count'])
            take = _cap_take_for_assign(rs, take, room_states)
            if take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if chunk:
                _assign_chunk_to_room(rs, pair_key, chunk)
                placed += len(chunk)
                moved = True
            if not moved:
                break
    return placed


def _fill_physical_spare_seats(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
) -> int:
    """
    Lấp ghế còn trống trong biên band_hi (TB+1) — không lấp tới sức chứa vật lý.
    Phòng neo khóa chỉ nhận cùng tổ hợp.
    """
    placed = 0
    if not queue:
        return 0
    max_rotations = len(queue) * 3
    rotations = 0
    while queue and rotations < max_rotations:
        vc = queue[0]
        pair_key = elective_pair_key(vc.candidate)
        candidates = sorted(
            [
                rs for rs in room_states
                if rs['remaining'] > 0
                and rs['assigned_count'] < _room_band_hi(rs)
            ],
            key=lambda rs: (
                1 if _room_is_anchor_locked(rs) else 0,
                rs['assigned_count'],
                0 if pair_key in rs['elective_keys'] else 1,
                -_room_subject_overlap(pair_key, rs['elective_keys']),
                len(rs['elective_keys']),
                rs['room'].sort_order,
                rs['room'].name,
            ),
        )
        picked: dict | None = None
        for rs in candidates:
            if not _room_may_accept_pair(rs, pair_key, room_states):
                continue
            if not _can_add_combo_emergency(rs, pair_key):
                continue
            picked = rs
            break
        if picked is None:
            queue.append(queue.pop(0))
            rotations += 1
            continue
        queue.pop(0)
        _assign_chunk_to_room(picked, pair_key, [vc])
        placed += 1
        rotations = 0
    return placed


def _emergency_combo_cap(rs: dict) -> int:
    """Trần tổ hợm khi gán khẩn cấp — phòng dưới band_lo được mở thêm tổ hợp."""
    if rs['assigned_count'] < _room_band_lo(rs):
        return MAX_COMBOS_HARD_CAP + 2
    return MAX_COMBOS_HARD_CAP


def _emergency_headroom(rs: dict) -> int:
    """Số ghế còn nhận (không vượt band_hi và sức chứa vật lý)."""
    ceiling = min(_room_band_hi(rs), rs['capacity'])
    return min(rs['remaining'], max(0, ceiling - rs['assigned_count']))


def _pick_emergency_room(
    room_states: list[dict],
    pair_key: tuple[str, str],
    *,
    shared_only: bool,
) -> dict | None:
    candidates: list[dict] = []
    for rs in room_states:
        if _emergency_headroom(rs) <= 0:
            continue
        if pair_key in rs['elective_keys']:
            candidates.append(rs)
            continue
        if len(rs['elective_keys']) >= _emergency_combo_cap(rs):
            continue
        if shared_only and _room_is_anchor_locked(rs):
            if not _room_accepts_shared_combo(rs, pair_key):
                continue
        candidates.append(rs)
    if not candidates:
        return None
    candidates.sort(
        key=lambda rs: (
            pair_key not in rs['elective_keys'],
            -_room_subject_overlap(pair_key, rs['elective_keys']),
            len(rs['elective_keys']),
            rs['assigned_count'],
            -rs['remaining'],
            rs['room'].sort_order,
        ),
    )
    return candidates[0]


def _emergency_assign_remaining_queue(
    room_states: list[dict],
    queue: list[ExamSort2VenueCandidate],
    errors: list[str],
    campus_code: str,
) -> int:
    """
    Gán nốt TS còn trong queue — theo từng tổ hợp (cụm), ưu tiên phòng trùng môn.
    Tránh gán lẻ từng em khiến phòng đủ 4 tổ hợp trước khi nhận cả cụm (vd. 6 em Hóa+Sinh).
    """
    placed = 0
    if not queue:
        return 0
    for pair_key, _n in _sorted_combo_keys_from_queue(queue, sort_by='size_desc'):
        while _venue_queue_count(queue, pair_key) > 0:
            rs = _pick_emergency_room(room_states, pair_key, shared_only=True)
            if rs is None:
                rs = _pick_emergency_room(room_states, pair_key, shared_only=False)
            if rs is None:
                break
            take = min(_venue_queue_count(queue, pair_key), _emergency_headroom(rs))
            if take <= 0:
                break
            chunk = _pop_from_venue_queue(queue, pair_key, take)
            if not chunk:
                break
            _assign_chunk_to_room(rs, pair_key, chunk)
            placed += len(chunk)

    for vc in list(queue):
        cand = vc.candidate
        pk = elective_pair_key(cand)
        errors.append(
            f'Không gán được {cand.full_name} — tổ hợp {elective_pair_label(pk)} '
            f'(cơ sở {campus_code}): hết ghế hoặc không có phòng đúng cơ sở.',
        )
    return placed


def _rebuild_room_elective_keys(rs: dict) -> None:
    keys: set[tuple[str, str]] = set()
    for a in rs['existing'].values():
        keys.add(elective_pair_key(a.venue_candidate.candidate))
    for vc in rs['new_assignments']:
        keys.add(elective_pair_key(vc.candidate))
    rs['elective_keys'] = keys
    counts: dict[tuple[str, str], int] = {}
    for a in rs['existing'].values():
        k = elective_pair_key(a.venue_candidate.candidate)
        counts[k] = counts.get(k, 0) + 1
    for vc in rs['new_assignments']:
        k = elective_pair_key(vc.candidate)
        counts[k] = counts.get(k, 0) + 1
    rs['combo_counts'] = counts


def assign_candidates_to_rooms(
    venue: ExamSort2Venue,
    *,
    clear_existing: bool = False,
) -> dict[str, Any]:
    """
    Xếp thí sinh vào phòng — hai tầng:
    1. Theo phòng: ưu tiên số TS ≤ TB (làm tròn), tối đa TB+1 khi mọi phòng đã đạt TB; tối thiểu số tổ hợm/phòng.
    2. Chọn HV cụ thể từ trên xuống danh sách điểm thi (tên TV) trong từng tổ hợp đã chọn.
    - Theo đối tượng cơ sở từng phòng (target_campus).

    Trả về: assigned_total, per_room, errors, target_avg_info.
    """
    from collections import defaultdict

    rooms = list(
        ExamSort2Room.objects.filter(venue=venue)
        .select_related('target_campus')
        .order_by('sort_order', 'name', 'pk')
    )
    if not rooms:
        return {
            'assigned_total': 0,
            'per_room': [],
            'errors': ['Chưa có phòng thi nào.'],
            'target_avg_info': [],
            'rooms_reordered': 0,
        }

    per_room: list[dict] = []
    errors: list[str] = []
    target_avg_info: list[str] = []
    rooms_reordered = 0

    with transaction.atomic():
        if clear_existing:
            ExamSort2SeatAssignment.objects.filter(room__venue=venue).delete()

        assigned_vc_ids = _venue_assigned_vc_ids(venue)
        unassigned = sort_venue_candidates_by_vi_name(
            list(
                ExamSort2VenueCandidate.objects.filter(venue=venue)
                .exclude(pk__in=assigned_vc_ids)
                .select_related('candidate', 'candidate__campus'),
            ),
        )

        rooms_by_campus: dict[int, list[ExamSort2Room]] = defaultdict(list)
        for room in rooms:
            if room.col_count * room.row_count <= 0:
                errors.append(f'Phòng {room.name}: kích thước không hợp lệ.')
                continue
            rooms_by_campus[room.target_campus_id].append(room)

        total_new = 0

        for campus_id, campus_rooms in rooms_by_campus.items():
            campus_students = [
                vc for vc in unassigned if vc.candidate.campus_id == campus_id
            ]
            if not campus_students:
                continue

            queue: list[ExamSort2VenueCandidate] = list(campus_students)

            n_rooms = len(campus_rooms)
            target_avg = len(campus_students) / n_rooms
            campus_code = campus_rooms[0].target_campus.code
            quotas = _fair_split_quotas(len(campus_students), n_rooms)
            count_mid, band_lo, band_hi = _count_band_from_avg(target_avg)
            rooms_min_cap = [
                r for r in campus_rooms
                if r.col_count * r.row_count >= MIN_STUDENTS_PER_ROOM
            ]
            n_campus = len(campus_students)
            max_rooms_ok = _max_rooms_for_min_students(n_campus)
            if (
                rooms_min_cap
                and n_campus < MIN_STUDENTS_PER_ROOM * len(rooms_min_cap)
            ):
                need_more = MIN_STUDENTS_PER_ROOM * len(rooms_min_cap) - n_campus
                errors.append(
                    f'{campus_code}: {n_campus} thí sinh / {len(rooms_min_cap)} phòng — '
                    f'không đủ cho tối thiểu {MIN_STUDENTS_PER_ROOM} TS/phòng '
                    f'(thiếu {need_more} TS, hoặc tối đa ~{max_rooms_ok} phòng).',
                )
            target_avg_info.append(
                f'{campus_code}: ~{target_avg:.1f} TS/phòng — ưu tiên ≤{count_mid} (TB), '
                f'cho phép tới {band_hi} (TB+1) khi mọi phòng đã đạt TB; '
                f'tối thiểu {max(band_lo, MIN_STUDENTS_PER_ROOM)} TS/phòng '
                f'(sàn {MIN_STUDENTS_PER_ROOM} nếu phòng đủ chỗ), '
                f'tối đa {MAX_COMBOS_HARD_CAP} tổ hợp/phòng (thường {MAX_COMBOS_PER_ROOM}); '
                f'chỉ tiêu: {quotas}.',
            )

            room_states: list[dict] = []
            for qi, room in enumerate(campus_rooms):
                capacity = room.col_count * room.row_count
                rm_mid, rm_lo, rm_hi = _room_count_bands_for_capacity(
                    capacity, count_mid, band_lo, band_hi,
                )
                existing, used_seats = _room_existing_assignments(room)
                elective_keys: set[tuple[str, str]] = set()
                combo_counts: dict[tuple[str, str], int] = {}
                assigned_count = len(used_seats)
                for a in existing.values():
                    k = elective_pair_key(a.venue_candidate.candidate)
                    elective_keys.add(k)
                    combo_counts[k] = combo_counts.get(k, 0) + 1

                room_states.append({
                    'room': room,
                    'capacity': capacity,
                    'remaining': capacity - assigned_count,
                    'used_seats': used_seats,
                    'existing': existing,
                    'new_assignments': [],
                    'elective_keys': elective_keys,
                    'combo_counts': combo_counts,
                    'assigned_count': assigned_count,
                    'ideal_target': quotas[qi],
                    'count_mid': rm_mid,
                    'band_lo': rm_lo,
                    'band_hi': rm_hi,
                    'soft_min': rm_lo,
                    'soft_max': capacity,
                    'anchor_locked': False,
                    'anchor_pair_key': None,
                })

            anchored = _anchor_large_combo_clusters(room_states, queue)
            if anchored:
                target_avg_info.append(
                    f'{campus_code}: neo cụm lớn (TB + tổ hợp lẻ trùng môn → TB+1): {anchored} TS.',
                )

            rr_placed = _distribute_combos_round_robin(room_states, queue)
            primary_placed = _assign_rooms_balance_min_combos(room_states, queue)
            if rr_placed or primary_placed:
                target_avg_info.append(
                    f'{campus_code}: chia vòng + lấp phòng: {rr_placed + primary_placed} TS.',
                )

            _fill_rooms_to_band(room_states, queue)
            _fill_rooms_to_mid(room_states, queue)
            _top_up_existing_combos(room_states, queue)
            _place_tiny_combos(room_states, queue)
            _spread_small_groups(room_states, queue)
            _fill_rooms_to_band(room_states, queue)
            _fill_rooms_to_mid(room_states, queue)
            third_mid = _fill_third_combo_for_under_mid_rooms(room_states, queue)
            if third_mid:
                target_avg_info.append(
                    f'{campus_code}: mở tổ hợp thứ 3 cho phòng dưới TB: {third_mid} TS.',
                )
            _top_up_existing_combos(room_states, queue)
            _finish_remaining_assignments(
                room_states, queue, errors, campus_code,
            )
            _balance_count_band(room_states)
            _pull_from_full_rooms(room_states)
            _place_tiny_combos(room_states, queue)
            _fill_rooms_to_band(room_states, queue)
            _fill_rooms_to_mid(room_states, queue)
            _fill_third_combo_for_under_mid_rooms(room_states, queue)
            _top_up_existing_combos(room_states, queue)
            _pull_from_full_rooms(room_states)
            _balance_count_band(room_states)
            _finish_remaining_assignments(
                room_states, queue, errors, campus_code,
            )
            _pull_from_full_rooms(room_states)
            _fill_rooms_to_band(room_states, queue)
            _fill_rooms_to_mid(room_states, queue)
            _fill_third_combo_for_under_mid_rooms(room_states, queue)
            _rebalance_surplus_to_deficit(room_states, n_campus)

            boosted = _boost_underfilled_rooms_from_queue(room_states, queue)
            if boosted:
                target_avg_info.append(
                    f'{campus_code}: đã bổ sung {boosted} TS vào phòng đang thiếu (kéo gần mức TB).',
                )

            spare_filled = _fill_physical_spare_seats(room_states, queue)
            if spare_filled:
                target_avg_info.append(
                    f'{campus_code}: đã lấp thêm {spare_filled} TS vào ghế trống còn lại.',
                )

            emergency_placed = _emergency_assign_remaining_queue(
                room_states, queue, errors, campus_code,
            )
            if emergency_placed:
                target_avg_info.append(
                    f'{campus_code}: gán khẩn cấp {emergency_placed} TS còn sót '
                    f'(tổ hợp hiếm / vượt giới hạn tổ hợp thường).',
                )

            rebalanced = _enforce_ideal_quotas(room_states, n_campus)
            if rebalanced:
                target_avg_info.append(
                    f'{campus_code}: cân bằng {rebalanced} TS theo chỉ tiêu phòng '
                    f'(phòng thiếu → ~chỉ tiêu).',
                )
            _fill_rooms_to_band(room_states, queue)
            _fill_rooms_to_mid(room_states, queue)
            _fill_third_combo_for_under_mid_rooms(room_states, queue)
            _fill_physical_spare_seats(room_states, queue)

            redistributed = _spread_tiny_combos_from_crowded_rooms(room_states)
            if redistributed:
                target_avg_info.append(
                    f'{campus_code}: đã chuyển {redistributed} TS tổ hợp lẻ '
                    f'(1–2 người) sang phòng có môn trùng (vd. 26→27).',
                )
            rebalanced2 = _enforce_ideal_quotas(room_states, n_campus)
            if rebalanced2:
                target_avg_info.append(
                    f'{campus_code}: cân bằng bổ sung {rebalanced2} TS sau rải tổ hợp lẻ.',
                )

            campus_shortage = _campus_cannot_fill_min_per_room(n_campus, room_states)
            for rs in room_states:
                _rebuild_room_elective_keys(rs)
                if rs['assigned_count'] < _room_band_lo(rs):
                    floor = _room_band_lo(rs)
                    ideal = rs.get('ideal_target', rs['count_mid'])
                    if campus_shortage and rs['assigned_count'] < ideal:
                        errors.append(
                            f'Phòng {rs["room"].name}: {rs["assigned_count"]} thí sinh '
                            f'(chỉ tiêu {ideal}; cơ sở thiếu TS — không thể mọi phòng '
                            f'≥{MIN_STUDENTS_PER_ROOM}, nên dùng ≤{max_rooms_ok} phòng).',
                        )
                    elif (
                        rs['capacity'] >= MIN_STUDENTS_PER_ROOM
                        and floor >= MIN_STUDENTS_PER_ROOM
                    ):
                        errors.append(
                            f'Phòng {rs["room"].name}: {rs["assigned_count"]} thí sinh '
                            f'(dưới tối thiểu {MIN_STUDENTS_PER_ROOM}, mức phòng {floor}).',
                        )
                    else:
                        errors.append(
                            f'Phòng {rs["room"].name}: {rs["assigned_count"]} thí sinh '
                            f'(dưới mức {floor}, trung bình {rs["count_mid"]}).',
                        )
                elif rs['assigned_count'] > _room_band_hi(rs):
                    errors.append(
                        f'Phòng {rs["room"].name}: {rs["assigned_count"]} thí sinh '
                        f'(vượt mức {rs["band_hi"]}, trung bình {rs["count_mid"]}).',
                    )
                if len(rs['elective_keys']) > MAX_COMBOS_HARD_CAP:
                    errors.append(
                        f'Phòng {rs["room"].name}: {len(rs["elective_keys"])} tổ hợp môn '
                        f'(vượt tối đa {MAX_COMBOS_HARD_CAP}) — cần thêm phòng.',
                    )

            for rs in room_states:
                room = rs['room']
                if not rs['new_assignments']:
                    combo_labels = sorted(
                        {elective_pair_label(k) for k in rs['elective_keys']},
                        key=lambda s: s.casefold(),
                    )
                    per_room.append({
                        'room_id': room.id,
                        'room_name': room.name,
                        'target_campus': room.target_campus.code,
                        'assigned': 0,
                        'capacity': rs['capacity'],
                        'filled': rs['assigned_count'],
                        'elective_combo_count': len(rs['elective_keys']),
                        'elective_combos': combo_labels,
                    })
                    continue

                free_seats = [
                    sn for sn in range(1, rs['capacity'] + 1) if sn not in rs['used_seats']
                ]
                bulk: list[ExamSort2SeatAssignment] = []
                for vc, seat_no in zip(rs['new_assignments'], free_seats):
                    bulk.append(ExamSort2SeatAssignment(
                        room=room,
                        venue_candidate=vc,
                        seat_number=seat_no,
                    ))
                    assigned_vc_ids.add(vc.pk)
                if bulk:
                    ExamSort2SeatAssignment.objects.bulk_create(bulk)
                total_new += len(bulk)

                combo_labels = sorted(
                    {elective_pair_label(k) for k in rs['elective_keys']},
                    key=lambda s: s.casefold(),
                )
                per_room.append({
                    'room_id': room.id,
                    'room_name': room.name,
                    'target_campus': room.target_campus.code,
                    'assigned': len(bulk),
                    'capacity': rs['capacity'],
                    'filled': rs['assigned_count'],
                    'elective_combo_count': len(rs['elective_keys']),
                    'elective_combos': combo_labels,
                })

            unassigned = [vc for vc in unassigned if vc.pk not in assigned_vc_ids]

        still_unassigned = len(unassigned)
        if still_unassigned > 0:
            spare_total = sum(
                max(0, r.col_count * r.row_count)
                - ExamSort2SeatAssignment.objects.filter(room=r).count()
                for r in rooms
            )
            errors.append(
                f'Còn {still_unassigned} thí sinh chưa xếp được. '
                f'Còn khoảng {spare_total} ghế trống toàn điểm thi — kiểm tra: '
                f'(1) phòng thi có đúng cơ sở (target_campus) với thí sinh; '
                f'(2) tổng sức chứa phòng ≥ số thí sinh; '
                f'(3) mỗi phòng đã đủ {MAX_COMBOS_HARD_CAP} tổ hợm — cần thêm phòng; '
                f'(4) tổ hợp môn TC đã nhập đủ 2 môn trong file import.',
            )

        rooms_reordered = reorder_venue_rooms_by_roster(venue)

    return {
        'assigned_total': total_new,
        'per_room': per_room,
        'errors': errors,
        'target_avg_info': target_avg_info,
        'rooms_reordered': rooms_reordered,
    }


# Sơ đồ môn tự chọn: điền cột (trên→dưới, trái→phải). Điểm bắt đầu STT để giám thị phát đề thuận tiện
# (môn đông nhất phát liền từ ô đầu; STT môn khác quay vòng xuống cuối — không phải bốc thăm ngẫu nhiên).
ELECTIVE_BOC_THAM_LAYOUTS: dict[int, dict[str, Any]] = {
    0: {
        'label': 'Thuận phát đề (đề xuất)',
        'description': (
            'Một điểm bắt đầu STT chung cho cả 2 môn TC (cùng ô ghế); '
            'tối ưu phát đề liền cho cả cột TC1 và TC2; vẫn in 2 sơ đồ riêng.'
        ),
        'row_dir': 'down',
        'col_dir': 'left',
        'start_mode': 'proctor',
    },
    1: {
        'label': 'Phương án sơ đồ 1',
        'description': 'Cột trên→dưới, trái→phải; ô đầu STT = 1 + số cột (mẫu 4×6: STT 5).',
        'row_dir': 'down',
        'col_dir': 'left',
        'start_mode': 'ref',
        'ref_scheme': 1,
    },
    2: {
        'label': 'Phương án sơ đồ 2',
        'description': 'Cột trên→dưới, trái→phải; ô đầu STT = 1 + 2×số cột (mẫu 4×6: STT 9).',
        'row_dir': 'down',
        'col_dir': 'left',
        'start_mode': 'ref',
        'ref_scheme': 2,
    },
    5: {
        'label': 'Phương án sơ đồ 5',
        'description': 'Cột trên→dưới, trái→phải; ô đầu STT = 1 + 2×số cột + 2 (mẫu 4×6: STT 11).',
        'row_dir': 'down',
        'col_dir': 'left',
        'start_mode': 'ref',
        'ref_scheme': 5,
    },
}

# Giữ alias để view cũ không lỗi import
ELECTIVE_LAYOUT_SCHEMES = ELECTIVE_BOC_THAM_LAYOUTS

# Viết tắt môn tự chọn trên sơ đồ (giống LI / SI / HO trong mẫu)
ELECTIVE_SUBJECT_ABBR: dict[str, str] = {
    'Lý': 'LI',
    'Hóa': 'HO',
    'Sinh': 'SI',
    'Tin': 'TI',
    'Sử': 'SU',
    'Địa': 'DI',
    'GDCD': 'GD',
    'KTPL': 'KT',
    'Văn': 'VA',
    'Toán': 'TO',
}


def _elective_continuous_fill_order(
    room: ExamSort2Room,
    layout_id: int = 1,
) -> list[int]:
    """
    Thứ tự ghế để gán STT liên tục (1, 2, 3…).
    Điền theo cột: trên → dưới trong cột, rồi cột tiếp theo trái → phải (nhìn từ ghế lên bảng).
    Không nhảy cóc kiểu 1,3,5… trên cùng một hàng.
    """
    layout = ELECTIVE_BOC_THAM_LAYOUTS.get(layout_id, ELECTIVE_BOC_THAM_LAYOUTS[1])
    rows = list(range(1, room.row_count + 1))
    cols = list(range(1, room.col_count + 1))
    if layout.get('row_dir') == 'up':
        rows = list(reversed(rows))
    if layout.get('col_dir') == 'right':
        cols = list(reversed(cols))
    order: list[int] = []
    for c in cols:
        for r in rows:
            order.append(grid_to_seat_number(r, c, room.col_count))
    return order


def _elective_reference_start_stt(
    room: ExamSort2Room,
    ref_scheme: int,
    n_students: int,
) -> int:
    """STT bắt đầu theo mẫu in (4 cột: PA1→5, PA2→9, PA5→11), quy đổi theo số cột phòng."""
    c = max(1, room.col_count)
    if ref_scheme == 1:
        raw = 1 + c
    elif ref_scheme == 2:
        raw = 1 + 2 * c
    elif ref_scheme == 5:
        raw = 1 + 2 * c + 2
    else:
        raw = 1
    if n_students <= 0:
        return 1
    return ((raw - 1) % n_students) + 1


def _prefix_len_for_slot_at_start(
    roster: list[ExamSort2SeatAssignment],
    slot: int,
    start_stt: int,
) -> int:
    """Số thí sinh cùng môn TC (theo slot) liên tiếp khi điền từ start_stt (quay vòng danh sách STT phòng)."""
    roster = sorted(roster, key=lambda a: a.seat_number)
    n = len(roster)
    if n <= 0:
        return 0
    subjects = [
        _elective_subject_for_slot(a.venue_candidate.candidate, slot)
        for a in roster
    ]
    subj_counts: dict[str, int] = {}
    for s in subjects:
        subj_counts[s] = subj_counts.get(s, 0) + 1
    plurality = max(subj_counts, key=lambda k: (subj_counts[k], k))
    start_idx = next(
        (i for i, a in enumerate(roster) if a.seat_number == start_stt),
        0,
    )
    prefix = 0
    for i in range(n):
        if subjects[(start_idx + i) % n] == plurality:
            prefix += 1
        else:
            break
    return prefix


def _pick_unified_elective_start_stt(
    roster: list[ExamSort2SeatAssignment],
) -> int:
    """
    Chọn STT bắt đầu chung cho cả 2 môn tự chọn (cùng sơ đồ ghế cả buổi).
    Tối ưu: cân bằng khối phát liền cho TC1 và TC2 (maximin rồi tổng).
    """
    roster = sorted(roster, key=lambda a: a.seat_number)
    if not roster:
        return 1
    best_stt = roster[0].seat_number
    best_score: tuple[int, int, int, int] = (-1, -1, -1, -1)
    for a in roster:
        stt = a.seat_number
        p1 = _prefix_len_for_slot_at_start(roster, 1, stt)
        p2 = _prefix_len_for_slot_at_start(roster, 2, stt)
        score = (min(p1, p2), p1 + p2, p1, p2)
        if score > best_score:
            best_score = score
            best_stt = stt
    return best_stt


def _resolve_unified_elective_start_stt(
    room: ExamSort2Room,
    layout_id: int,
    roster: list[ExamSort2SeatAssignment],
) -> int:
    """Điểm bắt đầu STT dùng chung cho sơ đồ TC1 và TC2."""
    layout = ELECTIVE_BOC_THAM_LAYOUTS.get(layout_id, ELECTIVE_BOC_THAM_LAYOUTS[0])
    n = len(roster)
    mode = layout.get('start_mode', 'fixed')
    if mode in ('proctor', 'auto'):
        return _pick_unified_elective_start_stt(roster)
    if mode == 'ref':
        return _elective_reference_start_stt(
            room, int(layout.get('ref_scheme', layout_id)), n,
        )
    return max(1, min(int(layout.get('start_stt', 1)), n or 1))


def _pick_proctor_friendly_start_stt(
    roster: list[ExamSort2SeatAssignment],
    slot: int,
) -> int:
    """
    STT bắt đầu điền ghế để giám thị phát đề thuận tiện:
    khối môn đông nhất (theo thứ tự danh sách phòng) đặt ở ô đầu; STT môn khác quay vòng cuối sơ đồ.
    Vd. STT 1 Sinh, 2 KTPL, 3–25 Sử → bắt đầu STT 3.
    """
    roster = sorted(roster, key=lambda a: a.seat_number)
    n = len(roster)
    if n <= 1:
        return roster[0].seat_number if roster else 1

    subjects = [
        _elective_subject_for_slot(a.venue_candidate.candidate, slot)
        for a in roster
    ]
    subj_counts: dict[str, int] = {}
    for s in subjects:
        subj_counts[s] = subj_counts.get(s, 0) + 1
    plurality = max(subj_counts, key=lambda k: (subj_counts[k], k))

    best_len = 0
    best_start_stt = roster[0].seat_number
    i = 0
    while i < n:
        if subjects[i] != plurality:
            i += 1
            continue
        j = i
        while j < n and subjects[j] == plurality:
            j += 1
        run_len = j - i
        if run_len > best_len:
            best_len = run_len
            best_start_stt = roster[i].seat_number
        i = j if j > i else i + 1

    if best_len > 0:
        return best_start_stt

    # Môn đông nhất không liền trong danh sách: xoay sao khối đó nằm đầu khi điền ghế
    best_prefix = -1
    best_start_stt = roster[0].seat_number
    for start_idx in range(n):
        prefix = 0
        for i in range(n):
            if subjects[(start_idx + i) % n] == plurality:
                prefix += 1
            else:
                break
        if prefix > best_prefix:
            best_prefix = prefix
            best_start_stt = roster[start_idx].seat_number
    return best_start_stt


def _elective_proctor_fill_summary(
    roster: list[ExamSort2SeatAssignment],
    slot: int,
    start_stt: int,
) -> dict[str, Any]:
    """Mô tả ngắn: môn phát liền + STT quay vòng cuối (cho giám thị)."""
    roster = sorted(roster, key=lambda a: a.seat_number)
    n = len(roster)
    seat_to_idx = {a.seat_number: i for i, a in enumerate(roster)}
    start_idx = seat_to_idx.get(start_stt, 0)

    subjects = [
        _elective_subject_for_slot(a.venue_candidate.candidate, slot)
        for a in roster
    ]
    subj_counts: dict[str, int] = {}
    for s in subjects:
        subj_counts[s] = subj_counts.get(s, 0) + 1
    plurality = max(subj_counts, key=lambda k: (subj_counts[k], k))

    prefix = 0
    for i in range(n):
        if subjects[(start_idx + i) % n] == plurality:
            prefix += 1
        else:
            break

    tail_stts: list[int] = []
    for i in range(prefix, n):
        tail_stts.append(roster[(start_idx + i) % n].seat_number)

    return {
        'primary_subject': plurality,
        'primary_abbr': ELECTIVE_SUBJECT_ABBR.get(
            plurality, (plurality[:2] if plurality else '—').upper(),
        ),
        'primary_count': prefix,
        'tail_stts': tail_stts,
    }


def _resolve_elective_start_stt(
    room: ExamSort2Room,
    layout_id: int,
    roster: list[ExamSort2SeatAssignment],
    slot: int,
) -> int:
    layout = ELECTIVE_BOC_THAM_LAYOUTS.get(layout_id, ELECTIVE_BOC_THAM_LAYOUTS[1])
    n = len(roster)
    mode = layout.get('start_mode', 'fixed')
    if mode in ('proctor', 'auto'):
        return _pick_proctor_friendly_start_stt(roster, slot)
    if mode == 'ref':
        return _elective_reference_start_stt(
            room, int(layout.get('ref_scheme', layout_id)), n,
        )
    return max(1, min(int(layout.get('start_stt', 1)), n or 1))


def _elective_subject_for_slot(candidate: ExamSort2Candidate, slot: int) -> str:
    """Tên môn tự chọn theo vị trí 1 hoặc 2."""
    if slot == 1:
        return (candidate.elective_subject_1 or '').strip()
    return (candidate.elective_subject_2 or '').strip()


def _elective_abbr_for_slot(candidate: ExamSort2Candidate, slot: int) -> str:
    s = _elective_subject_for_slot(candidate, slot)
    return ELECTIVE_SUBJECT_ABBR.get(s, (s[:2] if s else '—').upper())


def _elective_track_abbr(candidate: ExamSort2Candidate) -> str:
    """Mã môn (môn tự chọn 1)."""
    return _elective_abbr_for_slot(candidate, 1)


def _sort2_pattern_for_fixed_subject(room: ExamSort2Room, subject_name: str) -> int:
    """Chọn 1 trong 8 kiểu sơ đồ (ổn định theo phòng + môn) — giống xếp phòng thi cũ."""
    seed = (room.pk or 0) * 1009 + sum(ord(c) for c in subject_name)
    pool = [1, 2, 3, 4] if subject_name == 'Văn' else [5, 6, 7, 8]
    return pool[seed % len(pool)]


def _sort2_get_seat_order_for_pattern(
    room: ExamSort2Room,
    n_students: int,
    pattern_id: int,
) -> list[int]:
    from adminpage.views import (
        _seat_positions_column_balanced,
        _seat_positions_front_filled,
    )

    if pattern_id in (1, 2, 3, 4):
        return _seat_positions_column_balanced(room, n_students, pattern_id, shift=None)
    ff_dir = {5: 1, 6: 3, 7: 2, 8: 4}
    return _seat_positions_front_filled(
        room, n_students, ff_dir.get(pattern_id, 1), shift=None,
    )


def _grid_from_seat_cell_map(
    room: ExamSort2Room,
    seat_to_cell: dict[int, dict | None],
) -> list[list[dict | None]]:
    grid: list[list[dict | None]] = []
    for r in range(1, room.row_count + 1):
        row_cells: list[dict | None] = []
        for c in range(1, room.col_count + 1):
            sn = grid_to_seat_number(r, c, room.col_count)
            row_cells.append(seat_to_cell.get(sn))
        grid.append(row_cells)
    return grid


def _build_fixed_subject_diagram_grid(
    room: ExamSort2Room,
    assignments: list[ExamSort2SeatAssignment],
    subject_name: str,
) -> list[list[dict | None]]:
    """
    Sơ đồ một môn (Văn / Toán): xáo vị trí ghế theo 8 phương án như module xếp phòng thi cũ.
    Ô hiển thị SBD (số báo danh).
    """
    roster = _sort_assignments_by_roster(assignments)
    n = len(roster)
    if n == 0:
        return _grid_from_seat_cell_map(room, {})

    pattern = _sort2_pattern_for_fixed_subject(room, subject_name)
    fill_order = _sort2_get_seat_order_for_pattern(room, n, pattern)
    seat_to_cell: dict[int, dict | None] = {}
    for i, a in enumerate(roster):
        if i >= len(fill_order):
            break
        sn = fill_order[i]
        cand = a.venue_candidate.candidate
        seat_to_cell[sn] = {
            'seat_number': sn,
            'room_stt': i + 1,
            'sbd': cand.exam_number or '',
            'full_name': cand.full_name,
            'pattern_id': pattern,
        }
    return _grid_from_seat_cell_map(room, seat_to_cell)


def _build_elective_slot_diagram_grid(
    room: ExamSort2Room,
    assignments: list[ExamSort2SeatAssignment],
    slot: int,
    layout_id: int = 1,
    *,
    shared_start_stt: int | None = None,
) -> tuple[list[list[dict | None]], dict[str, Any]]:
    """
    Sơ đồ một môn tự chọn (1 hoặc 2):
    - Cùng ô ghế với môn TC kia khi ``shared_start_stt`` được truyền (một sơ đồ vật lý, hai bản in).
    - Điền cột: trên→dưới, trái→phải; STT liên tục từ điểm bắt đầu chung (nếu có) rồi quay vòng.
    - Đề STT đếm riêng theo từng môn thi ở vị trí slot đó.
    """
    base_layout = ELECTIVE_BOC_THAM_LAYOUTS.get(layout_id, ELECTIVE_BOC_THAM_LAYOUTS[0])
    layout = dict(base_layout)
    roster = sorted(assignments, key=lambda a: a.seat_number)
    n = len(roster)
    if n == 0:
        layout['start_stt'] = 1
        return _grid_from_seat_cell_map(room, {}), layout

    if shared_start_stt is not None:
        start_stt = shared_start_stt
        layout['shared_seating'] = True
        layout['shared_start_stt'] = start_stt
    else:
        start_stt = _resolve_elective_start_stt(room, layout_id, roster, slot)
    layout['start_stt'] = start_stt
    summary = _elective_proctor_fill_summary(roster, slot, start_stt)
    layout['proctor_summary'] = summary
    if summary['tail_stts']:
        tail = ', '.join(str(s) for s in summary['tail_stts'])
        layout['proctor_note'] = (
            f'Giám thị phát liền {summary["primary_abbr"]} ({summary["primary_count"]} thí sinh) '
            f'từ STT {start_stt}; STT {tail} (môn khác) ở cuối sơ đồ.'
        )
    else:
        layout['proctor_note'] = (
            f'Giám thị phát liền {summary["primary_abbr"]} từ STT {start_stt}.'
        )
    if layout.get('shared_seating'):
        other = 'TC2' if slot == 1 else 'TC1'
        layout['proctor_note'] = (
            f'{layout["proctor_note"]} '
            f'Cùng ô ghế với sơ đồ {other} (một sơ đồ phòng, hai môn tự chọn).'
        )

    fill_order = _elective_continuous_fill_order(room, layout_id)

    de_stt_by_roster_stt: dict[int, int] = {}
    track_de_counter: dict[str, int] = {}
    for a in roster:
        cand = a.venue_candidate.candidate
        subj = _elective_subject_for_slot(cand, slot)
        track = subj or '—'
        track_de_counter[track] = track_de_counter.get(track, 0) + 1
        de_stt_by_roster_stt[a.seat_number] = track_de_counter[track]

    seat_to_cell: dict[int, dict | None] = {}
    for i in range(min(n, len(fill_order))):
        a = roster[(start_stt - 1 + i) % n]
        sn = fill_order[i]
        cand = a.venue_candidate.candidate
        subj = _elective_subject_for_slot(cand, slot)
        seat_to_cell[sn] = {
            'seat_number': sn,
            'room_stt': a.seat_number,
            'sbd': cand.exam_number or '',
            'subject_code': _elective_abbr_for_slot(cand, slot),
            'subject_name': subj or '—',
            'de_stt': de_stt_by_roster_stt.get(a.seat_number, 1),
            'full_name': cand.full_name,
        }

    return _grid_from_seat_cell_map(room, seat_to_cell), layout


def build_room_diagram_blocks(
    room: ExamSort2Room,
    *,
    elective_layout_id: int = 0,
    elective_scheme_id: int | None = None,
) -> list[dict[str, Any]]:
    """
    Sơ đồ phòng thi SXPT II:
    - Văn, Toán: mỗi môn một sơ đồ (xáo ghế 8 phương án, giống xếp phòng thi cũ).
    - Môn tự chọn 1 & 2: hai bản in (môn / Đề STT khác nhau) nhưng **cùng ô ghế**;
      STT bắt đầu trên sơ đồ được chọn tối ưu cho cả hai môn.
    """
    if elective_scheme_id is not None:
        elective_layout_id = elective_scheme_id
    assignments = list(
        ExamSort2SeatAssignment.objects.filter(room=room)
        .select_related('venue_candidate__candidate', 'venue_candidate__candidate__campus'),
    )
    roster_sorted = sorted(assignments, key=lambda a: a.seat_number)
    shared_start = (
        _resolve_unified_elective_start_stt(room, elective_layout_id, roster_sorted)
        if roster_sorted
        else 1
    )
    el1_grid, layout1 = _build_elective_slot_diagram_grid(
        room, assignments, slot=1, layout_id=elective_layout_id,
        shared_start_stt=shared_start,
    )
    el2_grid, layout2 = _build_elective_slot_diagram_grid(
        room, assignments, slot=2, layout_id=elective_layout_id,
        shared_start_stt=shared_start,
    )
    van_pattern = _sort2_pattern_for_fixed_subject(room, 'Văn')
    toan_pattern = _sort2_pattern_for_fixed_subject(room, 'Toán')

    return [
        {
            'title': 'Văn',
            'subject_key': 'van',
            'layout_kind': 'fixed_subject',
            'pattern_id': van_pattern,
            'grid': _build_fixed_subject_diagram_grid(room, assignments, 'Văn'),
        },
        {
            'title': 'Toán',
            'subject_key': 'toan',
            'layout_kind': 'fixed_subject',
            'pattern_id': toan_pattern,
            'grid': _build_fixed_subject_diagram_grid(room, assignments, 'Toán'),
        },
        {
            'title': f'Môn tự chọn 1 — {layout1["label"]}',
            'subject_key': 'elective_1',
            'layout_kind': 'elective',
            'elective_slot': 1,
            'layout_id': elective_layout_id,
            'layout': layout1,
            'grid': el1_grid,
        },
        {
            'title': f'Môn tự chọn 2 — {layout2["label"]}',
            'subject_key': 'elective_2',
            'layout_kind': 'elective',
            'elective_slot': 2,
            'layout_id': elective_layout_id,
            'layout': layout2,
            'grid': el2_grid,
        },
    ]


def build_room_seat_grid(room: ExamSort2Room) -> list[list[dict | None]]:
    """Ma trận [hàng][cột] cho hiển thị sơ đồ phòng."""
    assignments = {
        a.seat_number: a
        for a in ExamSort2SeatAssignment.objects.filter(room=room)
        .select_related('venue_candidate__candidate', 'venue_candidate__candidate__campus')
    }
    grid: list[list[dict | None]] = []
    for r in range(1, room.row_count + 1):
        row_cells: list[dict | None] = []
        for c in range(1, room.col_count + 1):
            sn = grid_to_seat_number(r, c, room.col_count)
            a = assignments.get(sn)
            if a:
                cand = a.venue_candidate.candidate
                row_cells.append({
                    'seat_number': sn,
                    'full_name': cand.full_name,
                    'class_name': cand.class_name,
                    'campus_code': cand.campus.code,
                })
            else:
                row_cells.append(None)
        grid.append(row_cells)
    return grid


def get_room_assignments_ordered(room: ExamSort2Room) -> list[ExamSort2SeatAssignment]:
    """Danh sách thí sinh trong phòng: cặp môn (gom môn neo) → tên TV."""
    assignments = list(
        ExamSort2SeatAssignment.objects.filter(room=room)
        .select_related('venue_candidate__candidate', 'venue_candidate__candidate__campus')
    )
    return _sort_assignments_by_roster(assignments)


def build_room_four_subject_blocks(
    room: ExamSort2Room,
    *,
    elective_layout_id: int = 0,
    elective_scheme_id: int | None = None,
) -> list[dict]:
    """Giữ tên hàm cũ — gọi build_room_diagram_blocks."""
    return build_room_diagram_blocks(
        room,
        elective_layout_id=elective_layout_id,
        elective_scheme_id=elective_scheme_id,
    )
