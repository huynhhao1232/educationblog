"""
Export danh sách + sơ đồ phòng thi SXPT II (mẫu EXPORT_TEMPLATE như xếp phòng thi cũ).
"""
from __future__ import annotations

import re
from collections import OrderedDict
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from adminpage.exam_sort2_logic import (
    ELECTIVE_BOC_THAM_LAYOUTS,
    _roster_pair_context,
    build_room_diagram_blocks,
    elective_pair_label,
    roster_pair_display_label,
    roster_pair_group_key,
)
from homepage.models import (
    ExamSort2Candidate,
    ExamSort2Room,
    ExamSort2SeatAssignment,
    ExamSort2Venue,
)

# Reuse helpers từ module xếp phòng thi cũ
from adminpage.views import (
    _copy_worksheet_to_workbook,
    _detect_seat_cells_for_export,
    _detect_seat_cells_from_template,
    _exam_bulk_unique_sheet_name,
    _exam_export_workbook_paths,
    _exam_merged_list_prepare_sheet,
    _exam_sheet1_guess_stt_header_row,
    _exam_sheet1_template_has_th_column,
    _fallback_seat_cells_by_border,
    _fill_exam_giao_nhan_merged,
    _get_block_starts_from_template,
    _set_cell_value,
    natural_sort_key,
)

_SORT2_SUBJECTS_WS1 = ['VĂN', 'MÔN TC 1', 'TOÁN', 'MÔN TC 2']
_DIAGRAM_SUBJECT_KEYS = ('van', 'toan', 'elective_1', 'elective_2')

_THIN = Side(style='thin')
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_FONT_HDR = Font(bold=True, size=10)
_FONT_TITLE = Font(bold=True, size=12)
_FONT_BOARD = Font(bold=True, size=11)


class _Sort2ExportStudent:
    """Adapter để tái sử dụng _fill_exam_giao_nhan_merged."""

    def __init__(self, candidate: ExamSort2Candidate):
        self.full_name = candidate.full_name
        self.class_name = candidate.class_name
        self.exam_number = candidate.exam_number or ''
        self.student_code = self.exam_number
        self.is_integration = False
        self._candidate = candidate


class _Sort2ExportAssignment:
    def __init__(self, assignment: ExamSort2SeatAssignment):
        self.student = _Sort2ExportStudent(assignment.venue_candidate.candidate)
        self._assignment = assignment


def _sort2_th_label(stu: _Sort2ExportStudent | None, room: ExamSort2Room | None = None) -> str:
    if not stu:
        return '-'
    cand = getattr(stu, '_candidate', None)
    if not cand:
        return '-'
    if room is not None:
        assigns = list(
            ExamSort2SeatAssignment.objects.filter(room=room)
            .select_related('venue_candidate__candidate'),
        )
        cands = [a.venue_candidate.candidate for a in assigns]
        slots, weights = _roster_pair_context(cands)
        label = roster_pair_display_label(
            cand, subject_slots=slots, subject_weights=weights,
        )
    else:
        label = elective_pair_label(
            roster_pair_group_key(cand),
        )
    return label[:40] if label else '-'


def _sort2_combo_label(assignments: list[_Sort2ExportAssignment]) -> str:
    from collections import Counter

    if not assignments:
        return 'TỔ HỢP — (—)'
    candidates = [a.student._candidate for a in assignments]
    subject_slots, subject_weights = _roster_pair_context(candidates)
    counter: Counter[tuple[str, str]] = Counter()
    classes: list[str] = []
    seen_cls: set[str] = set()
    for a in assignments:
        cand = a.student._candidate
        gk = roster_pair_group_key(
            cand, subject_slots=subject_slots, subject_weights=subject_weights,
        )
        counter[gk] += 1
        cn = (cand.class_name or '').strip()
        if cn and cn not in seen_cls:
            seen_cls.add(cn)
            classes.append(cn)
    combo_parts = [
        elective_pair_label(k)
        for k, _ in sorted(
            counter.items(),
            key=lambda item: (
                -item[1],
                tuple(s.casefold() for s in item[0]),
            ),
        )
    ]
    classes.sort(key=natural_sort_key)
    combo_str = ', '.join(combo_parts) if combo_parts else '—'
    cls_str = ', '.join(classes) if classes else '—'
    return f'TỔ HỢP {combo_str} ({cls_str})'


def _venue_export_blocks(venue: ExamSort2Venue) -> list[tuple[ExamSort2Room, list[_Sort2ExportAssignment]]]:
    rooms = list(
        ExamSort2Room.objects.filter(venue=venue)
        .select_related('target_campus')
        .order_by('sort_order', 'name'),
    )
    blocks: list[tuple[ExamSort2Room, list[_Sort2ExportAssignment]]] = []
    for room in rooms:
        assigns = list(
            ExamSort2SeatAssignment.objects.filter(room=room)
            .select_related('venue_candidate__candidate')
            .order_by('seat_number'),
        )
        if not assigns:
            continue
        blocks.append((room, [_Sort2ExportAssignment(a) for a in assigns]))
    return blocks


def _diagram_sbd_by_seat_number(block: dict) -> dict[int, str]:
    """SBD theo số ghế vật lý — cùng dữ liệu ``block['grid']`` trên web."""
    out: dict[int, str] = {}
    for row in block.get('grid') or []:
        for cell in row:
            if not cell:
                continue
            sn = cell.get('seat_number')
            if sn is None:
                continue
            out[int(sn)] = str(cell.get('sbd') or '')
    return out


def _sbd_values_in_diagram_fill_order(
    room: ExamSort2Room,
    block: dict,
    assignments: list[ExamSort2SeatAssignment],
    ordered_len: int,
) -> list[str]:
    """
    SBD theo thứ tự ô ghế trên template Excel (row-major, ghế 1..n).
    Lấy từ ``block['grid']`` đã xáo theo từng môn — khớp sơ đồ trong hệ thống.
    """
    del room, assignments  # giữ chữ ký; dữ liệu lấy từ block
    by_seat = _diagram_sbd_by_seat_number(block)
    values: list[str] = []
    for idx in range(ordered_len):
        values.append(by_seat.get(idx + 1, ''))
    return values


def _fill_sort2_diagram_sheet(
    ws2,
    room: ExamSort2Room,
    assignments: list[ExamSort2SeatAssignment],
    *,
    elective_layout_id: int = 0,
    ws_src_style=None,
    ws_src_values=None,
) -> None:
    """Điền 4 khối sơ đồ (Văn, Toán, TC1, TC2) — cùng mẫu export phòng thi cũ."""
    blocks = build_room_diagram_blocks(room, elective_layout_id=elective_layout_id)
    block_by_key = {b['subject_key']: b for b in blocks}
    combo_label = _sort2_combo_label([_Sort2ExportAssignment(a) for a in assignments])

    if not (ws_src_style and ws_src_values):
        return

    block_starts = _get_block_starts_from_template(ws_src_style)
    if len(block_starts) >= 2:
        block_w = block_starts[1] - block_starts[0]
    else:
        block_w = 7

    block_seat_cells: dict[int, list[tuple[int, int]]] = {}
    for start in block_starts:
        seats = _detect_seat_cells_from_template(ws_src_values, start, block_w)
        if not seats:
            seats = _fallback_seat_cells_by_border(ws2, start, block_w)
        block_seat_cells[start] = seats

    for subj_idx, subject_key in enumerate(_DIAGRAM_SUBJECT_KEYS):
        if subj_idx >= len(block_starts):
            break
        block = block_by_key.get(subject_key)
        if not block:
            continue

        base_c = block_starts[subj_idx]
        base_r = 1
        seat_cells = block_seat_cells.get(base_c, [])
        ordered_cells = _detect_seat_cells_for_export(ws2, base_c, block_w) or seat_cells

        title = block.get('title', subject_key)
        subject_label = title.split('—')[0].strip().upper()
        if subject_key == 'van':
            subject_label = 'VĂN'
        elif subject_key == 'toan':
            subject_label = 'TOÁN'
        elif subject_key.startswith('elective'):
            subject_label = 'MÔN TỰ CHỌN'

        _set_cell_value(ws2, base_r, base_c, combo_label)
        _set_cell_value(ws2, base_r + 2, base_c, subject_label)
        _set_cell_value(ws2, base_r + 2, base_c + 4, f'P. {room.name}')

        for r, c in ordered_cells:
            _set_cell_value(ws2, r, c, '')

        values = _sbd_values_in_diagram_fill_order(
            room, block, assignments, len(ordered_cells),
        )
        for idx, (r, c) in enumerate(ordered_cells):
            if idx >= len(values):
                break
            val = values[idx]
            if isinstance(val, float) and val == int(val):
                val = int(val)
            _set_cell_value(ws2, r, c, val or '')

        if ordered_cells:
            grid_start = min(r for r, _ in ordered_cells)
            grid_end = max(r for r, _ in ordered_cells)
            seat_set = set(ordered_cells)
            for rr in range(grid_start, grid_end + 1):
                for cc in range(base_c, base_c + block_w):
                    if (rr, cc) in seat_set:
                        continue
                    cell = ws2.cell(row=rr, column=cc)
                    if cell.value is not None and str(cell.value).strip().isdigit():
                        _set_cell_value(ws2, rr, cc, '')


def _elective_layout_title(layout_id: int) -> str:
    lay = ELECTIVE_BOC_THAM_LAYOUTS.get(layout_id, ELECTIVE_BOC_THAM_LAYOUTS[0])
    if layout_id in (1, 2, 5):
        return f'Phương án sơ đồ {layout_id}'
    return lay.get('label', 'Thuận phát đề')


def _write_elective_formatted_sheet(
    ws,
    room: ExamSort2Room,
    venue: ExamSort2Venue,
    block: dict,
    *,
    elective_slot: int,
    layout_id: int,
) -> None:
    """
    Sơ đồ môn tự chọn dạng in mẫu: mỗi ghế có STT | SBD | Môn | Đề STT; điền theo cột.
    """
    cols = room.col_count
    rows = room.row_count
    subcols = 4
    total_cols = cols * subcols
    layout = block.get('layout') or {}
    start_stt = layout.get('start_stt', 1)

    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c1 = ws.cell(row=1, column=1, value='Bảng')
    c1.font = _FONT_BOARD
    c1.alignment = _ALIGN_CENTER

    ws.cell(row=2, column=1, value='Bàn GV').font = _FONT_HDR
    ws.cell(row=2, column=1).alignment = _ALIGN_CENTER
    if total_cols >= 4:
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=min(4, total_cols))
    ws.cell(row=2, column=2, value=f'P. {room.name} — {venue.code}').alignment = _ALIGN_CENTER

    title = _elective_layout_title(layout_id)
    slot_label = f'Môn tự chọn {elective_slot}'
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    tcell = ws.cell(
        row=3, column=1,
        value=f'{title} — {slot_label} (ô đầu STT {start_stt})',
    )
    tcell.font = _FONT_TITLE
    tcell.alignment = _ALIGN_CENTER

    hdr_row = 4
    for c_idx in range(cols):
        base = 1 + c_idx * subcols
        for off, label in enumerate(('STT', 'SBD', 'Môn', 'STT đề')):
            cell = ws.cell(row=hdr_row, column=base + off, value=label)
            cell.font = _FONT_HDR
            cell.alignment = _ALIGN_CENTER
            cell.border = _CELL_BORDER

    grid = block.get('grid') or []
    for r_idx in range(rows):
        data_row = hdr_row + 1 + r_idx
        ws.row_dimensions[data_row].height = 32
        for c_idx in range(cols):
            base = 1 + c_idx * subcols
            cell_data = None
            if r_idx < len(grid) and c_idx < len(grid[r_idx]):
                cell_data = grid[r_idx][c_idx]
            values = ('—', '—', '—', '—')
            if cell_data:
                values = (
                    f'{int(cell_data["room_stt"]):02d}',
                    cell_data.get('sbd') or '—',
                    cell_data.get('subject_code') or '—',
                    f'Đề STT {int(cell_data["de_stt"]):02d}',
                )
            for off, val in enumerate(values):
                cell = ws.cell(row=data_row, column=base + off, value=val)
                cell.alignment = _ALIGN_CENTER
                cell.border = _CELL_BORDER
                cell.font = Font(size=9)


def _append_elective_formatted_sheets(
    wb_out: Workbook,
    room: ExamSort2Room,
    venue: ExamSort2Venue,
    assignments: list[ExamSort2SeatAssignment],
    safe_code: str,
    *,
    layout_ids: list[int],
) -> None:
    """Thêm sheet TC1/TC2 (và có thể nhiều phương án PA) theo mẫu in."""
    for layout_id in layout_ids:
        blocks = build_room_diagram_blocks(room, elective_layout_id=layout_id)
        by_slot = {b['elective_slot']: b for b in blocks if b.get('elective_slot')}
        for slot in (1, 2):
            block = by_slot.get(slot)
            if not block:
                continue
            safe_room = re.sub(r'[^\w\-]', '_', (room.name or str(room.pk)).strip())[:8]
            pa_suffix = f'PA{layout_id}' if layout_id in (1, 2, 5) else 'TC'
            base = f'MTC{slot}_{safe_code}_{safe_room}_{pa_suffix}'
            sn = _exam_bulk_unique_sheet_name(wb_out, base[:28])
            ws = wb_out.create_sheet(sn)
            _write_elective_formatted_sheet(
                ws, room, venue, block,
                elective_slot=slot, layout_id=layout_id,
            )


def _build_sort2_room_diagram_workbook(
    room: ExamSort2Room,
    assignments: list[ExamSort2SeatAssignment],
    *,
    elective_layout_id: int = 0,
) -> Workbook:
    """Workbook một sheet sơ đồ (4 môn) cho một phòng."""
    tpl = _exam_export_workbook_paths()
    path_export = tpl['export_template']
    path_so_do = tpl['so_do']
    if not path_export.exists() and not path_so_do.exists():
        raise FileNotFoundError(
            'Không tìm thấy mẫu export trong templates_excel/ '
            '(EXPORT_TEMPLATE.xlsx hoặc SƠ ĐỒ CHỖ NGỒI PHÒNG THI HK I.xlsx).',
        )

    wb = Workbook()
    wb.remove(wb.active)

    ws_src_style = None
    ws_src_values = None
    if path_export.exists():
        tpl_wb = load_workbook(path_export, data_only=False)
        try:
            for name in ('4_Sơ đồ mẫu', '4_So do mau'):
                if name in tpl_wb.sheetnames:
                    ws_src_style = tpl_wb[name]
                    ws_src_values = tpl_wb[name]
                    break
        finally:
            tpl_wb.close()
    if ws_src_style is None and path_so_do.exists():
        wb_so = load_workbook(path_so_do, data_only=False)
        wb_vals = load_workbook(path_so_do, data_only=True)
        try:
            for name in ('4_Sơ đồ mẫu', '4_So do mau'):
                if name in wb_so.sheetnames:
                    ws_src_style = wb_so[name]
                    ws_src_values = (
                        wb_vals[name] if name in wb_vals.sheetnames else wb_vals.active
                    )
                    break
        finally:
            wb_so.close()
            wb_vals.close()

    if ws_src_style:
        ws2 = _copy_worksheet_to_workbook(ws_src_style, wb, sheet_name='4_Sơ đồ mẫu')
    else:
        ws2 = wb.create_sheet('4_Sơ đồ mẫu')

    _fill_sort2_diagram_sheet(
        ws2,
        room,
        assignments,
        elective_layout_id=elective_layout_id,
        ws_src_style=ws_src_style,
        ws_src_values=ws_src_values or ws_src_style,
    )
    return wb


def build_venue_export_workbook(
    venue: ExamSort2Venue,
    *,
    elective_layout_id: int = 0,
    elective_all_pa: bool = False,
) -> BytesIO:
    """
    Một điểm thi → một file Excel:
    - Sheet danh sách gộp (mẫu In_GIAO NHAN BAI)
    - Mỗi phòng: sheet sơ đồ 4 môn (SBD) + 2 sheet môn tự chọn (mẫu STT/SBD/Môn/Đề)
    - elective_all_pa: thêm sheet TC theo PA 1, 2, 5 (như mẫu in)
    """
    if elective_layout_id not in ELECTIVE_BOC_THAM_LAYOUTS:
        elective_layout_id = 0

    blocks = _venue_export_blocks(venue)
    if not blocks:
        raise ValueError('Điểm thi chưa có phòng nào có thí sinh được xếp ghế.')

    tpl = _exam_export_workbook_paths()
    if not tpl['export_template'].exists() and not tpl['ds_thi_tap_trung'].exists():
        raise FileNotFoundError(
            'Không tìm thấy mẫu EXPORT_TEMPLATE.xlsx trong thư mục templates_excel/.',
        )

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    safe_code = re.sub(r'[^\w\-]', '', (venue.code or '') or 'DT')[:12] or 'DT'
    list_sheet = _exam_merged_list_prepare_sheet(
        wb_out,
        blocks,
        'sang',
        f'DS_{safe_code}',
    )
    hdr_row = _exam_sheet1_guess_stt_header_row(list_sheet)
    use_th = _exam_sheet1_template_has_th_column(list_sheet, hdr_row)
    all_adapters = [a for _, arr in blocks for a in arr]
    _fill_exam_giao_nhan_merged(
        list_sheet,
        blocks,
        'sang',
        global_stt=False,
        include_th_column=use_th,
        combo_label_override=_sort2_combo_label(all_adapters),
        subjects_ws1_override=_SORT2_SUBJECTS_WS1,
        phong_thi_header_override=f'ĐIỂM THI: {venue.code} — {venue.name}',
        th_label_fn=_sort2_th_label,
    )

    for room, adapters in blocks:
        raw_assigns = [a._assignment for a in adapters]
        wb_part = _build_sort2_room_diagram_workbook(
            room,
            raw_assigns,
            elective_layout_id=elective_layout_id,
        )
        for title in wb_part.sheetnames:
            tit_up = (title or '').upper()
            if 'GIAO' in tit_up or (title or '').startswith('In_'):
                continue
            ws_src = wb_part[title]
            safe_room = re.sub(r'[^\w\-]', '_', (room.name or str(room.pk)).strip())[:12]
            sn = _exam_bulk_unique_sheet_name(wb_out, f'SD_{safe_code}_{safe_room}')
            _copy_worksheet_to_workbook(ws_src, wb_out, sheet_name=sn)
        wb_part.close()

        tc_layouts = [1, 2, 5] if elective_all_pa else [elective_layout_id]
        _append_elective_formatted_sheets(
            wb_out,
            room,
            venue,
            raw_assigns,
            safe_code,
            layout_ids=tc_layouts,
        )

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf
